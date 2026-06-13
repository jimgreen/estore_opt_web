#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Route 13: perspective I2R electro-thermal BESS MILP.

This route is a separate working copy derived from route 12. It keeps the
fuel-only day-ahead dispatch assumptions, but replaces the compact 3D
SOC/T/I PWL pack mapping with a perspective-style I^2R formulation:
SOC-T weights are retained, while current and squared-current loss are
represented by split weights over the SOC-T grid.

Useful test entry:
    python estore_opt_web/solve.py --mode test_1h --time-limit 120 --mip-gap 0.1
"""

from __future__ import annotations

import argparse
import copy
import json
import math
import os
import sys
import time
import warnings
from dataclasses import dataclass
from pathlib import Path
from types import SimpleNamespace
import importlib.util

import numpy as np

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager

warnings.filterwarnings("ignore")


SUFFIX = "perspective_i2r_20260530"
WEB_ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = WEB_ROOT.parent
ROOT = PROJECT_ROOT
DEFAULT_EXCEL_PATH = PROJECT_ROOT / "5.4模型定义_低温风光扩容_更多冷却液_20260512.xlsx"
DEFAULT_PARAM_PATH = DEFAULT_EXCEL_PATH if DEFAULT_EXCEL_PATH.exists() else PROJECT_ROOT / f"model_params_{SUFFIX}.json"
DEFAULT_TIGHT_TEMP_BOUNDS = {
    "T_bat_min": 0.0,
    "T_bat_max": 45.0,
    "T_tank_min": -20.0,
    "T_tank_max": 45.0,
    "T_cont_min": -25.0,
    "T_cont_max": 45.0,
}


def setup_chinese_font() -> None:
    candidates = [
        "Microsoft YaHei",
        "SimHei",
        "SimSun",
        "KaiTi",
        "FangSong",
        "Noto Sans CJK SC",
        "Source Han Sans SC",
    ]
    available = {f.name for f in font_manager.fontManager.ttflist}
    for name in candidates:
        if name in available:
            plt.rcParams["font.sans-serif"] = [name]
            plt.rcParams["axes.unicode_minus"] = False
            return
    plt.rcParams["axes.unicode_minus"] = False


setup_chinese_font()


def now_iso() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


def make_json_safe(value):
    if isinstance(value, dict):
        return {str(k): make_json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [make_json_safe(item) for item in value]
    if isinstance(value, np.ndarray):
        return make_json_safe(value.tolist())
    if isinstance(value, np.generic):
        return make_json_safe(value.item())
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, SimpleNamespace):
        return make_json_safe(vars(value))
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, (str, int, bool)) or value is None:
        return value
    return str(value)


def write_progress(path: Path | None, **updates) -> None:
    if path is None:
        return
    target = Path(path)
    payload = {}
    if target.exists():
        try:
            payload = json.loads(target.read_text(encoding="utf-8"))
            if not isinstance(payload, dict):
                payload = {}
        except Exception:
            payload = {}
    payload.update(make_json_safe(updates))
    payload["updated_at"] = now_iso()
    target.parent.mkdir(parents=True, exist_ok=True)
    tmp = target.with_name(f"{target.name}.tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(target)


def compact_result_for_progress(result: dict | None) -> dict:
    if not result:
        return {}
    checks = result.get("checks") or {}
    stats = result.get("model_stats") or {}
    return make_json_safe(
        {
            "success": bool(result.get("success", False)),
            "status": result.get("status"),
            "message": result.get("message", ""),
            "objective": result.get("objective"),
            "best_bound": result.get("best_bound"),
            "gap": result.get("gap"),
            "time_s": result.get("time_s"),
            "node_count": result.get("node_count"),
            "fuel_kg": result.get("fuel_kg"),
            "curt_kwh": result.get("curt_kwh"),
            "heat_kwh": result.get("heat_kwh"),
            "solver_used": result.get("solver_used"),
            "solver_backend": result.get("solver_backend"),
            "state_workbook": result.get("state_workbook"),
            "model_stats": {
                "solver_name": stats.get("solver_name"),
                "solver_backend": stats.get("solver_backend"),
                "variables_total": stats.get("variables_total"),
                "binary_variables": stats.get("binary_variables"),
                "constraints_total": stats.get("constraints_total"),
                "sos_constraints": stats.get("sos_constraints"),
                "nonzeros": stats.get("nonzeros"),
                "steps": stats.get("steps"),
                "dt_minutes": stats.get("dt_minutes"),
            },
            "checks": {
                "soc_min": checks.get("soc_min"),
                "soc_max": checks.get("soc_max"),
                "tbat_min_c": checks.get("tbat_min_c"),
                "tbat_max_c": checks.get("tbat_max_c"),
                "model_balance_max_kw": checks.get("model_balance_max_kw"),
                "pbess_physical_max_kw": checks.get("pbess_physical_max_kw"),
            },
        }
    )


@dataclass
class Breakpoints:
    soc: np.ndarray
    ocv: np.ndarray
    temp: np.ndarray
    r0: np.ndarray
    current: np.ndarray
    current2: np.ndarray
    charge_current_limit: np.ndarray
    discharge_current_limit: np.ndarray
    dt_hours: float
    mode: str

    @property
    def n_s(self) -> int:
        return len(self.soc)

    @property
    def n_t(self) -> int:
        return len(self.temp)

    @property
    def n_i(self) -> int:
        return len(self.current)


def load_params(path: Path) -> SimpleNamespace:
    path = Path(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        raw = load_raw_params_from_excel(path)
    else:
        raw = json.loads(path.read_text(encoding="utf-8"))

    r0_raw = raw["lookup_tables"]["r0"]
    ocv_raw = raw["lookup_tables"]["ocv"]
    soc_pts = np.asarray(r0_raw["soc"], dtype=float)
    temp_pts = np.asarray(r0_raw["temperatures_c"], dtype=float)
    r0_cell_table = np.asarray(r0_raw["values"], dtype=float)
    ocv_soc = np.asarray(ocv_raw["soc"], dtype=float)
    ocv_cell_1d = np.asarray(ocv_raw["values_1d"], dtype=float)

    cabinet = raw["battery_cabinet"]
    cell = raw["cell"]
    liquid = raw["liquid_cooling"]
    container = raw["container"]
    distribution = raw["distribution"]

    p = SimpleNamespace()
    p.raw = raw
    p.param_path = str(path)
    p.source_excel = raw.get("source_excel", "")

    p.g_std = float(raw["system"]["g_std_w_m2"])
    p.base_dt_s = float(raw["system"]["base_dt_s"])
    p.start_min = float(raw["system"]["start_min"])
    p.end_min = float(raw["system"]["end_min"])

    p.N_s = int(cabinet["n_s"])
    p.N_p = int(cabinet["n_p"])
    p.Q_cell = float(cell["q_nom_ah"])
    pack = raw.get("battery_pack", {})
    p.Q_nom = float(pack.get("q_pack_ah", p.N_p * p.Q_cell))
    p.SOC_min = float(cell["soc_min"])
    p.SOC_max = float(cell["soc_max"])
    pack_i_max = float(pack.get("i_pack_max_a", cabinet["pcs_i_max_a"]))
    pcs_i_max = float(cabinet["pcs_i_max_a"])
    cell_current_limits = raw.get("cell_current_limits") or {}
    if all(
        key in pack
        for key in (
            "current_limit_temperatures_c",
            "charge_current_limit_a",
            "discharge_current_limit_a",
        )
    ):
        limit_temps = np.asarray(pack["current_limit_temperatures_c"], dtype=float)
        charge_limit_pack = np.asarray(pack["charge_current_limit_a"], dtype=float)
        discharge_limit_pack = np.asarray(pack["discharge_current_limit_a"], dtype=float)
    elif all(key in cell_current_limits for key in ("temperatures_c", "charge_max_a", "discharge_max_a")):
        limit_temps = np.asarray(cell_current_limits["temperatures_c"], dtype=float)
        charge_limit_pack = np.asarray(cell_current_limits["charge_max_a"], dtype=float) * p.N_p
        discharge_limit_pack = np.asarray(cell_current_limits["discharge_max_a"], dtype=float) * p.N_p
    else:
        limit_temps = np.asarray(temp_pts, dtype=float)
        charge_limit_pack = np.full(limit_temps.shape, float(cell.get("i_charge_max_a", pack_i_max / max(1, p.N_p))) * p.N_p)
        discharge_limit_pack = np.full(limit_temps.shape, float(cell.get("i_discharge_max_a", pack_i_max / max(1, p.N_p))) * p.N_p)
    if not (len(limit_temps) and len(limit_temps) == len(charge_limit_pack) == len(discharge_limit_pack)):
        raise ValueError("Cell current-limit table must have matching temperature, charge, and discharge columns.")
    order = np.argsort(limit_temps)
    limit_temps = limit_temps[order]
    charge_limit_pack = charge_limit_pack[order]
    discharge_limit_pack = discharge_limit_pack[order]
    if len(limit_temps) > 1 and np.any(np.diff(limit_temps) <= 0):
        raise ValueError("Cell current-limit temperatures must be unique and strictly increasing.")
    if not (
        np.all(np.isfinite(limit_temps))
        and np.all(np.isfinite(charge_limit_pack))
        and np.all(np.isfinite(discharge_limit_pack))
    ):
        raise ValueError("Cell current-limit table contains non-finite values.")
    if np.any(charge_limit_pack <= 0.0) or np.any(discharge_limit_pack <= 0.0):
        raise ValueError("Cell current-limit values must be positive.")
    p.current_limit_temps = limit_temps
    p.charge_current_limit_pack = charge_limit_pack
    p.discharge_current_limit_pack = discharge_limit_pack
    p.pack_current_hardware_max = min(pack_i_max, pcs_i_max)
    p.charge_current_limit_effective_pack = np.minimum(charge_limit_pack, p.pack_current_hardware_max)
    p.discharge_current_limit_effective_pack = np.minimum(discharge_limit_pack, p.pack_current_hardware_max)
    p.I_charge_max = float(np.max(p.charge_current_limit_effective_pack))
    p.I_discharge_max = float(np.max(p.discharge_current_limit_effective_pack))
    p.I_abs_max = max(p.I_discharge_max, p.I_charge_max)

    p.T_bat_min = float(cell["temp_min_c"])
    p.T_bat_max = float(cell["temp_max_c"])
    p.T_tank_min = float(liquid["temp_min_c"])
    p.T_tank_max = float(liquid["temp_max_c"])
    p.T_cont_min = float(container["temp_min_c"])
    p.T_cont_max = float(container["temp_max_c"])

    p.C_bat = float(cabinet["c_th_kj_per_k"]) * 1000.0
    p.C_tank = float(liquid["c_th_kj_per_k"]) * 1000.0
    p.C_cont = float(container["c_th_kj_per_k"]) * 1000.0
    p.K_bt = float(cabinet["k_b_tank_kw_per_k"]) * 1000.0
    p.K_b_cont = float(cabinet["k_b_container_kw_per_k"]) * 1000.0
    p.K_t_cont = float(liquid["k_t_container_kw_per_k"]) * 1000.0
    p.K_t_amb = float(liquid["k_t_amb_kw_per_k"]) * 1000.0
    p.K_cont_amb = float(container["k_amb_kw_per_k"]) * 1000.0
    p.m_in = float(liquid.get("m_in_kg_s", 0.0))
    p.m_out = float(liquid.get("m_out_kg_s", 0.0))
    p.c_liq = float(liquid.get("c_liq_kj_kg_k", 0.0)) * 1000.0

    p.P_pump_in = float(distribution["pump_in_kw"]) * 1000.0
    p.P_pump_out = float(distribution["pump_out_kw"]) * 1000.0
    p.P_heat_liquid = float(distribution["heat_liquid_kw"]) * 1000.0
    p.P_heat_cont = float(distribution["heat_container_kw"]) * 1000.0
    p.mu_pcs = float(cabinet["pcs_loss_rate"])
    p.pcs_v_min = float(cabinet["pcs_v_min_v"])
    p.pcs_v_max = float(cabinet["pcs_v_max_v"])
    p.pcs_i_max = float(cabinet["pcs_i_max_a"])

    p.soc_pts = soc_pts
    p.temp_pts = temp_pts
    p.r0_cell_table = r0_cell_table
    p.r0_table = np.asarray(pack.get("r_pack_ohm", (p.N_s / p.N_p * r0_cell_table).tolist()), dtype=float)
    p.ocv_soc = ocv_soc
    p.ocv_cell_1d = ocv_cell_1d
    p.ocv_1d = np.asarray(pack.get("u_ocv_pack_v", (p.N_s * ocv_cell_1d).tolist()), dtype=float)
    p.OCV_min = float(np.min(p.ocv_1d))
    p.OCV_max = float(np.max(p.ocv_1d))
    p.R0_max = float(np.max(p.r0_table))
    p.pack_conversion_note = pack.get(
        "conversion_note",
        "Pack model uses U_pack=N_s*OCV_cell, R_pack=(N_s/N_p)*R_cell, Q_pack=N_p*Q_cell, I_pack=PCS DC current.",
    )

    p.diesel_units = [
        {
            "id": int(u["id"]),
            "rated_w": float(u["rated_kw"]) * 1000.0,
            "p_min_w": float(u["p_min_kw"]) * 1000.0,
            "p_max_w": float(u["p_max_kw"]) * 1000.0,
            "powers_w": np.asarray(u["powers_kw"], dtype=float) * 1000.0,
            "fuel_kg_h": np.asarray(u["fuel_kg_per_h"], dtype=float),
            "fuel_rate_kg_per_kwh": np.asarray(u["fuel_rate_kg_per_kwh"], dtype=float),
        }
        for u in raw["diesel_generators"]
    ]
    p.wind_units = raw["wind_turbines"]
    p.pv_units = raw["pv_units"]
    p.profiles = raw["profiles"]

    thermal_storage = raw.get("thermal_storage", {})
    p.T_tank_target = float(thermal_storage.get("tank_target_c", 35.0))
    p.T_cont_target = float(thermal_storage.get("cont_target_c", 10.0))
    p.T_bat_pref_low = float(thermal_storage.get("bat_pref_low_c", 15.0))
    p.T_bat_pref_high = float(thermal_storage.get("bat_pref_high_c", 35.0))
    p.T_tank_band_low = float(thermal_storage.get("tank_band_low_c", 30.0))
    p.T_tank_band_high = float(thermal_storage.get("tank_band_high_c", 40.0))
    p.T_tank_hot = float(thermal_storage.get("tank_hot_c", 45.0))
    p.T_cont_band_low = float(thermal_storage.get("cont_band_low_c", 5.0))
    p.T_cont_band_high = float(thermal_storage.get("cont_band_high_c", 15.0))
    p.T_cont_hot = float(thermal_storage.get("cont_hot_c", 25.0))
    p.T_tank_terminal_min = float(thermal_storage.get("tank_terminal_min_c", 30.0))
    p.T_cont_terminal_min = float(thermal_storage.get("cont_terminal_min_c", 5.0))
    p.T_tank_useful_min = float(thermal_storage.get("tank_useful_min_c", p.T_tank_band_low))
    p.T_cont_useful_min = float(thermal_storage.get("cont_useful_min_c", p.T_cont_band_low))
    p.thermal_storage_strategy = thermal_storage.get(
        "strategy",
        "thermal-value MILP with comfort bands, terminal heat reserve, and heat dumping penalty",
    )

    initial_state = raw.get("initial_state", {})
    p.SOC_init = float(initial_state.get("soc", 0.5))
    p.T_bat_init = float(initial_state.get("t_bat_c", 5.0))
    p.T_tank_init = float(initial_state.get("t_tank_c", 5.0))
    p.T_cont_init = float(initial_state.get("t_cont_c", 5.0))

    p.default_time_limit = 7200
    p.default_mip_gap = 0.02
    return p


def _sheet_row(ws, row: int = 2) -> list:
    max_col = 0
    for cells in ws.iter_rows():
        for cell in cells:
            if cell.value is not None:
                max_col = max(max_col, cell.column)
    return [ws.cell(row, col).value for col in range(1, max_col + 1)]


def _float(value, default: float | None = None) -> float:
    if value is None or value == "":
        if default is None:
            raise ValueError("Missing numeric value in Excel input.")
        return float(default)
    if isinstance(value, str):
        text = value.strip().replace("A", "").replace("a", "")
        return float(text)
    return float(value)


def _hour_value(value) -> float:
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return float(value.hour) + float(value.minute) / 60.0 + float(value.second) / 3600.0
    if isinstance(value, str):
        parts = value.strip().split(":")
        if len(parts) >= 2:
            return float(parts[0]) + float(parts[1]) / 60.0 + (float(parts[2]) / 3600.0 if len(parts) > 2 else 0.0)
        return float(value)
    return float(value)


def _temperature_from_header(value) -> float:
    text = str(value).replace("℃", "").replace("°C", "").replace("C", "").strip()
    return float(text)


def _table_extent(ws) -> tuple[int, int]:
    max_row = 0
    max_col = 0
    for row in ws.iter_rows():
        if any(cell.value is not None for cell in row):
            max_row = max(max_row, row[0].row)
            for cell in row:
                if cell.value is not None:
                    max_col = max(max_col, cell.column)
    return max_row, max_col


def _lookup_table(ws) -> tuple[list[float], list[float], list[list[float]]]:
    max_row, max_col = _table_extent(ws)
    temperatures = [_temperature_from_header(ws.cell(1, col).value) for col in range(2, max_col + 1)]
    soc = []
    values = []
    for row in range(2, max_row + 1):
        soc_value = ws.cell(row, 1).value
        if soc_value is None:
            continue
        soc.append(float(soc_value))
        values.append([_float(ws.cell(row, col).value) for col in range(2, max_col + 1)])
    return soc, temperatures, values


def _current_limit_table(ws) -> dict[str, list[float]]:
    max_row, _ = _table_extent(ws)
    rows: list[tuple[float, float, float]] = []
    for row in range(2, max_row + 1):
        temp_value = ws.cell(row, 1).value
        if temp_value is None or temp_value == "":
            continue
        rows.append(
            (
                _float(temp_value),
                _float(ws.cell(row, 2).value),
                _float(ws.cell(row, 3).value),
            )
        )
    if not rows:
        raise ValueError("Cell current-limit table must contain at least one data row.")
    rows.sort(key=lambda item: item[0])
    temperatures, charge_max, discharge_max = zip(*rows)
    return {
        "temperatures_c": [float(value) for value in temperatures],
        "charge_max_a": [float(value) for value in charge_max],
        "discharge_max_a": [float(value) for value in discharge_max],
    }


def load_raw_params_from_excel(path: Path) -> dict:
    """Convert the model-definition workbook into the raw JSON schema used by load_params."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("Reading Excel parameters requires openpyxl. Install openpyxl or pass a JSON parameter file.") from exc

    wb = openpyxl.load_workbook(path, data_only=True)
    required = [
        "系统定义",
        "配电模块",
        "电池柜",
        "电芯",
        "电芯内阻",
        "电芯OCV",
        "液冷系统",
        "舱体",
        "柴油发电机",
        "风电机组",
        "光伏机组",
        "运行曲线",
    ]
    missing = [name for name in required if name not in wb.sheetnames]
    if missing:
        raise ValueError(f"Excel parameter workbook is missing sheets: {', '.join(missing)}")

    system = _sheet_row(wb["系统定义"])
    distribution = _sheet_row(wb["配电模块"])
    cabinet = _sheet_row(wb["电池柜"])
    cell = _sheet_row(wb["电芯"])
    liquid = _sheet_row(wb["液冷系统"])
    container = _sheet_row(wb["舱体"])

    soc, r0_temps, r0_values = _lookup_table(wb["电芯内阻"])
    ocv_soc, ocv_temps, ocv_values = _lookup_table(wb["电芯OCV"])
    ocv_values_1d = [row[0] for row in ocv_values]

    n_s = int(_float(cabinet[2]))
    n_p = int(_float(cabinet[3]))
    q_cell = 40.0
    q_pack = n_p * q_cell
    pcs_i_max = _float(cabinet[10])
    current_limit_sheet = "电芯电流限值"
    if current_limit_sheet in wb.sheetnames:
        cell_current_limits = _current_limit_table(wb[current_limit_sheet])
    else:
        cell_current_limits = {
            "temperatures_c": [float(value) for value in r0_temps],
            "charge_max_a": [_float(cell[7]) for _ in r0_temps],
            "discharge_max_a": [_float(cell[6]) for _ in r0_temps],
        }
    charge_current_limit_pack = [n_p * float(value) for value in cell_current_limits["charge_max_a"]]
    discharge_current_limit_pack = [n_p * float(value) for value in cell_current_limits["discharge_max_a"]]
    r_pack = [[n_s / n_p * float(v) for v in row] for row in r0_values]
    u_pack = [n_s * float(v) for v in ocv_values_1d]

    diesel_units = []
    ws_dg = wb["柴油发电机"]
    dg_max_row, _ = _table_extent(ws_dg)
    for row in range(2, dg_max_row + 1):
        if ws_dg.cell(row, 1).value is None:
            continue
        point_count = int(_float(ws_dg.cell(row, 5).value))
        powers = []
        rates = []
        for idx in range(point_count):
            p_col = 6 + 2 * idx
            r_col = 7 + 2 * idx
            powers.append(_float(ws_dg.cell(row, p_col).value))
            rates.append(_float(ws_dg.cell(row, r_col).value))
        diesel_units.append(
            {
                "id": int(_float(ws_dg.cell(row, 1).value)),
                "rated_kw": _float(ws_dg.cell(row, 2).value),
                "p_min_kw": _float(ws_dg.cell(row, 3).value),
                "p_max_kw": _float(ws_dg.cell(row, 4).value),
                "powers_kw": powers,
                "fuel_rate_kg_per_kwh": rates,
                "fuel_kg_per_h": [pwr * rate for pwr, rate in zip(powers, rates)],
            }
        )

    wind_units = []
    ws_wind = wb["风电机组"]
    wind_max_row, _ = _table_extent(ws_wind)
    for row in range(2, wind_max_row + 1):
        if ws_wind.cell(row, 1).value is None:
            continue
        wind_units.append(
            {
                "id": int(_float(ws_wind.cell(row, 1).value)),
                "rated_kw": _float(ws_wind.cell(row, 2).value),
                "cut_in_m_s": _float(ws_wind.cell(row, 3).value),
                "rated_m_s": _float(ws_wind.cell(row, 4).value),
                "cut_out_m_s": _float(ws_wind.cell(row, 5).value),
            }
        )

    pv_units = []
    ws_pv = wb["光伏机组"]
    pv_max_row, _ = _table_extent(ws_pv)
    for row in range(2, pv_max_row + 1):
        if ws_pv.cell(row, 1).value is None:
            continue
        pv_units.append(
            {
                "id": int(_float(ws_pv.cell(row, 1).value)),
                "rated_kw": _float(ws_pv.cell(row, 2).value),
                "p_max_kw": _float(ws_pv.cell(row, 3).value),
            }
        )

    profiles = []
    ws_profile = wb["运行曲线"]
    profile_max_row, _ = _table_extent(ws_profile)
    for row in range(2, profile_max_row + 1):
        if ws_profile.cell(row, 1).value is None:
            continue
        profiles.append(
            {
                "hour": _hour_value(ws_profile.cell(row, 1).value),
                "wind_m_s": _float(ws_profile.cell(row, 2).value),
                "ambient_c": _float(ws_profile.cell(row, 3).value),
                "solar_w_m2": _float(ws_profile.cell(row, 4).value),
                "load_kw": _float(ws_profile.cell(row, 5).value),
            }
        )

    return {
        "source_excel": str(path),
        "generated_on": "loaded_from_excel",
        "variant": SUFFIX,
        "units_note": "Power input, result output, and stored result curves use kW.",
        "system": {
            "g_std_w_m2": _float(system[1]),
            "base_dt_s": _float(system[2]),
            "start_min": _float(system[3]),
            "end_min": _float(system[4]),
        },
        "distribution": {
            "pump_in_kw": _float(distribution[1]),
            "pump_out_kw": _float(distribution[2]),
            "heat_liquid_kw": _float(distribution[3]),
            "heat_container_kw": _float(distribution[4]),
        },
        "battery_cabinet": {
            "n_s": n_s,
            "n_p": n_p,
            "c_th_kj_per_k": _float(cabinet[4]),
            "k_b_tank_kw_per_k": _float(cabinet[5]),
            "k_b_container_kw_per_k": _float(cabinet[6]),
            "pcs_loss_rate": _float(cabinet[7]),
            "pcs_v_max_v": _float(cabinet[8]),
            "pcs_v_min_v": _float(cabinet[9]),
            "pcs_i_max_a": pcs_i_max,
        },
        "cell": {
            "q_nom_ah": q_cell,
            "soc_min": 0.1,
            "soc_max": 0.9,
            "temp_min_c": _float(cell[1]),
            "temp_max_c": _float(cell[2]),
            "ocv_min_v": _float(cell[3]),
            "ocv_max_v": _float(cell[4]),
            "i_discharge_max_a": _float(cell[6]),
            "i_charge_max_a": _float(cell[7]),
            "capacity_source_note": "Excel workbook does not expose cell Ah explicitly; 40 Ah/cell is kept from the existing pack model.",
        },
        "cell_current_limits": cell_current_limits,
        "liquid_cooling": {
            "c_th_kj_per_k": _float(liquid[1]),
            "k_t_container_kw_per_k": _float(liquid[2]),
            "k_t_amb_kw_per_k": _float(liquid[3]),
            "m_in_kg_s": _float(liquid[4]),
            "m_out_kg_s": _float(liquid[5]),
            "c_liq_kj_kg_k": _float(liquid[6]),
            "temp_max_c": _float(liquid[7]),
            "temp_min_c": _float(liquid[8]),
        },
        "container": {
            "c_th_kj_per_k": _float(container[1]),
            "k_amb_kw_per_k": _float(container[2]),
            "temp_max_c": _float(container[3]),
            "temp_min_c": _float(container[4]),
        },
        "diesel_generators": diesel_units,
        "wind_turbines": wind_units,
        "pv_units": pv_units,
        "lookup_tables": {
            "r0": {"soc": soc, "temperatures_c": r0_temps, "values": r0_values},
            "ocv": {"soc": ocv_soc, "temperatures_c": ocv_temps, "values": ocv_values, "values_1d": ocv_values_1d},
        },
        "battery_pack": {
            "q_pack_ah": q_pack,
            "i_pack_max_a": pcs_i_max,
            "current_limit_temperatures_c": cell_current_limits["temperatures_c"],
            "charge_current_limit_a": charge_current_limit_pack,
            "discharge_current_limit_a": discharge_current_limit_pack,
            "u_ocv_pack_v": u_pack,
            "r_pack_ohm": r_pack,
            "conversion_note": "Pack model uses U_pack=N_s*OCV_cell, R_pack=(N_s/N_p)*R_cell, Q_pack=N_p*Q_cell, I_pack=PCS DC current.",
        },
        "profiles": profiles,
        "thermal_storage": {
            "tank_target_c": 35.0,
            "cont_target_c": 10.0,
            "bat_pref_low_c": 15.0,
            "bat_pref_high_c": 35.0,
            "tank_band_low_c": 30.0,
            "tank_band_high_c": 40.0,
            "tank_hot_c": 45.0,
            "tank_terminal_min_c": 30.0,
            "tank_useful_min_c": 30.0,
            "cont_band_low_c": 5.0,
            "cont_band_high_c": 15.0,
            "cont_hot_c": 25.0,
            "cont_terminal_min_c": 5.0,
            "cont_useful_min_c": 5.0,
            "strategy": "thermal-value MILP with comfort bands, terminal heat reserve, and heat dumping penalty",
        },
        "initial_state": {"soc": 0.5, "t_bat_c": 5.0, "t_tank_c": 5.0, "t_cont_c": 5.0},
    }


def extend_temperature_table(p: SimpleNamespace) -> tuple[np.ndarray, np.ndarray]:
    temp = p.temp_pts.copy()
    table = p.r0_table.copy()
    if p.T_bat_min < temp[0] - 1e-9:
        temp = np.insert(temp, 0, p.T_bat_min)
        table = np.insert(table, 0, table[:, 0], axis=1)
    if p.T_bat_max > temp[-1] + 1e-9:
        temp = np.append(temp, p.T_bat_max)
        table = np.column_stack([table, table[:, -1]])
    return temp, table


def _nearest_unique_grid_indices(source: np.ndarray, targets: np.ndarray) -> np.ndarray:
    indices: list[int] = []
    for target in targets:
        idx = int(np.argmin(np.abs(source - target)))
        if not indices or indices[-1] != idx:
            indices.append(idx)
    return np.asarray(indices, dtype=int)


def select_soc_grid_indices(p: SimpleNamespace, soc_grid_width: float) -> np.ndarray:
    if float(soc_grid_width) <= 0.0:
        raise ValueError("soc_grid_width must be positive.")
    soc_source = np.asarray(p.soc_pts, dtype=float)
    soc_min = max(float(p.SOC_min), float(np.min(soc_source)))
    soc_max = min(float(p.SOC_max), float(np.max(soc_source)))
    if soc_min > soc_max:
        raise ValueError("SOC operating bounds do not overlap SOC lookup table.")

    width = float(soc_grid_width)
    targets = list(np.arange(soc_min, soc_max + width * 0.5, width, dtype=float))
    if not targets:
        targets = [soc_min]
    if abs(targets[0] - soc_min) > 1e-9:
        targets.insert(0, soc_min)
    if abs(targets[-1] - soc_max) > 1e-9:
        targets.append(soc_max)
    targets_arr = np.asarray(targets, dtype=float)
    targets_arr = np.clip(targets_arr, soc_min, soc_max)
    indices = _nearest_unique_grid_indices(soc_source, targets_arr)
    for bound in [soc_min, soc_max]:
        idx = int(np.argmin(np.abs(soc_source - bound)))
        if idx not in indices:
            indices = np.append(indices, idx)
    indices = np.unique(indices)
    indices.sort()
    return indices


def select_soc_breakpoints(p: SimpleNamespace, soc_grid_width: float) -> np.ndarray:
    if float(soc_grid_width) <= 0.0:
        raise ValueError("soc_grid_width must be positive.")
    soc_source = np.asarray(p.soc_pts, dtype=float)
    if soc_source.ndim != 1 or soc_source.size == 0:
        raise ValueError("SOC lookup table must contain at least one point.")
    if soc_source.size > 1 and np.any(np.diff(soc_source) <= 0.0):
        raise ValueError("SOC lookup table points must be strictly increasing.")
    soc_min = max(float(p.SOC_min), float(np.min(soc_source)))
    soc_max = min(float(p.SOC_max), float(np.max(soc_source)))
    if soc_min > soc_max:
        raise ValueError("SOC operating bounds do not overlap SOC lookup table.")

    width = float(soc_grid_width)
    values = [soc_min]
    next_value = soc_min + width
    while next_value < soc_max - 1e-9:
        values.append(next_value)
        next_value += width
    if soc_max - values[-1] > 1e-9:
        values.append(soc_max)
    return np.asarray(values, dtype=float)


def interpolate_soc_table(soc_targets: np.ndarray, soc_source: np.ndarray, table: np.ndarray) -> np.ndarray:
    source = np.asarray(soc_source, dtype=float)
    values = np.asarray(table, dtype=float)
    if values.ndim != 2 or values.shape[0] != source.size:
        raise ValueError("SOC-dependent lookup table shape does not match SOC points.")
    return np.column_stack([np.interp(soc_targets, source, values[:, col]) for col in range(values.shape[1])])


def make_breakpoints(
    p: SimpleNamespace,
    mode: str,
    i_points: int | None = None,
    dt_minutes: float = 15.0,
    current_segments: int | None = None,
    soc_grid_width: float = 0.1,
) -> Breakpoints:
    if float(dt_minutes) <= 0:
        raise ValueError("dt_minutes must be positive.")
    dt_hours = float(dt_minutes) / 60.0
    configs = {
        "full_more_coolant_20260512": {"soc_grid_width": 0.1, "current_segments": 10},
        "dayahead_24h": {"soc_grid_width": 0.1, "current_segments": 10},
        "minute_more_coolant_20260512": {"soc_grid_width": 0.1, "current_segments": 10},
        "test_1h": {"soc_grid_width": 0.1, "current_segments": 10},
        "test_4h": {"soc_grid_width": 0.1, "current_segments": 10},
    }
    cfg = configs.get(mode, configs["full_more_coolant_20260512"])
    if current_segments is not None:
        if int(current_segments) <= 0:
            raise ValueError("current_segments must be a positive integer.")
        n_i = int(current_segments) + 1
    elif i_points is not None:
        if int(i_points) < 2:
            raise ValueError("i_points/current_points must be at least 2.")
        n_i = int(i_points)
    else:
        n_i = int(cfg["current_segments"]) + 1

    grid_width = float(soc_grid_width if soc_grid_width is not None else cfg["soc_grid_width"])
    soc_bp = select_soc_breakpoints(p, grid_width)
    ocv_bp = np.interp(soc_bp, p.ocv_soc, p.ocv_1d)
    temp_full, r0_full = extend_temperature_table(p)
    r0_bp = interpolate_soc_table(soc_bp, p.soc_pts, r0_full)
    charge_current_limit_bp = np.interp(
        temp_full,
        p.current_limit_temps,
        p.charge_current_limit_pack,
        left=p.charge_current_limit_pack[0],
        right=p.charge_current_limit_pack[-1],
    )
    discharge_current_limit_bp = np.interp(
        temp_full,
        p.current_limit_temps,
        p.discharge_current_limit_pack,
        left=p.discharge_current_limit_pack[0],
        right=p.discharge_current_limit_pack[-1],
    )

    if getattr(p, "diag_discrete_current", False):
        levels = int(getattr(p, "diag_discrete_current_levels", 21))
        c_max = float(getattr(p, "diag_discrete_current_c_max", 1.0))
        i_max = min(float(p.I_charge_max), float(p.I_discharge_max), c_max * float(p.Q_nom))
        current_bp = np.linspace(-i_max, i_max, levels)
        n_i = len(current_bp)
    else:
        current_bp = np.linspace(-p.I_charge_max, p.I_discharge_max, n_i)
    return Breakpoints(
        soc=soc_bp,
        ocv=ocv_bp,
        temp=temp_full,
        r0=r0_bp,
        current=current_bp,
        current2=current_bp**2,
        charge_current_limit=charge_current_limit_bp,
        discharge_current_limit=discharge_current_limit_bp,
        dt_hours=dt_hours,
        mode=mode,
    )


def interpolate_profiles(p: SimpleNamespace, dt_hours: float, horizon_hours: float) -> dict[str, np.ndarray | float | int]:
    hours_raw = np.asarray([r["hour"] for r in p.profiles], dtype=float)
    wind_raw = np.asarray([r["wind_m_s"] for r in p.profiles], dtype=float)
    amb_raw = np.asarray([r["ambient_c"] for r in p.profiles], dtype=float)
    solar_raw = np.asarray([r["solar_w_m2"] for r in p.profiles], dtype=float)
    load_raw = np.asarray([r["load_kw"] for r in p.profiles], dtype=float)

    n = int(round(horizon_hours / dt_hours))
    hours = np.arange(n, dtype=float) * dt_hours
    wind = np.interp(hours, hours_raw, wind_raw)
    ambient = np.interp(hours, hours_raw, amb_raw)
    solar = np.interp(hours, hours_raw, solar_raw)
    load_w = np.interp(hours, hours_raw, load_raw) * 1000.0

    p_wt = np.zeros(n, dtype=float)
    for unit in p.wind_units:
        rated = float(unit["rated_kw"]) * 1000.0
        cut_in = float(unit["cut_in_m_s"])
        rated_v = float(unit["rated_m_s"])
        cut_out = float(unit["cut_out_m_s"])
        unit_power = np.zeros(n, dtype=float)
        for i, v in enumerate(wind):
            if v < cut_in or v > cut_out:
                unit_power[i] = 0.0
            elif v >= rated_v:
                unit_power[i] = rated
            else:
                unit_power[i] = rated * (v**3 - cut_in**3) / (rated_v**3 - cut_in**3)
        p_wt += unit_power

    p_pv = np.zeros(n, dtype=float)
    for unit in p.pv_units:
        rated = float(unit["rated_kw"]) * 1000.0
        max_w = float(unit["p_max_kw"]) * 1000.0
        p_pv += np.minimum(rated * solar / p.g_std, max_w)

    return {
        "N": n,
        "dt": dt_hours * 3600.0,
        "hours": hours,
        "wind_m_s": wind,
        "solar_w_m2": solar,
        "T_amb": ambient,
        "P_load": load_w,
        "P_pv": p_pv,
        "P_wt": p_wt,
    }


def estimate_model_size(
    bp: Breakpoints,
    n_steps: int,
    n_dg_units: int,
    dg_points: int,
    current_mode: str = "continuous",
) -> dict[str, int]:
    n_s, n_t, n_i = bp.n_s, bp.n_t, bp.n_i
    continuous = n_steps * (
        36
        + n_s
        + n_t
        + n_s * n_t
        + n_s * n_t * n_i
        + n_dg_units * (2 + dg_points)
    )
    binaries = n_steps * (4 + n_dg_units)
    if current_mode == "discrete":
        binaries += n_steps * n_i
    diesel_online_constraints = 1 if n_dg_units > 0 else 0
    constraints = n_steps * (
        46 + n_s + n_t + n_s + n_t + n_s * n_t + n_dg_units * 6 + diesel_online_constraints
    )
    if current_mode == "discrete":
        constraints += n_steps * (1 + n_i)
    return {
        "continuous_vars_est": int(continuous),
        "binary_vars_est": int(binaries),
        "constraints_est": int(constraints),
    }


def apply_diagnostic_experiment(p: SimpleNamespace, experiment: str) -> None:
    p.diag_experiment = experiment
    p.diag_thermal_block = 1
    p.diag_disable_storage_dispatch = False
    p.diag_discrete_current = False
    p.diag_discrete_current_levels = 0
    p.diag_discrete_current_c_max = 0.0
    p.diag_current_mode = "continuous"
    p.diag_perspective_i2r = True
    p.diag_strict_current_sos2 = False
    p.route13_objective = "fuel_only"
    # Thermal switch holding/blocking constraints were removed by request.
    # Historical experiment names such as "block20" are kept for CLI compatibility,
    # but no longer force u_pi/u_po/u_lh/u_ch to stay constant across multiple steps.
    if experiment == "fuel_only_no_storage_block20":
        p.diag_disable_storage_dispatch = True
    if experiment == "fuel_only_discrete_i21_block20":
        p.diag_discrete_current = True
        p.diag_discrete_current_levels = 21
        p.diag_discrete_current_c_max = 1.0
        p.diag_current_mode = "discrete"
    if experiment == "perspective_i2r_strict_block20":
        p.diag_strict_current_sos2 = True


def temperature_bounds_snapshot(p: SimpleNamespace) -> dict[str, float]:
    return {
        "T_bat_min": float(p.T_bat_min),
        "T_bat_max": float(p.T_bat_max),
        "T_tank_min": float(p.T_tank_min),
        "T_tank_max": float(p.T_tank_max),
        "T_cont_min": float(p.T_cont_min),
        "T_cont_max": float(p.T_cont_max),
    }


def validate_temperature_bounds(p: SimpleNamespace) -> None:
    checks = [
        ("battery", "T_bat_min", "T_bat_max", "T_bat_init"),
        ("liquid tank", "T_tank_min", "T_tank_max", "T_tank_init"),
        ("container", "T_cont_min", "T_cont_max", "T_cont_init"),
    ]
    for label, min_attr, max_attr, init_attr in checks:
        lb = float(getattr(p, min_attr))
        ub = float(getattr(p, max_attr))
        init = float(getattr(p, init_attr))
        if lb > ub:
            raise ValueError(f"{label} temperature lower bound {lb} is greater than upper bound {ub}.")
        if init < lb - 1e-9 or init > ub + 1e-9:
            raise ValueError(
                f"{label} initial temperature {init}C is outside [{lb}, {ub}]C. "
                "Relax the corresponding --*-temp-min/--*-temp-max setting."
            )


def terminal_deviation_ub(lb: float, ub: float, target: float) -> float:
    return max(0.0, abs(float(ub) - float(target)), abs(float(target) - float(lb)))


def apply_initial_state_config(p: SimpleNamespace, args: argparse.Namespace) -> None:
    overrides = {
        "soc": getattr(args, "initial_soc", None),
        "t_bat_c": getattr(args, "initial_t_bat_c", None),
        "t_tank_c": getattr(args, "initial_t_tank_c", None),
        "t_cont_c": getattr(args, "initial_t_cont_c", None),
    }
    if all(value is None for value in overrides.values()):
        return
    state = dict(getattr(p, "raw", {}).get("initial_state", {}) or {})
    state.update(
        {
            "soc": float(overrides["soc"]) if overrides["soc"] is not None else float(getattr(p, "SOC_init", state.get("soc", 0.5))),
            "t_bat_c": float(overrides["t_bat_c"]) if overrides["t_bat_c"] is not None else float(getattr(p, "T_bat_init", state.get("t_bat_c", 5.0))),
            "t_tank_c": float(overrides["t_tank_c"]) if overrides["t_tank_c"] is not None else float(getattr(p, "T_tank_init", state.get("t_tank_c", 5.0))),
            "t_cont_c": float(overrides["t_cont_c"]) if overrides["t_cont_c"] is not None else float(getattr(p, "T_cont_init", state.get("t_cont_c", 5.0))),
        }
    )
    p.SOC_init = float(state["soc"])
    p.T_bat_init = float(state["t_bat_c"])
    p.T_tank_init = float(state["t_tank_c"])
    p.T_cont_init = float(state["t_cont_c"])
    if p.SOC_init < float(p.SOC_min) - 1e-9 or p.SOC_init > float(p.SOC_max) + 1e-9:
        raise ValueError(f"Battery initial SOC {p.SOC_init} is outside [{p.SOC_min}, {p.SOC_max}].")
    if hasattr(p, "raw"):
        p.raw["initial_state"] = state


def apply_tight_temp_bounds(
    p: SimpleNamespace,
    *,
    use_defaults: bool = True,
    bat_min: float | None = None,
    bat_max: float | None = None,
    tank_min: float | None = None,
    tank_max: float | None = None,
    cont_min: float | None = None,
    cont_max: float | None = None,
) -> None:
    """Shrink temperature bounds to reduce Big-M ranges without widening Excel bounds."""
    requested = {
        "T_bat_min": bat_min,
        "T_bat_max": bat_max,
        "T_tank_min": tank_min,
        "T_tank_max": tank_max,
        "T_cont_min": cont_min,
        "T_cont_max": cont_max,
    }
    for attr, value in requested.items():
        if value is None and use_defaults:
            value = DEFAULT_TIGHT_TEMP_BOUNDS[attr]
        if value is None:
            continue
        if attr.endswith("_min"):
            setattr(p, attr, max(float(getattr(p, attr)), float(value)))
        else:
            setattr(p, attr, min(float(getattr(p, attr)), float(value)))
    validate_temperature_bounds(p)
    p.diag_tight_temp_bounds = bool(use_defaults or any(v is not None for v in requested.values()))
    p.diag_temperature_bounds = temperature_bounds_snapshot(p)


def big_m_diagnostics(p: SimpleNamespace, data: dict[str, np.ndarray | float | int]) -> list[dict[str, object]]:
    t_amb = np.asarray(data["T_amb"], dtype=float)
    dT_bt_lb = p.T_bat_min - p.T_tank_max
    dT_bt_ub = p.T_bat_max - p.T_tank_min
    dT_ta_lb = p.T_tank_min - float(np.max(t_amb))
    dT_ta_ub = p.T_tank_max - float(np.min(t_amb))
    rows = [
        {
            "name": "dT_bt = T_bat - T_tank",
            "lb": dT_bt_lb,
            "ub": dT_bt_ub,
            "range": dT_bt_ub - dT_bt_lb,
            "used_by": "u_pi 与 v_bt/Q_bt 线性化",
            "suggestion": "若液冷罐正式运行温度可限定在更窄区间，可优先收紧 T_tank 上下界。",
        },
        {
            "name": "dT_ta = T_tank - T_amb",
            "lb": dT_ta_lb,
            "ub": dT_ta_ub,
            "range": dT_ta_ub - dT_ta_lb,
            "used_by": "u_po 与 v_ta/Q_tamb 线性化",
            "suggestion": "该 M 同时受环境最低温和液冷罐最高温影响；可按场景温度和储热上限分情景收紧。",
        },
        {
            "name": "tank_low_dev",
            "lb": 0.0,
            "ub": max(0.0, p.T_tank_band_low - p.T_tank_min),
            "range": max(0.0, p.T_tank_band_low - p.T_tank_min),
            "used_by": "液冷罐储热区下偏差",
            "suggestion": "若保留储热目标，可考虑把低温偏差变成少数时段或终端约束，减少目标项耦合。",
        },
        {
            "name": "tank_high_dev/hot_dev",
            "lb": 0.0,
            "ub": max(0.0, p.T_tank_max - min(p.T_tank_band_high, p.T_tank_hot)),
            "range": max(0.0, p.T_tank_max - min(p.T_tank_band_high, p.T_tank_hot)),
            "used_by": "液冷罐高温区/过热偏差",
            "suggestion": "当前 T_tank_max=120C 时该范围很宽；若工程上不允许长期高温，可收紧 T_tank_max。",
        },
        {
            "name": "cont_low/high/hot_dev",
            "lb": 0.0,
            "ub": max(0.0, p.T_cont_max - p.T_cont_min),
            "range": max(0.0, p.T_cont_max - p.T_cont_min),
            "used_by": "舱体温度区间偏差",
            "suggestion": "舱体热容小，温度变化快；建议检查舱体温度上下界是否过宽。",
        },
    ]
    return rows


def current_limit_stats(bp: Breakpoints, p: SimpleNamespace) -> dict[str, object]:
    return {
        "temperature_current_limits": True,
        "current_limit_temperature_points_c": [float(value) for value in p.current_limit_temps],
        "charge_current_limit_pack_min_a": float(np.min(p.charge_current_limit_pack)),
        "charge_current_limit_pack_max_a": float(np.max(p.charge_current_limit_pack)),
        "discharge_current_limit_pack_min_a": float(np.min(p.discharge_current_limit_pack)),
        "discharge_current_limit_pack_max_a": float(np.max(p.discharge_current_limit_pack)),
        "effective_charge_current_max_a": float(p.I_charge_max),
        "effective_discharge_current_max_a": float(p.I_discharge_max),
        "pack_current_hardware_max_a": float(p.pack_current_hardware_max),
        "breakpoint_charge_current_limit_a": [float(value) for value in bp.charge_current_limit],
        "breakpoint_discharge_current_limit_a": [float(value) for value in bp.discharge_current_limit],
    }


def model_stats_from_gurobi_model(m, bp: Breakpoints, n_steps: int, p: SimpleNamespace, data: dict) -> dict[str, object]:
    vars_by_prefix: dict[str, int] = {}
    for var in m.getVars():
        prefix = var.VarName.split("[", 1)[0]
        vars_by_prefix[prefix] = vars_by_prefix.get(prefix, 0) + 1
    lam_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("lam"))
    w_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("w"))
    rho_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("rho_i2"))
    current_choice_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("z_current"))
    balance_slack_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("P_balance_"))
    thermal_binary_effective = int(math.ceil(n_steps / max(1, int(getattr(p, "diag_thermal_block", 1)))) * 4)
    current_mode = str(getattr(p, "diag_current_mode", "continuous"))
    return {
        "experiment": getattr(p, "diag_experiment", "baseline"),
        "variables_total": int(m.NumVars),
        "binary_variables": int(m.NumBinVars),
        "continuous_variables": int(m.NumVars - m.NumBinVars),
        "constraints_total": int(m.NumConstrs),
        "sos_constraints": int(m.NumSOS),
        "nonzeros": int(m.NumNZs),
        "lam_variables_count": int(lam_count),
        "pwl_weight_variables_count": int(w_count),
        "rho_i2_variables_count": int(rho_count),
        "pwl_sos_or_lam_variables_count": int(lam_count + w_count + rho_count),
        "dg_ramp_constraints": 0,
        "strict_power_balance": True,
        "power_balance_slack_variables": int(balance_slack_count),
        "per_time_step_binary_count": float(m.NumBinVars) / max(1, n_steps),
        "thermal_binary_effective_count": thermal_binary_effective,
        "thermal_binary_block_steps": int(getattr(p, "diag_thermal_block", 1)),
        "steps": int(n_steps),
        "dt_minutes": float(data["dt"]) / 60.0,
        "breakpoints_soc": int(bp.n_s),
        "soc_grid_width": float(np.min(np.diff(bp.soc))) if bp.n_s > 1 else 0.0,
        "breakpoints_temp": int(bp.n_t),
        "breakpoints_current": int(bp.n_i),
        "current_segments": int(max(0, bp.n_i - 1)),
        "current_mode": current_mode,
        "discrete_current": current_mode == "discrete",
        "current_discrete_variables": int(current_choice_count),
        "perspective_i2r": bool(getattr(p, "diag_perspective_i2r", True)),
        "strict_current_sos2": bool(getattr(p, "diag_strict_current_sos2", False)),
        "perspective_mode": (
            "discrete_current"
            if current_mode == "discrete"
            else ("strict_current_sos2" if getattr(p, "diag_strict_current_sos2", False) else "convex_hull")
        ),
        "legacy_discrete_current_grid_override": bool(getattr(p, "diag_discrete_current", False)),
        "discrete_current_levels": int(getattr(p, "diag_discrete_current_levels", 0)),
        "discrete_current_c_max": float(getattr(p, "diag_discrete_current_c_max", 0.0)),
        "current_breakpoints_a": [float(x) for x in bp.current],
        **current_limit_stats(bp, p),
        "tight_temperature_bounds": bool(getattr(p, "diag_tight_temp_bounds", False)),
        "temperature_bounds": temperature_bounds_snapshot(p),
        "big_m_diagnostics": big_m_diagnostics(p, data),
    }


def model_stats_from_cplex_model(m, bp: Breakpoints, n_steps: int, p: SimpleNamespace, data: dict) -> dict[str, object]:
    var_names = list(m.variables.get_names())
    vars_by_prefix: dict[str, int] = {}
    for name in var_names:
        prefix = name.split("[", 1)[0]
        vars_by_prefix[prefix] = vars_by_prefix.get(prefix, 0) + 1
    var_types = list(m.variables.get_types())
    binary_count = sum(1 for kind in var_types if str(kind).upper() == "B")
    lam_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("lam"))
    w_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("w"))
    rho_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("rho_i2"))
    current_choice_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("z_current"))
    balance_slack_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("P_balance_"))
    thermal_binary_effective = int(math.ceil(n_steps / max(1, int(getattr(p, "diag_thermal_block", 1)))) * 4)
    current_mode = str(getattr(p, "diag_current_mode", "continuous"))
    try:
        nonzeros = int(m.linear_constraints.get_num_nonzeros())
    except Exception:
        nonzeros = 0
    return {
        "solver_backend": "cplex_native",
        "solver_name": "cplex",
        "uses_cvxpy": False,
        "native_sos2": True,
        "sos2_formulation": "native_sos2",
        "experiment": getattr(p, "diag_experiment", "baseline"),
        "variables_total": int(m.variables.get_num()),
        "binary_variables": int(binary_count),
        "continuous_variables": int(m.variables.get_num() - binary_count),
        "constraints_total": int(m.linear_constraints.get_num()),
        "sos_constraints": int(m.SOS.get_num()),
        "nonzeros": nonzeros,
        "lam_variables_count": int(lam_count),
        "pwl_weight_variables_count": int(w_count),
        "rho_i2_variables_count": int(rho_count),
        "pwl_sos_or_lam_variables_count": int(lam_count + w_count + rho_count),
        "dg_ramp_constraints": 0,
        "strict_power_balance": True,
        "power_balance_slack_variables": int(balance_slack_count),
        "per_time_step_binary_count": float(binary_count) / max(1, n_steps),
        "thermal_binary_effective_count": thermal_binary_effective,
        "thermal_binary_block_steps": int(getattr(p, "diag_thermal_block", 1)),
        "steps": int(n_steps),
        "dt_minutes": float(data["dt"]) / 60.0,
        "breakpoints_soc": int(bp.n_s),
        "soc_grid_width": float(np.min(np.diff(bp.soc))) if bp.n_s > 1 else 0.0,
        "breakpoints_temp": int(bp.n_t),
        "breakpoints_current": int(bp.n_i),
        "current_segments": int(max(0, bp.n_i - 1)),
        "current_mode": current_mode,
        "discrete_current": current_mode == "discrete",
        "current_discrete_variables": int(current_choice_count),
        "perspective_i2r": bool(getattr(p, "diag_perspective_i2r", True)),
        "strict_current_sos2": bool(getattr(p, "diag_strict_current_sos2", False)),
        "perspective_mode": (
            "discrete_current"
            if current_mode == "discrete"
            else ("strict_current_sos2" if getattr(p, "diag_strict_current_sos2", False) else "convex_hull")
        ),
        "legacy_discrete_current_grid_override": bool(getattr(p, "diag_discrete_current", False)),
        "discrete_current_levels": int(getattr(p, "diag_discrete_current_levels", 0)),
        "discrete_current_c_max": float(getattr(p, "diag_discrete_current_c_max", 0.0)),
        "current_breakpoints_a": [float(x) for x in bp.current],
        **current_limit_stats(bp, p),
        "tight_temperature_bounds": bool(getattr(p, "diag_tight_temp_bounds", False)),
        "temperature_bounds": temperature_bounds_snapshot(p),
        "big_m_diagnostics": big_m_diagnostics(p, data),
    }


def interp_weights(value: float, grid: np.ndarray) -> list[tuple[int, float]]:
    value = float(np.clip(value, grid[0], grid[-1]))
    hi = int(np.searchsorted(grid, value, side="right"))
    if hi <= 0:
        return [(0, 1.0)]
    if hi >= len(grid):
        return [(len(grid) - 1, 1.0)]
    lo = hi - 1
    span = grid[hi] - grid[lo]
    if abs(span) < 1e-12:
        return [(lo, 1.0)]
    w_hi = (value - grid[lo]) / span
    return [(lo, 1.0 - w_hi), (hi, w_hi)]


def bilinear_interp(x: float, y: float, x_grid: np.ndarray, y_grid: np.ndarray, values: np.ndarray) -> float:
    x = float(np.clip(x, x_grid[0], x_grid[-1]))
    y = float(np.clip(y, y_grid[0], y_grid[-1]))
    ix = int(np.searchsorted(x_grid, x, side="right")) - 1
    iy = int(np.searchsorted(y_grid, y, side="right")) - 1
    ix = max(0, min(ix, len(x_grid) - 2))
    iy = max(0, min(iy, len(y_grid) - 2))
    x0, x1 = x_grid[ix], x_grid[ix + 1]
    y0, y1 = y_grid[iy], y_grid[iy + 1]
    tx = 0.0 if abs(x1 - x0) < 1e-12 else (x - x0) / (x1 - x0)
    ty = 0.0 if abs(y1 - y0) < 1e-12 else (y - y0) / (y1 - y0)
    v00 = values[ix, iy]
    v10 = values[ix + 1, iy]
    v01 = values[ix, iy + 1]
    v11 = values[ix + 1, iy + 1]
    return float((1 - tx) * (1 - ty) * v00 + tx * (1 - ty) * v10 + (1 - tx) * ty * v01 + tx * ty * v11)


def set_var_start(var, value: float) -> None:
    try:
        var.Start = float(value)
    except Exception:
        pass


def set_var_hint(var, value: float) -> None:
    try:
        var.VarHintVal = float(value)
    except Exception:
        pass


def resolve_solver_backend(solver: str) -> tuple[str, str]:
    requested = solver.lower()
    if requested == "auto":
        if importlib.util.find_spec("gurobipy") is not None:
            return "gurobi", "gurobi"
        if importlib.util.find_spec("cplex") is not None:
            return "cplex", "cplex_native"
        if importlib.util.find_spec("mosek") is not None:
            return "mosek", "mosek_native"
        return "gurobi", "gurobi"
    if requested == "gurobi":
        return "gurobi", "gurobi"
    if requested == "cplex":
        return "cplex", "cplex_native"
    if requested == "mosek":
        return "mosek", "mosek_native"
    raise ValueError(f"Unsupported solver: {solver}")


def solver_candidates(solver: str) -> list[tuple[str, str]]:
    requested = solver.lower()
    if requested == "auto":
        return [
            ("gurobi", "gurobi"),
            ("cplex", "cplex_native"),
            ("mosek", "mosek_native"),
        ]
    return [resolve_solver_backend(requested)]


def solve_milp_cplex_native(
    p: SimpleNamespace,
    data: dict[str, np.ndarray | float | int],
    bp: Breakpoints,
    time_limit: float,
    mip_gap: float,
    build_only: bool = False,
) -> dict:
    try:
        import cplex
        from cplex import SparsePair
    except Exception as exc:
        return {
            "success": False,
            "status": "SOLVER_NOT_AVAILABLE",
            "time_s": 0.0,
            "message": f"CPLEX Python API is not available: {exc}",
        }

    n = int(data["N"])
    dt_s = float(data["dt"])
    hours = np.asarray(data["hours"], dtype=float)
    t_amb = np.asarray(data["T_amb"], dtype=float)
    renewable_surplus_w = np.maximum(
        0.0, np.asarray(data["P_pv"], dtype=float) + np.asarray(data["P_wt"], dtype=float) - np.asarray(data["P_load"], dtype=float)
    )
    renewable_surplus_max_w = float(np.max(renewable_surplus_w)) if renewable_surplus_w.size else 0.0
    renewable_surplus_weight = renewable_surplus_w / renewable_surplus_max_w if renewable_surplus_max_w > 1e-9 else np.zeros(n, dtype=float)

    n_s, n_t, n_i = bp.n_s, bp.n_t, bp.n_i
    dg_units = p.diesel_units
    n_g = len(dg_units)
    dg_points = [len(u["powers_w"]) for u in dg_units]
    pdc_pack_table = np.array(
        [
            [[bp.ocv[a] * bp.current[c] - bp.current2[c] * bp.r0[a, b] for c in range(n_i)] for b in range(n_t)]
            for a in range(n_s)
        ],
        dtype=float,
    )
    qgen_pack_table = np.array(
        [
            [[bp.current2[c] * bp.r0[a, b] for c in range(n_i)] for b in range(n_t)]
            for a in range(n_s)
        ],
        dtype=float,
    )
    r0_flat = bp.r0.reshape(-1)

    m = cplex.Cplex()
    m.set_problem_name("BESS_MILP_perspective_i2r_20260530_cplex_native")
    m.objective.set_sense(m.objective.sense.minimize)
    if float(time_limit) > 0:
        m.parameters.timelimit.set(float(time_limit))
    m.parameters.mip.tolerances.mipgap.set(float(mip_gap))
    threads = int(getattr(p, "diag_threads", 0))
    if threads > 0:
        m.parameters.threads.set(threads)
    log_handle = None
    if getattr(p, "diag_log_file", None):
        Path(p.diag_log_file).parent.mkdir(parents=True, exist_ok=True)
        log_handle = open(p.diag_log_file, "w", encoding="utf-8")
        m.set_log_stream(log_handle)
        m.set_results_stream(log_handle)
        m.set_warning_stream(log_handle)
        m.set_error_stream(log_handle)

    inf = cplex.infinity

    def add_vars(keys, name_fn, lb_fn=0.0, ub_fn=inf, type_fn="C", obj_fn=0.0):
        names: list[str] = []
        lbs: list[float] = []
        ubs: list[float] = []
        objs: list[float] = []
        types: list[str] = []
        for key in keys:
            names.append(name_fn(key))
            lbs.append(float(lb_fn(key) if callable(lb_fn) else lb_fn))
            ubs.append(float(ub_fn(key) if callable(ub_fn) else ub_fn))
            objs.append(float(obj_fn(key) if callable(obj_fn) else obj_fn))
            types.append(str(type_fn(key) if callable(type_fn) else type_fn))
        if names:
            m.variables.add(obj=objs, lb=lbs, ub=ubs, types="".join(types), names=names)
        return {key: names[i] for i, key in enumerate(keys)}

    def add_scalar(name: str, lb: float = 0.0, ub: float = inf, vtype: str = "C", obj: float = 0.0) -> str:
        m.variables.add(obj=[float(obj)], lb=[float(lb)], ub=[float(ub)], types=vtype, names=[name])
        return name

    lin_exprs: list[SparsePair] = []
    senses: list[str] = []
    rhs_values: list[float] = []
    row_names: list[str] = []

    def add_constr(terms: list[tuple[str, float]], sense: str, rhs: float, name: str) -> None:
        acc: dict[str, float] = {}
        for var, coef in terms:
            coef = float(coef)
            if abs(coef) <= 1e-12:
                continue
            acc[var] = acc.get(var, 0.0) + coef
        ind = []
        val = []
        for var, coef in acc.items():
            if abs(coef) > 1e-12:
                ind.append(var)
                val.append(coef)
        lin_exprs.append(SparsePair(ind=ind, val=val))
        senses.append(sense)
        rhs_values.append(float(rhs))
        row_names.append(name)

    def add_sos2(names: list[str], name: str) -> None:
        if len(names) <= 1:
            return
        weights = [float(i + 1) for i in range(len(names))]
        m.SOS.add(type=m.SOS.type.SOS2, SOS=SparsePair(ind=names, val=weights), name=name)

    I_bat = add_vars(range(n), lambda t: f"I_bat[{t}]", -p.I_charge_max, p.I_discharge_max)
    SOC = add_vars(range(n), lambda t: f"SOC[{t}]", p.SOC_min, p.SOC_max)
    T_bat = add_vars(range(n), lambda t: f"T_bat[{t}]", p.T_bat_min, p.T_bat_max)
    T_tank = add_vars(range(n), lambda t: f"T_tank[{t}]", p.T_tank_min, p.T_tank_max)
    T_cont = add_vars(range(n), lambda t: f"T_cont[{t}]", p.T_cont_min, p.T_cont_max)

    P_dg: dict[tuple[int, int], str] = {}
    M_dg: dict[tuple[int, int], str] = {}
    u_dg: dict[tuple[int, int], str] = {}
    lam_dg: dict[tuple[int, int, int], str] = {}
    for g, unit in enumerate(dg_units):
        P_dg.update(add_vars([(g, t) for t in range(n)], lambda key: f"P_dg[{key[0]},{key[1]}]", 0.0, float(unit["p_max_w"])))
        M_dg.update(
            add_vars(
                [(g, t) for t in range(n)],
                lambda key: f"M_dg[{key[0]},{key[1]}]",
                0.0,
                float(np.max(unit["fuel_kg_h"])),
                obj_fn=lambda key: dt_s / 3600.0,
            )
        )
        u_dg.update(add_vars([(g, t) for t in range(n)], lambda key: f"u_dg[{key[0]},{key[1]}]", 0.0, 1.0, "B"))
        lam_dg.update(
            add_vars(
                [(g, t, k) for t in range(n) for k in range(dg_points[g])],
                lambda key: f"lam_dg[{key[0]},{key[1]},{key[2]}]",
                0.0,
                1.0,
            )
        )

    P_BESS = add_vars(range(n), lambda t: f"P_BESS[{t}]", -350e3, 350e3)
    P_pv_use = add_vars(range(n), lambda t: f"P_pv_use[{t}]", 0.0, inf)
    P_wt_use = add_vars(range(n), lambda t: f"P_wt_use[{t}]", 0.0, inf)
    P_pv_curt = add_vars(range(n), lambda t: f"P_pv_curt[{t}]", 0.0, inf)
    P_wt_curt = add_vars(range(n), lambda t: f"P_wt_curt[{t}]", 0.0, inf)
    tank_preheat_short = add_vars(range(n), lambda t: f"tank_preheat_short[{t}]", 0.0, max(0.0, p.T_tank_target - p.T_tank_min))
    cont_preheat_short = add_vars(range(n), lambda t: f"cont_preheat_short[{t}]", 0.0, max(0.0, p.T_cont_target - p.T_cont_min))
    bat_low_dev = add_vars(range(n), lambda t: f"bat_low_dev[{t}]", 0.0, max(0.0, p.T_bat_pref_low - p.T_bat_min))
    bat_high_dev = add_vars(range(n), lambda t: f"bat_high_dev[{t}]", 0.0, max(0.0, p.T_bat_max - p.T_bat_pref_high))
    tank_low_dev = add_vars(range(n), lambda t: f"tank_low_dev[{t}]", 0.0, max(0.0, p.T_tank_band_low - p.T_tank_min))
    tank_high_dev = add_vars(range(n), lambda t: f"tank_high_dev[{t}]", 0.0, max(0.0, p.T_tank_max - p.T_tank_band_high))
    tank_hot_dev = add_vars(range(n), lambda t: f"tank_hot_dev[{t}]", 0.0, max(0.0, p.T_tank_max - p.T_tank_hot))
    cont_low_dev = add_vars(range(n), lambda t: f"cont_low_dev[{t}]", 0.0, max(0.0, p.T_cont_band_low - p.T_cont_min))
    cont_high_dev = add_vars(range(n), lambda t: f"cont_high_dev[{t}]", 0.0, max(0.0, p.T_cont_max - p.T_cont_band_high))
    cont_hot_dev = add_vars(range(n), lambda t: f"cont_hot_dev[{t}]", 0.0, max(0.0, p.T_cont_max - p.T_cont_hot))
    tank_terminal_short = add_scalar(
        "tank_terminal_short",
        0.0,
        terminal_deviation_ub(p.T_tank_min, p.T_tank_max, p.T_tank_init),
    )
    cont_terminal_short = add_scalar(
        "cont_terminal_short",
        0.0,
        terminal_deviation_ub(p.T_cont_min, p.T_cont_max, p.T_cont_init),
    )

    u_pi = add_vars(range(n), lambda t: f"u_pi[{t}]", 0.0, 1.0, "B")
    u_po = add_vars(range(n), lambda t: f"u_po[{t}]", 0.0, 1.0, "B")
    u_lh = add_vars(range(n), lambda t: f"u_lh[{t}]", 0.0, 1.0, "B")
    u_ch = add_vars(range(n), lambda t: f"u_ch[{t}]", 0.0, 1.0, "B")
    P_heat_liquid_ctrl = add_vars(range(n), lambda t: f"P_heat_liquid_ctrl[{t}]", 0.0, p.P_heat_liquid)
    P_heat_cont_ctrl = add_vars(range(n), lambda t: f"P_heat_cont_ctrl[{t}]", 0.0, p.P_heat_cont)
    p_heat_liquid_min_w = min(1.0, max(0.0, float(p.P_heat_liquid)))
    p_heat_cont_min_w = min(1.0, max(0.0, float(p.P_heat_cont)))
    u_pi_change = add_vars(range(max(0, n - 1)), lambda t: f"u_pi_change[{t}]", 0.0, 1.0)
    u_po_change = add_vars(range(max(0, n - 1)), lambda t: f"u_po_change[{t}]", 0.0, 1.0)
    u_lh_change = add_vars(range(max(0, n - 1)), lambda t: f"u_lh_change[{t}]", 0.0, 1.0)
    u_ch_change = add_vars(range(max(0, n - 1)), lambda t: f"u_ch_change[{t}]", 0.0, 1.0)
    tank_temp_step = add_vars(range(max(0, n - 1)), lambda t: f"tank_temp_step[{t}]", 0.0, p.T_tank_max - p.T_tank_min)
    cont_temp_step = add_vars(range(max(0, n - 1)), lambda t: f"cont_temp_step[{t}]", 0.0, p.T_cont_max - p.T_cont_min)

    dT_bt_lb = p.T_bat_min - p.T_tank_max
    dT_bt_ub = p.T_bat_max - p.T_tank_min
    dT_ta_lb = p.T_tank_min - float(np.max(t_amb))
    dT_ta_ub = p.T_tank_max - float(np.min(t_amb))
    Q_bt = add_vars(range(n), lambda t: f"Q_bt[{t}]", min(0.0, p.K_bt * dT_bt_lb), max(0.0, p.K_bt * dT_bt_ub))
    Q_tamb = add_vars(range(n), lambda t: f"Q_tamb[{t}]", min(0.0, p.K_t_amb * dT_ta_lb), max(0.0, p.K_t_amb * dT_ta_ub))
    Q_tamb_dump = add_vars(range(n), lambda t: f"Q_tamb_dump[{t}]", 0.0, max(0.0, p.K_t_amb * dT_ta_ub))
    v_bt = add_vars(range(n), lambda t: f"v_bt[{t}]", min(0.0, dT_bt_lb), max(0.0, dT_bt_ub))
    v_ta = add_vars(range(n), lambda t: f"v_ta[{t}]", min(0.0, dT_ta_lb), max(0.0, dT_ta_ub))

    lam_s = add_vars([(t, a) for t in range(n) for a in range(n_s)], lambda key: f"lam_s[{key[0]},{key[1]}]", 0.0, 1.0)
    lam_t = add_vars([(t, b) for t in range(n) for b in range(n_t)], lambda key: f"lam_t[{key[0]},{key[1]}]", 0.0, 1.0)
    pdc_lb = float(np.min(pdc_pack_table))
    pdc_ub = float(np.max(pdc_pack_table))
    qgen_ub = float(np.max(qgen_pack_table))
    P_dc_pack = add_vars(range(n), lambda t: f"P_dc_pack[{t}]", pdc_lb, pdc_ub)
    Q_gen_pack = add_vars(range(n), lambda t: f"Q_gen_pack[{t}]", 0.0, qgen_ub)
    P_dc_abs = add_vars(range(n), lambda t: f"P_dc_abs[{t}]", 0.0, max(abs(pdc_lb), abs(pdc_ub)))
    w_st = add_vars([(t, a, b) for t in range(n) for a in range(n_s) for b in range(n_t)], lambda key: f"w_st[{key[0]},{key[1]},{key[2]}]", 0.0, 1.0)
    rho_i2 = add_vars(
        [(t, a, b, c) for t in range(n) for a in range(n_s) for b in range(n_t) for c in range(n_i)],
        lambda key: f"rho_i2[{key[0]},{key[1]},{key[2]},{key[3]}]",
        0.0,
        1.0,
    )
    current_mode = str(getattr(p, "diag_current_mode", "continuous"))
    z_current = None
    if current_mode == "discrete":
        z_current = add_vars([(t, c) for t in range(n) for c in range(n_i)], lambda key: f"z_current[{key[0]},{key[1]}]", 0.0, 1.0, "B")

    for t in range(n):
        add_constr([(lam_s[t, a], 1.0) for a in range(n_s)], "E", 1.0, f"sum_lam_s_{t}")
        add_sos2([lam_s[t, a] for a in range(n_s)], f"sos_lam_s_{t}")
        add_constr([(lam_t[t, b], 1.0) for b in range(n_t)], "E", 1.0, f"sum_lam_t_{t}")
        add_sos2([lam_t[t, b] for b in range(n_t)], f"sos_lam_t_{t}")
        add_constr([(SOC[t], 1.0)] + [(lam_s[t, a], -float(bp.soc[a])) for a in range(n_s)], "E", 0.0, f"soc_map_{t}")
        add_constr([(T_bat[t], 1.0)] + [(lam_t[t, b], -float(bp.temp[b])) for b in range(n_t)], "E", 0.0, f"tbat_map_{t}")
        add_constr(
            [(I_bat[t], 1.0)] + [(lam_t[t, b], float(bp.charge_current_limit[b])) for b in range(n_t)],
            "G",
            0.0,
            f"ibat_charge_temp_limit_{t}",
        )
        add_constr(
            [(I_bat[t], 1.0)] + [(lam_t[t, b], -float(bp.discharge_current_limit[b])) for b in range(n_t)],
            "L",
            0.0,
            f"ibat_discharge_temp_limit_{t}",
        )
        for a in range(n_s):
            add_constr([(w_st[t, a, b], 1.0) for b in range(n_t)] + [(lam_s[t, a], -1.0)], "E", 0.0, f"w_st_s_{t}_{a}")
        for b in range(n_t):
            add_constr([(w_st[t, a, b], 1.0) for a in range(n_s)] + [(lam_t[t, b], -1.0)], "E", 0.0, f"w_st_t_{t}_{b}")
        for a in range(n_s):
            for b in range(n_t):
                add_constr(
                    [(rho_i2[t, a, b, c], 1.0) for c in range(n_i)] + [(w_st[t, a, b], -1.0)],
                    "E",
                    0.0,
                    f"rho_st_{t}_{a}_{b}",
                )
                if getattr(p, "diag_strict_current_sos2", False):
                    add_sos2([rho_i2[t, a, b, c] for c in range(n_i)], f"sos_rho_i2_{t}_{a}_{b}")
        if z_current is not None:
            add_constr([(z_current[t, c], 1.0) for c in range(n_i)], "E", 1.0, f"current_choice_sum_{t}")
            for c in range(n_i):
                add_constr(
                    [(rho_i2[t, a, b, c], 1.0) for a in range(n_s) for b in range(n_t)] + [(z_current[t, c], -1.0)],
                    "E",
                    0.0,
                    f"current_choice_link_{t}_{c}",
                )
        add_constr(
            [(I_bat[t], 1.0)] + [(rho_i2[t, a, b, c], -float(bp.current[c])) for a in range(n_s) for b in range(n_t) for c in range(n_i)],
            "E",
            0.0,
            f"ibat_perspective_map_{t}",
        )
        add_constr(
            [(P_dc_pack[t], 1.0)]
            + [(rho_i2[t, a, b, c], -float(pdc_pack_table[a, b, c])) for a in range(n_s) for b in range(n_t) for c in range(n_i)],
            "E",
            0.0,
            f"pdc_pack_map_{t}",
        )
        add_constr(
            [(Q_gen_pack[t], 1.0)]
            + [(rho_i2[t, a, b, c], -float(qgen_pack_table[a, b, c])) for a in range(n_s) for b in range(n_t) for c in range(n_i)],
            "E",
            0.0,
            f"qgen_pack_map_{t}",
        )

        for g, unit in enumerate(dg_units):
            add_constr([(lam_dg[g, t, k], 1.0) for k in range(dg_points[g])] + [(u_dg[g, t], -1.0)], "E", 0.0, f"dg_lam_sum_{g}_{t}")
            add_sos2([lam_dg[g, t, k] for k in range(dg_points[g])], f"sos_lam_dg_{g}_{t}")
            add_constr(
                [(P_dg[g, t], 1.0)] + [(lam_dg[g, t, k], -float(unit["powers_w"][k])) for k in range(dg_points[g])],
                "E",
                0.0,
                f"dg_p_map_{g}_{t}",
            )
            add_constr(
                [(M_dg[g, t], 1.0)] + [(lam_dg[g, t, k], -float(unit["fuel_kg_h"][k])) for k in range(dg_points[g])],
                "E",
                0.0,
                f"dg_f_map_{g}_{t}",
            )
            add_constr([(P_dg[g, t], 1.0), (u_dg[g, t], -float(unit["p_min_w"]))], "G", 0.0, f"dg_min_{g}_{t}")
            add_constr([(P_dg[g, t], 1.0), (u_dg[g, t], -float(unit["p_max_w"]))], "L", 0.0, f"dg_max_{g}_{t}")
        if n_g:
            add_constr([(u_dg[g, t], 1.0) for g in range(n_g)], "G", 1.0, f"dg_at_least_one_on_{t}")

        add_constr([(v_bt[t], 1.0), (u_pi[t], -dT_bt_lb)], "G", 0.0, f"vbt_lb0_{t}")
        add_constr([(v_bt[t], 1.0), (u_pi[t], -dT_bt_ub)], "L", 0.0, f"vbt_ub0_{t}")
        add_constr([(v_bt[t], 1.0), (T_bat[t], -1.0), (T_tank[t], 1.0), (u_pi[t], -dT_bt_ub)], "G", -dT_bt_ub, f"vbt_lb1_{t}")
        add_constr([(v_bt[t], 1.0), (T_bat[t], -1.0), (T_tank[t], 1.0), (u_pi[t], -dT_bt_lb)], "L", -dT_bt_lb, f"vbt_ub1_{t}")
        add_constr([(Q_bt[t], 1.0), (v_bt[t], -p.K_bt)], "E", 0.0, f"qbt_map_{t}")

        add_constr([(v_ta[t], 1.0), (u_po[t], -dT_ta_lb)], "G", 0.0, f"vta_lb0_{t}")
        add_constr([(v_ta[t], 1.0), (u_po[t], -dT_ta_ub)], "L", 0.0, f"vta_ub0_{t}")
        add_constr([(v_ta[t], 1.0), (T_tank[t], -1.0), (u_po[t], -dT_ta_ub)], "G", -float(t_amb[t]) - dT_ta_ub, f"vta_lb1_{t}")
        add_constr([(v_ta[t], 1.0), (T_tank[t], -1.0), (u_po[t], -dT_ta_lb)], "L", -float(t_amb[t]) - dT_ta_lb, f"vta_ub1_{t}")
        add_constr([(Q_tamb[t], 1.0), (v_ta[t], -p.K_t_amb)], "E", 0.0, f"qtamb_map_{t}")
        add_constr([(Q_tamb_dump[t], 1.0), (Q_tamb[t], -1.0)], "G", 0.0, f"qtamb_dump_pos_{t}")

        add_constr([(P_heat_liquid_ctrl[t], 1.0), (u_lh[t], -p.P_heat_liquid)], "L", 0.0, f"p_heat_liquid_ctrl_max_{t}")
        add_constr([(P_heat_cont_ctrl[t], 1.0), (u_ch[t], -p.P_heat_cont)], "L", 0.0, f"p_heat_cont_ctrl_max_{t}")
        add_constr([(P_heat_liquid_ctrl[t], 1.0), (u_lh[t], -p_heat_liquid_min_w)], "G", 0.0, f"p_heat_liquid_ctrl_min_{t}")
        add_constr([(P_heat_cont_ctrl[t], 1.0), (u_ch[t], -p_heat_cont_min_w)], "G", 0.0, f"p_heat_cont_ctrl_min_{t}")
        aux_terms = [
            (u_pi[t], p.P_pump_in),
            (u_po[t], p.P_pump_out),
            (P_heat_liquid_ctrl[t], 1.0),
            (P_heat_cont_ctrl[t], 1.0),
        ]
        add_constr([(P_dc_abs[t], 1.0), (P_dc_pack[t], -1.0)], "G", 0.0, f"pdc_abs_pos_{t}")
        add_constr([(P_dc_abs[t], 1.0), (P_dc_pack[t], 1.0)], "G", 0.0, f"pdc_abs_neg_{t}")
        add_constr([(P_BESS[t], 1.0), (P_dc_pack[t], -1.0), (P_dc_abs[t], p.mu_pcs)] + aux_terms, "E", 0.0, f"pbess_{t}")
        if getattr(p, "diag_disable_storage_dispatch", False):
            add_constr([(I_bat[t], 1.0)], "E", 0.0, f"no_storage_current_{t}")

        add_constr([(P_pv_use[t], 1.0), (P_pv_curt[t], 1.0)], "E", float(data["P_pv"][t]), f"pv_split_{t}")
        add_constr([(P_wt_use[t], 1.0), (P_wt_curt[t], 1.0)], "E", float(data["P_wt"][t]), f"wt_split_{t}")
        add_constr(
            [(P_pv_use[t], 1.0), (P_wt_use[t], 1.0), (P_BESS[t], 1.0)] + [(P_dg[g, t], 1.0) for g in range(n_g)],
            "E",
            float(data["P_load"][t]),
            f"power_balance_{t}",
        )
        add_constr([(tank_preheat_short[t], 1.0), (T_tank[t], 1.0)], "G", p.T_tank_target, f"tank_preheat_short_{t}")
        add_constr([(cont_preheat_short[t], 1.0), (T_cont[t], 1.0)], "G", p.T_cont_target, f"cont_preheat_short_{t}")
        add_constr([(bat_low_dev[t], 1.0), (T_bat[t], 1.0)], "G", p.T_bat_pref_low, f"bat_low_dev_{t}")
        add_constr([(bat_high_dev[t], 1.0), (T_bat[t], -1.0)], "G", -p.T_bat_pref_high, f"bat_high_dev_{t}")
        add_constr([(tank_low_dev[t], 1.0), (T_tank[t], 1.0)], "G", p.T_tank_band_low, f"tank_low_dev_{t}")
        add_constr([(tank_high_dev[t], 1.0), (T_tank[t], -1.0)], "G", -p.T_tank_band_high, f"tank_high_dev_{t}")
        add_constr([(tank_hot_dev[t], 1.0), (T_tank[t], -1.0)], "G", -p.T_tank_hot, f"tank_hot_dev_{t}")
        add_constr([(cont_low_dev[t], 1.0), (T_cont[t], 1.0)], "G", p.T_cont_band_low, f"cont_low_dev_{t}")
        add_constr([(cont_high_dev[t], 1.0), (T_cont[t], -1.0)], "G", -p.T_cont_band_high, f"cont_high_dev_{t}")
        add_constr([(cont_hot_dev[t], 1.0), (T_cont[t], -1.0)], "G", -p.T_cont_hot, f"cont_hot_dev_{t}")

    add_constr([(SOC[0], 1.0)], "E", p.SOC_init, "soc_initial")
    add_constr([(T_bat[0], 1.0)], "E", p.T_bat_init, "tbat_initial")
    add_constr([(T_tank[0], 1.0)], "E", p.T_tank_init, "ttank_initial")
    add_constr([(T_cont[0], 1.0)], "E", p.T_cont_init, "tcont_initial")

    for t in range(1, n):
        prev = t - 1
        add_constr([(SOC[t], 1.0), (SOC[prev], -1.0), (I_bat[prev], dt_s / (p.Q_nom * 3600.0))], "E", 0.0, f"soc_dyn_{t}")
        add_constr([(u_pi_change[prev], 1.0), (u_pi[t], -1.0), (u_pi[prev], 1.0)], "G", 0.0, f"upi_change_pos_{t}")
        add_constr([(u_pi_change[prev], 1.0), (u_pi[prev], -1.0), (u_pi[t], 1.0)], "G", 0.0, f"upi_change_neg_{t}")
        add_constr([(u_po_change[prev], 1.0), (u_po[t], -1.0), (u_po[prev], 1.0)], "G", 0.0, f"upo_change_pos_{t}")
        add_constr([(u_po_change[prev], 1.0), (u_po[prev], -1.0), (u_po[t], 1.0)], "G", 0.0, f"upo_change_neg_{t}")
        add_constr([(u_lh_change[prev], 1.0), (u_lh[t], -1.0), (u_lh[prev], 1.0)], "G", 0.0, f"ulh_change_pos_{t}")
        add_constr([(u_lh_change[prev], 1.0), (u_lh[prev], -1.0), (u_lh[t], 1.0)], "G", 0.0, f"ulh_change_neg_{t}")
        add_constr([(u_ch_change[prev], 1.0), (u_ch[t], -1.0), (u_ch[prev], 1.0)], "G", 0.0, f"uch_change_pos_{t}")
        add_constr([(u_ch_change[prev], 1.0), (u_ch[prev], -1.0), (u_ch[t], 1.0)], "G", 0.0, f"uch_change_neg_{t}")
        add_constr([(tank_temp_step[prev], 1.0), (T_tank[t], -1.0), (T_tank[prev], 1.0)], "G", 0.0, f"tank_temp_step_pos_{t}")
        add_constr([(tank_temp_step[prev], 1.0), (T_tank[prev], -1.0), (T_tank[t], 1.0)], "G", 0.0, f"tank_temp_step_neg_{t}")
        add_constr([(cont_temp_step[prev], 1.0), (T_cont[t], -1.0), (T_cont[prev], 1.0)], "G", 0.0, f"cont_temp_step_pos_{t}")
        add_constr([(cont_temp_step[prev], 1.0), (T_cont[prev], -1.0), (T_cont[t], 1.0)], "G", 0.0, f"cont_temp_step_neg_{t}")

        add_constr(
            [
                (T_bat[t], 1.0),
                (T_bat[prev], -1.0),
                (Q_gen_pack[prev], -dt_s / p.C_bat),
                (Q_bt[prev], dt_s / p.C_bat),
                (T_bat[t], p.K_b_cont * dt_s / p.C_bat),
                (T_cont[t], -p.K_b_cont * dt_s / p.C_bat),
            ],
            "E",
            0.0,
            f"tbat_dyn_{t}",
        )
        add_constr(
            [
                (T_tank[t], 1.0),
                (T_tank[prev], -1.0),
                (Q_bt[prev], -dt_s / p.C_tank),
                (P_heat_liquid_ctrl[prev], -dt_s / p.C_tank),
                (Q_tamb[prev], dt_s / p.C_tank),
                (T_tank[t], p.K_t_cont * dt_s / p.C_tank),
                (T_cont[t], -p.K_t_cont * dt_s / p.C_tank),
            ],
            "E",
            0.0,
            f"ttank_dyn_{t}",
        )
        add_constr(
            [
                (T_cont[t], 1.0),
                (T_cont[prev], -1.0),
                (T_bat[t], -p.K_b_cont * dt_s / p.C_cont),
                (T_cont[t], p.K_b_cont * dt_s / p.C_cont),
                (T_tank[t], -p.K_t_cont * dt_s / p.C_cont),
                (T_cont[t], p.K_t_cont * dt_s / p.C_cont),
                (P_heat_cont_ctrl[prev], -dt_s / p.C_cont),
                (T_cont[t], p.K_cont_amb * dt_s / p.C_cont),
            ],
            "E",
            p.K_cont_amb * float(t_amb[prev]) * dt_s / p.C_cont,
            f"tcont_dyn_{t}",
        )

    SOC_end = add_scalar("SOC_end", p.SOC_min, p.SOC_max)
    add_constr([(SOC_end, 1.0), (SOC[n - 1], -1.0), (I_bat[n - 1], dt_s / (p.Q_nom * 3600.0))], "E", 0.0, "soc_terminal")
    soc_dev = add_scalar("soc_dev", 0.0, terminal_deviation_ub(p.SOC_min, p.SOC_max, p.SOC_init))
    add_constr([(SOC[n - 1], 1.0)], "E", p.SOC_init, "soc_schedule_terminal_target")
    add_constr([(soc_dev, 1.0), (SOC_end, -1.0)], "G", -p.SOC_init, "soc_dev_pos")
    add_constr([(soc_dev, 1.0), (SOC_end, 1.0)], "G", p.SOC_init, "soc_dev_neg")

    T_bat_end = add_scalar("T_bat_end", p.T_bat_min, p.T_bat_max)
    T_tank_end = add_scalar("T_tank_end", p.T_tank_min, p.T_tank_max)
    T_cont_end = add_scalar("T_cont_end", p.T_cont_min, p.T_cont_max)
    last = n - 1
    add_constr(
        [
            (T_bat_end, 1.0),
            (T_bat[last], -1.0),
            (Q_gen_pack[last], -dt_s / p.C_bat),
            (Q_bt[last], dt_s / p.C_bat),
            (T_bat_end, p.K_b_cont * dt_s / p.C_bat),
            (T_cont_end, -p.K_b_cont * dt_s / p.C_bat),
        ],
        "E",
        0.0,
        "tbat_terminal",
    )
    add_constr(
        [
            (T_tank_end, 1.0),
            (T_tank[last], -1.0),
            (Q_bt[last], -dt_s / p.C_tank),
            (P_heat_liquid_ctrl[last], -dt_s / p.C_tank),
            (Q_tamb[last], dt_s / p.C_tank),
            (T_tank_end, p.K_t_cont * dt_s / p.C_tank),
            (T_cont_end, -p.K_t_cont * dt_s / p.C_tank),
        ],
        "E",
        0.0,
        "ttank_terminal",
    )
    add_constr(
        [
            (T_cont_end, 1.0),
            (T_cont[last], -1.0),
            (T_bat_end, -p.K_b_cont * dt_s / p.C_cont),
            (T_cont_end, p.K_b_cont * dt_s / p.C_cont),
            (T_tank_end, -p.K_t_cont * dt_s / p.C_cont),
            (T_cont_end, p.K_t_cont * dt_s / p.C_cont),
            (P_heat_cont_ctrl[last], -dt_s / p.C_cont),
            (T_cont_end, p.K_cont_amb * dt_s / p.C_cont),
        ],
        "E",
        p.K_cont_amb * float(t_amb[last]) * dt_s / p.C_cont,
        "tcont_terminal",
    )
    add_constr([(T_tank[n - 1], 1.0)], "E", p.T_tank_init, "ttank_schedule_terminal_target")
    add_constr([(T_cont[n - 1], 1.0)], "E", p.T_cont_init, "tcont_schedule_terminal_target")
    add_constr([(tank_terminal_short, 1.0), (T_tank_end, -1.0)], "G", -p.T_tank_init, "tank_terminal_dev_pos")
    add_constr([(tank_terminal_short, 1.0), (T_tank_end, 1.0)], "G", p.T_tank_init, "tank_terminal_dev_neg")
    add_constr([(cont_terminal_short, 1.0), (T_cont_end, -1.0)], "G", -p.T_cont_init, "cont_terminal_dev_pos")
    add_constr([(cont_terminal_short, 1.0), (T_cont_end, 1.0)], "G", p.T_cont_init, "cont_terminal_dev_neg")

    m.linear_constraints.add(lin_expr=lin_exprs, senses="".join(senses), rhs=rhs_values, names=row_names)
    model_stats = model_stats_from_cplex_model(m, bp, n, p, data)
    if getattr(p, "diag_model_stats_json", None):
        Path(p.diag_model_stats_json).parent.mkdir(parents=True, exist_ok=True)
        Path(p.diag_model_stats_json).write_text(json.dumps(model_stats, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        f"Model size: {model_stats['variables_total']} vars ({model_stats['binary_variables']} binaries), "
        f"{model_stats['constraints_total']} constraints, {model_stats['sos_constraints']} native SOS, "
        f"{model_stats['nonzeros']} nonzeros"
    )
    print(
        f"Breakpoints: SOC={n_s}, T={n_t}, I={n_i}; "
        f"steps={n}, dt={dt_s / 60:.1f} min; diesel units={n_g}"
    )
    if build_only:
        if log_handle is not None:
            log_handle.close()
        return {
            "success": False,
            "status": "BUILD_ONLY",
            "time_s": 0.0,
            "message": "CPLEX native model built but not optimized.",
            "model_stats": model_stats,
        }

    started = time.time()
    try:
        m.solve()
    except Exception as exc:
        if log_handle is not None:
            log_handle.close()
        return {
            "success": False,
            "status": "SOLVER_ERROR",
            "time_s": time.time() - started,
            "message": str(exc),
            "model_stats": model_stats,
        }
    solve_time = time.time() - started
    if log_handle is not None:
        log_handle.close()

    try:
        objective_value = float(m.solution.get_objective_value())
    except Exception:
        return {
            "success": False,
            "status": str(m.solution.get_status_string()).upper().replace(" ", "_"),
            "time_s": solve_time,
            "message": "No feasible solution returned by CPLEX native backend.",
            "model_stats": model_stats,
        }

    def val(name: str) -> float:
        return float(m.solution.get_values(name))

    def arr(names: dict[int, str] | list[str]) -> np.ndarray:
        if isinstance(names, dict):
            ordered = [names[i] for i in range(len(names))]
        else:
            ordered = names
        if not ordered:
            return np.zeros(0, dtype=float)
        return np.asarray(m.solution.get_values(ordered), dtype=float)

    def arr1(var_dict: dict[int, str], length: int) -> np.ndarray:
        return np.asarray(m.solution.get_values([var_dict[i] for i in range(length)]), dtype=float)

    P_dg_units = np.array([[val(P_dg[g, t]) for t in range(n)] for g in range(n_g)], dtype=float)
    M_dg_units = np.array([[val(M_dg[g, t]) for t in range(n)] for g in range(n_g)], dtype=float)
    u_dg_units = np.array([[val(u_dg[g, t]) for t in range(n)] for g in range(n_g)], dtype=float)
    w_st_val = np.array([[val(w_st[t, a, b]) for a in range(n_s) for b in range(n_t)] for t in range(n)], dtype=float)
    try:
        reported_gap = float(m.solution.MIP.get_mip_relative_gap())
    except Exception:
        reported_gap = 0.0
    try:
        best_bound = float(m.solution.MIP.get_best_objective())
    except Exception:
        best_bound = objective_value
    try:
        node_count = float(m.solution.progress.get_num_nodes_processed())
    except Exception:
        node_count = 0.0

    result = {
        "success": True,
        "status": str(m.solution.get_status_string()).upper().replace(" ", "_"),
        "objective": objective_value,
        "gap": reported_gap,
        "best_bound": best_bound,
        "node_count": node_count,
        "time_s": solve_time,
        "model_stats": model_stats,
        "method": f"{bp.mode} perspective I2R MILP ({bp.n_s}x{bp.n_t}x{bp.n_i}) via CPLEX native SOS2",
        "hours": hours,
        "wind_m_s": np.asarray(data["wind_m_s"], dtype=float),
        "solar_w_m2": np.asarray(data["solar_w_m2"], dtype=float),
        "T_amb": t_amb,
        "load_kw": np.asarray(data["P_load"], dtype=float) / 1000.0,
        "pv_kw": np.asarray(data["P_pv"], dtype=float) / 1000.0,
        "wt_kw": np.asarray(data["P_wt"], dtype=float) / 1000.0,
        "I_bat": arr1(I_bat, n),
        "SOC": arr1(SOC, n),
        "SOC_end": val(SOC_end),
        "T_bat": arr1(T_bat, n),
        "T_tank": arr1(T_tank, n),
        "T_cont": arr1(T_cont, n),
        "T_bat_end": val(T_bat_end),
        "T_tank_end": val(T_tank_end),
        "T_cont_end": val(T_cont_end),
        "P_BESS": arr1(P_BESS, n),
        "P_dg_units": P_dg_units,
        "M_dg_units": M_dg_units,
        "u_dg": u_dg_units,
        "R0": np.array([float(w_st_val[t, :] @ r0_flat) for t in range(n)]),
        "P_dc_pack": arr1(P_dc_pack, n),
        "P_dc_abs": arr1(P_dc_abs, n),
        "Q_gen_pack": arr1(Q_gen_pack, n),
        "pv_use_kw": arr1(P_pv_use, n) / 1000.0,
        "wt_use_kw": arr1(P_wt_use, n) / 1000.0,
        "pv_curt_kw": arr1(P_pv_curt, n) / 1000.0,
        "wt_curt_kw": arr1(P_wt_curt, n) / 1000.0,
        "u_pi": arr1(u_pi, n),
        "u_po": arr1(u_po, n),
        "u_lh": arr1(u_lh, n),
        "u_ch": arr1(u_ch, n),
        "P_heat_liquid_w": arr1(P_heat_liquid_ctrl, n),
        "P_heat_cont_w": arr1(P_heat_cont_ctrl, n),
        "u_pi_change": arr1(u_pi_change, max(0, n - 1)),
        "u_po_change": arr1(u_po_change, max(0, n - 1)),
        "u_lh_change": arr1(u_lh_change, max(0, n - 1)),
        "u_ch_change": arr1(u_ch_change, max(0, n - 1)),
        "tank_temp_step": arr1(tank_temp_step, max(0, n - 1)),
        "cont_temp_step": arr1(cont_temp_step, max(0, n - 1)),
        "tank_preheat_short": arr1(tank_preheat_short, n),
        "cont_preheat_short": arr1(cont_preheat_short, n),
        "bat_low_dev": arr1(bat_low_dev, n),
        "bat_high_dev": arr1(bat_high_dev, n),
        "tank_low_dev": arr1(tank_low_dev, n),
        "tank_high_dev": arr1(tank_high_dev, n),
        "tank_hot_dev": arr1(tank_hot_dev, n),
        "cont_low_dev": arr1(cont_low_dev, n),
        "cont_high_dev": arr1(cont_high_dev, n),
        "cont_hot_dev": arr1(cont_hot_dev, n),
        "tank_terminal_short": val(tank_terminal_short),
        "cont_terminal_short": val(cont_terminal_short),
        "Q_bt": arr1(Q_bt, n),
        "Q_tamb": arr1(Q_tamb, n),
        "Q_tamb_dump": arr1(Q_tamb_dump, n),
        "renewable_surplus_kw": renewable_surplus_w / 1000.0,
        "renewable_surplus_weight": renewable_surplus_weight,
        "fuel_kg": objective_value,
        "curt_kwh": float(np.sum((arr1(P_pv_curt, n) + arr1(P_wt_curt, n)) * dt_s / 3600.0 / 1000.0)),
        "heat_kwh": float(np.sum((arr1(P_heat_liquid_ctrl, n) + arr1(P_heat_cont_ctrl, n)) * dt_s / 3600.0 / 1000.0)),
        "preheat_short_score": float(np.sum(renewable_surplus_weight * (arr1(tank_preheat_short, n) + arr1(cont_preheat_short, n)) * dt_s / 3600.0)),
        "bat_band_score": float(np.sum((arr1(bat_low_dev, n) + arr1(bat_high_dev, n)) * dt_s / 3600.0)),
        "tank_band_score": float(np.sum((arr1(tank_low_dev, n) + arr1(tank_high_dev, n)) * dt_s / 3600.0)),
        "cont_band_score": float(np.sum((arr1(cont_low_dev, n) + arr1(cont_high_dev, n)) * dt_s / 3600.0)),
        "hot_score": float(np.sum((arr1(tank_hot_dev, n) + arr1(cont_hot_dev, n)) * dt_s / 3600.0)),
        "terminal_heat_short_score": val(tank_terminal_short) + val(cont_terminal_short),
        "heat_dump_kwh_th": float(np.sum(arr1(Q_tamb_dump, n) * dt_s / 3.6e6)),
        "switch_score": float(np.sum(arr1(u_pi_change, max(0, n - 1)) + arr1(u_po_change, max(0, n - 1)) + arr1(u_lh_change, max(0, n - 1)) + arr1(u_ch_change, max(0, n - 1)))),
        "temp_ramp_score": float(np.sum(arr1(tank_temp_step, max(0, n - 1)) + arr1(cont_temp_step, max(0, n - 1)))),
        "dt_s": dt_s,
        "bp_soc": bp.soc,
        "bp_temp": bp.temp,
        "bp_r0": bp.r0,
        "bp_current": bp.current,
    }
    result["P_dg"] = result["P_dg_units"].sum(axis=0)
    result["M_dg"] = result["M_dg_units"].sum(axis=0)
    result["tank_storage_from_amb_kwh_th"] = p.C_tank * (result["T_tank"] - result["T_amb"]) / 3.6e6
    result["cont_storage_from_amb_kwh_th"] = p.C_cont * (result["T_cont"] - result["T_amb"]) / 3.6e6
    result["tank_storage_over_target_kwh_th"] = p.C_tank * np.maximum(0.0, result["T_tank"] - p.T_tank_target) / 3.6e6
    result["cont_storage_over_target_kwh_th"] = p.C_cont * np.maximum(0.0, result["T_cont"] - p.T_cont_target) / 3.6e6
    result["tank_useful_storage_kwh_th"] = p.C_tank * np.maximum(0.0, result["T_tank"] - p.T_tank_useful_min) / 3.6e6
    result["cont_useful_storage_kwh_th"] = p.C_cont * np.maximum(0.0, result["T_cont"] - p.T_cont_useful_min) / 3.6e6
    result["heat_dump_kw_th"] = result["Q_tamb_dump"] / 1000.0
    result["objective_breakdown"] = {
        "formal_objective": "fuel_kg_only",
        "fuel": float(result["fuel_kg"]),
        "nonfuel_terms_are_diagnostics_only": True,
        "solver_backend": "cplex_native",
        "solver_name": "cplex",
    }
    result["q_bc_w"] = p.K_b_cont * (result["T_bat"] - result["T_cont"])
    result["q_tc_w"] = p.K_t_cont * (result["T_tank"] - result["T_cont"])
    t_amb_prev = np.concatenate(([result["T_amb"][0]], result["T_amb"][:-1]))
    result["q_camb_w"] = p.K_cont_amb * (result["T_cont"] - t_amb_prev)
    result["P_heat_liquid_w"] = np.asarray(result["P_heat_liquid_w"], dtype=float)
    result["P_heat_cont_w"] = np.asarray(result["P_heat_cont_w"], dtype=float)
    result["P_pump_in_w"] = result["u_pi"] * p.P_pump_in
    result["P_pump_out_w"] = result["u_po"] * p.P_pump_out
    result["checks"] = compute_checks(p, result)
    return result


def model_stats_from_mosek_task(
    task,
    var_names: list[str],
    bp: Breakpoints,
    n_steps: int,
    p: SimpleNamespace,
    data: dict,
) -> dict[str, object]:
    vars_by_prefix: dict[str, int] = {}
    for name in var_names:
        prefix = name.split("[", 1)[0]
        vars_by_prefix[prefix] = vars_by_prefix.get(prefix, 0) + 1
    lam_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("lam"))
    w_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("w"))
    rho_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("rho_i2"))
    current_choice_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("z_current"))
    balance_slack_count = sum(count for name, count in vars_by_prefix.items() if name.startswith("P_balance_"))
    binary_count = int(task.getnumintvar())
    current_mode = str(getattr(p, "diag_current_mode", "continuous"))
    return {
        "solver_backend": "mosek_native",
        "solver_name": "mosek",
        "uses_cvxpy": False,
        "native_sos2": False,
        "sos2_formulation": "binary_adjacency",
        "experiment": getattr(p, "diag_experiment", "baseline"),
        "variables_total": int(task.getnumvar()),
        "binary_variables": binary_count,
        "continuous_variables": int(task.getnumvar() - binary_count),
        "constraints_total": int(task.getnumcon()),
        "sos_constraints": 0,
        "nonzeros": int(task.getnumanz()),
        "lam_variables_count": int(lam_count),
        "pwl_weight_variables_count": int(w_count),
        "rho_i2_variables_count": int(rho_count),
        "pwl_sos_or_lam_variables_count": int(lam_count + w_count + rho_count),
        "dg_ramp_constraints": 0,
        "strict_power_balance": True,
        "power_balance_slack_variables": int(balance_slack_count),
        "per_time_step_binary_count": float(binary_count) / max(1, n_steps),
        "thermal_binary_effective_count": int(n_steps * 4),
        "thermal_binary_block_steps": 1,
        "steps": int(n_steps),
        "dt_minutes": float(data["dt"]) / 60.0,
        "breakpoints_soc": int(bp.n_s),
        "soc_grid_width": float(np.min(np.diff(bp.soc))) if bp.n_s > 1 else 0.0,
        "breakpoints_temp": int(bp.n_t),
        "breakpoints_current": int(bp.n_i),
        "current_segments": int(max(0, bp.n_i - 1)),
        "current_mode": current_mode,
        "discrete_current": current_mode == "discrete",
        "current_discrete_variables": int(current_choice_count),
        "perspective_i2r": bool(getattr(p, "diag_perspective_i2r", True)),
        "strict_current_sos2": bool(getattr(p, "diag_strict_current_sos2", False)),
        "perspective_mode": "mosek_discrete_current" if current_mode == "discrete" else "mosek_binary_adjacency_sos2",
        "legacy_discrete_current_grid_override": bool(getattr(p, "diag_discrete_current", False)),
        "discrete_current_levels": int(getattr(p, "diag_discrete_current_levels", 0)),
        "discrete_current_c_max": float(getattr(p, "diag_discrete_current_c_max", 0.0)),
        "current_breakpoints_a": [float(x) for x in bp.current],
        **current_limit_stats(bp, p),
        "tight_temperature_bounds": bool(getattr(p, "diag_tight_temp_bounds", False)),
        "temperature_bounds": temperature_bounds_snapshot(p),
        "big_m_diagnostics": big_m_diagnostics(p, data),
    }


def solve_milp_mosek_native(
    p: SimpleNamespace,
    data: dict[str, np.ndarray | float | int],
    bp: Breakpoints,
    time_limit: float,
    mip_gap: float,
    build_only: bool = False,
) -> dict:
    try:
        import mosek
    except Exception as exc:
        return {
            "success": False,
            "status": "SOLVER_NOT_AVAILABLE",
            "time_s": 0.0,
            "message": f"MOSEK Python API is not available: {exc}",
        }

    n = int(data["N"])
    dt_s = float(data["dt"])
    hours = np.asarray(data["hours"], dtype=float)
    t_amb = np.asarray(data["T_amb"], dtype=float)
    renewable_surplus_w = np.maximum(
        0.0, np.asarray(data["P_pv"], dtype=float) + np.asarray(data["P_wt"], dtype=float) - np.asarray(data["P_load"], dtype=float)
    )
    renewable_surplus_max_w = float(np.max(renewable_surplus_w)) if renewable_surplus_w.size else 0.0
    renewable_surplus_weight = renewable_surplus_w / renewable_surplus_max_w if renewable_surplus_max_w > 1e-9 else np.zeros(n, dtype=float)

    n_s, n_t, n_i = bp.n_s, bp.n_t, bp.n_i
    dg_units = p.diesel_units
    n_g = len(dg_units)
    dg_points = [len(u["powers_w"]) for u in dg_units]
    pdc_pack_table = np.array(
        [
            [[bp.ocv[a] * bp.current[c] - bp.current2[c] * bp.r0[a, b] for c in range(n_i)] for b in range(n_t)]
            for a in range(n_s)
        ],
        dtype=float,
    )
    qgen_pack_table = np.array(
        [
            [[bp.current2[c] * bp.r0[a, b] for c in range(n_i)] for b in range(n_t)]
            for a in range(n_s)
        ],
        dtype=float,
    )
    r0_flat = bp.r0.reshape(-1)

    task = mosek.Task()
    task.putobjsense(mosek.objsense.minimize)
    if float(time_limit) > 0:
        task.putdouparam(mosek.dparam.optimizer_max_time, float(time_limit))
        try:
            task.putdouparam(mosek.dparam.mio_max_time, float(time_limit))
        except Exception:
            pass
    task.putdouparam(mosek.dparam.mio_tol_rel_gap, float(mip_gap))
    threads = int(getattr(p, "diag_threads", 0))
    if threads > 0:
        task.putintparam(mosek.iparam.num_threads, threads)
    if getattr(p, "diag_log_file", None):
        Path(p.diag_log_file).parent.mkdir(parents=True, exist_ok=True)
        for stream in [mosek.streamtype.log, mosek.streamtype.msg, mosek.streamtype.wrn, mosek.streamtype.err]:
            task.linkfiletostream(stream, str(p.diag_log_file), 1)

    inf = 1.0e30
    var_names: list[str] = []
    var_lbs: list[float] = []
    var_ubs: list[float] = []

    def bound_key(lb: float, ub: float):
        if lb <= -0.5 * inf and ub >= 0.5 * inf:
            return mosek.boundkey.fr, -inf, inf
        if lb <= -0.5 * inf:
            return mosek.boundkey.up, -inf, float(ub)
        if ub >= 0.5 * inf:
            return mosek.boundkey.lo, float(lb), inf
        if abs(float(lb) - float(ub)) <= 1e-12:
            return mosek.boundkey.fx, float(lb), float(ub)
        return mosek.boundkey.ra, float(lb), float(ub)

    def add_scalar(name: str, lb: float = 0.0, ub: float = inf, vtype: str = "C", obj: float = 0.0) -> int:
        idx = task.getnumvar()
        task.appendvars(1)
        var_names.append(name)
        var_lbs.append(float(lb))
        var_ubs.append(float(ub))
        try:
            task.putvarname(idx, name)
        except Exception:
            pass
        bk, bl, bu = bound_key(float(lb), float(ub))
        task.putvarbound(idx, bk, bl, bu)
        if abs(float(obj)) > 1e-12:
            task.putcj(idx, float(obj))
        if vtype == "B":
            task.putvartype(idx, mosek.variabletype.type_int)
        return idx

    def add_vars(keys, name_fn, lb_fn=0.0, ub_fn=inf, type_fn="C", obj_fn=0.0):
        result = {}
        for key in keys:
            result[key] = add_scalar(
                name_fn(key),
                lb_fn(key) if callable(lb_fn) else lb_fn,
                ub_fn(key) if callable(ub_fn) else ub_fn,
                type_fn(key) if callable(type_fn) else type_fn,
                obj_fn(key) if callable(obj_fn) else obj_fn,
            )
        return result

    rows: list[tuple[list[int], list[float], object, float, float, str]] = []

    def add_constr(terms: list[tuple[int, float]], sense: str, rhs: float, name: str) -> None:
        acc: dict[int, float] = {}
        for var, coef in terms:
            coef = float(coef)
            if abs(coef) <= 1e-12:
                continue
            acc[int(var)] = acc.get(int(var), 0.0) + coef
        ind = []
        val = []
        for var, coef in acc.items():
            if abs(coef) > 1e-12:
                ind.append(var)
                val.append(coef)
        if sense == "E":
            bk, bl, bu = mosek.boundkey.fx, float(rhs), float(rhs)
        elif sense == "L":
            bk, bl, bu = mosek.boundkey.up, -inf, float(rhs)
        elif sense == "G":
            bk, bl, bu = mosek.boundkey.lo, float(rhs), inf
        else:
            raise ValueError(f"Unsupported constraint sense {sense}")
        rows.append((ind, val, bk, bl, bu, name))

    def add_sos2_binary(lam_vars: list[int], active: float | int, name: str) -> None:
        if len(lam_vars) == 1:
            if isinstance(active, (int, np.integer)):
                add_constr([(lam_vars[0], 1.0), (int(active), -1.0)], "E", 0.0, f"{name}_single")
            else:
                add_constr([(lam_vars[0], 1.0)], "E", float(active), f"{name}_single")
            return
        z_vars = [add_scalar(f"z_sos2_{name}[{i}]", 0.0, 1.0, "B") for i in range(len(lam_vars) - 1)]
        if isinstance(active, (int, np.integer)):
            add_constr([(z, 1.0) for z in z_vars] + [(int(active), -1.0)], "E", 0.0, f"{name}_z_sum")
        else:
            add_constr([(z, 1.0) for z in z_vars], "E", float(active), f"{name}_z_sum")
        add_constr([(lam_vars[0], 1.0), (z_vars[0], -1.0)], "L", 0.0, f"{name}_lam0")
        for i in range(1, len(lam_vars) - 1):
            add_constr([(lam_vars[i], 1.0), (z_vars[i - 1], -1.0), (z_vars[i], -1.0)], "L", 0.0, f"{name}_lam{i}")
        add_constr([(lam_vars[-1], 1.0), (z_vars[-1], -1.0)], "L", 0.0, f"{name}_lam_last")

    I_bat = add_vars(range(n), lambda t: f"I_bat[{t}]", -p.I_charge_max, p.I_discharge_max)
    SOC = add_vars(range(n), lambda t: f"SOC[{t}]", p.SOC_min, p.SOC_max)
    T_bat = add_vars(range(n), lambda t: f"T_bat[{t}]", p.T_bat_min, p.T_bat_max)
    T_tank = add_vars(range(n), lambda t: f"T_tank[{t}]", p.T_tank_min, p.T_tank_max)
    T_cont = add_vars(range(n), lambda t: f"T_cont[{t}]", p.T_cont_min, p.T_cont_max)

    P_dg: dict[tuple[int, int], int] = {}
    M_dg: dict[tuple[int, int], int] = {}
    u_dg: dict[tuple[int, int], int] = {}
    lam_dg: dict[tuple[int, int, int], int] = {}
    for g, unit in enumerate(dg_units):
        P_dg.update(add_vars([(g, t) for t in range(n)], lambda key: f"P_dg[{key[0]},{key[1]}]", 0.0, float(unit["p_max_w"])))
        M_dg.update(
            add_vars(
                [(g, t) for t in range(n)],
                lambda key: f"M_dg[{key[0]},{key[1]}]",
                0.0,
                float(np.max(unit["fuel_kg_h"])),
                obj_fn=lambda key: dt_s / 3600.0,
            )
        )
        u_dg.update(add_vars([(g, t) for t in range(n)], lambda key: f"u_dg[{key[0]},{key[1]}]", 0.0, 1.0, "B"))
        lam_dg.update(
            add_vars(
                [(g, t, k) for t in range(n) for k in range(dg_points[g])],
                lambda key: f"lam_dg[{key[0]},{key[1]},{key[2]}]",
                0.0,
                1.0,
            )
        )

    P_BESS = add_vars(range(n), lambda t: f"P_BESS[{t}]", -350e3, 350e3)
    P_pv_use = add_vars(range(n), lambda t: f"P_pv_use[{t}]", 0.0, inf)
    P_wt_use = add_vars(range(n), lambda t: f"P_wt_use[{t}]", 0.0, inf)
    P_pv_curt = add_vars(range(n), lambda t: f"P_pv_curt[{t}]", 0.0, inf)
    P_wt_curt = add_vars(range(n), lambda t: f"P_wt_curt[{t}]", 0.0, inf)
    tank_preheat_short = add_vars(range(n), lambda t: f"tank_preheat_short[{t}]", 0.0, max(0.0, p.T_tank_target - p.T_tank_min))
    cont_preheat_short = add_vars(range(n), lambda t: f"cont_preheat_short[{t}]", 0.0, max(0.0, p.T_cont_target - p.T_cont_min))
    bat_low_dev = add_vars(range(n), lambda t: f"bat_low_dev[{t}]", 0.0, max(0.0, p.T_bat_pref_low - p.T_bat_min))
    bat_high_dev = add_vars(range(n), lambda t: f"bat_high_dev[{t}]", 0.0, max(0.0, p.T_bat_max - p.T_bat_pref_high))
    tank_low_dev = add_vars(range(n), lambda t: f"tank_low_dev[{t}]", 0.0, max(0.0, p.T_tank_band_low - p.T_tank_min))
    tank_high_dev = add_vars(range(n), lambda t: f"tank_high_dev[{t}]", 0.0, max(0.0, p.T_tank_max - p.T_tank_band_high))
    tank_hot_dev = add_vars(range(n), lambda t: f"tank_hot_dev[{t}]", 0.0, max(0.0, p.T_tank_max - p.T_tank_hot))
    cont_low_dev = add_vars(range(n), lambda t: f"cont_low_dev[{t}]", 0.0, max(0.0, p.T_cont_band_low - p.T_cont_min))
    cont_high_dev = add_vars(range(n), lambda t: f"cont_high_dev[{t}]", 0.0, max(0.0, p.T_cont_max - p.T_cont_band_high))
    cont_hot_dev = add_vars(range(n), lambda t: f"cont_hot_dev[{t}]", 0.0, max(0.0, p.T_cont_max - p.T_cont_hot))
    tank_terminal_short = add_scalar(
        "tank_terminal_short",
        0.0,
        terminal_deviation_ub(p.T_tank_min, p.T_tank_max, p.T_tank_init),
    )
    cont_terminal_short = add_scalar(
        "cont_terminal_short",
        0.0,
        terminal_deviation_ub(p.T_cont_min, p.T_cont_max, p.T_cont_init),
    )

    u_pi = add_vars(range(n), lambda t: f"u_pi[{t}]", 0.0, 1.0, "B")
    u_po = add_vars(range(n), lambda t: f"u_po[{t}]", 0.0, 1.0, "B")
    u_lh = add_vars(range(n), lambda t: f"u_lh[{t}]", 0.0, 1.0, "B")
    u_ch = add_vars(range(n), lambda t: f"u_ch[{t}]", 0.0, 1.0, "B")
    P_heat_liquid_ctrl = add_vars(range(n), lambda t: f"P_heat_liquid_ctrl[{t}]", 0.0, p.P_heat_liquid)
    P_heat_cont_ctrl = add_vars(range(n), lambda t: f"P_heat_cont_ctrl[{t}]", 0.0, p.P_heat_cont)
    p_heat_liquid_min_w = min(1.0, max(0.0, float(p.P_heat_liquid)))
    p_heat_cont_min_w = min(1.0, max(0.0, float(p.P_heat_cont)))
    u_pi_change = add_vars(range(max(0, n - 1)), lambda t: f"u_pi_change[{t}]", 0.0, 1.0)
    u_po_change = add_vars(range(max(0, n - 1)), lambda t: f"u_po_change[{t}]", 0.0, 1.0)
    u_lh_change = add_vars(range(max(0, n - 1)), lambda t: f"u_lh_change[{t}]", 0.0, 1.0)
    u_ch_change = add_vars(range(max(0, n - 1)), lambda t: f"u_ch_change[{t}]", 0.0, 1.0)
    tank_temp_step = add_vars(range(max(0, n - 1)), lambda t: f"tank_temp_step[{t}]", 0.0, p.T_tank_max - p.T_tank_min)
    cont_temp_step = add_vars(range(max(0, n - 1)), lambda t: f"cont_temp_step[{t}]", 0.0, p.T_cont_max - p.T_cont_min)

    dT_bt_lb = p.T_bat_min - p.T_tank_max
    dT_bt_ub = p.T_bat_max - p.T_tank_min
    dT_ta_lb = p.T_tank_min - float(np.max(t_amb))
    dT_ta_ub = p.T_tank_max - float(np.min(t_amb))
    Q_bt = add_vars(range(n), lambda t: f"Q_bt[{t}]", min(0.0, p.K_bt * dT_bt_lb), max(0.0, p.K_bt * dT_bt_ub))
    Q_tamb = add_vars(range(n), lambda t: f"Q_tamb[{t}]", min(0.0, p.K_t_amb * dT_ta_lb), max(0.0, p.K_t_amb * dT_ta_ub))
    Q_tamb_dump = add_vars(range(n), lambda t: f"Q_tamb_dump[{t}]", 0.0, max(0.0, p.K_t_amb * dT_ta_ub))
    v_bt = add_vars(range(n), lambda t: f"v_bt[{t}]", min(0.0, dT_bt_lb), max(0.0, dT_bt_ub))
    v_ta = add_vars(range(n), lambda t: f"v_ta[{t}]", min(0.0, dT_ta_lb), max(0.0, dT_ta_ub))

    lam_s = add_vars([(t, a) for t in range(n) for a in range(n_s)], lambda key: f"lam_s[{key[0]},{key[1]}]", 0.0, 1.0)
    lam_t = add_vars([(t, b) for t in range(n) for b in range(n_t)], lambda key: f"lam_t[{key[0]},{key[1]}]", 0.0, 1.0)
    pdc_lb = float(np.min(pdc_pack_table))
    pdc_ub = float(np.max(pdc_pack_table))
    qgen_ub = float(np.max(qgen_pack_table))
    P_dc_pack = add_vars(range(n), lambda t: f"P_dc_pack[{t}]", pdc_lb, pdc_ub)
    Q_gen_pack = add_vars(range(n), lambda t: f"Q_gen_pack[{t}]", 0.0, qgen_ub)
    P_dc_abs = add_vars(range(n), lambda t: f"P_dc_abs[{t}]", 0.0, max(abs(pdc_lb), abs(pdc_ub)))
    w_st = add_vars([(t, a, b) for t in range(n) for a in range(n_s) for b in range(n_t)], lambda key: f"w_st[{key[0]},{key[1]},{key[2]}]", 0.0, 1.0)
    rho_i2 = add_vars(
        [(t, a, b, c) for t in range(n) for a in range(n_s) for b in range(n_t) for c in range(n_i)],
        lambda key: f"rho_i2[{key[0]},{key[1]},{key[2]},{key[3]}]",
        0.0,
        1.0,
    )
    current_mode = str(getattr(p, "diag_current_mode", "continuous"))
    z_current = None
    if current_mode == "discrete":
        z_current = add_vars([(t, c) for t in range(n) for c in range(n_i)], lambda key: f"z_current[{key[0]},{key[1]}]", 0.0, 1.0, "B")

    for t in range(n):
        add_constr([(lam_s[t, a], 1.0) for a in range(n_s)], "E", 1.0, f"sum_lam_s_{t}")
        add_sos2_binary([lam_s[t, a] for a in range(n_s)], 1.0, f"s_{t}")
        add_constr([(lam_t[t, b], 1.0) for b in range(n_t)], "E", 1.0, f"sum_lam_t_{t}")
        add_sos2_binary([lam_t[t, b] for b in range(n_t)], 1.0, f"t_{t}")
        add_constr([(SOC[t], 1.0)] + [(lam_s[t, a], -float(bp.soc[a])) for a in range(n_s)], "E", 0.0, f"soc_map_{t}")
        add_constr([(T_bat[t], 1.0)] + [(lam_t[t, b], -float(bp.temp[b])) for b in range(n_t)], "E", 0.0, f"tbat_map_{t}")
        add_constr(
            [(I_bat[t], 1.0)] + [(lam_t[t, b], float(bp.charge_current_limit[b])) for b in range(n_t)],
            "G",
            0.0,
            f"ibat_charge_temp_limit_{t}",
        )
        add_constr(
            [(I_bat[t], 1.0)] + [(lam_t[t, b], -float(bp.discharge_current_limit[b])) for b in range(n_t)],
            "L",
            0.0,
            f"ibat_discharge_temp_limit_{t}",
        )
        for a in range(n_s):
            add_constr([(w_st[t, a, b], 1.0) for b in range(n_t)] + [(lam_s[t, a], -1.0)], "E", 0.0, f"w_st_s_{t}_{a}")
        for b in range(n_t):
            add_constr([(w_st[t, a, b], 1.0) for a in range(n_s)] + [(lam_t[t, b], -1.0)], "E", 0.0, f"w_st_t_{t}_{b}")
        for a in range(n_s):
            for b in range(n_t):
                add_constr(
                    [(rho_i2[t, a, b, c], 1.0) for c in range(n_i)] + [(w_st[t, a, b], -1.0)],
                    "E",
                    0.0,
                    f"rho_st_{t}_{a}_{b}",
                )
                if getattr(p, "diag_strict_current_sos2", False):
                    add_sos2_binary([rho_i2[t, a, b, c] for c in range(n_i)], w_st[t, a, b], f"rho_i2_{t}_{a}_{b}")
        if z_current is not None:
            add_constr([(z_current[t, c], 1.0) for c in range(n_i)], "E", 1.0, f"current_choice_sum_{t}")
            for c in range(n_i):
                add_constr(
                    [(rho_i2[t, a, b, c], 1.0) for a in range(n_s) for b in range(n_t)] + [(z_current[t, c], -1.0)],
                    "E",
                    0.0,
                    f"current_choice_link_{t}_{c}",
                )
        add_constr(
            [(I_bat[t], 1.0)] + [(rho_i2[t, a, b, c], -float(bp.current[c])) for a in range(n_s) for b in range(n_t) for c in range(n_i)],
            "E",
            0.0,
            f"ibat_perspective_map_{t}",
        )
        add_constr(
            [(P_dc_pack[t], 1.0)]
            + [(rho_i2[t, a, b, c], -float(pdc_pack_table[a, b, c])) for a in range(n_s) for b in range(n_t) for c in range(n_i)],
            "E",
            0.0,
            f"pdc_pack_map_{t}",
        )
        add_constr(
            [(Q_gen_pack[t], 1.0)]
            + [(rho_i2[t, a, b, c], -float(qgen_pack_table[a, b, c])) for a in range(n_s) for b in range(n_t) for c in range(n_i)],
            "E",
            0.0,
            f"qgen_pack_map_{t}",
        )
        for g, unit in enumerate(dg_units):
            add_constr([(lam_dg[g, t, k], 1.0) for k in range(dg_points[g])] + [(u_dg[g, t], -1.0)], "E", 0.0, f"dg_lam_sum_{g}_{t}")
            add_sos2_binary([lam_dg[g, t, k] for k in range(dg_points[g])], u_dg[g, t], f"dg_{g}_{t}")
            add_constr(
                [(P_dg[g, t], 1.0)] + [(lam_dg[g, t, k], -float(unit["powers_w"][k])) for k in range(dg_points[g])],
                "E",
                0.0,
                f"dg_p_map_{g}_{t}",
            )
            add_constr(
                [(M_dg[g, t], 1.0)] + [(lam_dg[g, t, k], -float(unit["fuel_kg_h"][k])) for k in range(dg_points[g])],
                "E",
                0.0,
                f"dg_f_map_{g}_{t}",
            )
            add_constr([(P_dg[g, t], 1.0), (u_dg[g, t], -float(unit["p_min_w"]))], "G", 0.0, f"dg_min_{g}_{t}")
            add_constr([(P_dg[g, t], 1.0), (u_dg[g, t], -float(unit["p_max_w"]))], "L", 0.0, f"dg_max_{g}_{t}")
        if n_g:
            add_constr([(u_dg[g, t], 1.0) for g in range(n_g)], "G", 1.0, f"dg_at_least_one_on_{t}")

        add_constr([(v_bt[t], 1.0), (u_pi[t], -dT_bt_lb)], "G", 0.0, f"vbt_lb0_{t}")
        add_constr([(v_bt[t], 1.0), (u_pi[t], -dT_bt_ub)], "L", 0.0, f"vbt_ub0_{t}")
        add_constr([(v_bt[t], 1.0), (T_bat[t], -1.0), (T_tank[t], 1.0), (u_pi[t], -dT_bt_ub)], "G", -dT_bt_ub, f"vbt_lb1_{t}")
        add_constr([(v_bt[t], 1.0), (T_bat[t], -1.0), (T_tank[t], 1.0), (u_pi[t], -dT_bt_lb)], "L", -dT_bt_lb, f"vbt_ub1_{t}")
        add_constr([(Q_bt[t], 1.0), (v_bt[t], -p.K_bt)], "E", 0.0, f"qbt_map_{t}")

        add_constr([(v_ta[t], 1.0), (u_po[t], -dT_ta_lb)], "G", 0.0, f"vta_lb0_{t}")
        add_constr([(v_ta[t], 1.0), (u_po[t], -dT_ta_ub)], "L", 0.0, f"vta_ub0_{t}")
        add_constr([(v_ta[t], 1.0), (T_tank[t], -1.0), (u_po[t], -dT_ta_ub)], "G", -float(t_amb[t]) - dT_ta_ub, f"vta_lb1_{t}")
        add_constr([(v_ta[t], 1.0), (T_tank[t], -1.0), (u_po[t], -dT_ta_lb)], "L", -float(t_amb[t]) - dT_ta_lb, f"vta_ub1_{t}")
        add_constr([(Q_tamb[t], 1.0), (v_ta[t], -p.K_t_amb)], "E", 0.0, f"qtamb_map_{t}")
        add_constr([(Q_tamb_dump[t], 1.0), (Q_tamb[t], -1.0)], "G", 0.0, f"qtamb_dump_pos_{t}")

        add_constr([(P_heat_liquid_ctrl[t], 1.0), (u_lh[t], -p.P_heat_liquid)], "L", 0.0, f"p_heat_liquid_ctrl_max_{t}")
        add_constr([(P_heat_cont_ctrl[t], 1.0), (u_ch[t], -p.P_heat_cont)], "L", 0.0, f"p_heat_cont_ctrl_max_{t}")
        add_constr([(P_heat_liquid_ctrl[t], 1.0), (u_lh[t], -p_heat_liquid_min_w)], "G", 0.0, f"p_heat_liquid_ctrl_min_{t}")
        add_constr([(P_heat_cont_ctrl[t], 1.0), (u_ch[t], -p_heat_cont_min_w)], "G", 0.0, f"p_heat_cont_ctrl_min_{t}")
        aux_terms = [
            (u_pi[t], p.P_pump_in),
            (u_po[t], p.P_pump_out),
            (P_heat_liquid_ctrl[t], 1.0),
            (P_heat_cont_ctrl[t], 1.0),
        ]
        add_constr([(P_dc_abs[t], 1.0), (P_dc_pack[t], -1.0)], "G", 0.0, f"pdc_abs_pos_{t}")
        add_constr([(P_dc_abs[t], 1.0), (P_dc_pack[t], 1.0)], "G", 0.0, f"pdc_abs_neg_{t}")
        add_constr([(P_BESS[t], 1.0), (P_dc_pack[t], -1.0), (P_dc_abs[t], p.mu_pcs)] + aux_terms, "E", 0.0, f"pbess_{t}")
        if getattr(p, "diag_disable_storage_dispatch", False):
            add_constr([(I_bat[t], 1.0)], "E", 0.0, f"no_storage_current_{t}")

        add_constr([(P_pv_use[t], 1.0), (P_pv_curt[t], 1.0)], "E", float(data["P_pv"][t]), f"pv_split_{t}")
        add_constr([(P_wt_use[t], 1.0), (P_wt_curt[t], 1.0)], "E", float(data["P_wt"][t]), f"wt_split_{t}")
        add_constr(
            [(P_pv_use[t], 1.0), (P_wt_use[t], 1.0), (P_BESS[t], 1.0)] + [(P_dg[g, t], 1.0) for g in range(n_g)],
            "E",
            float(data["P_load"][t]),
            f"power_balance_{t}",
        )
        add_constr([(tank_preheat_short[t], 1.0), (T_tank[t], 1.0)], "G", p.T_tank_target, f"tank_preheat_short_{t}")
        add_constr([(cont_preheat_short[t], 1.0), (T_cont[t], 1.0)], "G", p.T_cont_target, f"cont_preheat_short_{t}")
        add_constr([(bat_low_dev[t], 1.0), (T_bat[t], 1.0)], "G", p.T_bat_pref_low, f"bat_low_dev_{t}")
        add_constr([(bat_high_dev[t], 1.0), (T_bat[t], -1.0)], "G", -p.T_bat_pref_high, f"bat_high_dev_{t}")
        add_constr([(tank_low_dev[t], 1.0), (T_tank[t], 1.0)], "G", p.T_tank_band_low, f"tank_low_dev_{t}")
        add_constr([(tank_high_dev[t], 1.0), (T_tank[t], -1.0)], "G", -p.T_tank_band_high, f"tank_high_dev_{t}")
        add_constr([(tank_hot_dev[t], 1.0), (T_tank[t], -1.0)], "G", -p.T_tank_hot, f"tank_hot_dev_{t}")
        add_constr([(cont_low_dev[t], 1.0), (T_cont[t], 1.0)], "G", p.T_cont_band_low, f"cont_low_dev_{t}")
        add_constr([(cont_high_dev[t], 1.0), (T_cont[t], -1.0)], "G", -p.T_cont_band_high, f"cont_high_dev_{t}")
        add_constr([(cont_hot_dev[t], 1.0), (T_cont[t], -1.0)], "G", -p.T_cont_hot, f"cont_hot_dev_{t}")

    add_constr([(SOC[0], 1.0)], "E", p.SOC_init, "soc_initial")
    add_constr([(T_bat[0], 1.0)], "E", p.T_bat_init, "tbat_initial")
    add_constr([(T_tank[0], 1.0)], "E", p.T_tank_init, "ttank_initial")
    add_constr([(T_cont[0], 1.0)], "E", p.T_cont_init, "tcont_initial")

    for t in range(1, n):
        prev = t - 1
        add_constr([(SOC[t], 1.0), (SOC[prev], -1.0), (I_bat[prev], dt_s / (p.Q_nom * 3600.0))], "E", 0.0, f"soc_dyn_{t}")
        add_constr([(u_pi_change[prev], 1.0), (u_pi[t], -1.0), (u_pi[prev], 1.0)], "G", 0.0, f"upi_change_pos_{t}")
        add_constr([(u_pi_change[prev], 1.0), (u_pi[prev], -1.0), (u_pi[t], 1.0)], "G", 0.0, f"upi_change_neg_{t}")
        add_constr([(u_po_change[prev], 1.0), (u_po[t], -1.0), (u_po[prev], 1.0)], "G", 0.0, f"upo_change_pos_{t}")
        add_constr([(u_po_change[prev], 1.0), (u_po[prev], -1.0), (u_po[t], 1.0)], "G", 0.0, f"upo_change_neg_{t}")
        add_constr([(u_lh_change[prev], 1.0), (u_lh[t], -1.0), (u_lh[prev], 1.0)], "G", 0.0, f"ulh_change_pos_{t}")
        add_constr([(u_lh_change[prev], 1.0), (u_lh[prev], -1.0), (u_lh[t], 1.0)], "G", 0.0, f"ulh_change_neg_{t}")
        add_constr([(u_ch_change[prev], 1.0), (u_ch[t], -1.0), (u_ch[prev], 1.0)], "G", 0.0, f"uch_change_pos_{t}")
        add_constr([(u_ch_change[prev], 1.0), (u_ch[prev], -1.0), (u_ch[t], 1.0)], "G", 0.0, f"uch_change_neg_{t}")
        add_constr([(tank_temp_step[prev], 1.0), (T_tank[t], -1.0), (T_tank[prev], 1.0)], "G", 0.0, f"tank_temp_step_pos_{t}")
        add_constr([(tank_temp_step[prev], 1.0), (T_tank[prev], -1.0), (T_tank[t], 1.0)], "G", 0.0, f"tank_temp_step_neg_{t}")
        add_constr([(cont_temp_step[prev], 1.0), (T_cont[t], -1.0), (T_cont[prev], 1.0)], "G", 0.0, f"cont_temp_step_pos_{t}")
        add_constr([(cont_temp_step[prev], 1.0), (T_cont[prev], -1.0), (T_cont[t], 1.0)], "G", 0.0, f"cont_temp_step_neg_{t}")
        add_constr(
            [
                (T_bat[t], 1.0),
                (T_bat[prev], -1.0),
                (Q_gen_pack[prev], -dt_s / p.C_bat),
                (Q_bt[prev], dt_s / p.C_bat),
                (T_bat[t], p.K_b_cont * dt_s / p.C_bat),
                (T_cont[t], -p.K_b_cont * dt_s / p.C_bat),
            ],
            "E",
            0.0,
            f"tbat_dyn_{t}",
        )
        add_constr(
            [
                (T_tank[t], 1.0),
                (T_tank[prev], -1.0),
                (Q_bt[prev], -dt_s / p.C_tank),
                (P_heat_liquid_ctrl[prev], -dt_s / p.C_tank),
                (Q_tamb[prev], dt_s / p.C_tank),
                (T_tank[t], p.K_t_cont * dt_s / p.C_tank),
                (T_cont[t], -p.K_t_cont * dt_s / p.C_tank),
            ],
            "E",
            0.0,
            f"ttank_dyn_{t}",
        )
        add_constr(
            [
                (T_cont[t], 1.0),
                (T_cont[prev], -1.0),
                (T_bat[t], -p.K_b_cont * dt_s / p.C_cont),
                (T_cont[t], p.K_b_cont * dt_s / p.C_cont),
                (T_tank[t], -p.K_t_cont * dt_s / p.C_cont),
                (T_cont[t], p.K_t_cont * dt_s / p.C_cont),
                (P_heat_cont_ctrl[prev], -dt_s / p.C_cont),
                (T_cont[t], p.K_cont_amb * dt_s / p.C_cont),
            ],
            "E",
            p.K_cont_amb * float(t_amb[prev]) * dt_s / p.C_cont,
            f"tcont_dyn_{t}",
        )

    SOC_end = add_scalar("SOC_end", p.SOC_min, p.SOC_max)
    add_constr([(SOC_end, 1.0), (SOC[n - 1], -1.0), (I_bat[n - 1], dt_s / (p.Q_nom * 3600.0))], "E", 0.0, "soc_terminal")
    soc_dev = add_scalar("soc_dev", 0.0, terminal_deviation_ub(p.SOC_min, p.SOC_max, p.SOC_init))
    add_constr([(SOC[n - 1], 1.0)], "E", p.SOC_init, "soc_schedule_terminal_target")
    add_constr([(soc_dev, 1.0), (SOC_end, -1.0)], "G", -p.SOC_init, "soc_dev_pos")
    add_constr([(soc_dev, 1.0), (SOC_end, 1.0)], "G", p.SOC_init, "soc_dev_neg")
    T_bat_end = add_scalar("T_bat_end", p.T_bat_min, p.T_bat_max)
    T_tank_end = add_scalar("T_tank_end", p.T_tank_min, p.T_tank_max)
    T_cont_end = add_scalar("T_cont_end", p.T_cont_min, p.T_cont_max)
    last = n - 1
    add_constr(
        [
            (T_bat_end, 1.0),
            (T_bat[last], -1.0),
            (Q_gen_pack[last], -dt_s / p.C_bat),
            (Q_bt[last], dt_s / p.C_bat),
            (T_bat_end, p.K_b_cont * dt_s / p.C_bat),
            (T_cont_end, -p.K_b_cont * dt_s / p.C_bat),
        ],
        "E",
        0.0,
        "tbat_terminal",
    )
    add_constr(
        [
            (T_tank_end, 1.0),
            (T_tank[last], -1.0),
            (Q_bt[last], -dt_s / p.C_tank),
            (P_heat_liquid_ctrl[last], -dt_s / p.C_tank),
            (Q_tamb[last], dt_s / p.C_tank),
            (T_tank_end, p.K_t_cont * dt_s / p.C_tank),
            (T_cont_end, -p.K_t_cont * dt_s / p.C_tank),
        ],
        "E",
        0.0,
        "ttank_terminal",
    )
    add_constr(
        [
            (T_cont_end, 1.0),
            (T_cont[last], -1.0),
            (T_bat_end, -p.K_b_cont * dt_s / p.C_cont),
            (T_cont_end, p.K_b_cont * dt_s / p.C_cont),
            (T_tank_end, -p.K_t_cont * dt_s / p.C_cont),
            (T_cont_end, p.K_t_cont * dt_s / p.C_cont),
            (P_heat_cont_ctrl[last], -dt_s / p.C_cont),
            (T_cont_end, p.K_cont_amb * dt_s / p.C_cont),
        ],
        "E",
        p.K_cont_amb * float(t_amb[last]) * dt_s / p.C_cont,
        "tcont_terminal",
    )
    add_constr([(T_tank[n - 1], 1.0)], "E", p.T_tank_init, "ttank_schedule_terminal_target")
    add_constr([(T_cont[n - 1], 1.0)], "E", p.T_cont_init, "tcont_schedule_terminal_target")
    add_constr([(tank_terminal_short, 1.0), (T_tank_end, -1.0)], "G", -p.T_tank_init, "tank_terminal_dev_pos")
    add_constr([(tank_terminal_short, 1.0), (T_tank_end, 1.0)], "G", p.T_tank_init, "tank_terminal_dev_neg")
    add_constr([(cont_terminal_short, 1.0), (T_cont_end, -1.0)], "G", -p.T_cont_init, "cont_terminal_dev_pos")
    add_constr([(cont_terminal_short, 1.0), (T_cont_end, 1.0)], "G", p.T_cont_init, "cont_terminal_dev_neg")

    task.appendcons(len(rows))
    for row_idx, (ind, val, bk, bl, bu, name) in enumerate(rows):
        if ind:
            task.putarow(row_idx, ind, val)
        task.putconbound(row_idx, bk, bl, bu)
        try:
            task.putconname(row_idx, name)
        except Exception:
            pass

    model_stats = model_stats_from_mosek_task(task, var_names, bp, n, p, data)
    if getattr(p, "diag_model_stats_json", None):
        Path(p.diag_model_stats_json).parent.mkdir(parents=True, exist_ok=True)
        Path(p.diag_model_stats_json).write_text(json.dumps(model_stats, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        f"Model size: {model_stats['variables_total']} vars ({model_stats['binary_variables']} binaries), "
        f"{model_stats['constraints_total']} constraints, {model_stats['sos_constraints']} native SOS, "
        f"{model_stats['nonzeros']} nonzeros"
    )
    print(
        f"Breakpoints: SOC={n_s}, T={n_t}, I={n_i}; "
        f"steps={n}, dt={dt_s / 60:.1f} min; diesel units={n_g}"
    )
    if build_only:
        return {
            "success": False,
            "status": "BUILD_ONLY",
            "time_s": 0.0,
            "message": "MOSEK native model built but not optimized.",
            "model_stats": model_stats,
        }

    started = time.time()
    try:
        task.optimize()
    except Exception as exc:
        return {
            "success": False,
            "status": "SOLVER_ERROR",
            "time_s": time.time() - started,
            "message": str(exc),
            "model_stats": model_stats,
        }
    solve_time = time.time() - started
    try:
        solsta = task.getsolsta(mosek.soltype.itg)
    except Exception as exc:
        return {
            "success": False,
            "status": "NO_INTEGER_SOLUTION",
            "time_s": solve_time,
            "message": str(exc),
            "model_stats": model_stats,
        }
    feasible_states = {mosek.solsta.integer_optimal, mosek.solsta.prim_feas, mosek.solsta.optimal}
    status_name = str(solsta).split(".")[-1].upper()
    if solsta not in feasible_states:
        return {
            "success": False,
            "status": status_name,
            "time_s": solve_time,
            "message": "No feasible integer solution returned by MOSEK native backend.",
            "model_stats": model_stats,
        }
    xx = [0.0] * task.getnumvar()
    task.getxx(mosek.soltype.itg, xx)

    def val(idx: int) -> float:
        return float(xx[int(idx)])

    def arr1(var_dict: dict[int, int], length: int) -> np.ndarray:
        return np.asarray([val(var_dict[i]) for i in range(length)], dtype=float)

    P_dg_units = np.array([[val(P_dg[g, t]) for t in range(n)] for g in range(n_g)], dtype=float)
    M_dg_units = np.array([[val(M_dg[g, t]) for t in range(n)] for g in range(n_g)], dtype=float)
    u_dg_units = np.array([[val(u_dg[g, t]) for t in range(n)] for g in range(n_g)], dtype=float)
    w_st_val = np.array([[val(w_st[t, a, b]) for a in range(n_s) for b in range(n_t)] for t in range(n)], dtype=float)
    try:
        objective_value = float(task.getprimalobj(mosek.soltype.itg))
    except Exception:
        objective_value = float(sum(val(M_dg[g, t]) * dt_s / 3600.0 for g in range(n_g) for t in range(n)))
    try:
        reported_gap = float(task.getdouinf(mosek.dinfitem.mio_obj_rel_gap))
    except Exception:
        reported_gap = 0.0
    try:
        best_bound = float(task.getdouinf(mosek.dinfitem.mio_obj_bound))
    except Exception:
        best_bound = objective_value
    try:
        node_count = float(task.getintinf(mosek.iinfitem.mio_num_solved_nodes))
    except Exception:
        node_count = 0.0

    result = {
        "success": True,
        "status": status_name,
        "objective": objective_value,
        "gap": reported_gap,
        "best_bound": best_bound,
        "node_count": node_count,
        "time_s": solve_time,
        "model_stats": model_stats,
        "method": f"{bp.mode} perspective I2R MILP ({bp.n_s}x{bp.n_t}x{bp.n_i}) via MOSEK native",
        "hours": hours,
        "wind_m_s": np.asarray(data["wind_m_s"], dtype=float),
        "solar_w_m2": np.asarray(data["solar_w_m2"], dtype=float),
        "T_amb": t_amb,
        "load_kw": np.asarray(data["P_load"], dtype=float) / 1000.0,
        "pv_kw": np.asarray(data["P_pv"], dtype=float) / 1000.0,
        "wt_kw": np.asarray(data["P_wt"], dtype=float) / 1000.0,
        "I_bat": arr1(I_bat, n),
        "SOC": arr1(SOC, n),
        "SOC_end": val(SOC_end),
        "T_bat": arr1(T_bat, n),
        "T_tank": arr1(T_tank, n),
        "T_cont": arr1(T_cont, n),
        "T_bat_end": val(T_bat_end),
        "T_tank_end": val(T_tank_end),
        "T_cont_end": val(T_cont_end),
        "P_BESS": arr1(P_BESS, n),
        "P_dg_units": P_dg_units,
        "M_dg_units": M_dg_units,
        "u_dg": u_dg_units,
        "R0": np.array([float(w_st_val[t, :] @ r0_flat) for t in range(n)]),
        "P_dc_pack": arr1(P_dc_pack, n),
        "P_dc_abs": arr1(P_dc_abs, n),
        "Q_gen_pack": arr1(Q_gen_pack, n),
        "pv_use_kw": arr1(P_pv_use, n) / 1000.0,
        "wt_use_kw": arr1(P_wt_use, n) / 1000.0,
        "pv_curt_kw": arr1(P_pv_curt, n) / 1000.0,
        "wt_curt_kw": arr1(P_wt_curt, n) / 1000.0,
        "u_pi": arr1(u_pi, n),
        "u_po": arr1(u_po, n),
        "u_lh": arr1(u_lh, n),
        "u_ch": arr1(u_ch, n),
        "P_heat_liquid_w": arr1(P_heat_liquid_ctrl, n),
        "P_heat_cont_w": arr1(P_heat_cont_ctrl, n),
        "u_pi_change": arr1(u_pi_change, max(0, n - 1)),
        "u_po_change": arr1(u_po_change, max(0, n - 1)),
        "u_lh_change": arr1(u_lh_change, max(0, n - 1)),
        "u_ch_change": arr1(u_ch_change, max(0, n - 1)),
        "tank_temp_step": arr1(tank_temp_step, max(0, n - 1)),
        "cont_temp_step": arr1(cont_temp_step, max(0, n - 1)),
        "tank_preheat_short": arr1(tank_preheat_short, n),
        "cont_preheat_short": arr1(cont_preheat_short, n),
        "bat_low_dev": arr1(bat_low_dev, n),
        "bat_high_dev": arr1(bat_high_dev, n),
        "tank_low_dev": arr1(tank_low_dev, n),
        "tank_high_dev": arr1(tank_high_dev, n),
        "tank_hot_dev": arr1(tank_hot_dev, n),
        "cont_low_dev": arr1(cont_low_dev, n),
        "cont_high_dev": arr1(cont_high_dev, n),
        "cont_hot_dev": arr1(cont_hot_dev, n),
        "tank_terminal_short": val(tank_terminal_short),
        "cont_terminal_short": val(cont_terminal_short),
        "Q_bt": arr1(Q_bt, n),
        "Q_tamb": arr1(Q_tamb, n),
        "Q_tamb_dump": arr1(Q_tamb_dump, n),
        "renewable_surplus_kw": renewable_surplus_w / 1000.0,
        "renewable_surplus_weight": renewable_surplus_weight,
        "fuel_kg": objective_value,
        "curt_kwh": float(np.sum((arr1(P_pv_curt, n) + arr1(P_wt_curt, n)) * dt_s / 3600.0 / 1000.0)),
        "heat_kwh": float(np.sum((arr1(P_heat_liquid_ctrl, n) + arr1(P_heat_cont_ctrl, n)) * dt_s / 3600.0 / 1000.0)),
        "preheat_short_score": float(np.sum(renewable_surplus_weight * (arr1(tank_preheat_short, n) + arr1(cont_preheat_short, n)) * dt_s / 3600.0)),
        "bat_band_score": float(np.sum((arr1(bat_low_dev, n) + arr1(bat_high_dev, n)) * dt_s / 3600.0)),
        "tank_band_score": float(np.sum((arr1(tank_low_dev, n) + arr1(tank_high_dev, n)) * dt_s / 3600.0)),
        "cont_band_score": float(np.sum((arr1(cont_low_dev, n) + arr1(cont_high_dev, n)) * dt_s / 3600.0)),
        "hot_score": float(np.sum((arr1(tank_hot_dev, n) + arr1(cont_hot_dev, n)) * dt_s / 3600.0)),
        "terminal_heat_short_score": val(tank_terminal_short) + val(cont_terminal_short),
        "heat_dump_kwh_th": float(np.sum(arr1(Q_tamb_dump, n) * dt_s / 3.6e6)),
        "switch_score": float(np.sum(arr1(u_pi_change, max(0, n - 1)) + arr1(u_po_change, max(0, n - 1)) + arr1(u_lh_change, max(0, n - 1)) + arr1(u_ch_change, max(0, n - 1)))),
        "temp_ramp_score": float(np.sum(arr1(tank_temp_step, max(0, n - 1)) + arr1(cont_temp_step, max(0, n - 1)))),
        "dt_s": dt_s,
        "bp_soc": bp.soc,
        "bp_temp": bp.temp,
        "bp_r0": bp.r0,
        "bp_current": bp.current,
    }
    result["P_dg"] = result["P_dg_units"].sum(axis=0)
    result["M_dg"] = result["M_dg_units"].sum(axis=0)
    result["tank_storage_from_amb_kwh_th"] = p.C_tank * (result["T_tank"] - result["T_amb"]) / 3.6e6
    result["cont_storage_from_amb_kwh_th"] = p.C_cont * (result["T_cont"] - result["T_amb"]) / 3.6e6
    result["tank_storage_over_target_kwh_th"] = p.C_tank * np.maximum(0.0, result["T_tank"] - p.T_tank_target) / 3.6e6
    result["cont_storage_over_target_kwh_th"] = p.C_cont * np.maximum(0.0, result["T_cont"] - p.T_cont_target) / 3.6e6
    result["tank_useful_storage_kwh_th"] = p.C_tank * np.maximum(0.0, result["T_tank"] - p.T_tank_useful_min) / 3.6e6
    result["cont_useful_storage_kwh_th"] = p.C_cont * np.maximum(0.0, result["T_cont"] - p.T_cont_useful_min) / 3.6e6
    result["heat_dump_kw_th"] = result["Q_tamb_dump"] / 1000.0
    result["objective_breakdown"] = {
        "formal_objective": "fuel_kg_only",
        "fuel": float(result["fuel_kg"]),
        "nonfuel_terms_are_diagnostics_only": True,
        "solver_backend": "mosek_native",
        "solver_name": "mosek",
    }
    result["q_bc_w"] = p.K_b_cont * (result["T_bat"] - result["T_cont"])
    result["q_tc_w"] = p.K_t_cont * (result["T_tank"] - result["T_cont"])
    t_amb_prev = np.concatenate(([result["T_amb"][0]], result["T_amb"][:-1]))
    result["q_camb_w"] = p.K_cont_amb * (result["T_cont"] - t_amb_prev)
    result["P_heat_liquid_w"] = np.asarray(result["P_heat_liquid_w"], dtype=float)
    result["P_heat_cont_w"] = np.asarray(result["P_heat_cont_w"], dtype=float)
    result["P_pump_in_w"] = result["u_pi"] * p.P_pump_in
    result["P_pump_out_w"] = result["u_po"] * p.P_pump_out
    result["checks"] = compute_checks(p, result)
    return result


def solve_milp(
    p: SimpleNamespace,
    data: dict[str, np.ndarray | float | int],
    bp: Breakpoints,
    time_limit: float,
    mip_gap: float,
    feasibility_focus: bool = True,
    build_only: bool = False,
) -> dict:
    import gurobipy as gp
    from gurobipy import GRB

    n = int(data["N"])
    dt_s = float(data["dt"])
    hours = np.asarray(data["hours"], dtype=float)
    t_amb = np.asarray(data["T_amb"], dtype=float)
    renewable_surplus_w = np.maximum(
        0.0, np.asarray(data["P_pv"], dtype=float) + np.asarray(data["P_wt"], dtype=float) - np.asarray(data["P_load"], dtype=float)
    )
    renewable_surplus_max_w = float(np.max(renewable_surplus_w)) if renewable_surplus_w.size else 0.0
    if renewable_surplus_max_w > 1e-9:
        renewable_surplus_weight = renewable_surplus_w / renewable_surplus_max_w
    else:
        renewable_surplus_weight = np.zeros(n, dtype=float)

    n_s, n_t, n_i = bp.n_s, bp.n_t, bp.n_i
    dg_units = p.diesel_units
    n_g = len(dg_units)
    dg_points = [len(u["powers_w"]) for u in dg_units]

    pdc_pack_table = np.array(
        [
            [[bp.ocv[a] * bp.current[c] - bp.current2[c] * bp.r0[a, b] for c in range(n_i)] for b in range(n_t)]
            for a in range(n_s)
        ]
    )
    qgen_pack_table = np.array(
        [
            [[bp.current2[c] * bp.r0[a, b] for c in range(n_i)] for b in range(n_t)]
            for a in range(n_s)
        ]
    )

    m = gp.Model("BESS_MILP_perspective_i2r_20260530")
    if float(time_limit) > 0:
        m.setParam("TimeLimit", float(time_limit))
    m.setParam("MIPGap", float(mip_gap))
    m.setParam("OutputFlag", 1)
    m.setParam("Threads", int(getattr(p, "diag_threads", 4)))
    if getattr(p, "diag_log_file", None):
        m.setParam("LogFile", str(p.diag_log_file))
    if getattr(p, "diag_cuts", None) is not None:
        m.setParam("Cuts", int(p.diag_cuts))
    if feasibility_focus:
        m.setParam("MIPFocus", int(getattr(p, "diag_mip_focus", 1)))
        m.setParam("Heuristics", float(getattr(p, "diag_heuristics", 0.45)))
        if float(time_limit) > 0:
            m.setParam("NoRelHeurTime", min(60.0, max(5.0, 0.05 * float(time_limit))))
        else:
            m.setParam("NoRelHeurTime", 60.0)
        m.setParam("PumpPasses", 50)

    I_bat = m.addVars(n, lb=-p.I_charge_max, ub=p.I_discharge_max, name="I_bat")
    SOC = m.addVars(n, lb=p.SOC_min, ub=p.SOC_max, name="SOC")
    T_bat = m.addVars(n, lb=p.T_bat_min, ub=p.T_bat_max, name="T_bat")
    T_tank = m.addVars(n, lb=p.T_tank_min, ub=p.T_tank_max, name="T_tank")
    T_cont = m.addVars(n, lb=p.T_cont_min, ub=p.T_cont_max, name="T_cont")

    P_dg = {}
    M_dg = {}
    u_dg = {}
    lam_dg = {}
    for g, unit in enumerate(dg_units):
        for t in range(n):
            P_dg[g, t] = m.addVar(lb=0.0, ub=float(unit["p_max_w"]), name=f"P_dg_{g+1}_{t}")
            M_dg[g, t] = m.addVar(lb=0.0, ub=float(np.max(unit["fuel_kg_h"])), name=f"M_dg_{g+1}_{t}")
            u_dg[g, t] = m.addVar(vtype=GRB.BINARY, name=f"u_dg_{g+1}_{t}")
            for k in range(dg_points[g]):
                lam_dg[g, t, k] = m.addVar(lb=0.0, ub=1.0, name=f"lam_dg_{g+1}_{t}_{k}")

    P_BESS = m.addVars(n, lb=-350e3, ub=350e3, name="P_BESS")
    P_pv_use = m.addVars(n, lb=0.0, name="P_pv_use")
    P_wt_use = m.addVars(n, lb=0.0, name="P_wt_use")
    P_pv_curt = m.addVars(n, lb=0.0, name="P_pv_curt")
    P_wt_curt = m.addVars(n, lb=0.0, name="P_wt_curt")
    tank_preheat_short = m.addVars(
        n, lb=0.0, ub=max(0.0, p.T_tank_target - p.T_tank_min), name="tank_preheat_short"
    )
    cont_preheat_short = m.addVars(
        n, lb=0.0, ub=max(0.0, p.T_cont_target - p.T_cont_min), name="cont_preheat_short"
    )
    bat_low_dev = m.addVars(n, lb=0.0, ub=max(0.0, p.T_bat_pref_low - p.T_bat_min), name="bat_low_dev")
    bat_high_dev = m.addVars(n, lb=0.0, ub=max(0.0, p.T_bat_max - p.T_bat_pref_high), name="bat_high_dev")
    tank_low_dev = m.addVars(n, lb=0.0, ub=max(0.0, p.T_tank_band_low - p.T_tank_min), name="tank_low_dev")
    tank_high_dev = m.addVars(n, lb=0.0, ub=max(0.0, p.T_tank_max - p.T_tank_band_high), name="tank_high_dev")
    tank_hot_dev = m.addVars(n, lb=0.0, ub=max(0.0, p.T_tank_max - p.T_tank_hot), name="tank_hot_dev")
    cont_low_dev = m.addVars(n, lb=0.0, ub=max(0.0, p.T_cont_band_low - p.T_cont_min), name="cont_low_dev")
    cont_high_dev = m.addVars(n, lb=0.0, ub=max(0.0, p.T_cont_max - p.T_cont_band_high), name="cont_high_dev")
    cont_hot_dev = m.addVars(n, lb=0.0, ub=max(0.0, p.T_cont_max - p.T_cont_hot), name="cont_hot_dev")
    tank_terminal_short = m.addVar(
        lb=0.0,
        ub=terminal_deviation_ub(p.T_tank_min, p.T_tank_max, p.T_tank_init),
        name="tank_terminal_short",
    )
    cont_terminal_short = m.addVar(
        lb=0.0,
        ub=terminal_deviation_ub(p.T_cont_min, p.T_cont_max, p.T_cont_init),
        name="cont_terminal_short",
    )

    u_pi = m.addVars(n, vtype=GRB.BINARY, name="u_pi")
    u_po = m.addVars(n, vtype=GRB.BINARY, name="u_po")
    u_lh = m.addVars(n, vtype=GRB.BINARY, name="u_lh")
    u_ch = m.addVars(n, vtype=GRB.BINARY, name="u_ch")
    P_heat_liquid_ctrl = m.addVars(n, lb=0.0, ub=p.P_heat_liquid, name="P_heat_liquid_ctrl")
    P_heat_cont_ctrl = m.addVars(n, lb=0.0, ub=p.P_heat_cont, name="P_heat_cont_ctrl")
    p_heat_liquid_min_w = min(1.0, max(0.0, float(p.P_heat_liquid)))
    p_heat_cont_min_w = min(1.0, max(0.0, float(p.P_heat_cont)))
    u_pi_change = m.addVars(max(0, n - 1), lb=0.0, ub=1.0, name="u_pi_change")
    u_po_change = m.addVars(max(0, n - 1), lb=0.0, ub=1.0, name="u_po_change")
    u_lh_change = m.addVars(max(0, n - 1), lb=0.0, ub=1.0, name="u_lh_change")
    u_ch_change = m.addVars(max(0, n - 1), lb=0.0, ub=1.0, name="u_ch_change")
    tank_temp_step = m.addVars(
        max(0, n - 1), lb=0.0, ub=p.T_tank_max - p.T_tank_min, name="tank_temp_step"
    )
    cont_temp_step = m.addVars(
        max(0, n - 1), lb=0.0, ub=p.T_cont_max - p.T_cont_min, name="cont_temp_step"
    )

    dT_bt_lb = p.T_bat_min - p.T_tank_max
    dT_bt_ub = p.T_bat_max - p.T_tank_min
    dT_ta_lb = p.T_tank_min - float(np.max(t_amb))
    dT_ta_ub = p.T_tank_max - float(np.min(t_amb))
    Q_bt = m.addVars(n, lb=min(0.0, p.K_bt * dT_bt_lb), ub=max(0.0, p.K_bt * dT_bt_ub), name="Q_bt")
    Q_tamb = m.addVars(n, lb=min(0.0, p.K_t_amb * dT_ta_lb), ub=max(0.0, p.K_t_amb * dT_ta_ub), name="Q_tamb")
    Q_tamb_dump = m.addVars(n, lb=0.0, ub=max(0.0, p.K_t_amb * dT_ta_ub), name="Q_tamb_dump")
    v_bt = m.addVars(n, lb=min(0.0, dT_bt_lb), ub=max(0.0, dT_bt_ub), name="v_bt")
    v_ta = m.addVars(n, lb=min(0.0, dT_ta_lb), ub=max(0.0, dT_ta_ub), name="v_ta")

    lam_s = m.addVars(n, n_s, lb=0.0, ub=1.0, name="lam_s")
    lam_t = m.addVars(n, n_t, lb=0.0, ub=1.0, name="lam_t")
    pdc_lb = float(np.min(pdc_pack_table))
    pdc_ub = float(np.max(pdc_pack_table))
    qgen_ub = float(np.max(qgen_pack_table))
    P_dc_pack = m.addVars(n, lb=pdc_lb, ub=pdc_ub, name="P_dc_pack")
    Q_gen_pack = m.addVars(n, lb=0.0, ub=qgen_ub, name="Q_gen_pack")
    P_dc_abs = m.addVars(n, lb=0.0, ub=max(abs(pdc_lb), abs(pdc_ub)), name="P_dc_abs")

    w_st = m.addVars(n, n_s, n_t, lb=0.0, ub=1.0, name="w_st")
    rho_i2 = m.addVars(n, n_s, n_t, n_i, lb=0.0, ub=1.0, name="rho_i2")
    current_mode = str(getattr(p, "diag_current_mode", "continuous"))
    z_current = None
    if current_mode == "discrete":
        z_current = m.addVars(n, n_i, vtype=GRB.BINARY, name="z_current")
    R0 = {}
    for t in range(n):
        R0[t] = gp.quicksum(w_st[t, a, b] * bp.r0[a, b] for a in range(n_s) for b in range(n_t))

    for t in range(n):
        m.addConstr(sum(lam_s[t, a] for a in range(n_s)) == 1.0, name=f"sum_lam_s_{t}")
        m.addSOS(GRB.SOS_TYPE2, [lam_s[t, a] for a in range(n_s)])
        m.addConstr(sum(lam_t[t, b] for b in range(n_t)) == 1.0, name=f"sum_lam_t_{t}")
        m.addSOS(GRB.SOS_TYPE2, [lam_t[t, b] for b in range(n_t)])

        m.addConstr(SOC[t] == sum(lam_s[t, a] * bp.soc[a] for a in range(n_s)), name=f"soc_map_{t}")
        m.addConstr(T_bat[t] == sum(lam_t[t, b] * bp.temp[b] for b in range(n_t)), name=f"tbat_map_{t}")
        m.addConstr(
            I_bat[t] + gp.quicksum(bp.charge_current_limit[b] * lam_t[t, b] for b in range(n_t)) >= 0.0,
            name=f"ibat_charge_temp_limit_{t}",
        )
        m.addConstr(
            I_bat[t] - gp.quicksum(bp.discharge_current_limit[b] * lam_t[t, b] for b in range(n_t)) <= 0.0,
            name=f"ibat_discharge_temp_limit_{t}",
        )
        for a in range(n_s):
            m.addConstr(
                sum(w_st[t, a, b] for b in range(n_t)) == lam_s[t, a],
                name=f"w_st_s_{t}_{a}",
            )
        for b in range(n_t):
            m.addConstr(
                sum(w_st[t, a, b] for a in range(n_s)) == lam_t[t, b],
                name=f"w_st_t_{t}_{b}",
            )
        for a in range(n_s):
            for b in range(n_t):
                m.addConstr(
                    sum(rho_i2[t, a, b, c] for c in range(n_i)) == w_st[t, a, b],
                    name=f"rho_st_{t}_{a}_{b}",
                )
                if getattr(p, "diag_strict_current_sos2", False):
                    m.addSOS(GRB.SOS_TYPE2, [rho_i2[t, a, b, c] for c in range(n_i)])
        if z_current is not None:
            m.addConstr(sum(z_current[t, c] for c in range(n_i)) == 1.0, name=f"current_choice_sum_{t}")
            for c in range(n_i):
                m.addConstr(
                    sum(rho_i2[t, a, b, c] for a in range(n_s) for b in range(n_t)) == z_current[t, c],
                    name=f"current_choice_link_{t}_{c}",
                )
        m.addConstr(
            I_bat[t]
            == sum(rho_i2[t, a, b, c] * bp.current[c] for a in range(n_s) for b in range(n_t) for c in range(n_i)),
            name=f"ibat_perspective_map_{t}",
        )
        m.addConstr(
            P_dc_pack[t]
            == sum(rho_i2[t, a, b, c] * pdc_pack_table[a, b, c] for a in range(n_s) for b in range(n_t) for c in range(n_i)),
            name=f"pdc_pack_map_{t}",
        )
        m.addConstr(
            Q_gen_pack[t]
            == sum(rho_i2[t, a, b, c] * qgen_pack_table[a, b, c] for a in range(n_s) for b in range(n_t) for c in range(n_i)),
            name=f"qgen_pack_map_{t}",
        )

        for g, unit in enumerate(dg_units):
            m.addConstr(
                sum(lam_dg[g, t, k] for k in range(dg_points[g])) == u_dg[g, t],
                name=f"dg_lam_sum_{g}_{t}",
            )
            m.addSOS(GRB.SOS_TYPE2, [lam_dg[g, t, k] for k in range(dg_points[g])])
            m.addConstr(
                P_dg[g, t] == sum(lam_dg[g, t, k] * unit["powers_w"][k] for k in range(dg_points[g])),
                name=f"dg_p_map_{g}_{t}",
            )
            m.addConstr(
                M_dg[g, t] == sum(lam_dg[g, t, k] * unit["fuel_kg_h"][k] for k in range(dg_points[g])),
                name=f"dg_f_map_{g}_{t}",
            )
            m.addConstr(P_dg[g, t] >= u_dg[g, t] * unit["p_min_w"], name=f"dg_min_{g}_{t}")
            m.addConstr(P_dg[g, t] <= u_dg[g, t] * unit["p_max_w"], name=f"dg_max_{g}_{t}")
        if n_g:
            m.addConstr(sum(u_dg[g, t] for g in range(n_g)) >= 1.0, name=f"dg_at_least_one_on_{t}")

        x_bt = T_bat[t] - T_tank[t]
        m.addConstr(v_bt[t] >= dT_bt_lb * u_pi[t], name=f"vbt_lb0_{t}")
        m.addConstr(v_bt[t] <= dT_bt_ub * u_pi[t], name=f"vbt_ub0_{t}")
        m.addConstr(v_bt[t] >= x_bt - dT_bt_ub * (1 - u_pi[t]), name=f"vbt_lb1_{t}")
        m.addConstr(v_bt[t] <= x_bt - dT_bt_lb * (1 - u_pi[t]), name=f"vbt_ub1_{t}")
        m.addConstr(Q_bt[t] == p.K_bt * v_bt[t], name=f"qbt_map_{t}")

        x_ta = T_tank[t] - t_amb[t]
        m.addConstr(v_ta[t] >= dT_ta_lb * u_po[t], name=f"vta_lb0_{t}")
        m.addConstr(v_ta[t] <= dT_ta_ub * u_po[t], name=f"vta_ub0_{t}")
        m.addConstr(v_ta[t] >= x_ta - dT_ta_ub * (1 - u_po[t]), name=f"vta_lb1_{t}")
        m.addConstr(v_ta[t] <= x_ta - dT_ta_lb * (1 - u_po[t]), name=f"vta_ub1_{t}")
        m.addConstr(Q_tamb[t] == p.K_t_amb * v_ta[t], name=f"qtamb_map_{t}")
        m.addConstr(Q_tamb_dump[t] >= Q_tamb[t], name=f"qtamb_dump_pos_{t}")

        m.addConstr(P_heat_liquid_ctrl[t] <= u_lh[t] * p.P_heat_liquid, name=f"p_heat_liquid_ctrl_max_{t}")
        m.addConstr(P_heat_cont_ctrl[t] <= u_ch[t] * p.P_heat_cont, name=f"p_heat_cont_ctrl_max_{t}")
        m.addConstr(P_heat_liquid_ctrl[t] >= u_lh[t] * p_heat_liquid_min_w, name=f"p_heat_liquid_ctrl_min_{t}")
        m.addConstr(P_heat_cont_ctrl[t] >= u_ch[t] * p_heat_cont_min_w, name=f"p_heat_cont_ctrl_min_{t}")
        aux = u_pi[t] * p.P_pump_in + u_po[t] * p.P_pump_out + P_heat_liquid_ctrl[t] + P_heat_cont_ctrl[t]
        m.addConstr(P_dc_abs[t] >= P_dc_pack[t], name=f"pdc_abs_pos_{t}")
        m.addConstr(P_dc_abs[t] >= -P_dc_pack[t], name=f"pdc_abs_neg_{t}")
        m.addConstr(P_BESS[t] == P_dc_pack[t] - p.mu_pcs * P_dc_abs[t] - aux, name=f"pbess_{t}")
        if getattr(p, "diag_disable_storage_dispatch", False):
            # Same exogenous assumptions and thermal model, but the battery is not
            # allowed to shift energy. Auxiliary thermal loads remain represented
            # through P_BESS via the existing port-power equation.
            m.addConstr(I_bat[t] == 0.0, name=f"no_storage_current_{t}")

        m.addConstr(P_pv_use[t] + P_pv_curt[t] == float(data["P_pv"][t]), name=f"pv_split_{t}")
        m.addConstr(P_wt_use[t] + P_wt_curt[t] == float(data["P_wt"][t]), name=f"wt_split_{t}")
        m.addConstr(
            P_pv_use[t]
            + P_wt_use[t]
            + sum(P_dg[g, t] for g in range(n_g))
            + P_BESS[t]
            == float(data["P_load"][t]),
            name=f"power_balance_{t}",
        )
        m.addConstr(tank_preheat_short[t] >= p.T_tank_target - T_tank[t], name=f"tank_preheat_short_{t}")
        m.addConstr(cont_preheat_short[t] >= p.T_cont_target - T_cont[t], name=f"cont_preheat_short_{t}")
        m.addConstr(bat_low_dev[t] >= p.T_bat_pref_low - T_bat[t], name=f"bat_low_dev_{t}")
        m.addConstr(bat_high_dev[t] >= T_bat[t] - p.T_bat_pref_high, name=f"bat_high_dev_{t}")
        m.addConstr(tank_low_dev[t] >= p.T_tank_band_low - T_tank[t], name=f"tank_low_dev_{t}")
        m.addConstr(tank_high_dev[t] >= T_tank[t] - p.T_tank_band_high, name=f"tank_high_dev_{t}")
        m.addConstr(tank_hot_dev[t] >= T_tank[t] - p.T_tank_hot, name=f"tank_hot_dev_{t}")
        m.addConstr(cont_low_dev[t] >= p.T_cont_band_low - T_cont[t], name=f"cont_low_dev_{t}")
        m.addConstr(cont_high_dev[t] >= T_cont[t] - p.T_cont_band_high, name=f"cont_high_dev_{t}")
        m.addConstr(cont_hot_dev[t] >= T_cont[t] - p.T_cont_hot, name=f"cont_hot_dev_{t}")

    m.addConstr(SOC[0] == p.SOC_init, name="soc_initial")
    m.addConstr(T_bat[0] == p.T_bat_init, name="tbat_initial")
    m.addConstr(T_tank[0] == p.T_tank_init, name="ttank_initial")
    m.addConstr(T_cont[0] == p.T_cont_init, name="tcont_initial")

    for t in range(1, n):
        prev = t - 1
        m.addConstr(SOC[t] == SOC[t - 1] - I_bat[prev] * dt_s / (p.Q_nom * 3600.0), name=f"soc_dyn_{t}")
        m.addConstr(u_pi_change[prev] >= u_pi[t] - u_pi[prev], name=f"upi_change_pos_{t}")
        m.addConstr(u_pi_change[prev] >= u_pi[prev] - u_pi[t], name=f"upi_change_neg_{t}")
        m.addConstr(u_po_change[prev] >= u_po[t] - u_po[prev], name=f"upo_change_pos_{t}")
        m.addConstr(u_po_change[prev] >= u_po[prev] - u_po[t], name=f"upo_change_neg_{t}")
        m.addConstr(u_lh_change[prev] >= u_lh[t] - u_lh[prev], name=f"ulh_change_pos_{t}")
        m.addConstr(u_lh_change[prev] >= u_lh[prev] - u_lh[t], name=f"ulh_change_neg_{t}")
        m.addConstr(u_ch_change[prev] >= u_ch[t] - u_ch[prev], name=f"uch_change_pos_{t}")
        m.addConstr(u_ch_change[prev] >= u_ch[prev] - u_ch[t], name=f"uch_change_neg_{t}")
        m.addConstr(tank_temp_step[prev] >= T_tank[t] - T_tank[prev], name=f"tank_temp_step_pos_{t}")
        m.addConstr(tank_temp_step[prev] >= T_tank[prev] - T_tank[t], name=f"tank_temp_step_neg_{t}")
        m.addConstr(cont_temp_step[prev] >= T_cont[t] - T_cont[prev], name=f"cont_temp_step_pos_{t}")
        m.addConstr(cont_temp_step[prev] >= T_cont[prev] - T_cont[t], name=f"cont_temp_step_neg_{t}")

        q_gen = Q_gen_pack[prev]
        q_bc = p.K_b_cont * (T_bat[t] - T_cont[t])
        q_tc = p.K_t_cont * (T_tank[t] - T_cont[t])
        q_camb = p.K_cont_amb * (T_cont[t] - t_amb[prev])
        m.addConstr(T_bat[t] == T_bat[t - 1] + (q_gen - Q_bt[prev] - q_bc) * dt_s / p.C_bat, name=f"tbat_dyn_{t}")
        m.addConstr(
            T_tank[t]
            == T_tank[t - 1]
            + (Q_bt[prev] + P_heat_liquid_ctrl[prev] - Q_tamb[prev] - q_tc) * dt_s / p.C_tank,
            name=f"ttank_dyn_{t}",
        )
        m.addConstr(
            T_cont[t]
            == T_cont[t - 1] + (q_bc + q_tc + P_heat_cont_ctrl[prev] - q_camb) * dt_s / p.C_cont,
            name=f"tcont_dyn_{t}",
        )

    SOC_end = m.addVar(lb=p.SOC_min, ub=p.SOC_max, name="SOC_end")
    m.addConstr(SOC_end == SOC[n - 1] - I_bat[n - 1] * dt_s / (p.Q_nom * 3600.0), name="soc_terminal")
    soc_dev = m.addVar(lb=0.0, ub=terminal_deviation_ub(p.SOC_min, p.SOC_max, p.SOC_init), name="soc_dev")
    m.addConstr(SOC[n - 1] == p.SOC_init, name="soc_schedule_terminal_target")
    m.addConstr(soc_dev >= SOC_end - p.SOC_init, name="soc_dev_pos")
    m.addConstr(soc_dev >= p.SOC_init - SOC_end, name="soc_dev_neg")

    T_bat_end = m.addVar(lb=p.T_bat_min, ub=p.T_bat_max, name="T_bat_end")
    T_tank_end = m.addVar(lb=p.T_tank_min, ub=p.T_tank_max, name="T_tank_end")
    T_cont_end = m.addVar(lb=p.T_cont_min, ub=p.T_cont_max, name="T_cont_end")
    last = n - 1
    q_bc_end = p.K_b_cont * (T_bat_end - T_cont_end)
    q_tc_end = p.K_t_cont * (T_tank_end - T_cont_end)
    q_camb_end = p.K_cont_amb * (T_cont_end - t_amb[last])
    m.addConstr(
        T_bat_end
        == T_bat[last] + (Q_gen_pack[last] - Q_bt[last] - q_bc_end) * dt_s / p.C_bat,
        name="tbat_terminal",
    )
    m.addConstr(
        T_tank_end
        == T_tank[last]
        + (Q_bt[last] + P_heat_liquid_ctrl[last] - Q_tamb[last] - q_tc_end) * dt_s / p.C_tank,
        name="ttank_terminal",
    )
    m.addConstr(
        T_cont_end
        == T_cont[last]
        + (q_bc_end + q_tc_end + P_heat_cont_ctrl[last] - q_camb_end) * dt_s / p.C_cont,
        name="tcont_terminal",
    )
    m.addConstr(T_tank[n - 1] == p.T_tank_init, name="ttank_schedule_terminal_target")
    m.addConstr(T_cont[n - 1] == p.T_cont_init, name="tcont_schedule_terminal_target")
    m.addConstr(tank_terminal_short >= T_tank_end - p.T_tank_init, name="tank_terminal_dev_pos")
    m.addConstr(tank_terminal_short >= p.T_tank_init - T_tank_end, name="tank_terminal_dev_neg")
    m.addConstr(cont_terminal_short >= T_cont_end - p.T_cont_init, name="cont_terminal_dev_pos")
    m.addConstr(cont_terminal_short >= p.T_cont_init - T_cont_end, name="cont_terminal_dev_neg")

    fuel_kg = sum(M_dg[g, t] * dt_s / 3600.0 for g in range(n_g) for t in range(n))
    heat_kwh = sum((P_heat_liquid_ctrl[t] + P_heat_cont_ctrl[t]) * dt_s / 3600.0 / 1000.0 for t in range(n))
    curt_kwh = sum((P_pv_curt[t] + P_wt_curt[t]) * dt_s / 3600.0 / 1000.0 for t in range(n))
    dt_h = dt_s / 3600.0
    preheat_short_score = sum(
        float(renewable_surplus_weight[t]) * (tank_preheat_short[t] + cont_preheat_short[t]) * dt_h
        for t in range(n)
    )
    bat_band_score = sum((bat_low_dev[t] + bat_high_dev[t]) * dt_h for t in range(n))
    tank_band_score = sum((tank_low_dev[t] + tank_high_dev[t]) * dt_h for t in range(n))
    cont_band_score = sum((cont_low_dev[t] + cont_high_dev[t]) * dt_h for t in range(n))
    hot_score = sum((tank_hot_dev[t] + cont_hot_dev[t]) * dt_h for t in range(n))
    terminal_heat_short_score = tank_terminal_short + cont_terminal_short
    heat_dump_kwh_th = sum(Q_tamb_dump[t] * dt_s / 3.6e6 for t in range(n))
    switch_score = sum(
        u_pi_change[t] + u_po_change[t] + u_lh_change[t] + u_ch_change[t] for t in range(max(0, n - 1))
    )
    temp_ramp_score = sum((tank_temp_step[t] + cont_temp_step[t]) for t in range(max(0, n - 1)))
    # Formal route-12 objective: diesel fuel only. All non-fuel terms below
    # remain computed for diagnostics, but they do not influence optimization.
    m.setObjective(fuel_kg, GRB.MINIMIZE)

    apply_mip_start(
        p,
        data,
        bp,
        I_bat,
        SOC,
        T_bat,
        T_tank,
        T_cont,
        lam_s,
        lam_t,
        u_pi,
        u_po,
        u_lh,
        u_ch,
        P_heat_liquid_ctrl,
        P_heat_cont_ctrl,
        u_pi_change,
        u_po_change,
        u_lh_change,
        u_ch_change,
        tank_temp_step,
        cont_temp_step,
        u_dg,
        P_dg,
        M_dg,
        lam_dg,
        P_BESS,
        P_pv_use,
        P_wt_use,
        P_pv_curt,
        P_wt_curt,
        tank_preheat_short,
        cont_preheat_short,
        bat_low_dev,
        bat_high_dev,
        tank_low_dev,
        tank_high_dev,
        tank_hot_dev,
        cont_low_dev,
        cont_high_dev,
        cont_hot_dev,
        tank_terminal_short,
        cont_terminal_short,
        Q_bt,
        Q_tamb,
        Q_tamb_dump,
        v_bt,
        v_ta,
        P_dc_pack,
        P_dc_abs,
        Q_gen_pack,
        w_st,
        rho_i2,
    )

    m.update()
    model_stats = model_stats_from_gurobi_model(m, bp, n, p, data)
    if getattr(p, "diag_model_stats_json", None):
        Path(p.diag_model_stats_json).parent.mkdir(parents=True, exist_ok=True)
        Path(p.diag_model_stats_json).write_text(
            json.dumps(model_stats, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    print(
        f"Model size: {m.NumVars} vars ({m.NumBinVars} binaries), "
        f"{m.NumConstrs} constraints, {m.NumSOS} SOS, {m.NumNZs} nonzeros"
    )
    print(
        f"Breakpoints: SOC={n_s}, T={n_t}, I={n_i}; "
        f"steps={n}, dt={dt_s / 60:.1f} min; diesel units={n_g}"
    )
    if build_only:
        return {
            "success": False,
            "status": "BUILD_ONLY",
            "time_s": 0.0,
            "message": "Model built but not optimized.",
            "model_stats": model_stats,
        }

    started = time.time()
    m.optimize()
    solve_time = time.time() - started

    status_names = {
        GRB.OPTIMAL: "OPTIMAL",
        GRB.INFEASIBLE: "INFEASIBLE",
        GRB.INF_OR_UNBD: "INF_OR_UNBD",
        GRB.UNBOUNDED: "UNBOUNDED",
        GRB.TIME_LIMIT: "TIME_LIMIT",
        GRB.NODE_LIMIT: "NODE_LIMIT",
        GRB.SOLUTION_LIMIT: "SOLUTION_LIMIT",
        GRB.INTERRUPTED: "INTERRUPTED",
        GRB.NUMERIC: "NUMERIC",
        GRB.SUBOPTIMAL: "SUBOPTIMAL",
        GRB.LOADED: "LOADED",
        GRB.CUTOFF: "CUTOFF",
    }
    status_name = status_names.get(m.Status, f"UNKNOWN({m.Status})")
    if m.SolCount == 0:
        if m.Status == GRB.INFEASIBLE:
            try:
                m.computeIIS()
                m.write(str(ROOT / f"bess_iis_{SUFFIX}.ilp"))
            except Exception:
                pass
        return {
            "success": False,
            "status": status_name,
            "time_s": solve_time,
            "message": "No feasible solution returned by Gurobi.",
            "model_stats": model_stats,
        }

    def xv(var) -> float:
        return float(var.X)

    try:
        reported_gap = float(m.MIPGap)
    except Exception:
        reported_gap = 0.0

    result = {
        "success": True,
        "status": status_name,
        "objective": float(m.ObjVal),
        "gap": reported_gap,
        "best_bound": float(m.ObjBound),
        "node_count": float(m.NodeCount),
        "time_s": solve_time,
        "model_stats": model_stats,
        "method": f"{bp.mode} perspective I2R MILP ({bp.n_s}x{bp.n_t}x{bp.n_i})",
        "hours": hours,
        "wind_m_s": np.asarray(data["wind_m_s"], dtype=float),
        "solar_w_m2": np.asarray(data["solar_w_m2"], dtype=float),
        "T_amb": t_amb,
        "load_kw": np.asarray(data["P_load"], dtype=float) / 1000.0,
        "pv_kw": np.asarray(data["P_pv"], dtype=float) / 1000.0,
        "wt_kw": np.asarray(data["P_wt"], dtype=float) / 1000.0,
        "I_bat": np.array([xv(I_bat[t]) for t in range(n)]),
        "SOC": np.array([xv(SOC[t]) for t in range(n)]),
        "SOC_end": xv(SOC_end),
        "T_bat": np.array([xv(T_bat[t]) for t in range(n)]),
        "T_tank": np.array([xv(T_tank[t]) for t in range(n)]),
        "T_cont": np.array([xv(T_cont[t]) for t in range(n)]),
        "T_bat_end": xv(T_bat_end),
        "T_tank_end": xv(T_tank_end),
        "T_cont_end": xv(T_cont_end),
        "P_BESS": np.array([xv(P_BESS[t]) for t in range(n)]),
        "P_dg_units": np.array([[xv(P_dg[g, t]) for t in range(n)] for g in range(n_g)]),
        "M_dg_units": np.array([[xv(M_dg[g, t]) for t in range(n)] for g in range(n_g)]),
        "u_dg": np.array([[xv(u_dg[g, t]) for t in range(n)] for g in range(n_g)]),
        "R0": np.array([float(R0[t].getValue()) for t in range(n)]),
        "P_dc_pack": np.array([xv(P_dc_pack[t]) for t in range(n)]),
        "P_dc_abs": np.array([xv(P_dc_abs[t]) for t in range(n)]),
        "Q_gen_pack": np.array([xv(Q_gen_pack[t]) for t in range(n)]),
        "pv_use_kw": np.array([xv(P_pv_use[t]) for t in range(n)]) / 1000.0,
        "wt_use_kw": np.array([xv(P_wt_use[t]) for t in range(n)]) / 1000.0,
        "pv_curt_kw": np.array([xv(P_pv_curt[t]) for t in range(n)]) / 1000.0,
        "wt_curt_kw": np.array([xv(P_wt_curt[t]) for t in range(n)]) / 1000.0,
        "u_pi": np.array([xv(u_pi[t]) for t in range(n)]),
        "u_po": np.array([xv(u_po[t]) for t in range(n)]),
        "u_lh": np.array([xv(u_lh[t]) for t in range(n)]),
        "u_ch": np.array([xv(u_ch[t]) for t in range(n)]),
        "P_heat_liquid_w": np.array([xv(P_heat_liquid_ctrl[t]) for t in range(n)]),
        "P_heat_cont_w": np.array([xv(P_heat_cont_ctrl[t]) for t in range(n)]),
        "u_pi_change": np.array([xv(u_pi_change[t]) for t in range(max(0, n - 1))]),
        "u_po_change": np.array([xv(u_po_change[t]) for t in range(max(0, n - 1))]),
        "u_lh_change": np.array([xv(u_lh_change[t]) for t in range(max(0, n - 1))]),
        "u_ch_change": np.array([xv(u_ch_change[t]) for t in range(max(0, n - 1))]),
        "tank_temp_step": np.array([xv(tank_temp_step[t]) for t in range(max(0, n - 1))]),
        "cont_temp_step": np.array([xv(cont_temp_step[t]) for t in range(max(0, n - 1))]),
        "tank_preheat_short": np.array([xv(tank_preheat_short[t]) for t in range(n)]),
        "cont_preheat_short": np.array([xv(cont_preheat_short[t]) for t in range(n)]),
        "bat_low_dev": np.array([xv(bat_low_dev[t]) for t in range(n)]),
        "bat_high_dev": np.array([xv(bat_high_dev[t]) for t in range(n)]),
        "tank_low_dev": np.array([xv(tank_low_dev[t]) for t in range(n)]),
        "tank_high_dev": np.array([xv(tank_high_dev[t]) for t in range(n)]),
        "tank_hot_dev": np.array([xv(tank_hot_dev[t]) for t in range(n)]),
        "cont_low_dev": np.array([xv(cont_low_dev[t]) for t in range(n)]),
        "cont_high_dev": np.array([xv(cont_high_dev[t]) for t in range(n)]),
        "cont_hot_dev": np.array([xv(cont_hot_dev[t]) for t in range(n)]),
        "tank_terminal_short": xv(tank_terminal_short),
        "cont_terminal_short": xv(cont_terminal_short),
        "Q_bt": np.array([xv(Q_bt[t]) for t in range(n)]),
        "Q_tamb": np.array([xv(Q_tamb[t]) for t in range(n)]),
        "Q_tamb_dump": np.array([xv(Q_tamb_dump[t]) for t in range(n)]),
        "renewable_surplus_kw": renewable_surplus_w / 1000.0,
        "renewable_surplus_weight": renewable_surplus_weight,
        "fuel_kg": sum(xv(M_dg[g, t]) * dt_s / 3600.0 for g in range(n_g) for t in range(n)),
        "curt_kwh": sum((xv(P_pv_curt[t]) + xv(P_wt_curt[t])) * dt_s / 3600.0 / 1000.0 for t in range(n)),
        "heat_kwh": sum((xv(P_heat_liquid_ctrl[t]) + xv(P_heat_cont_ctrl[t])) * dt_s / 3600.0 / 1000.0 for t in range(n)),
        "preheat_short_score": sum(
            float(renewable_surplus_weight[t]) * (xv(tank_preheat_short[t]) + xv(cont_preheat_short[t])) * dt_s / 3600.0
            for t in range(n)
        ),
        "bat_band_score": sum((xv(bat_low_dev[t]) + xv(bat_high_dev[t])) * dt_s / 3600.0 for t in range(n)),
        "tank_band_score": sum((xv(tank_low_dev[t]) + xv(tank_high_dev[t])) * dt_s / 3600.0 for t in range(n)),
        "cont_band_score": sum((xv(cont_low_dev[t]) + xv(cont_high_dev[t])) * dt_s / 3600.0 for t in range(n)),
        "hot_score": sum((xv(tank_hot_dev[t]) + xv(cont_hot_dev[t])) * dt_s / 3600.0 for t in range(n)),
        "terminal_heat_short_score": xv(tank_terminal_short) + xv(cont_terminal_short),
        "heat_dump_kwh_th": sum(xv(Q_tamb_dump[t]) * dt_s / 3.6e6 for t in range(n)),
        "switch_score": sum(
            xv(u_pi_change[t]) + xv(u_po_change[t]) + xv(u_lh_change[t]) + xv(u_ch_change[t])
            for t in range(max(0, n - 1))
        ),
        "temp_ramp_score": sum(xv(tank_temp_step[t]) + xv(cont_temp_step[t]) for t in range(max(0, n - 1))),
        "dt_s": dt_s,
        "bp_soc": bp.soc,
        "bp_temp": bp.temp,
        "bp_r0": bp.r0,
        "bp_current": bp.current,
    }
    result["P_dg"] = result["P_dg_units"].sum(axis=0)
    result["M_dg"] = result["M_dg_units"].sum(axis=0)
    result["tank_storage_from_amb_kwh_th"] = p.C_tank * (result["T_tank"] - result["T_amb"]) / 3.6e6
    result["cont_storage_from_amb_kwh_th"] = p.C_cont * (result["T_cont"] - result["T_amb"]) / 3.6e6
    result["tank_storage_over_target_kwh_th"] = p.C_tank * np.maximum(0.0, result["T_tank"] - p.T_tank_target) / 3.6e6
    result["cont_storage_over_target_kwh_th"] = p.C_cont * np.maximum(0.0, result["T_cont"] - p.T_cont_target) / 3.6e6
    result["tank_useful_storage_kwh_th"] = p.C_tank * np.maximum(0.0, result["T_tank"] - p.T_tank_useful_min) / 3.6e6
    result["cont_useful_storage_kwh_th"] = p.C_cont * np.maximum(0.0, result["T_cont"] - p.T_cont_useful_min) / 3.6e6
    result["heat_dump_kw_th"] = result["Q_tamb_dump"] / 1000.0
    result["objective_breakdown"] = {
        "formal_objective": "fuel_kg_only",
        "fuel": float(result["fuel_kg"]),
        "nonfuel_terms_are_diagnostics_only": True,
        "soc_diagnostic": float(abs(result["SOC"][-1] - p.SOC_init)),
        "post_step_soc_diagnostic": float(abs(result["SOC_end"] - p.SOC_init)),
        "heat_kwh_diagnostic": float(result["heat_kwh"]),
        "curtailment_kwh_diagnostic": float(result["curt_kwh"]),
        "preheat_short_diagnostic": float(result["preheat_short_score"]),
        "battery_temp_band_diagnostic": float(result["bat_band_score"]),
        "tank_temp_band_diagnostic": float(result["tank_band_score"]),
        "container_temp_band_diagnostic": float(result["cont_band_score"]),
        "hot_temp_diagnostic": float(result["hot_score"]),
        "terminal_heat_short_diagnostic": float(result["terminal_heat_short_score"]),
        "heat_dump_kwh_th_diagnostic": float(result["heat_dump_kwh_th"]),
        "switch_diagnostic": float(result["switch_score"]),
        "temp_ramp_diagnostic": float(result["temp_ramp_score"]),
    }
    # 各传热功率 (W) —— 与 MILP 内 IMEX 离散一致：
    #   主动项 Q_gen_pack / Q_bt / Q_tamb / 电加热 已是变量解，被动项 q_* 用当前 T (隐式) 复算。
    result["q_bc_w"] = p.K_b_cont * (result["T_bat"] - result["T_cont"])
    result["q_tc_w"] = p.K_t_cont * (result["T_tank"] - result["T_cont"])
    t_amb_prev = np.concatenate(([result["T_amb"][0]], result["T_amb"][:-1]))
    result["q_camb_w"] = p.K_cont_amb * (result["T_cont"] - t_amb_prev)
    result["P_heat_liquid_w"] = np.asarray(result["P_heat_liquid_w"], dtype=float)
    result["P_heat_cont_w"] = np.asarray(result["P_heat_cont_w"], dtype=float)
    result["P_pump_in_w"] = result["u_pi"] * p.P_pump_in
    result["P_pump_out_w"] = result["u_po"] * p.P_pump_out
    result["checks"] = compute_checks(p, result)
    return result


def apply_mip_start(
    p,
    data,
    bp,
    I_bat,
    SOC,
    T_bat,
    T_tank,
    T_cont,
    lam_s,
    lam_t,
    u_pi,
    u_po,
    u_lh,
    u_ch,
    P_heat_liquid_ctrl,
    P_heat_cont_ctrl,
    u_pi_change,
    u_po_change,
    u_lh_change,
    u_ch_change,
    tank_temp_step,
    cont_temp_step,
    u_dg,
    P_dg,
    M_dg,
    lam_dg,
    P_BESS,
    P_pv_use,
    P_wt_use,
    P_pv_curt,
    P_wt_curt,
    tank_preheat_short,
    cont_preheat_short,
    bat_low_dev,
    bat_high_dev,
    tank_low_dev,
    tank_high_dev,
    tank_hot_dev,
    cont_low_dev,
    cont_high_dev,
    cont_hot_dev,
    tank_terminal_short,
    cont_terminal_short,
    Q_bt,
    Q_tamb,
    Q_tamb_dump,
    v_bt,
    v_ta,
    P_dc_pack,
    P_dc_abs,
    Q_gen_pack,
    w_st,
    rho_i2,
) -> None:
    n = int(data["N"])
    dt_s = float(data["dt"])
    soc_guess = p.SOC_init
    n_g = len(p.diesel_units)
    thermal = idle_thermal_start(p, data)

    for t in range(n):
        load = float(data["P_load"][t])
        pv = float(data["P_pv"][t])
        wt = float(data["P_wt"][t])
        net = load - pv - wt
        remaining = max(0.0, net)
        p_dg_guess = np.zeros(n_g, dtype=float)
        for g, unit in enumerate(p.diesel_units):
            if remaining <= 0.0:
                continue
            val = min(float(unit["p_max_w"]), max(float(unit["p_min_w"]), remaining))
            p_dg_guess[g] = val
            remaining -= val
        total_dg = float(p_dg_guess.sum())
        pv_use = pv
        wt_use = wt
        if total_dg > net:
            excess = total_dg - net
            pv_curt = min(pv_use, excess)
            pv_use -= pv_curt
            excess -= pv_curt
            wt_curt = min(wt_use, excess)
            wt_use -= wt_curt
        else:
            pv_curt = 0.0
            wt_curt = 0.0
        pbess_guess = 0.0
        i_guess = 0.0
        tbat_guess = float(thermal["T_bat"][t])
        ttank_guess = float(thermal["T_tank"][t])
        tcont_guess = float(thermal["T_cont"][t])

        set_var_start(SOC[t], soc_guess)
        set_var_start(T_bat[t], tbat_guess)
        set_var_start(T_tank[t], ttank_guess)
        set_var_start(T_cont[t], tcont_guess)
        set_var_start(I_bat[t], i_guess)
        set_var_start(P_BESS[t], pbess_guess)
        set_var_start(P_pv_use[t], pv_use)
        set_var_start(P_pv_curt[t], pv_curt)
        set_var_start(P_wt_use[t], wt_use)
        set_var_start(P_wt_curt[t], wt_curt)
        set_var_start(tank_preheat_short[t], max(0.0, p.T_tank_target - ttank_guess))
        set_var_start(cont_preheat_short[t], max(0.0, p.T_cont_target - tcont_guess))
        set_var_start(bat_low_dev[t], max(0.0, p.T_bat_pref_low - tbat_guess))
        set_var_start(bat_high_dev[t], max(0.0, tbat_guess - p.T_bat_pref_high))
        set_var_start(tank_low_dev[t], max(0.0, p.T_tank_band_low - ttank_guess))
        set_var_start(tank_high_dev[t], max(0.0, ttank_guess - p.T_tank_band_high))
        set_var_start(tank_hot_dev[t], max(0.0, ttank_guess - p.T_tank_hot))
        set_var_start(cont_low_dev[t], max(0.0, p.T_cont_band_low - tcont_guess))
        set_var_start(cont_high_dev[t], max(0.0, tcont_guess - p.T_cont_band_high))
        set_var_start(cont_hot_dev[t], max(0.0, tcont_guess - p.T_cont_hot))
        set_var_start(Q_bt[t], 0.0)
        set_var_start(Q_tamb[t], 0.0)
        set_var_start(Q_tamb_dump[t], 0.0)
        set_var_start(v_bt[t], 0.0)
        set_var_start(v_ta[t], 0.0)

        for var, val in ((u_pi[t], 0.0), (u_po[t], 0.0), (u_lh[t], 0.0), (u_ch[t], 0.0)):
            set_var_start(var, val)
            set_var_hint(var, val)
        set_var_start(P_heat_liquid_ctrl[t], 0.0)
        set_var_start(P_heat_cont_ctrl[t], 0.0)
        if t > 0:
            prev = t - 1
            set_var_start(u_pi_change[prev], 0.0)
            set_var_start(u_po_change[prev], 0.0)
            set_var_start(u_lh_change[prev], 0.0)
            set_var_start(u_ch_change[prev], 0.0)
            set_var_start(tank_temp_step[prev], abs(ttank_guess - float(thermal["T_tank"][prev])))
            set_var_start(cont_temp_step[prev], abs(tcont_guess - float(thermal["T_cont"][prev])))

        lam_s_guess = np.zeros(bp.n_s, dtype=float)
        for a in range(bp.n_s):
            set_var_start(lam_s[t, a], 0.0)
        for idx, weight in interp_weights(soc_guess, bp.soc):
            lam_s_guess[idx] = weight
            set_var_start(lam_s[t, idx], weight)
        lam_t_guess = np.zeros(bp.n_t, dtype=float)
        for b in range(bp.n_t):
            set_var_start(lam_t[t, b], 0.0)
        for idx, weight in interp_weights(tbat_guess, bp.temp):
            lam_t_guess[idx] = weight
            set_var_start(lam_t[t, idx], weight)
        current_guess = np.zeros(bp.n_i, dtype=float)
        for idx, weight in interp_weights(i_guess, bp.current):
            current_guess[idx] = weight

        pdc_guess = 0.0
        qgen_guess = 0.0
        for a in range(bp.n_s):
            for b in range(bp.n_t):
                w_st_val = lam_s_guess[a] * lam_t_guess[b]
                set_var_start(w_st[t, a, b], w_st_val)
                for c in range(bp.n_i):
                    w = w_st_val * current_guess[c]
                    set_var_start(rho_i2[t, a, b, c], w)
                    qgen_guess += w * bp.current2[c] * bp.r0[a, b]
                    pdc_guess += w * (bp.ocv[a] * bp.current[c] - bp.current2[c] * bp.r0[a, b])
        set_var_start(P_dc_pack[t], pdc_guess)
        set_var_start(P_dc_abs[t], abs(pdc_guess))
        set_var_start(Q_gen_pack[t], qgen_guess)

        for g, unit in enumerate(p.diesel_units):
            on = 1.0 if p_dg_guess[g] > 1e-6 else 0.0
            set_var_start(u_dg[g, t], on)
            set_var_hint(u_dg[g, t], on)
            set_var_start(P_dg[g, t], p_dg_guess[g])
            if on < 0.5:
                set_var_start(M_dg[g, t], 0.0)
                for k in range(len(unit["powers_w"])):
                    set_var_start(lam_dg[g, t, k], 0.0)
            else:
                pgrid = unit["powers_w"]
                fgrid = unit["fuel_kg_h"]
                fuel_guess = float(np.interp(p_dg_guess[g], pgrid, fgrid))
                set_var_start(M_dg[g, t], fuel_guess)
                for k in range(len(pgrid)):
                    set_var_start(lam_dg[g, t, k], 0.0)
                for idx, weight in interp_weights(p_dg_guess[g], pgrid):
                    set_var_start(lam_dg[g, t, idx], weight)

        soc_guess = float(np.clip(soc_guess - i_guess * dt_s / (p.Q_nom * 3600.0), p.SOC_min, p.SOC_max))

    set_var_start(tank_terminal_short, abs(float(thermal["T_tank"][-1]) - p.T_tank_init))
    set_var_start(cont_terminal_short, abs(float(thermal["T_cont"][-1]) - p.T_cont_init))


def idle_thermal_start(p: SimpleNamespace, data: dict[str, np.ndarray | float | int]) -> dict[str, np.ndarray]:
    n = int(data["N"])
    dt_s = float(data["dt"])
    t_amb = np.asarray(data["T_amb"], dtype=float)
    tbat = np.zeros(n, dtype=float)
    ttank = np.zeros(n, dtype=float)
    tcont = np.zeros(n, dtype=float)
    tbat[0] = p.T_bat_init
    ttank[0] = p.T_tank_init
    tcont[0] = p.T_cont_init

    ab = dt_s / p.C_bat
    at = dt_s / p.C_tank
    ac = dt_s / p.C_cont
    for t in range(1, n):
        mat = np.array(
            [
                [1.0 + ab * p.K_b_cont, 0.0, -ab * p.K_b_cont],
                [0.0, 1.0 + at * p.K_t_cont, -at * p.K_t_cont],
                [-ac * p.K_b_cont, -ac * p.K_t_cont, 1.0 + ac * (p.K_b_cont + p.K_t_cont + p.K_cont_amb)],
            ],
            dtype=float,
        )
        rhs = np.array(
            [
                tbat[t - 1],
                ttank[t - 1],
                tcont[t - 1] + ac * p.K_cont_amb * t_amb[t - 1],
            ],
            dtype=float,
        )
        vals = np.linalg.solve(mat, rhs)
        tbat[t] = float(np.clip(vals[0], p.T_bat_min, p.T_bat_max))
        ttank[t] = float(np.clip(vals[1], p.T_tank_min, p.T_tank_max))
        tcont[t] = float(np.clip(vals[2], p.T_cont_min, p.T_cont_max))
    return {"T_bat": tbat, "T_tank": ttank, "T_cont": tcont}


def compute_checks(p: SimpleNamespace, result: dict) -> dict:
    soc = result["SOC"]
    tbat = result["T_bat"]
    ibat = result["I_bat"]
    r0_true = np.array(
        [bilinear_interp(soc[i], tbat[i], result["bp_soc"], result["bp_temp"], result["bp_r0"]) for i in range(len(soc))]
    )
    uocv_true = np.interp(soc, p.ocv_soc, p.ocv_1d)
    qgen_true = ibat**2 * r0_true
    pdc_true = uocv_true * ibat - qgen_true
    p_heat_liquid_w = np.asarray(result.get("P_heat_liquid_w", result["u_lh"] * p.P_heat_liquid), dtype=float)
    p_heat_cont_w = np.asarray(result.get("P_heat_cont_w", result["u_ch"] * p.P_heat_cont), dtype=float)
    aux = result["u_pi"] * p.P_pump_in + result["u_po"] * p.P_pump_out + p_heat_liquid_w + p_heat_cont_w
    pbess_true = pdc_true - p.mu_pcs * np.abs(pdc_true) - aux
    model_balance_kw = (
        result["pv_use_kw"]
        + result["wt_use_kw"]
        + result["P_dg"] / 1000.0
        + result["P_BESS"] / 1000.0
        - result["load_kw"]
    )
    physical_balance_kw = (
        result["pv_use_kw"]
        + result["wt_use_kw"]
        + result["P_dg"] / 1000.0
        + pbess_true / 1000.0
        - result["load_kw"]
    )
    qgen_error_kw = np.abs(result["Q_gen_pack"] - qgen_true) / 1000.0
    pdc_error_kw = np.abs(result["P_dc_pack"] - pdc_true) / 1000.0
    p_error_kw = np.abs((result["P_BESS"] - pbess_true) / 1000.0)
    r0_error = np.abs(result["R0"] - r0_true)
    soc_for_bounds = np.append(result["SOC"], result.get("SOC_end", result["SOC"][-1]))
    dg_ramp_kw = np.abs(np.diff(result["P_dg_units"], axis=1)) / 1000.0 if result["P_dg_units"].shape[1] > 1 else np.zeros((result["P_dg_units"].shape[0], 0))
    pbess_deviation_max_kw = float(np.max(p_error_kw))
    pbess_deviation_avg_kw = float(np.mean(p_error_kw))
    physical_balance_deviation_max_kw = float(np.max(np.abs(physical_balance_kw)))
    charge_limit_at_t = np.interp(
        tbat,
        p.current_limit_temps,
        p.charge_current_limit_pack,
        left=p.charge_current_limit_pack[0],
        right=p.charge_current_limit_pack[-1],
    )
    discharge_limit_at_t = np.interp(
        tbat,
        p.current_limit_temps,
        p.discharge_current_limit_pack,
        left=p.discharge_current_limit_pack[0],
        right=p.discharge_current_limit_pack[-1],
    )
    charge_current_violation = np.maximum(-ibat - charge_limit_at_t, 0.0)
    discharge_current_violation = np.maximum(ibat - discharge_limit_at_t, 0.0)
    return {
        "r0_max_mohm": float(np.max(r0_error) * 1000.0),
        "r0_avg_mohm": float(np.mean(r0_error) * 1000.0),
        "qgen_pack_max_kw": float(np.max(qgen_error_kw)),
        "qgen_pack_avg_kw": float(np.mean(qgen_error_kw)),
        "pdc_pack_max_kw": float(np.max(pdc_error_kw)),
        "pdc_pack_avg_kw": float(np.mean(pdc_error_kw)),
        "pbess_physical_max_kw": pbess_deviation_max_kw,
        "pbess_physical_avg_kw": pbess_deviation_avg_kw,
        "pwl_physical_pbess_deviation_max_kw": pbess_deviation_max_kw,
        "pwl_physical_pbess_deviation_avg_kw": pbess_deviation_avg_kw,
        "model_balance_max_kw": float(np.max(np.abs(model_balance_kw))),
        "physical_balance_max_kw": physical_balance_deviation_max_kw,
        "pwl_physical_balance_deviation_max_kw": physical_balance_deviation_max_kw,
        "dg_ramp_max_kw_per_step": float(np.max(dg_ramp_kw)) if dg_ramp_kw.size else 0.0,
        "soc_min": float(np.min(soc_for_bounds)),
        "soc_max": float(np.max(soc_for_bounds)),
        "i_pack_min_a": float(np.min(result["I_bat"])),
        "i_pack_max_a": float(np.max(result["I_bat"])),
        "charge_current_limit_min_a": float(np.min(charge_limit_at_t)),
        "charge_current_limit_max_a": float(np.max(charge_limit_at_t)),
        "discharge_current_limit_min_a": float(np.min(discharge_limit_at_t)),
        "discharge_current_limit_max_a": float(np.max(discharge_limit_at_t)),
        "charge_current_limit_violation_max_a": float(np.max(charge_current_violation)),
        "discharge_current_limit_violation_max_a": float(np.max(discharge_current_violation)),
        "tbat_min_c": float(np.min(result["T_bat"])),
        "tbat_max_c": float(np.max(result["T_bat"])),
        "ttank_min_c": float(np.min(result["T_tank"])),
        "ttank_max_c": float(np.max(result["T_tank"])),
        "tcont_min_c": float(np.min(result["T_cont"])),
        "tcont_max_c": float(np.max(result["T_cont"])),
        "tank_preheat_short_max_c": float(np.max(result["tank_preheat_short"])),
        "cont_preheat_short_max_c": float(np.max(result["cont_preheat_short"])),
        "bat_band_dev_avg_c": float(np.mean(result["bat_low_dev"] + result["bat_high_dev"])),
        "bat_band_dev_max_c": float(np.max(result["bat_low_dev"] + result["bat_high_dev"])),
        "tank_band_dev_avg_c": float(np.mean(result["tank_low_dev"] + result["tank_high_dev"])),
        "tank_band_dev_max_c": float(np.max(result["tank_low_dev"] + result["tank_high_dev"])),
        "cont_band_dev_avg_c": float(np.mean(result["cont_low_dev"] + result["cont_high_dev"])),
        "cont_band_dev_max_c": float(np.max(result["cont_low_dev"] + result["cont_high_dev"])),
        "tank_terminal_short_c": float(result["tank_terminal_short"]),
        "cont_terminal_short_c": float(result["cont_terminal_short"]),
        "heat_dump_kwh_th": float(result["heat_dump_kwh_th"]),
        "heat_dump_max_kw_th": float(np.max(result["heat_dump_kw_th"])),
        "switch_score": float(result["switch_score"]),
        "tank_temp_step_max_c": float(np.max(result["tank_temp_step"])) if result["tank_temp_step"].size else 0.0,
        "cont_temp_step_max_c": float(np.max(result["cont_temp_step"])) if result["cont_temp_step"].size else 0.0,
        "temp_ramp_score": float(result["temp_ramp_score"]),
        "tank_useful_storage_max_kwh_th": float(np.max(result["tank_useful_storage_kwh_th"])),
        "tank_useful_storage_end_kwh_th": float(result["tank_useful_storage_kwh_th"][-1]),
        "cont_useful_storage_max_kwh_th": float(np.max(result["cont_useful_storage_kwh_th"])),
        "cont_useful_storage_end_kwh_th": float(result["cont_useful_storage_kwh_th"][-1]),
        "tank_storage_from_amb_max_kwh_th": float(np.max(result["tank_storage_from_amb_kwh_th"])),
        "cont_storage_from_amb_max_kwh_th": float(np.max(result["cont_storage_from_amb_kwh_th"])),
        "tank_storage_over_target_max_kwh_th": float(np.max(result["tank_storage_over_target_kwh_th"])),
        "cont_storage_over_target_max_kwh_th": float(np.max(result["cont_storage_over_target_kwh_th"])),
        "renewable_surplus_hours": float(np.sum(result["renewable_surplus_kw"] > 1e-6) * result["dt_s"] / 3600.0),
        "heated_during_surplus_hours": float(
            np.sum(((result["u_lh"] + result["u_ch"]) > 0.5) & (result["renewable_surplus_kw"] > 1e-6))
            * result["dt_s"]
            / 3600.0
        ),
    }


def _series(result: dict, key: str, n: int, default: float = 0.0) -> np.ndarray:
    value = result.get(key)
    if value is None:
        default_arr = np.asarray(default, dtype=float)
        if default_arr.ndim == 0:
            return np.full(n, float(default_arr), dtype=float)
        default_arr = default_arr.reshape(-1)
        if default_arr.size >= n:
            return default_arr[:n].astype(float)
        if default_arr.size == 0:
            return np.zeros(n, dtype=float)
        return np.concatenate([default_arr.astype(float), np.full(n - default_arr.size, float(default_arr[-1]), dtype=float)])
    arr = np.asarray(value, dtype=float)
    if arr.ndim == 0:
        return np.full(n, float(arr), dtype=float)
    arr = arr.reshape(-1)
    if arr.size >= n:
        return arr[:n].astype(float)
    if arr.size == 0:
        return np.full(n, float(default), dtype=float)
    return np.concatenate([arr.astype(float), np.full(n - arr.size, float(default), dtype=float)])


def _safe_divide(numerator, denominator, fallback=0.0):
    num = np.asarray(numerator, dtype=float)
    den = np.asarray(denominator, dtype=float)
    return np.divide(num, den, out=np.full(np.broadcast_shapes(num.shape, den.shape), fallback, dtype=float), where=np.abs(den) > 1e-9)


def _round_float(value):
    if value is None:
        return None
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    return value


def _result_table(headers: list[str], rows: list[dict[str, object]]) -> dict[str, object]:
    return {"headers": headers, "rows": rows}


def build_detailed_state_tables(p: SimpleNamespace, result: dict) -> dict[str, dict[str, object]]:
    """Build per-time-step state tables for the optimization result workbook."""
    hours = np.asarray(result.get("hours", []), dtype=float).reshape(-1)
    n = int(hours.size)
    if n <= 0:
        return {
            "求解摘要": _result_table(["项目", "值"], [{"项目": "状态", "值": result.get("status", "NO_TIMESERIES")}]),
            "电芯状态": _result_table(["时刻序号"], []),
            "PCS状态": _result_table(["时刻序号"], []),
            "液冷系统状态": _result_table(["时刻序号"], []),
            "舱体状态": _result_table(["时刻序号"], []),
            "配电系统状态": _result_table(["时刻序号"], []),
        }

    soc = _series(result, "SOC", n)
    tbat = _series(result, "T_bat", n)
    ttank = _series(result, "T_tank", n)
    tcont = _series(result, "T_cont", n)
    tamb = _series(result, "T_amb", n)
    ibat = _series(result, "I_bat", n)
    r0_pack = _series(result, "R0", n)
    pdc_pack_w = _series(result, "P_dc_pack", n)
    pdc_abs_w = _series(result, "P_dc_abs", n, np.abs(pdc_pack_w))
    qgen_pack_w = _series(result, "Q_gen_pack", n)
    pbess_w = _series(result, "P_BESS", n)
    qbt_w = _series(result, "Q_bt", n)
    qtamb_w = _series(result, "Q_tamb", n)
    qtamb_dump_w = _series(result, "Q_tamb_dump", n)
    qbc_w = _series(result, "q_bc_w", n, 0.0)
    qtc_w = _series(result, "q_tc_w", n, 0.0)
    qcamb_w = _series(result, "q_camb_w", n, 0.0)
    u_pi = _series(result, "u_pi", n)
    u_po = _series(result, "u_po", n)
    u_lh = _series(result, "u_lh", n)
    u_ch = _series(result, "u_ch", n)
    p_pump_in_w = _series(result, "P_pump_in_w", n, u_pi * float(getattr(p, "P_pump_in", 0.0)))
    p_pump_out_w = _series(result, "P_pump_out_w", n, u_po * float(getattr(p, "P_pump_out", 0.0)))
    p_heat_liquid_w = _series(result, "P_heat_liquid_w", n, u_lh * float(getattr(p, "P_heat_liquid", 0.0)))
    p_heat_cont_w = _series(result, "P_heat_cont_w", n, u_ch * float(getattr(p, "P_heat_cont", 0.0)))

    n_s = max(1, int(getattr(p, "N_s", 1)))
    n_p = max(1, int(getattr(p, "N_p", 1)))
    total_cells = max(1, n_s * n_p)
    pack_ocv_v = np.interp(soc, np.asarray(getattr(p, "ocv_soc", [0.0, 1.0]), dtype=float), np.asarray(getattr(p, "ocv_1d", [0.0, 0.0]), dtype=float))
    if hasattr(p, "ocv_cell_1d"):
        cell_ocv_v = np.interp(soc, np.asarray(getattr(p, "ocv_soc", [0.0, 1.0]), dtype=float), np.asarray(p.ocv_cell_1d, dtype=float))
    else:
        cell_ocv_v = pack_ocv_v / n_s
    pack_terminal_v = pack_ocv_v - ibat * r0_pack
    cell_terminal_v = pack_terminal_v / n_s
    cell_current_a = ibat / n_p
    cell_r0_ohm = r0_pack * n_p / n_s
    cell_power_kw = cell_terminal_v * cell_current_a / 1000.0
    cell_heat_kw = qgen_pack_w / total_cells / 1000.0
    pcs_loss_w = float(getattr(p, "mu_pcs", 0.0)) * pdc_abs_w
    pcs_ac_output_w = pdc_pack_w - pcs_loss_w
    aux_power_w = p_pump_in_w + p_pump_out_w + p_heat_liquid_w + p_heat_cont_w
    pcs_ac_current_a = _safe_divide(pcs_ac_output_w, pack_terminal_v, 0.0)
    external_current_a = _safe_divide(pbess_w, pack_terminal_v, 0.0)

    m_in = float(getattr(p, "m_in", 0.0))
    m_out = float(getattr(p, "m_out", 0.0))
    c_liq = float(getattr(p, "c_liq", 0.0))
    battery_flow = u_pi * m_in
    ambient_flow = u_po * m_out
    battery_return_temp = np.where(
        (battery_flow > 1e-9) & (c_liq > 1e-9),
        ttank + qbt_w / np.maximum(battery_flow * c_liq, 1e-9),
        ttank,
    )
    ambient_return_temp = np.where(
        (ambient_flow > 1e-9) & (c_liq > 1e-9),
        ttank - qtamb_w / np.maximum(ambient_flow * c_liq, 1e-9),
        ttank,
    )

    cell_headers = [
        "时刻序号",
        "时刻(h)",
        "SOC",
        "电芯温度(℃)",
        "电芯产热功率(kW)",
        "电芯OCV电压(V)",
        "电芯端口电压(V)",
        "电芯电流(A)",
        "电芯电功率(kW)",
        "电芯等效电阻(Ω)",
        "电池包OCV电压(V)",
        "电池包端口电压(V)",
        "电池包电流(A)",
        "电池包电功率(kW)",
        "电池包产热功率(kW)",
        "电池包等效电阻(Ω)",
    ]
    pcs_headers = [
        "时刻序号",
        "时刻(h)",
        "PCS直流电压(V)",
        "PCS直流电流(A)",
        "PCS直流功率(kW)",
        "PCS损耗(kW)",
        "PCS交流输出电压(V)",
        "PCS交流输出电流(A)",
        "PCS交流输出功率(kW)",
        "BESS并网端口功率(kW)",
        "辅助用电功率(kW)",
    ]
    liquid_headers = [
        "时刻序号",
        "时刻(h)",
        "液冷罐液体温度(℃)",
        "环境温度(℃)",
        "电池侧回水温度(℃)",
        "电池侧流量(kg/s)",
        "电池侧换热功率(kW)",
        "环境侧回水温度(℃)",
        "环境侧流量(kg/s)",
        "环境侧换热功率(kW)",
        "与环境交换热功率(kW)",
        "液冷电加热启停状态",
        "液冷加热功率(kW)",
        "内循环泵启停状态",
        "外循环泵启停状态",
    ]
    container_headers = [
        "时刻序号",
        "时刻(h)",
        "舱体温度(℃)",
        "环境温度(℃)",
        "电池-舱体换热功率(kW)",
        "液冷罐-舱体换热功率(kW)",
        "舱体-环境换热功率(kW)",
        "舱体电制热启停状态",
        "舱体电加热功率(kW)",
        "舱体净热功率(kW)",
    ]
    distribution_headers = [
        "时刻序号",
        "时刻(h)",
        "内循环泵启停状态",
        "内循环泵用电功率(kW)",
        "外循环泵启停状态",
        "外循环泵用电功率(kW)",
        "液冷电制热启停状态",
        "液冷电制热用电功率(kW)",
        "舱体电制热启停状态",
        "舱体电制热用电功率(kW)",
        "总辅助用电功率(kW)",
        "PCS输出电压(V)",
        "PCS输出电流(A)",
        "PCS输出功率(kW)",
        "与外界交换电压(V)",
        "与外界交换电流(A)",
        "与外界交换功率(kW)",
        "光伏利用功率(kW)",
        "风电利用功率(kW)",
        "柴油机总功率(kW)",
        "负荷功率(kW)",
    ]

    cell_rows = []
    pcs_rows = []
    liquid_rows = []
    container_rows = []
    distribution_rows = []
    for idx in range(n):
        base = {"时刻序号": idx, "时刻(h)": float(hours[idx])}
        cell_rows.append(
            {
                **base,
                "SOC": float(soc[idx]),
                "电芯温度(℃)": float(tbat[idx]),
                "电芯产热功率(kW)": float(cell_heat_kw[idx]),
                "电芯OCV电压(V)": float(cell_ocv_v[idx]),
                "电芯端口电压(V)": float(cell_terminal_v[idx]),
                "电芯电流(A)": float(cell_current_a[idx]),
                "电芯电功率(kW)": float(cell_power_kw[idx]),
                "电芯等效电阻(Ω)": float(cell_r0_ohm[idx]),
                "电池包OCV电压(V)": float(pack_ocv_v[idx]),
                "电池包端口电压(V)": float(pack_terminal_v[idx]),
                "电池包电流(A)": float(ibat[idx]),
                "电池包电功率(kW)": float(pdc_pack_w[idx] / 1000.0),
                "电池包产热功率(kW)": float(qgen_pack_w[idx] / 1000.0),
                "电池包等效电阻(Ω)": float(r0_pack[idx]),
            }
        )
        pcs_rows.append(
            {
                **base,
                "PCS直流电压(V)": float(pack_terminal_v[idx]),
                "PCS直流电流(A)": float(ibat[idx]),
                "PCS直流功率(kW)": float(pdc_pack_w[idx] / 1000.0),
                "PCS损耗(kW)": float(pcs_loss_w[idx] / 1000.0),
                "PCS交流输出电压(V)": float(pack_terminal_v[idx]),
                "PCS交流输出电流(A)": float(pcs_ac_current_a[idx]),
                "PCS交流输出功率(kW)": float(pcs_ac_output_w[idx] / 1000.0),
                "BESS并网端口功率(kW)": float(pbess_w[idx] / 1000.0),
                "辅助用电功率(kW)": float(aux_power_w[idx] / 1000.0),
            }
        )
        liquid_rows.append(
            {
                **base,
                "液冷罐液体温度(℃)": float(ttank[idx]),
                "环境温度(℃)": float(tamb[idx]),
                "电池侧回水温度(℃)": float(battery_return_temp[idx]),
                "电池侧流量(kg/s)": float(battery_flow[idx]),
                "电池侧换热功率(kW)": float(qbt_w[idx] / 1000.0),
                "环境侧回水温度(℃)": float(ambient_return_temp[idx]),
                "环境侧流量(kg/s)": float(ambient_flow[idx]),
                "环境侧换热功率(kW)": float(qtamb_w[idx] / 1000.0),
                "与环境交换热功率(kW)": float(qtamb_dump_w[idx] / 1000.0),
                "液冷电加热启停状态": float(u_lh[idx]),
                "液冷加热功率(kW)": float(p_heat_liquid_w[idx] / 1000.0),
                "内循环泵启停状态": float(u_pi[idx]),
                "外循环泵启停状态": float(u_po[idx]),
            }
        )
        container_rows.append(
            {
                **base,
                "舱体温度(℃)": float(tcont[idx]),
                "环境温度(℃)": float(tamb[idx]),
                "电池-舱体换热功率(kW)": float(qbc_w[idx] / 1000.0),
                "液冷罐-舱体换热功率(kW)": float(qtc_w[idx] / 1000.0),
                "舱体-环境换热功率(kW)": float(qcamb_w[idx] / 1000.0),
                "舱体电制热启停状态": float(u_ch[idx]),
                "舱体电加热功率(kW)": float(p_heat_cont_w[idx] / 1000.0),
                "舱体净热功率(kW)": float((qbc_w[idx] + qtc_w[idx] + p_heat_cont_w[idx] - qcamb_w[idx]) / 1000.0),
            }
        )
        distribution_rows.append(
            {
                **base,
                "内循环泵启停状态": float(u_pi[idx]),
                "内循环泵用电功率(kW)": float(p_pump_in_w[idx] / 1000.0),
                "外循环泵启停状态": float(u_po[idx]),
                "外循环泵用电功率(kW)": float(p_pump_out_w[idx] / 1000.0),
                "液冷电制热启停状态": float(u_lh[idx]),
                "液冷电制热用电功率(kW)": float(p_heat_liquid_w[idx] / 1000.0),
                "舱体电制热启停状态": float(u_ch[idx]),
                "舱体电制热用电功率(kW)": float(p_heat_cont_w[idx] / 1000.0),
                "总辅助用电功率(kW)": float(aux_power_w[idx] / 1000.0),
                "PCS输出电压(V)": float(pack_terminal_v[idx]),
                "PCS输出电流(A)": float(pcs_ac_current_a[idx]),
                "PCS输出功率(kW)": float(pcs_ac_output_w[idx] / 1000.0),
                "与外界交换电压(V)": float(pack_terminal_v[idx]),
                "与外界交换电流(A)": float(external_current_a[idx]),
                "与外界交换功率(kW)": float(pbess_w[idx] / 1000.0),
                "光伏利用功率(kW)": float(_series(result, "pv_use_kw", n)[idx]),
                "风电利用功率(kW)": float(_series(result, "wt_use_kw", n)[idx]),
                "柴油机总功率(kW)": float(_series(result, "P_dg", n)[idx] / 1000.0),
                "负荷功率(kW)": float(_series(result, "load_kw", n)[idx]),
            }
        )

    summary_rows = [
        {"项目": "状态", "值": result.get("status")},
        {"项目": "求解器", "值": result.get("solver_used")},
        {"项目": "求解后端", "值": result.get("solver_backend")},
        {"项目": "目标值", "值": _round_float(result.get("objective"))},
        {"项目": "MIP Gap", "值": _round_float(result.get("gap"))},
        {"项目": "求解耗时(s)", "值": _round_float(result.get("time_s"))},
        {"项目": "时段数", "值": n},
        {"项目": "时间步长(min)", "值": float(result.get("dt_s", 0.0)) / 60.0},
    ]
    return {
        "求解摘要": _result_table(["项目", "值"], summary_rows),
        "电芯状态": _result_table(cell_headers, cell_rows),
        "PCS状态": _result_table(pcs_headers, pcs_rows),
        "液冷系统状态": _result_table(liquid_headers, liquid_rows),
        "舱体状态": _result_table(container_headers, container_rows),
        "配电系统状态": _result_table(distribution_headers, distribution_rows),
    }


def _write_rows_sheet(ws, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([_round_float(row.get(header)) for header in headers])


def _style_table_sheet(ws, header_fill, header_font) -> None:
    from openpyxl.styles import Alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            if isinstance(cell.value, float):
                cell.number_format = "0.000000"
    for column_cells in ws.columns:
        max_len = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 34))
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 28))


def _flatten_statistics_rows(stats: dict[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for key, value in stats.items():
        if isinstance(value, (dict, list)):
            display_value = json.dumps(make_json_safe(value), ensure_ascii=False)
        else:
            display_value = value
        rows.append({"项目": key, "值": display_value})
    return rows


def _write_result_payload_sheet(wb, payload: dict[str, object]) -> None:
    ws = wb.create_sheet("__result_payload")
    ws.sheet_state = "hidden"
    ws.append(["chunk_index", "json_chunk"])
    payload_text = json.dumps(make_json_safe(payload), ensure_ascii=False)
    chunk_size = 30000
    for index, start in enumerate(range(0, len(payload_text), chunk_size), start=1):
        ws.append([index, payload_text[start : start + chunk_size]])


def write_detailed_results_workbook(p: SimpleNamespace, result: dict, path: Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise ImportError("Writing detailed optimization results requires openpyxl.") from exc

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    result["state_workbook"] = str(path)
    payload = build_result_data_payload(p, result)
    payload["files"] = make_json_safe({"state_workbook": str(path)})
    statistics = dict(payload.get("statistics") or {})
    statistics["files"] = make_json_safe({"state_workbook": str(path)})
    payload["statistics"] = make_json_safe(statistics)
    tables = dict(payload.get("tables") or {})
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    header_fill = PatternFill("solid", fgColor="0F3A3A")
    header_font = Font(bold=True, color="FFFFFF")

    stats_ws = wb.create_sheet("统计信息")
    _write_rows_sheet(stats_ws, ["项目", "值"], _flatten_statistics_rows(statistics))
    _style_table_sheet(stats_ws, header_fill, header_font)

    series_headers = ["key", "source_key", "label", "unit", "min", "max", "avg", "point_count"]
    series_ws = wb.create_sheet("曲线元数据")
    _write_rows_sheet(series_ws, series_headers, list(payload.get("series") or []))
    _style_table_sheet(series_ws, header_fill, header_font)

    curve_headers = ["step", "hour"] + [str(item.get("key")) for item in payload.get("series", [])]
    curve_ws = wb.create_sheet("调度曲线")
    _write_rows_sheet(curve_ws, curve_headers, list(payload.get("rows") or []))
    _style_table_sheet(curve_ws, header_fill, header_font)

    for sheet_name, table in tables.items():
        ws = wb.create_sheet(sheet_name[:31])
        headers = list(table["headers"])
        rows = list(table["rows"])
        _write_rows_sheet(ws, headers, rows)
        _style_table_sheet(ws, header_fill, header_font)

    _write_result_payload_sheet(wb, payload)
    wb.save(path)
    return path


RESULT_SERIES_META: dict[str, tuple[str, str]] = {
    "wind_m_s": ("风速", "m/s"),
    "solar_w_m2": ("太阳辐照度", "W/m2"),
    "T_amb": ("环境温度", "℃"),
    "load_kw": ("负荷功率", "kW"),
    "pv_kw": ("光伏可用功率", "kW"),
    "wt_kw": ("风电可用功率", "kW"),
    "I_bat": ("电池包电流", "A"),
    "SOC": ("SOC", ""),
    "T_bat": ("电芯温度", "℃"),
    "T_tank": ("液冷罐温度", "℃"),
    "T_cont": ("舱体温度", "℃"),
    "P_BESS": ("BESS并网端口功率", "kW"),
    "P_dg": ("柴油机总功率", "kW"),
    "M_dg": ("柴油消耗率", "kg/h"),
    "P_dg_units": ("柴油机单机功率", "kW"),
    "M_dg_units": ("柴油机单机油耗率", "kg/h"),
    "u_dg": ("柴油机启停状态", ""),
    "R0": ("电池包等效内阻", "Ω"),
    "P_dc_pack": ("电池包直流功率", "kW"),
    "P_dc_abs": ("电池包直流功率绝对值", "kW"),
    "Q_gen_pack": ("电池包产热功率", "kW"),
    "pv_use_kw": ("光伏利用功率", "kW"),
    "wt_use_kw": ("风电利用功率", "kW"),
    "pv_curt_kw": ("弃光功率", "kW"),
    "wt_curt_kw": ("弃风功率", "kW"),
    "u_pi": ("内循环泵启停状态", ""),
    "u_po": ("外循环泵启停状态", ""),
    "u_lh": ("液冷加热启停状态", ""),
    "u_ch": ("舱体加热启停状态", ""),
    "Q_bt": ("电池侧换热功率", "kW"),
    "Q_tamb": ("环境侧换热功率", "kW"),
    "Q_tamb_dump": ("对环境散热功率", "kW"),
    "renewable_surplus_kw": ("风光富余功率", "kW"),
    "renewable_surplus_weight": ("风光富余权重", ""),
    "tank_storage_from_amb_kwh_th": ("液冷罐相对环境储热量", "kWh_th"),
    "cont_storage_from_amb_kwh_th": ("舱体相对环境储热量", "kWh_th"),
    "tank_storage_over_target_kwh_th": ("液冷罐高于目标储热量", "kWh_th"),
    "cont_storage_over_target_kwh_th": ("舱体高于目标储热量", "kWh_th"),
    "tank_useful_storage_kwh_th": ("液冷罐可用储热量", "kWh_th"),
    "cont_useful_storage_kwh_th": ("舱体可用储热量", "kWh_th"),
    "heat_dump_kw_th": ("散热功率", "kW"),
    "q_bc_w": ("电池-舱体换热功率", "kW"),
    "q_tc_w": ("液冷罐-舱体换热功率", "kW"),
    "q_camb_w": ("舱体-环境换热功率", "kW"),
    "P_heat_liquid_w": ("液冷电加热功率", "kW"),
    "P_heat_cont_w": ("舱体电加热功率", "kW"),
    "P_pump_in_w": ("内循环泵功率", "kW"),
    "P_pump_out_w": ("外循环泵功率", "kW"),
}

RESULT_SERIES_SKIP_KEYS = {"hours", "bp_soc", "bp_temp", "bp_r0", "bp_current"}
RESULT_SERIES_POWER_W_KEYS = {
    "P_BESS",
    "P_dg",
    "P_dg_units",
    "P_dc_pack",
    "P_dc_abs",
    "Q_gen_pack",
    "Q_bt",
    "Q_tamb",
    "Q_tamb_dump",
    "q_bc_w",
    "q_tc_w",
    "q_camb_w",
    "P_heat_liquid_w",
    "P_heat_cont_w",
    "P_pump_in_w",
    "P_pump_out_w",
}


def _json_number(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, np.generic):
        value = value.item()
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _result_series_display_base(source_key: str) -> str:
    if source_key.endswith("_w"):
        return f"{source_key[:-2]}_kw"
    return source_key


def _result_series_scale(source_key: str) -> float:
    return 0.001 if source_key in RESULT_SERIES_POWER_W_KEYS or source_key.endswith("_w") else 1.0


def _series_label_and_unit(source_key: str, display_key: str, index: int | None = None) -> tuple[str, str]:
    label, unit = RESULT_SERIES_META.get(source_key, (source_key, ""))
    if index is not None:
        label = f"{label} {index + 1}"
    if source_key in RESULT_SERIES_POWER_W_KEYS or source_key.endswith("_w"):
        unit = "kW"
    if display_key.endswith("_kw") and not unit:
        unit = "kW"
    elif display_key.endswith("_w") and not unit:
        unit = "kW"
    elif display_key.endswith("_c") and not unit:
        unit = "℃"
    return label, unit


def _values_for_result_series(values, n: int, scale: float = 1.0) -> list[float | None]:
    arr = np.asarray(values, dtype=float).reshape(-1) * float(scale)
    if arr.size >= n:
        arr = arr[:n]
    elif arr.size == n - 1 and n > 0:
        arr = np.concatenate([arr, np.asarray([np.nan], dtype=float)])
    elif arr.size < n:
        arr = np.concatenate([arr, np.full(n - arr.size, np.nan, dtype=float)])
    return [_json_number(value) for value in arr]


def _series_summary(values: list[float | None]) -> dict[str, float | int | None]:
    valid = np.asarray([value for value in values if value is not None], dtype=float)
    if valid.size == 0:
        return {"min": None, "max": None, "avg": None, "point_count": 0}
    return {
        "min": float(np.min(valid)),
        "max": float(np.max(valid)),
        "avg": float(np.mean(valid)),
        "point_count": int(valid.size),
    }


def collect_result_series(result: dict, n: int, max_components: int = 32) -> list[dict[str, object]]:
    series: list[dict[str, object]] = []
    used_keys: set[str] = set()
    for key, raw_value in result.items():
        if key in RESULT_SERIES_SKIP_KEYS:
            continue
        try:
            arr = np.asarray(raw_value, dtype=float)
        except (TypeError, ValueError):
            continue
        if arr.ndim == 1 and arr.size in {n, max(0, n - 1)}:
            display_key = _result_series_display_base(str(key))
            label, unit = _series_label_and_unit(str(key), display_key)
            values = _values_for_result_series(arr, n, _result_series_scale(str(key)))
            series.append(
                {
                    "key": display_key,
                    "source_key": str(key),
                    "label": label,
                    "unit": unit,
                    "values": values,
                    **_series_summary(values),
                }
            )
            used_keys.add(display_key)
        elif arr.ndim == 2:
            source_key = str(key)
            if arr.shape[1] in {n, max(0, n - 1)} and 0 < arr.shape[0] <= max_components:
                for index in range(arr.shape[0]):
                    display_key = f"{_result_series_display_base(source_key)}_{index + 1}"
                    if display_key in used_keys:
                        continue
                    label, unit = _series_label_and_unit(source_key, display_key, index)
                    values = _values_for_result_series(arr[index, :], n, _result_series_scale(source_key))
                    series.append(
                        {
                            "key": display_key,
                            "source_key": source_key,
                            "label": label,
                            "unit": unit,
                            "values": values,
                            **_series_summary(values),
                        }
                    )
                    used_keys.add(display_key)
            elif arr.shape[0] in {n, max(0, n - 1)} and 0 < arr.shape[1] <= max_components:
                for index in range(arr.shape[1]):
                    display_key = f"{_result_series_display_base(source_key)}_{index + 1}"
                    if display_key in used_keys:
                        continue
                    label, unit = _series_label_and_unit(source_key, display_key, index)
                    values = _values_for_result_series(arr[:, index], n, _result_series_scale(source_key))
                    series.append(
                        {
                            "key": display_key,
                            "source_key": source_key,
                            "label": label,
                            "unit": unit,
                            "values": values,
                            **_series_summary(values),
                        }
                    )
                    used_keys.add(display_key)
    series.sort(key=lambda item: (str(item.get("source_key", "")), str(item.get("key", ""))))
    return series


def build_result_statistics(result: dict, row_count: int, series_count: int) -> dict[str, object]:
    scalar_keys = [
        "success",
        "status",
        "message",
        "objective",
        "best_bound",
        "gap",
        "time_s",
        "node_count",
        "fuel_kg",
        "curt_kwh",
        "heat_kwh",
        "solver_requested",
        "solver_used",
        "solver_backend",
        "state_workbook",
    ]
    statistics = {key: result.get(key) for key in scalar_keys if key in result}
    statistics.update(
        {
            "row_count": row_count,
            "series_count": series_count,
            "checks": result.get("checks") or {},
            "model_stats": result.get("model_stats") or {},
            "objective_breakdown": result.get("objective_breakdown") or {},
        }
    )
    return make_json_safe(statistics)


def build_result_data_payload(p: SimpleNamespace, result: dict) -> dict[str, object]:
    hours = np.asarray(result.get("hours", []), dtype=float).reshape(-1)
    n = int(hours.size)
    series = collect_result_series(result, n) if n > 0 else []
    rows: list[dict[str, object]] = []
    for idx in range(n):
        row: dict[str, object] = {"step": idx, "hour": _json_number(hours[idx])}
        for item in series:
            values = item.get("values") or []
            if idx < len(values):
                row[str(item["key"])] = values[idx]
        rows.append(row)
    statistics = build_result_statistics(result, n, len(series))
    payload = {
        "version": 1,
        "row_count": n,
        "time_axis": {"key": "hour", "label": "时刻", "unit": "h"},
        "statistics": statistics,
        "series": series,
        "rows": rows,
        "tables": build_detailed_state_tables(p, result),
    }
    return make_json_safe(payload)


def plot_results(result: dict, path: Path) -> None:
    hrs = result["hours"]
    fig, axes = plt.subplots(5, 2, figsize=(16, 18))
    fig.suptitle("低温风光扩容电热耦合优化结果（热价值强化版，2026-05-07）", fontsize=16)

    axes[0, 0].plot(hrs, result["P_dg"] / 1000.0, color="#c0392b", lw=1.8, label="柴油机总出力")
    soc_hrs = np.append(hrs, hrs[-1] + result["dt_s"] / 3600.0)
    soc_vals = np.append(result["SOC"], result.get("SOC_end", result["SOC"][-1]))
    axes[0, 1].plot(soc_hrs, soc_vals * 100.0, color="#2471a3", lw=1.8, label="SOC")
    axes[1, 0].plot(hrs, result["T_bat"], color="#d35400", lw=1.8, label="电池温度")
    axes[1, 1].plot(hrs, result["P_BESS"] / 1000.0, color="#16a085", lw=1.8, label="BESS端口功率")
    axes[2, 0].plot(hrs, result["I_bat"], color="#7d3c98", lw=1.8, label="电池包电流")
    axes[2, 1].plot(hrs, result["R0"] * 1000.0, color="#6e2c00", lw=1.8, label="R_pack")
    axes[3, 0].plot(hrs, result["pv_curt_kw"], color="#f1c40f", lw=1.6, label="弃光")
    axes[3, 0].plot(hrs, result["wt_curt_kw"], color="#27ae60", lw=1.6, label="弃风")
    axes[3, 1].plot(hrs, result["pv_curt_kw"] + result["wt_curt_kw"], color="#8e44ad", lw=1.8, label="总弃电")
    axes[4, 0].plot(hrs, result["pv_kw"], color="#f39c12", ls="--", lw=1.3, label="光伏可用")
    axes[4, 0].plot(hrs, result["pv_use_kw"], color="#f39c12", lw=1.8, label="光伏利用")
    axes[4, 0].plot(hrs, result["wt_kw"], color="#1e8449", ls="--", lw=1.3, label="风电可用")
    axes[4, 0].plot(hrs, result["wt_use_kw"], color="#1e8449", lw=1.8, label="风电利用")
    axes[4, 1].plot(hrs, result["load_kw"], color="#2c3e50", lw=1.9, label="负荷曲线")

    titles = [
        "柴油机总功率 (kW)",
        "SOC (%)",
        "电池温度 (℃)",
        "BESS功率 (kW)",
        "电池包电流 (A)",
        "电池包等效内阻R_pack (mΩ)",
        "弃风弃光功率 (kW)",
        "总弃电功率 (kW)",
        "风光出力曲线 (kW)",
        "负荷曲线 (kW)",
    ]
    for ax, title in zip(axes.flat, titles):
        ax.set_title(title)
        ax.set_xlabel("时间 (h)")
        ax.grid(True, alpha=0.3)
        ax.legend(fontsize=8)
    plt.tight_layout(rect=(0, 0, 1, 0.98))
    plt.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def plot_switches(result: dict, path: Path) -> None:
    hrs = result["hours"]
    series = [
        ("内循环泵 u_pi", result["u_pi"]),
        ("外循环泵 u_po", result["u_po"]),
        ("液冷加热 u_lh", result["u_lh"]),
        ("舱体加热 u_ch", result["u_ch"]),
    ]
    for g in range(result["u_dg"].shape[0]):
        series.append((f"柴油机{g + 1} u_dg_{g + 1}", result["u_dg"][g]))

    n_total = len(series)
    n_cols = 2
    n_rows = math.ceil(n_total / n_cols)
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(14, 2.5 * n_rows + 1), sharex=True, squeeze=False)
    fig.suptitle("启停变量曲线", fontsize=15)

    palette = ["#2471a3", "#16a085", "#d35400", "#8e44ad", "#c0392b", "#7f8c8d", "#2e86c1", "#884ea0", "#117a65", "#922b21"]
    flat_axes = axes.flat
    for idx, ax in enumerate(flat_axes):
        if idx < n_total:
            name, values = series[idx]
            color = palette[idx % len(palette)]
            ax.step(hrs, values, where="post", color=color, lw=1.8)
            ax.set_title(name)
            ax.set_ylim(-0.1, 1.1)
            ax.set_yticks([0, 1])
            ax.grid(True, alpha=0.3)
        else:
            ax.set_visible(False)
    for ax in axes[-1, :]:
        ax.set_xlabel("时间 (h)")
    plt.tight_layout(rect=(0, 0, 1, 0.96))
    plt.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def plot_temperatures(p: SimpleNamespace, result: dict, path: Path) -> None:
    hrs = result["hours"]
    fig, ax = plt.subplots(figsize=(14, 7))
    ax.plot(hrs, result["T_bat"], color="#c0392b", lw=2.0, label="电池温度 T_bat")
    ax.plot(hrs, result["T_tank"], color="#2471a3", lw=2.0, label="液冷罐温度 T_tank")
    ax.plot(hrs, result["T_cont"], color="#16a085", lw=2.0, label="舱体温度 T_cont")
    ax.plot(hrs, result["T_amb"], color="#2c3e50", lw=1.8, ls="--", label="环境温度 T_amb")
    ax.axhspan(p.T_bat_pref_low, p.T_bat_pref_high, color="#fadbd8", alpha=0.22, label="电池舒适区")
    ax.axhspan(p.T_tank_band_low, p.T_tank_band_high, color="#d6eaf8", alpha=0.22, label="液冷罐储热区")
    ax.axhspan(p.T_cont_band_low, p.T_cont_band_high, color="#d1f2eb", alpha=0.22, label="舱体/内壁储热区")
    ax.axhline(p.T_bat_min, color="#c0392b", lw=1.0, ls=":", label="电池温度下限")
    ax.axhline(p.T_bat_max, color="#c0392b", lw=1.0, ls="-.", label="电池温度上限")
    ax.axhline(p.T_cont_min, color="#16a085", lw=1.0, ls=":", label="舱体温度下限")
    ax.axhline(p.T_cont_max, color="#16a085", lw=1.0, ls="-.", label="舱体温度上限")
    ax.axhline(p.T_tank_terminal_min, color="#2471a3", lw=1.0, ls="--", label="液冷罐终端储备")
    ax.axhline(p.T_cont_terminal_min, color="#16a085", lw=1.0, ls="--", label="舱体终端储备")
    ax.set_title("低温工况各元件温度变化曲线（含舒适区/储热区）")
    ax.set_xlabel("时间 (h)")
    ax.set_ylabel("温度 (℃)")
    ax.grid(True, alpha=0.3)
    ax.legend(ncol=2, fontsize=9)
    plt.tight_layout()
    plt.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def plot_thermal_storage(p: SimpleNamespace, result: dict, path: Path) -> None:
    hrs = result["hours"]
    fig, axes = plt.subplots(3, 1, figsize=(14, 11), sharex=True)
    fig.suptitle("储能水箱与舱体内壁储热利用（热价值强化版）", fontsize=15)

    axes[0].plot(hrs, result["tank_storage_from_amb_kwh_th"], color="#2471a3", lw=2.0, label="液冷罐相对环境储热")
    axes[0].plot(hrs, result["cont_storage_from_amb_kwh_th"], color="#16a085", lw=2.0, label="舱体/内壁相对环境储热")
    axes[0].plot(hrs, result["tank_useful_storage_kwh_th"], color="#154360", lw=1.7, ls="--", label="液冷罐可用储热")
    axes[0].plot(hrs, result["cont_useful_storage_kwh_th"], color="#0e6251", lw=1.7, ls="--", label="舱体/内壁可用储热")
    axes[0].set_ylabel("kWh_th")
    axes[0].legend()

    axes[1].plot(hrs, result["tank_low_dev"] + result["tank_high_dev"], color="#2471a3", lw=1.8, label="液冷罐储热区偏差")
    axes[1].plot(hrs, result["cont_low_dev"] + result["cont_high_dev"], color="#16a085", lw=1.8, label="舱体储热区偏差")
    axes[1].plot(hrs, result["bat_low_dev"] + result["bat_high_dev"], color="#c0392b", lw=1.5, label="电池舒适区偏差")
    axes[1].set_ylabel("°C")
    axes[1].legend()

    heat_kw = (result["u_lh"] * p.P_heat_liquid + result["u_ch"] * p.P_heat_cont) / 1000.0
    axes[2].plot(hrs, result["renewable_surplus_kw"], color="#8e44ad", lw=1.7, label="可用风光相对负荷富余")
    axes[2].step(hrs, heat_kw, where="post", color="#d35400", lw=1.7, label="电加热功率")
    axes[2].plot(hrs, result["heat_dump_kw_th"], color="#2c3e50", lw=1.5, ls="--", label="外循环散热")
    axes[2].set_ylabel("kW")
    axes[2].set_xlabel("时间 (h)")
    axes[2].legend()

    for ax in axes:
        ax.grid(True, alpha=0.3)
    plt.tight_layout(rect=(0, 0, 1, 0.96))
    plt.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def plot_thermal_value(p: SimpleNamespace, result: dict, path: Path) -> None:
    hrs = result["hours"]
    fig, axes = plt.subplots(4, 1, figsize=(14, 13), sharex=True)
    fig.suptitle("热调度价值函数与储热保持检查", fontsize=15)

    axes[0].plot(hrs, result["T_tank"], color="#2471a3", lw=1.9, label="液冷罐温度")
    axes[0].axhspan(p.T_tank_band_low, p.T_tank_band_high, color="#d6eaf8", alpha=0.35, label="推荐储热区")
    axes[0].axhline(p.T_tank_terminal_min, color="#154360", lw=1.0, ls="--", label="终端储热下限")
    axes[0].set_ylabel("℃")
    axes[0].legend(ncol=3, fontsize=8)

    axes[1].plot(hrs, result["T_cont"], color="#16a085", lw=1.9, label="舱体/内壁温度")
    axes[1].axhspan(p.T_cont_band_low, p.T_cont_band_high, color="#d1f2eb", alpha=0.35, label="推荐储热区")
    axes[1].axhline(p.T_cont_terminal_min, color="#0e6251", lw=1.0, ls="--", label="终端储热下限")
    axes[1].set_ylabel("℃")
    axes[1].legend(ncol=3, fontsize=8)

    axes[2].plot(hrs, result["bat_low_dev"] + result["bat_high_dev"], color="#c0392b", lw=1.6, label="电池舒适区偏差")
    axes[2].plot(hrs, result["tank_low_dev"] + result["tank_high_dev"], color="#2471a3", lw=1.6, label="水箱储热区偏差")
    axes[2].plot(hrs, result["cont_low_dev"] + result["cont_high_dev"], color="#16a085", lw=1.6, label="舱体储热区偏差")
    axes[2].set_ylabel("℃")
    axes[2].legend(ncol=3, fontsize=8)

    heat_kw = (result["u_lh"] * p.P_heat_liquid + result["u_ch"] * p.P_heat_cont) / 1000.0
    axes[3].plot(hrs, result["renewable_surplus_kw"], color="#8e44ad", lw=1.6, label="风光富余")
    axes[3].step(hrs, heat_kw, where="post", color="#d35400", lw=1.6, label="电加热")
    axes[3].plot(hrs, result["heat_dump_kw_th"], color="#2c3e50", lw=1.5, ls="--", label="外循环散热")
    axes[3].set_ylabel("kW")
    axes[3].set_xlabel("时间 (h)")
    axes[3].legend(ncol=3, fontsize=8)

    for ax in axes:
        ax.grid(True, alpha=0.3)
    plt.tight_layout(rect=(0, 0, 1, 0.96))
    plt.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def plot_heat_flows(p: SimpleNamespace, result: dict, path: Path) -> None:
    hrs = result["hours"]
    fig, axes = plt.subplots(3, 1, figsize=(14, 12), sharex=True)  
    fig.suptitle("各部分传热功率曲线（kW，正值=热量流出/释放）", fontsize=15)

    # 1) 电池节点收支
    axes[0].plot(hrs, result["Q_gen_pack"] / 1000.0, color="#c0392b", lw=1.8, label="Q_gen 电池内热生成")
    axes[0].plot(hrs, result["Q_bt"] / 1000.0, color="#2471a3", lw=1.6, label="Q_bt 电池→液冷罐 (内泵)")
    axes[0].plot(hrs, result["q_bc_w"] / 1000.0, color="#7d3c98", lw=1.4, ls="--", label="q_bc 电池↔舱体 (被动)")
    bat_net = (result["Q_gen_pack"] - result["Q_bt"] - result["q_bc_w"]) / 1000.0
    axes[0].plot(hrs, bat_net, color="#1b2631", lw=1.2, ls=":", label="电池净热流入")
    axes[0].set_ylabel("kW")
    axes[0].set_title("电池节点")
    axes[0].legend(ncol=2, fontsize=8)

    # 2) 液冷罐节点收支
    axes[1].plot(hrs, result["Q_bt"] / 1000.0, color="#2471a3", lw=1.6, label="Q_bt 来自电池")
    axes[1].plot(hrs, result["P_heat_liquid_w"] / 1000.0, color="#d35400", lw=1.6, label="电加热 P_LiquidHeat")
    axes[1].plot(hrs, result["Q_tamb"] / 1000.0, color="#16a085", lw=1.4, ls="--", label="Q_tamb 罐↔环境 (外泵)")
    axes[1].plot(hrs, result["q_tc_w"] / 1000.0, color="#7d3c98", lw=1.4, ls=":", label="q_tc 罐↔舱体 (被动)")
    tank_net = (result["Q_bt"] + result["P_heat_liquid_w"] - result["Q_tamb"] - result["q_tc_w"]) / 1000.0
    axes[1].plot(hrs, tank_net, color="#1b2631", lw=1.2, ls=":", label="液冷罐净热流入")
    axes[1].set_ylabel("kW")
    axes[1].set_title("液冷罐节点")
    axes[1].legend(ncol=2, fontsize=8)

    # 3) 舱体节点收支
    axes[2].plot(hrs, result["q_bc_w"] / 1000.0, color="#7d3c98", lw=1.4, label="q_bc 来自电池")
    axes[2].plot(hrs, result["q_tc_w"] / 1000.0, color="#2471a3", lw=1.4, label="q_tc 来自液冷罐")
    axes[2].plot(hrs, result["P_heat_cont_w"] / 1000.0, color="#d35400", lw=1.6, label="电加热 P_ContainerHeat")
    axes[2].plot(hrs, result["q_camb_w"] / 1000.0, color="#16a085", lw=1.4, ls="--", label="q_camb 舱体→环境 (被动)")
    cont_net = (result["q_bc_w"] + result["q_tc_w"] + result["P_heat_cont_w"] - result["q_camb_w"]) / 1000.0
    axes[2].plot(hrs, cont_net, color="#1b2631", lw=1.2, ls=":", label="舱体净热流入")
    axes[2].set_ylabel("kW")
    axes[2].set_title("舱体/内壁节点")
    axes[2].set_xlabel("时间 (h)")
    axes[2].legend(ncol=2, fontsize=8)

    for ax in axes:
        ax.grid(True, alpha=0.3)
        ax.axhline(0.0, color="#566573", lw=0.6)
    plt.tight_layout(rect=(0, 0, 1, 0.96))
    plt.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def write_summary(p: SimpleNamespace, result: dict | None, args, path: Path) -> None:
    # Markdown sidecar generation has been removed; the result workbook is the UI artifact.
    return
    results_png = getattr(args, "results_png", RESULTS_PNG)
    switches_png = getattr(args, "switches_png", SWITCHES_PNG)
    temperatures_png = getattr(args, "temperatures_png", TEMPERATURES_PNG)
    thermal_storage_png = getattr(args, "thermal_storage_png", THERMAL_STORAGE_PNG)
    thermal_value_png = getattr(args, "thermal_value_png", THERMAL_VALUE_PNG)
    heat_flows_png = getattr(args, "heat_flows_png", HEAT_FLOWS_PNG)
    lines = [
        f"# 路线 13 Perspective I2R 模型结果说明（20260530）",
        "",
        f"- 参数 JSON: `{args.params}`",
        f"- Excel 来源: `{p.source_excel}`",
        f"- 运行模式: `{args.mode}`",
        f"- 时间粒度: `{args.dt_minutes:.3g}` min" if hasattr(args, "dt_minutes") else "",
        f"- 电流线性化: `{getattr(args, 'current_segments', 10)}` 段，模式 `{getattr(args, 'current_mode', getattr(p, 'diag_current_mode', 'continuous'))}`",
        f"- SOC 查表宽网格: `{getattr(args, 'soc_grid_width', 0.1):.3g}`",
        f"- 柴油机数量: `{len(p.diesel_units)}`",
        f"- 风机额定容量: `{sum(u['rated_kw'] for u in p.wind_units):.1f}` kW",
        f"- 光伏额定容量: `{sum(u['rated_kw'] for u in p.pv_units):.1f}` kW",
        f"- 电池舒适区: `{p.T_bat_pref_low:.1f} - {p.T_bat_pref_high:.1f}` ℃",
        f"- 液冷罐储热区: `{p.T_tank_band_low:.1f} - {p.T_tank_band_high:.1f}` ℃，高温惩罚阈值 `{p.T_tank_hot:.1f}` ℃",
        f"- 舱体/内壁储热区: `{p.T_cont_band_low:.1f} - {p.T_cont_band_high:.1f}` ℃，高温惩罚阈值 `{p.T_cont_hot:.1f}` ℃",
        f"- 终端储热软约束: 液冷罐 >= `{p.T_tank_terminal_min:.1f}` ℃，舱体/内壁 >= `{p.T_cont_terminal_min:.1f}` ℃",
        "",
        "## 建模方案",
        "",
        "- `u_pi*(T_bat-T_tank)` 与 `u_po*(T_tank-T_amb)` 使用由变量上下限自动给出的紧 Big-M。",
        "- 电芯 OCV/R0 已折算到电池包层级：`U_pack=N_s*OCV_cell`、`R_pack=(N_s/N_p)*R_cell`、`Q_pack=N_p*Q_cell`。",
        "- `SOC/T_bat` 保留 SOS2 断点并由二维权重 `w_st` 映射；电流不再使用全局三维 `w_pack`，改用 `rho_i2[t,soc,temp,current]` 将电流分摊到 SOC-T 网格。",
        "- `I_bat = sum(I_k*rho_i2)`，`Q_gen_pack = sum(R0_ab*I_k^2*rho_i2)`，`P_dc_pack = sum((OCV_a*I_k - R0_ab*I_k^2)*rho_i2)`；`--current-mode continuous` 允许电流在断点间线性组合，`--current-mode discrete` 强制每个时段只选择一个电流断点。",
        "- PCS 损耗按文档定义为 `P_PCS,Loss = μ·|P_bat|`，引入 `P_dc_abs >= ±P_dc_pack`，`P_BESS = P_dc_pack - μ·P_dc_abs - aux`，充放电对称损耗。",
        "- 调度曲线最后一点按初始 SOC/液冷罐温度/舱体温度设置硬约束；`SOC_end/T_*_end` 保留为末步动力学诊断值。",
        "- 电加热和泵耗进入功率平衡；原富余风光预热项保留，但新增全时段温度区间、终端储热和外循环散热惩罚。",
        "- 外循环 `Q_tamb>0` 被计入热量倾倒诊断；当电池或水箱确实过热时仍可散热。",
        "",
    ]

    if result is None or not result.get("success"):
        fail = result or {}
        lines += [
            "## 求解状态",
            "",
            f"- 状态: `{fail.get('status', 'NOT_RUN')}`",
            f"- 耗时: {fail.get('time_s', 0.0):.2f} s",
            f"- 说明: {fail.get('message', '未得到可行解。')}",
        ]
    else:
        checks = result["checks"]
        lines += [
            "## 求解状态",
            "",
            f"- 状态: `{result['status']}`",
            f"- 目标值: {result['objective']:.6g}",
            f"- MIP gap: {result['gap'] * 100:.3f}%",
            f"- 求解耗时: {result['time_s']:.2f} s",
            f"- 求解器: `{getattr(args, 'solver', 'gurobi')}`",
            f"- 时间限制: {args.time_limit:.0f} s",
            f"- 柴油消耗: {result['fuel_kg']:.3f} kg",
            f"- 弃风弃光电量: {result['curt_kwh']:.3f} kWh",
            f"- 热控耗电: {result['heat_kwh']:.3f} kWh",
            f"- 外循环散热/热量倾倒: {result['heat_dump_kwh_th']:.3f} kWh_th",
            f"- 富余风光时段: {checks['renewable_surplus_hours']:.3f} h",
            f"- 富余风光时段内启用电加热: {checks['heated_during_surplus_hours']:.3f} h",
            f"- 调度末端 SOC: {result['SOC'][-1] * 100:.3f}%",
            f"- 末步后 SOC 诊断值: {result['SOC_end'] * 100:.3f}%",
            f"- 富余风光预热短缺项: {result['preheat_short_score']:.6f} ℃·h",
            f"- 电池舒适区偏差项: {result['bat_band_score']:.6f} ℃·h",
            f"- 液冷罐储热区偏差项: {result['tank_band_score']:.6f} ℃·h",
            f"- 舱体/内壁储热区偏差项: {result['cont_band_score']:.6f} ℃·h",
            f"- 过热偏差项: {result['hot_score']:.6f} ℃·h",
            f"- 终端储热短缺项: {result['terminal_heat_short_score']:.6f} ℃",
            f"- 启停变化惩罚项: {result['switch_score']:.6f}",
            f"- 温度跳变惩罚项: {result['temp_ramp_score']:.6f} ℃",
            "",
            "## 线性化和物理残差",
            "",
            f"- R0 最大/平均误差: {checks['r0_max_mohm']:.6f} / {checks['r0_avg_mohm']:.6f} mΩ",
            f"- P_dc_pack 最大/平均误差: {checks['pdc_pack_max_kw']:.6f} / {checks['pdc_pack_avg_kw']:.6f} kW",
            f"- Q_gen_pack 最大/平均误差: {checks['qgen_pack_max_kw']:.6f} / {checks['qgen_pack_avg_kw']:.6f} kW",
            f"- PWL 物理复算偏差(P_BESS) 最大/平均: {checks['pbess_physical_max_kw']:.6f} / {checks['pbess_physical_avg_kw']:.6f} kW",
            f"- 模型严格功率平衡最大残差: {checks['model_balance_max_kw']:.9f} kW",
            f"- PWL 物理复算偏差传导到功率平衡的最大值: {checks['physical_balance_max_kw']:.6f} kW",
            f"- 柴油机最大相邻出力变化: {checks['dg_ramp_max_kw_per_step']:.6f} kW/步",
            "",
            "## 变量范围检查",
            "",
            f"- SOC 范围: {checks['soc_min']:.4f} - {checks['soc_max']:.4f}",
            f"- 电池包电流范围: {checks['i_pack_min_a']:.3f} - {checks['i_pack_max_a']:.3f} A",
            f"- 温度电流限值范围(充/放): {checks['charge_current_limit_min_a']:.3f} - {checks['charge_current_limit_max_a']:.3f} A / {checks['discharge_current_limit_min_a']:.3f} - {checks['discharge_current_limit_max_a']:.3f} A",
            f"- 温度电流限值最大越限(充/放): {checks['charge_current_limit_violation_max_a']:.6f} / {checks['discharge_current_limit_violation_max_a']:.6f} A",
            f"- 电池温度范围: {checks['tbat_min_c']:.3f} - {checks['tbat_max_c']:.3f} ℃",
            f"- 液冷罐温度范围: {checks['ttank_min_c']:.3f} - {checks['ttank_max_c']:.3f} ℃",
            f"- 舱体温度范围: {checks['tcont_min_c']:.3f} - {checks['tcont_max_c']:.3f} ℃",
            f"- 液冷罐最大预热短缺: {checks['tank_preheat_short_max_c']:.3f} ℃",
            f"- 舱体/内壁最大预热短缺: {checks['cont_preheat_short_max_c']:.3f} ℃",
            f"- 电池舒适区平均/最大偏差: {checks['bat_band_dev_avg_c']:.3f} / {checks['bat_band_dev_max_c']:.3f} ℃",
            f"- 液冷罐储热区平均/最大偏差: {checks['tank_band_dev_avg_c']:.3f} / {checks['tank_band_dev_max_c']:.3f} ℃",
            f"- 舱体/内壁储热区平均/最大偏差: {checks['cont_band_dev_avg_c']:.3f} / {checks['cont_band_dev_max_c']:.3f} ℃",
            f"- 终端液冷罐/舱体储热短缺: {checks['tank_terminal_short_c']:.3f} / {checks['cont_terminal_short_c']:.3f} ℃",
            f"- 液冷罐/舱体最大相邻温度跳变: {checks['tank_temp_step_max_c']:.3f} / {checks['cont_temp_step_max_c']:.3f} ℃",
            f"- 启停变化总次数代理值: {checks['switch_score']:.3f}",
            "",
            "## 储热利用",
            "",
            f"- 液冷罐相对环境最大储热: {checks['tank_storage_from_amb_max_kwh_th']:.3f} kWh_th",
            f"- 舱体/内壁相对环境最大储热: {checks['cont_storage_from_amb_max_kwh_th']:.3f} kWh_th",
            f"- 液冷罐超过预热目标最大储热: {checks['tank_storage_over_target_max_kwh_th']:.3f} kWh_th",
            f"- 舱体/内壁超过预热目标最大储热: {checks['cont_storage_over_target_max_kwh_th']:.3f} kWh_th",
            f"- 液冷罐可用储热最大/末端: {checks['tank_useful_storage_max_kwh_th']:.3f} / {checks['tank_useful_storage_end_kwh_th']:.3f} kWh_th",
            f"- 舱体/内壁可用储热最大/末端: {checks['cont_useful_storage_max_kwh_th']:.3f} / {checks['cont_useful_storage_end_kwh_th']:.3f} kWh_th",
            f"- 外循环散热总量/峰值: {checks['heat_dump_kwh_th']:.3f} kWh_th / {checks['heat_dump_max_kw_th']:.3f} kW_th",
            "",
            "## 输出文件",
            "",
            f"- 结果工作簿: `{result.get('state_workbook', '')}`",
            "",
            "## 结果观察与旧版对比",
            "",
            "- 原 `20260505` 版本只在富余风光时段惩罚预热短缺，全日 7200 秒结果为 `TIME_LIMIT`、gap `6.720%`、柴油 `303.112 kg`、弃风弃光 `0.295 kWh`、热控耗电 `140.000 kWh`。",
            f"- 本 `20260512` 更多冷却液版全日结果为 `{result['status']}`、gap `{result['gap'] * 100:.3f}%`、柴油 `{result['fuel_kg']:.3f} kg`、弃风弃光 `{result['curt_kwh']:.3f} kWh`、热控耗电 `{result['heat_kwh']:.3f} kWh`。",
            "- 储热保持方面有改善：旧版液冷罐温度范围约 `-44.444 - 57.026 ℃`，舱体温度范围约 `-25.742 - 59.874 ℃`；新版温度范围见上文变量检查，终端储热短缺也已单独列出。",
            "- 经济性方面可能变差：新版为了保留终端热储备并减少无价值散热，可能允许更多弃风弃光，柴油消耗也可能上升。因此该版本更适合作为“热安全/储热保持优先”的策略验证，不应直接视为最终经济最优方案。",
            "- 若全日模型以 `TIME_LIMIT` 结束，当前图表应按可行解解释，而不是按严格全局最优解释。后续工程结论建议进一步加密或改造包层级 PWL，并用分解式电-热协调/MPC 降低单体 MILP 难度。",
            f"- 若温度曲线仍存在局部尖峰，说明当前舱体热容较小、{args.dt_minutes:.0f} 分钟离散步长和二进制加热/泵控制会造成跳变。进一步工程化时应增加热惯性辨识、限制温度变率，或把舱体/内壁拆成独立热容状态。",
        ]
    path.write_text("\n".join(lines), encoding="utf-8")


def get_output_paths(mode: str, output_dir: Path | None = None, dt_minutes: float = 15.0) -> dict[str, Path]:
    base = output_dir or ROOT
    base.mkdir(parents=True, exist_ok=True)
    if mode in {"test_1h", "test_4h"}:
        tag = mode
    else:
        tag = f"{float(dt_minutes):g}min"
    return {
        "state_workbook": base / f"optimization_results_perspective_i2r_{tag}_20260530.xlsx",
    }


def solve_with_backend(
    solver_backend: str,
    p: SimpleNamespace,
    data: dict[str, np.ndarray | float | int],
    bp: Breakpoints,
    time_limit: float,
    mip_gap: float,
    *,
    build_only: bool = False,
) -> dict:
    if solver_backend == "gurobi":
        return solve_milp(
            p,
            data,
            bp,
            time_limit,
            mip_gap,
            feasibility_focus=True,
            build_only=build_only,
        )
    if solver_backend == "cplex_native":
        return solve_milp_cplex_native(
            p,
            data,
            bp,
            time_limit,
            mip_gap,
            build_only=build_only,
        )
    if solver_backend == "mosek_native":
        return solve_milp_mosek_native(
            p,
            data,
            bp,
            time_limit,
            mip_gap,
            build_only=build_only,
        )
    raise ValueError(f"Unsupported solver backend: {solver_backend}")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Route 13 perspective I2R BESS electro-thermal MILP.")
    parser.add_argument(
        "--mode",
        choices=["full_more_coolant_20260512", "dayahead_24h", "minute_more_coolant_20260512", "test_1h", "test_4h"],
        default="dayahead_24h",
        help="Solve mode. Default is dayahead_24h: 15-minute, 24-hour single MILP.",
    )
    parser.add_argument("--params", type=Path, default=DEFAULT_PARAM_PATH, help="JSON or Excel parameter file.")
    parser.add_argument("--config-file", type=Path, default=None, help="JSON compute-parameter file. Command-line options override values from this file.")
    parser.add_argument("--time-limit", type=float, default=300.0, help="Solver time limit in seconds. Use <=0 for no TimeLimit where supported.")
    parser.add_argument("--mip-gap", type=float, default=0.05, help="Accepted MIP gap.")
    parser.add_argument(
        "--solver",
        choices=["auto", "gurobi", "cplex", "mosek"],
        default="auto",
        help="MILP solver. auto uses native Gurobi when available, then native CPLEX, then native MOSEK.",
    )
    parser.add_argument("--hours", type=float, default=None, help="Optional horizon override in hours.")
    parser.add_argument("--dt-minutes", type=float, default=15.0, help="Dispatch interval in minutes. Default is 15; use 5 or 1 for finer dispatch.")
    parser.add_argument("--current-segments", type=int, default=10, help="Current segmentation count for I^2R/Pdc PWL. Default is 10 segments, which creates 11 current breakpoints.")
    parser.add_argument("--current-mode", choices=["continuous", "discrete"], default="continuous", help="Treat battery current as a continuous PWL variable or force it to one discrete breakpoint per period.")
    parser.add_argument("--soc-grid-width", type=float, default=0.1, help="SOC lookup grid width for R0/OCV linearization. Default is 0.1; use 0.05 for finer SOC lookup.")
    parser.add_argument("--initial-soc", type=float, default=None, help="Battery initial SOC. The dispatch terminal SOC is constrained to the same value.")
    parser.add_argument("--initial-t-bat-c", type=float, default=None, help="Initial cell temperature in C.")
    parser.add_argument("--initial-t-tank-c", type=float, default=None, help="Initial liquid tank temperature in C. The dispatch terminal liquid tank temperature is constrained to the same value.")
    parser.add_argument("--initial-t-cont-c", type=float, default=None, help="Initial container temperature in C. The dispatch terminal container temperature is constrained to the same value.")
    parser.add_argument("--current-points", type=int, default=None, help="Deprecated: current breakpoint count. Prefer --current-segments.")
    parser.add_argument("--i-points", type=int, default=None, help="Deprecated alias for --current-points.")
    parser.add_argument("--strict-current-sos2", action="store_true", help="Add SOS2 constraints to each SOC-T current split; intended for short tests.")
    parser.add_argument("--build-only", action="store_true", help="Build model and write model stats without optimizing.")
    parser.add_argument("--no-plots", action="store_true", help="Skip plot generation.")
    parser.add_argument(
        "--experiment",
        choices=[
            "baseline",
            "reduced_pwl_points",
            "thermal_binary_20min",
            "thermal_binary_30min",
            "no_temp_ramp_penalty",
            "no_switch_penalty",
            "simplified_storage_objective",
            "fuel_only_hifi_full",
            "fuel_only_reduced_i2",
            "fuel_only_block20",
            "fuel_only_block30",
            "fuel_only_tight",
            "fuel_only_tight_block20",
            "fuel_only_simplified_storage",
            "fuel_only_no_storage_block20",
            "fuel_only_discrete_i21_block20",
            "perspective_i2r_block20",
            "perspective_i2r_strict_block20",
        ],
        default="perspective_i2r_block20",
        help="Diagnostic experiment switch.",
    )
    parser.add_argument("--mip-focus", type=int, choices=[1, 2, 3], default=1, help="Gurobi-native MIPFocus setting.")
    parser.add_argument("--cuts", type=int, choices=[0, 1, 2], default=None, help="Gurobi-native Cuts setting.")
    parser.add_argument("--heuristics", type=float, default=0.3, help="Gurobi-native Heuristics setting.")
    parser.add_argument("--threads", type=int, default=os.cpu_count() or 0, help="Solver thread count. Default requests all logical CPUs where supported.")
    parser.add_argument("--log-file", type=Path, default=None, help="Gurobi-native log file path.")
    parser.add_argument("--diagnostics-json", type=Path, default=None, help="Write compact result diagnostics JSON.")
    parser.add_argument("--model-stats-json", type=Path, default=None, help="Write model size diagnostics JSON after model build.")
    parser.add_argument("--progress-json", type=Path, default=None, help="Write incremental progress JSON for Web UI polling.")
    parser.add_argument(
        "--tight-temp-bounds",
        action="store_true",
        help="Use tighter safe temperature bounds to reduce Big-M ranges; defaults emphasize liquid tank and container bounds.",
    )
    parser.add_argument("--bat-temp-min", type=float, default=None, help="Optional tightened battery lower temperature bound in C.")
    parser.add_argument("--bat-temp-max", type=float, default=None, help="Optional tightened battery upper temperature bound in C.")
    parser.add_argument("--tank-temp-min", type=float, default=None, help="Optional tightened liquid tank lower temperature bound in C.")
    parser.add_argument("--tank-temp-max", type=float, default=None, help="Optional tightened liquid tank upper temperature bound in C.")
    parser.add_argument("--cont-temp-min", type=float, default=None, help="Optional tightened container lower temperature bound in C.")
    parser.add_argument("--cont-temp-max", type=float, default=None, help="Optional tightened container upper temperature bound in C.")
    parser.add_argument("--output-dir", type=Path, default=None, help="Directory for summary and figures.")
    args = parser.parse_args(argv)
    apply_compute_config_file(args, argv)
    return args


def apply_compute_config_file(args: argparse.Namespace, argv: list[str]) -> None:
    if not args.config_file:
        return
    config_path = Path(args.config_file)
    if not config_path.exists():
        raise FileNotFoundError(f"Compute config file not found: {config_path}")
    config = json.loads(config_path.read_text(encoding="utf-8-sig"))
    if not isinstance(config, dict):
        raise ValueError(f"Compute config file must contain a JSON object: {config_path}")
    provided = {item.split("=", 1)[0] for item in argv if item.startswith("--")}
    mapping = {
        "mode": ("--mode",),
        "time_limit": ("--time-limit",),
        "mip_gap": ("--mip-gap",),
        "solver": ("--solver",),
        "hours": ("--hours",),
        "dt_minutes": ("--dt-minutes",),
        "current_segments": ("--current-segments",),
        "current_mode": ("--current-mode",),
        "soc_grid_width": ("--soc-grid-width",),
        "initial_soc": ("--initial-soc",),
        "initial_t_bat_c": ("--initial-t-bat-c",),
        "initial_t_tank_c": ("--initial-t-tank-c",),
        "initial_t_cont_c": ("--initial-t-cont-c",),
        "strict_current_sos2": ("--strict-current-sos2",),
        "build_only": ("--build-only",),
        "no_plots": ("--no-plots",),
        "experiment": ("--experiment",),
        "mip_focus": ("--mip-focus",),
        "cuts": ("--cuts",),
        "heuristics": ("--heuristics",),
        "threads": ("--threads",),
        "tight_temp_bounds": ("--tight-temp-bounds",),
        "bat_temp_min": ("--bat-temp-min",),
        "bat_temp_max": ("--bat-temp-max",),
        "tank_temp_min": ("--tank-temp-min",),
        "tank_temp_max": ("--tank-temp-max",),
        "cont_temp_min": ("--cont-temp-min",),
        "cont_temp_max": ("--cont-temp-max",),
    }
    for key, flags in mapping.items():
        if key not in config or any(flag in provided for flag in flags):
            continue
        setattr(args, key, config[key])


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    progress_json = args.progress_json
    write_progress(
        progress_json,
        stage="starting",
        message="开始读取输入参数",
        solver_requested=args.solver,
        params_path=args.params,
        config_file=args.config_file,
        output_dir=args.output_dir,
    )
    p = load_params(args.params)
    apply_initial_state_config(p, args)
    write_progress(
        progress_json,
        stage="parameters_loaded",
        message="已读取输入参数",
        params_path=args.params,
        source_excel=getattr(p, "source_excel", ""),
    )
    apply_diagnostic_experiment(p, args.experiment)
    if args.experiment in {"fuel_only_tight", "fuel_only_tight_block20"}:
        args.tight_temp_bounds = True
    explicit_temp_bounds = {
        "bat_min": args.bat_temp_min,
        "bat_max": args.bat_temp_max,
        "tank_min": args.tank_temp_min,
        "tank_max": args.tank_temp_max,
        "cont_min": args.cont_temp_min,
        "cont_max": args.cont_temp_max,
    }
    if args.tight_temp_bounds or any(v is not None for v in explicit_temp_bounds.values()):
        apply_tight_temp_bounds(p, use_defaults=args.tight_temp_bounds, **explicit_temp_bounds)
    else:
        validate_temperature_bounds(p)
        p.diag_tight_temp_bounds = False
        p.diag_temperature_bounds = temperature_bounds_snapshot(p)
    if args.dt_minutes <= 0:
        raise ValueError("--dt-minutes must be positive.")
    if args.current_segments <= 0:
        raise ValueError("--current-segments must be positive.")
    if args.soc_grid_width <= 0:
        raise ValueError("--soc-grid-width must be positive.")
    p.diag_mip_focus = args.mip_focus
    p.diag_cuts = args.cuts
    p.diag_heuristics = args.heuristics
    p.diag_threads = args.threads
    if args.strict_current_sos2:
        p.diag_strict_current_sos2 = True
    p.diag_current_mode = args.current_mode
    p.diag_log_file = args.log_file
    p.diag_model_stats_json = args.model_stats_json
    if args.experiment in {"reduced_pwl_points", "fuel_only_reduced_i2"} and args.i_points is None:
        args.i_points = 2
    current_points = args.i_points if args.i_points is not None else args.current_points
    current_segments = None if current_points is not None else args.current_segments
    bp = make_breakpoints(
        p,
        args.mode,
        current_points,
        dt_minutes=args.dt_minutes,
        current_segments=current_segments,
        soc_grid_width=args.soc_grid_width,
    )

    if args.hours is not None:
        horizon = float(args.hours)
    elif args.mode == "test_1h":
        horizon = 1.0
    elif args.mode == "test_4h":
        horizon = 4.0
    else:
        horizon = 24.0
    if args.dt_minutes <= 1.0 and horizon == 24.0:
        print("Warning: a 24h horizon with <=1 minute dispatch creates at least 1440 periods and may take a long time.")

    data = interpolate_profiles(p, bp.dt_hours, horizon)
    args.dt_minutes = float(data["dt"]) / 60.0
    paths = get_output_paths(args.mode, args.output_dir, args.dt_minutes)

    max_dg_points = max(len(u["powers_w"]) for u in p.diesel_units)
    est = estimate_model_size(bp, int(data["N"]), len(p.diesel_units), max_dg_points, current_mode=args.current_mode)
    candidates = solver_candidates(args.solver)
    solver_order = [name for name, _ in candidates]
    write_progress(
        progress_json,
        stage="model_prepared",
        message="已完成断点、时序数据和模型规模估算",
        solver_requested=args.solver,
        solver_order=solver_order,
        experiment=args.experiment,
        mode=args.mode,
        hours=horizon,
        steps=int(data["N"]),
        dt_minutes=float(data["dt"]) / 60.0,
        estimated_scale=est,
        current_segments=bp.n_i - 1,
        current_breakpoints=bp.n_i,
        soc_grid_width=args.soc_grid_width,
        current_mode=args.current_mode,
        strict_current_sos2=bool(p.diag_strict_current_sos2),
    )
    print("=" * 72)
    print("Route 13 perspective I2R BESS MILP")
    print("=" * 72)
    print(f"Parameter file: {args.params}")
    print(f"Experiment: {args.experiment}; objective=fuel_kg_only; strict_power_balance=True")
    print(f"MIPFocus={args.mip_focus}; Cuts={args.cuts}; Heuristics={args.heuristics:g}; tight_temp_bounds={args.tight_temp_bounds}")
    print(f"Solver requested={args.solver}; order={', '.join(solver_order)}")
    print(
        f"Threads requested={args.threads}; current_segments={bp.n_i - 1}; "
        f"current_breakpoints={bp.n_i}; soc_grid_width={args.soc_grid_width:g}; "
        f"current_mode={args.current_mode}; strict_current_sos2={p.diag_strict_current_sos2}"
    )
    print(f"Mode: {args.mode}; horizon={horizon:g} h; steps={data['N']}; dt={data['dt'] / 60:g} min")
    print(f"Diesel units={len(p.diesel_units)}, wind rated={sum(u['rated_kw'] for u in p.wind_units):.1f} kW, pv rated={sum(u['rated_kw'] for u in p.pv_units):.1f} kW")
    print(f"Estimated scale: {est}")

    solver_attempts: list[dict[str, object]] = []
    result: dict | None = None
    selected_solver_name = ""
    selected_solver_backend = ""
    for attempt_index, (solver_name, solver_backend) in enumerate(candidates, start=1):
        selected_solver_name = solver_name
        selected_solver_backend = solver_backend
        attempt_started = time.time()
        attempt: dict[str, object] = {
            "attempt": attempt_index,
            "solver": solver_name,
            "backend": solver_backend,
            "started_at": now_iso(),
            "status": "STARTED",
            "success": False,
        }
        solver_attempts.append(attempt)
        print("-" * 72)
        print(f"Solver attempt {attempt_index}/{len(candidates)}: {solver_name} ({solver_backend})")
        write_progress(
            progress_json,
            stage="solver_attempt_started",
            message=f"启动求解器：{solver_name}",
            active_solver=solver_name,
            active_backend=solver_backend,
            solver_attempts=solver_attempts,
        )
        try:
            write_progress(
                progress_json,
                stage="solving",
                message=f"{solver_name} 正在优化求解主模型",
                active_solver=solver_name,
                active_backend=solver_backend,
                solver_attempts=solver_attempts,
            )
            result = solve_with_backend(
                solver_backend,
                p,
                data,
                bp,
                args.time_limit,
                args.mip_gap,
                build_only=args.build_only,
            )
            if result is None:
                result = {
                    "success": False,
                    "status": "NO_RESULT",
                    "time_s": time.time() - attempt_started,
                    "message": f"{solver_name} did not return a result.",
                }
        except Exception as exc:
            result = {
                "success": False,
                "status": "SOLVER_EXCEPTION",
                "time_s": time.time() - attempt_started,
                "message": str(exc),
            }
            print(f"Solver attempt failed with exception: {solver_name} - {exc}")

        result["solver_requested"] = args.solver
        result["solver_used"] = solver_name
        result["solver_backend"] = solver_backend
        attempt.update(
            {
                "finished_at": now_iso(),
                "elapsed_s": round(time.time() - attempt_started, 3),
                "status": result.get("status"),
                "success": bool(result.get("success", False)),
                "message": result.get("message", ""),
                "result": compact_result_for_progress(result),
            }
        )
        terminal_stage = "solver_attempt_succeeded" if result.get("success") or result.get("status") == "BUILD_ONLY" else "solver_attempt_failed"
        write_progress(
            progress_json,
            stage=terminal_stage,
            message=f"{solver_name} 求解状态：{result.get('status')}",
            active_solver=solver_name,
            active_backend=solver_backend,
            solver_attempts=solver_attempts,
            last_result=compact_result_for_progress(result),
        )
        if result.get("success") or result.get("status") == "BUILD_ONLY":
            break
        if attempt_index < len(candidates):
            print(f"Solver {solver_name} failed: {result.get('status')} - {result.get('message', '')}")
            print("Trying next solver backend.")

    if result is None:
        result = {
            "success": False,
            "status": "NO_SOLVER_ATTEMPT",
            "message": "No solver attempt was executed.",
            "time_s": 0.0,
        }
    result["solver_requested"] = args.solver
    result["solver_order"] = solver_order
    result["solver_used"] = selected_solver_name
    result["solver_backend"] = selected_solver_backend
    result["solver_attempts"] = make_json_safe(solver_attempts)
    if result.get("success"):
        write_progress(
            progress_json,
            stage="writing_outputs",
            message="正在写入结果工作簿（含统计信息和调度曲线）",
            active_solver=result.get("solver_used"),
            active_backend=result.get("solver_backend"),
            output_paths=paths,
            last_result=compact_result_for_progress(result),
        )
        result["state_workbook"] = str(paths["state_workbook"])
        state_workbook = write_detailed_results_workbook(p, result, paths["state_workbook"])
        result["state_workbook"] = str(state_workbook)

    if args.diagnostics_json:
        args.diagnostics_json.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "experiment": args.experiment,
            "mode": args.mode,
            "hours": horizon,
            "dt_minutes": args.dt_minutes,
            "time_limit": args.time_limit,
            "mip_gap_target": args.mip_gap,
            "solver_requested": args.solver,
            "solver_order": solver_order,
            "solver_used": result.get("solver_used"),
            "solver_backend": result.get("solver_backend"),
            "solver_attempts": result.get("solver_attempts"),
            "status": result.get("status"),
            "success": result.get("success", False),
            "objective": result.get("objective"),
            "best_bound": result.get("best_bound"),
            "gap": result.get("gap"),
            "time_s": result.get("time_s"),
            "node_count": result.get("node_count"),
            "fuel_kg": result.get("fuel_kg"),
            "curt_kwh": result.get("curt_kwh"),
            "heat_kwh": result.get("heat_kwh"),
            "objective_breakdown": result.get("objective_breakdown"),
            "checks": result.get("checks"),
            "model_stats": result.get("model_stats"),
            "state_workbook": result.get("state_workbook"),
            "output_paths": make_json_safe(paths),
        }
        args.diagnostics_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    if result.get("success"):
        write_progress(
            progress_json,
            stage="completed",
            message="优化求解完成",
            active_solver=result.get("solver_used"),
            active_backend=result.get("solver_backend"),
            output_paths=paths,
            final_result=compact_result_for_progress(result),
            solver_attempts=result.get("solver_attempts"),
        )
        checks = result["checks"]
        print("-" * 72)
        print(f"Status: {result['status']}, gap={result['gap'] * 100:.3f}%, bound={result['best_bound']:.6g}, time={result['time_s']:.1f}s")
        print(
            f"Fuel={result['fuel_kg']:.2f} kg, curtailment={result['curt_kwh']:.2f} kWh, "
            f"dispatch final SOC={result['SOC'][-1] * 100:.2f}%, post-step SOC={result['SOC_end'] * 100:.2f}%"
        )
        print(f"Model balance residual max={checks['model_balance_max_kw']:.9f} kW")
        print(f"PWL physical replay deviation max/avg={checks['pbess_physical_max_kw']:.4f}/{checks['pbess_physical_avg_kw']:.4f} kW")
        print(f"Thermal surplus hours={checks['renewable_surplus_hours']:.2f} h, heated in surplus={checks['heated_during_surplus_hours']:.2f} h")
        print(f"Wrote: {paths['state_workbook']}")
        return 0

    print(f"Solve failed: {result.get('status')} - {result.get('message')}")
    if result.get("status") == "BUILD_ONLY":
        write_progress(
            progress_json,
            stage="build_only",
            message="模型构建完成，未执行优化",
            active_solver=result.get("solver_used"),
            active_backend=result.get("solver_backend"),
            final_result=compact_result_for_progress(result),
            solver_attempts=result.get("solver_attempts"),
        )
        return 0
    write_progress(
        progress_json,
        stage="failed",
        message=f"优化求解失败：{result.get('status')}",
        active_solver=result.get("solver_used"),
        active_backend=result.get("solver_backend"),
        final_result=compact_result_for_progress(result),
        solver_attempts=result.get("solver_attempts"),
    )
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
