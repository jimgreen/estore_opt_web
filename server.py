#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Local web system for BESS electro-thermal optimization scenarios."""

from __future__ import annotations

import argparse
import cgi
import csv
import hashlib
import hmac
import importlib.util
import io
import json
import math
import os
import queue
import secrets
import shutil
import sqlite3
import subprocess
import sys
import threading
import time
import traceback
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from http import HTTPStatus
from http.cookies import SimpleCookie
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote, urlparse

import numpy as np

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.workbook import Workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"openpyxl is required to run estore_opt_web: {exc}") from exc


WEB_ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = WEB_ROOT.parent
DEFAULT_PARAMS = PROJECT_ROOT / "5.4模型定义_低温风光扩容_更多冷却液_20260512.xlsx"
SOLVER_SCRIPT = WEB_ROOT / "solve.py"
SCHEME_ROOT = WEB_ROOT / "estore_schemes"
RUN_ROOT = WEB_ROOT / "estore_runs"
VERIFY_ROOT = WEB_ROOT / "estore_verifications"
VERIFY_WORKER_SCRIPT = WEB_ROOT / "verification_worker.py"
USER_DB_PATH = Path(os.environ.get("ESTORE_OPT_USER_DB", WEB_ROOT / "estore_opt_users.sqlite3"))
SESSION_COOKIE_NAME = "estore_opt_session"
SESSION_MAX_AGE_SECONDS = int(os.environ.get("ESTORE_OPT_SESSION_MAX_AGE", "604800"))
LOCAL_AUTH_BYPASS_ENABLED = os.environ.get("ESTORE_OPT_LOCAL_AUTH_BYPASS", "").lower() in {"1", "true", "yes"}
PARAM_FILE_NAME = "params.xlsx"
COMPUTE_CONFIG_FILE_NAME = "compute_config.json"
RESULT_DATA_FILE_NAME = "optimization_result_data.json"
RESULT_STATISTICS_FILE_NAME = "optimization_statistics.json"
RESULT_TIMESERIES_CSV_FILE_NAME = "optimization_timeseries.csv"
DISPATCH_SCHEDULE_FILE_NAME = "dispatch_schedule.xlsx"
COMPUTE_CONFIG_SHEET_NAME = "计算参数"
DISPATCH_SCHEDULE_SHEET_NAME = "调度控制曲线"
CELL_SHEET_NAME = "电芯"
CELL_RESISTANCE_SHEET_NAME = "电芯内阻"
CELL_CURRENT_LIMIT_SHEET_NAME = "电芯电流限值"
CELL_CURRENT_LIMIT_HEADERS = ["温度(℃)", "最大充电电流(A)", "最大放电电流(A)"]
META_FILE_NAME = "scheme.json"
OVERVIEW_HEADER_COLS = 32
IS_WORKER_IMPORT = os.environ.get("ESTORE_OPT_WORKER", "").lower() in {"1", "true", "yes"}

DEFAULT_COMPUTE_CONFIG: dict[str, Any] = {
    "solver": "auto",
    "mode": "dayahead_24h",
    "experiment": "perspective_i2r_block20",
    "dt_minutes": 60,
    "time_limit": 300,
    "mip_gap": 0.05,
    "current_segments": 5,
    "current_mode": "continuous",
    "soc_grid_width": 0.2,
    "threads": 8,
    "mip_focus": 1,
    "cuts": None,
    "heuristics": 0.3,
    "hours": None,
    "initial_soc": 0.5,
    "initial_t_bat_c": 5.0,
    "initial_t_tank_c": 5.0,
    "initial_t_cont_c": 5.0,
    "strict_current_sos2": False,
    "tight_temp_bounds": True,
    "build_only": False,
    "no_plots": True,
}

COMPUTE_CONFIG_LABELS: dict[str, str] = {
    "solver": "求解器",
    "mode": "模式",
    "experiment": "实验配置",
    "dt_minutes": "时间步长(分钟)",
    "time_limit": "时间上限(秒)",
    "mip_gap": "MIP Gap",
    "current_segments": "电流分段数",
    "current_mode": "电流模式",
    "soc_grid_width": "SOC网格宽度",
    "threads": "线程数",
    "mip_focus": "MIP Focus",
    "cuts": "Cuts",
    "heuristics": "Heuristics",
    "hours": "时域覆盖(小时)",
    "initial_soc": "电池初始SOC",
    "initial_t_bat_c": "电芯初始温度(℃)",
    "initial_t_tank_c": "液冷罐初始温度(℃)",
    "initial_t_cont_c": "舱体初始温度(℃)",
    "strict_current_sos2": "严格电流SOS2",
    "tight_temp_bounds": "使用紧温度边界",
    "build_only": "仅建模检查",
    "no_plots": "跳过图形输出",
}


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def timestamp_log_line(message: Any) -> str:
    return f"[{now_text()}] {str(message).rstrip()}"


def truncate_log_line(line: str, limit: int = 300) -> str:
    if len(line) <= limit:
        return line
    prefix_len = 22 if line.startswith("[") and len(line) > 22 and line[20:22] == "] " else 0
    if prefix_len and limit > prefix_len + 8:
        tail_len = max(0, limit - prefix_len - 3)
        return f"{line[:prefix_len]}...{line[-tail_len:]}"
    return line[-limit:]


def json_response(handler: SimpleHTTPRequestHandler, payload: Any, status: int = 200) -> None:
    data = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def read_json(path: Path, default: Any = None) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


class UserStore:
    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path
        self.lock = threading.RLock()
        self._init_db()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path, timeout=20)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self) -> None:
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        with self._connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL,
                    salt TEXT NOT NULL,
                    role TEXT NOT NULL DEFAULT 'user',
                    created_at TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS sessions (
                    token TEXT PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    created_at REAL NOT NULL,
                    expires_at REAL NOT NULL,
                    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
                )
                """
            )
            conn.commit()

    def _hash_password(self, password: str, salt: str) -> str:
        digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), bytes.fromhex(salt), 180_000)
        return digest.hex()

    def _public_user(self, row: sqlite3.Row) -> dict[str, Any]:
        return {"id": row["id"], "username": row["username"], "role": row["role"], "created_at": row["created_at"]}

    def register(self, username: str, password: str) -> dict[str, Any]:
        clean = str(username or "").strip()
        if len(clean) < 3:
            raise ValueError("用户名至少需要 3 个字符")
        if len(str(password or "")) < 6:
            raise ValueError("密码至少需要 6 个字符")
        salt = secrets.token_hex(16)
        password_hash = self._hash_password(password, salt)
        with self.lock, self._connect() as conn:
            existing_count = conn.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"]
            role = "admin" if existing_count == 0 else "user"
            try:
                cursor = conn.execute(
                    "INSERT INTO users(username, password_hash, salt, role, created_at) VALUES(?,?,?,?,?)",
                    (clean, password_hash, salt, role, now_text()),
                )
                conn.commit()
            except sqlite3.IntegrityError as exc:
                raise ValueError("用户名已存在") from exc
            row = conn.execute("SELECT * FROM users WHERE id = ?", (cursor.lastrowid,)).fetchone()
            return self._public_user(row)

    def authenticate(self, username: str, password: str) -> dict[str, Any] | None:
        with self.lock, self._connect() as conn:
            row = conn.execute("SELECT * FROM users WHERE username = ?", (str(username or "").strip(),)).fetchone()
            if not row:
                return None
            candidate = self._hash_password(str(password or ""), row["salt"])
            if not hmac.compare_digest(candidate, row["password_hash"]):
                return None
            return self._public_user(row)

    def create_session(self, user_id: int) -> str:
        token = secrets.token_urlsafe(32)
        now = time.time()
        with self.lock, self._connect() as conn:
            conn.execute(
                "INSERT INTO sessions(token, user_id, created_at, expires_at) VALUES(?,?,?,?)",
                (token, user_id, now, now + SESSION_MAX_AGE_SECONDS),
            )
            conn.execute("DELETE FROM sessions WHERE expires_at < ?", (now,))
            conn.commit()
        return token

    def delete_session(self, token: str) -> None:
        if not token:
            return
        with self.lock, self._connect() as conn:
            conn.execute("DELETE FROM sessions WHERE token = ?", (token,))
            conn.commit()

    def user_for_session(self, token: str) -> dict[str, Any] | None:
        if not token:
            return None
        now = time.time()
        with self.lock, self._connect() as conn:
            row = conn.execute(
                """
                SELECT users.* FROM sessions
                JOIN users ON users.id = sessions.user_id
                WHERE sessions.token = ? AND sessions.expires_at >= ?
                """,
                (token, now),
            ).fetchone()
            if not row:
                conn.execute("DELETE FROM sessions WHERE token = ? OR expires_at < ?", (token, now))
                conn.commit()
                return None
            return self._public_user(row)


USER_STORE = None if IS_WORKER_IMPORT else UserStore(USER_DB_PATH)


def json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_safe(v) for v in value]
    if isinstance(value, np.ndarray):
        return json_safe(value.tolist())
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        v = float(value)
        return v if math.isfinite(v) else None
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    return value


def safe_name(value: str) -> str:
    text = "".join(ch for ch in str(value or "").strip() if ch not in '\\/:*?"<>|')
    text = text.replace("..", "").strip(" .")
    return text[:80]


def cell_to_json(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return value


def positive_int(value: Any, default: int, min_value: int = 1, max_value: int | None = None) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError):
        number = default
    number = max(min_value, number)
    if max_value is not None:
        number = min(number, max_value)
    return number


def scheme_dir(name: str) -> Path:
    clean = safe_name(name)
    if not clean:
        raise ValueError("方案名称不能为空")
    path = (SCHEME_ROOT / clean).resolve()
    if SCHEME_ROOT.resolve() not in path.parents and path != SCHEME_ROOT.resolve():
        raise ValueError("方案路径无效")
    return path


def scheme_params_path(name: str) -> Path:
    return scheme_dir(name) / PARAM_FILE_NAME


def scheme_compute_config_path(name: str) -> Path:
    return scheme_dir(name) / COMPUTE_CONFIG_FILE_NAME


def scheme_dispatch_schedule_path(name: str) -> Path:
    return scheme_dir(name) / DISPATCH_SCHEDULE_FILE_NAME


def normalize_compute_config(config: dict[str, Any] | None = None) -> dict[str, Any]:
    cfg = dict(DEFAULT_COMPUTE_CONFIG)
    for key, value in (config or {}).items():
        if key in cfg:
            cfg[key] = value
    int_keys = {"current_segments", "threads", "mip_focus", "cuts"}
    float_keys = {
        "dt_minutes",
        "time_limit",
        "mip_gap",
        "soc_grid_width",
        "heuristics",
        "hours",
        "initial_soc",
        "initial_t_bat_c",
        "initial_t_tank_c",
        "initial_t_cont_c",
    }
    bool_keys = {"strict_current_sos2", "tight_temp_bounds", "build_only", "no_plots"}
    for key in int_keys:
        if cfg.get(key) in ("", None):
            cfg[key] = None
        else:
            cfg[key] = int(cfg[key])
    for key in float_keys:
        if cfg.get(key) in ("", None):
            cfg[key] = None
        else:
            cfg[key] = float(cfg[key])
    for key in bool_keys:
        cfg[key] = bool(cfg.get(key))
    if cfg["dt_minutes"] is None or cfg["dt_minutes"] <= 0:
        raise ValueError("时间步长必须大于 0")
    if cfg["time_limit"] is None:
        cfg["time_limit"] = 300.0
    if cfg["mip_gap"] is None or cfg["mip_gap"] < 0:
        raise ValueError("MIP Gap 不能小于 0")
    if cfg["current_segments"] is None or cfg["current_segments"] <= 0:
        raise ValueError("电流分段数必须大于 0")
    if cfg["soc_grid_width"] is None or cfg["soc_grid_width"] <= 0:
        raise ValueError("SOC网格宽度必须大于 0")
    if cfg["initial_soc"] is None or cfg["initial_soc"] < 0.0 or cfg["initial_soc"] > 1.0:
        raise ValueError("电池初始SOC必须在 0 到 1 之间")
    if cfg["threads"] is None or cfg["threads"] <= 0:
        cfg["threads"] = max(1, os.cpu_count() or 1)
    if cfg["mip_focus"] not in {1, 2, 3}:
        cfg["mip_focus"] = 1
    if cfg["cuts"] not in {None, 0, 1, 2}:
        cfg["cuts"] = None
    if cfg["solver"] not in {"auto", "gurobi", "cplex", "mosek"}:
        cfg["solver"] = "auto"
    if cfg["current_mode"] not in {"continuous", "discrete"}:
        cfg["current_mode"] = "continuous"
    return cfg


def parse_compute_sheet_value(value: Any) -> Any:
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        if text.lower() in {"true", "yes", "是", "启用", "1"}:
            return True
        if text.lower() in {"false", "no", "否", "禁用", "0"}:
            return False
        try:
            if "." in text:
                return float(text)
            return int(text)
        except ValueError:
            return text
    return value


def excel_number(value: Any, default: float | None = None) -> float:
    if value is None or value == "":
        if default is None:
            raise ValueError("Excel 数值不能为空")
        return float(default)
    if isinstance(value, str):
        text = value.strip().replace("A", "").replace("a", "")
        return float(text)
    return float(value)


def temperature_from_header(value: Any) -> float:
    text = str(value).replace("℃", "").replace("°C", "").replace("C", "").strip()
    return float(text)


def ensure_cell_current_limit_sheet(path: Path) -> bool:
    if not path.exists():
        return False
    wb = load_workbook(path)
    try:
        if CELL_CURRENT_LIMIT_SHEET_NAME in wb.sheetnames:
            return False
        if CELL_SHEET_NAME not in wb.sheetnames or CELL_RESISTANCE_SHEET_NAME not in wb.sheetnames:
            return False
        cell_ws = wb[CELL_SHEET_NAME]
        r0_ws = wb[CELL_RESISTANCE_SHEET_NAME]
        temperatures: list[float] = []
        for col in range(2, r0_ws.max_column + 1):
            value = r0_ws.cell(1, col).value
            if value is not None:
                temperatures.append(temperature_from_header(value))
        if not temperatures:
            return False
        discharge_max = excel_number(cell_ws.cell(2, 7).value)
        charge_max = excel_number(cell_ws.cell(2, 8).value)
        insert_at = wb.sheetnames.index(CELL_SHEET_NAME) + 1
        ws = wb.create_sheet(CELL_CURRENT_LIMIT_SHEET_NAME, insert_at)
        ws.append(CELL_CURRENT_LIMIT_HEADERS)
        for temp in temperatures:
            ws.append([temp, charge_max, discharge_max])
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        wb.save(path)
        return True
    finally:
        wb.close()


def compute_config_from_workbook(name: str) -> dict[str, Any] | None:
    path = scheme_params_path(name)
    if not path.exists():
        return None
    wb = load_workbook(path, read_only=True, data_only=True)
    if COMPUTE_CONFIG_SHEET_NAME not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[COMPUTE_CONFIG_SHEET_NAME]
    payload: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0] if row else None
        if key in DEFAULT_COMPUTE_CONFIG:
            payload[str(key)] = parse_compute_sheet_value(row[1] if len(row) > 1 else None)
    wb.close()
    return payload


def write_compute_config_sheet(name: str, config: dict[str, Any]) -> None:
    path = scheme_params_path(name)
    wb = load_workbook(path)
    if COMPUTE_CONFIG_SHEET_NAME in wb.sheetnames:
        ws = wb[COMPUTE_CONFIG_SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(COMPUTE_CONFIG_SHEET_NAME)
    ws.append(["参数", "值", "说明"])
    for key, default_value in DEFAULT_COMPUTE_CONFIG.items():
        ws.append([key, config.get(key, default_value), COMPUTE_CONFIG_LABELS.get(key, key)])
    wb.save(path)
    wb.close()


def ensure_scheme_compute_config(name: str) -> dict[str, Any]:
    path = scheme_compute_config_path(name)
    sheet_config = compute_config_from_workbook(name)
    if sheet_config is None:
        current = read_json(path, None)
        cfg = normalize_compute_config(current if isinstance(current, dict) else {})
        write_compute_config_sheet(name, cfg)
    else:
        cfg = normalize_compute_config(sheet_config)
        if any(key not in sheet_config for key in DEFAULT_COMPUTE_CONFIG):
            write_compute_config_sheet(name, cfg)
    current_json = read_json(path, None)
    if current_json != cfg:
        write_json(path, cfg)
    return cfg


def read_scheme_compute_config(name: str) -> dict[str, Any]:
    return ensure_scheme_compute_config(name)


def write_scheme_compute_config(name: str, config: dict[str, Any]) -> dict[str, Any]:
    cfg = normalize_compute_config(config)
    write_compute_config_sheet(name, cfg)
    write_json(scheme_compute_config_path(name), cfg)
    meta_path = scheme_dir(name) / META_FILE_NAME
    meta = read_json(meta_path, {}) or {}
    meta["updated_at"] = now_text()
    write_json(meta_path, meta)
    return cfg


def ensure_default_scheme() -> None:
    SCHEME_ROOT.mkdir(parents=True, exist_ok=True)
    RUN_ROOT.mkdir(parents=True, exist_ok=True)
    VERIFY_ROOT.mkdir(parents=True, exist_ok=True)
    if not DEFAULT_PARAMS.exists():
        return
    if any(path.is_dir() for path in SCHEME_ROOT.iterdir()):
        return
    create_scheme("默认方案", "由附件模型定义工作簿导入", DEFAULT_PARAMS)


def create_scheme(name: str, description: str = "", source: Path | None = None) -> dict[str, Any]:
    clean = safe_name(name)
    target = scheme_dir(clean)
    if target.exists():
        raise ValueError(f"方案已存在：{clean}")
    target.mkdir(parents=True, exist_ok=True)
    source_path = source or DEFAULT_PARAMS
    if not source_path.exists():
        wb = Workbook()
        wb.active.title = "系统定义"
        wb.save(target / PARAM_FILE_NAME)
        wb.close()
    else:
        shutil.copy2(source_path, target / PARAM_FILE_NAME)
    ensure_cell_current_limit_sheet(target / PARAM_FILE_NAME)
    write_scheme_compute_config(clean, normalize_compute_config())
    meta = {
        "name": clean,
        "description": description or "",
        "created_at": now_text(),
        "updated_at": now_text(),
        "source": str(source_path),
    }
    write_json(target / META_FILE_NAME, meta)
    return scheme_summary(clean)


def copy_scheme(source_name: str, target_name: str, description: str = "") -> dict[str, Any]:
    source = scheme_params_path(source_name)
    if not source.exists():
        raise FileNotFoundError(f"源方案不存在：{source_name}")
    scheme = create_scheme(target_name, description, source)
    write_scheme_compute_config(scheme["name"], read_scheme_compute_config(source_name))
    source_dispatch = scheme_dispatch_schedule_path(source_name)
    if source_dispatch.exists():
        shutil.copy2(source_dispatch, scheme_dispatch_schedule_path(scheme["name"]))
    return scheme


def update_scheme_meta(old_name: str, new_name: str = "", description: str | None = None) -> dict[str, Any]:
    source = scheme_dir(old_name)
    if not source.exists():
        raise FileNotFoundError(f"方案不存在：{old_name}")
    clean_new = safe_name(new_name or old_name)
    if not clean_new:
        raise ValueError("方案名称不能为空")
    target = scheme_dir(clean_new)
    if target != source and target.exists():
        raise ValueError(f"方案已存在：{clean_new}")
    if target != source:
        source.rename(target)
    meta_path = target / META_FILE_NAME
    meta = read_json(meta_path, {}) or {}
    meta["name"] = clean_new
    if description is not None:
        meta["description"] = str(description or "")
    meta["updated_at"] = now_text()
    write_json(meta_path, meta)
    return scheme_summary(clean_new)


def scheme_summary(name: str) -> dict[str, Any]:
    path = scheme_dir(name)
    meta = read_json(path / META_FILE_NAME, {}) or {}
    params = path / PARAM_FILE_NAME
    dispatch = path / DISPATCH_SCHEDULE_FILE_NAME
    ensure_cell_current_limit_sheet(params)
    stat = params.stat() if params.exists() else None
    dispatch_stat = dispatch.stat() if dispatch.exists() else None
    config = ensure_scheme_compute_config(path.name)
    return {
        "name": meta.get("name") or path.name,
        "description": meta.get("description", ""),
        "created_at": meta.get("created_at", ""),
        "updated_at": meta.get("updated_at", ""),
        "source": meta.get("source", ""),
        "file": str(params),
        "file_size": stat.st_size if stat else 0,
        "file_mtime": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S") if stat else "",
        "compute_config_file": str(path / COMPUTE_CONFIG_FILE_NAME),
        "compute_config": config,
        "dispatch_schedule_file": str(dispatch),
        "dispatch_schedule_exists": bool(dispatch_stat),
        "dispatch_schedule_mtime": datetime.fromtimestamp(dispatch_stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S") if dispatch_stat else "",
    }


def list_schemes() -> list[dict[str, Any]]:
    ensure_default_scheme()
    items = [scheme_summary(path.name) for path in SCHEME_ROOT.iterdir() if path.is_dir()]
    return sorted(items, key=lambda item: item["name"].lower())


def workbook_overview(name: str) -> dict[str, Any]:
    path = scheme_params_path(name)
    ensure_cell_current_limit_sheet(path)
    ensure_scheme_compute_config(name)
    wb = load_workbook(path, read_only=True, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        headers = [cell_to_json(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
        sheets.append(
            {
                "name": ws.title,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "headers": headers[:OVERVIEW_HEADER_COLS],
                "kind": "compute_config" if ws.title == COMPUTE_CONFIG_SHEET_NAME else "workbook",
            }
        )
    wb.close()
    return {"scheme": scheme_summary(name), "sheets": sheets}


def read_sheet(name: str, sheet_name: str, page: int = 1, page_size: int = 200) -> dict[str, Any]:
    path = scheme_params_path(name)
    ensure_cell_current_limit_sheet(path)
    wb = load_workbook(path, read_only=True, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise FileNotFoundError(f"工作表不存在：{sheet_name}")
    ws = wb[sheet_name]
    page_size = positive_int(page_size, 200, 1, 2000)
    total_body_rows = max(ws.max_row - 1, 0)
    total_pages = max(1, math.ceil(total_body_rows / page_size)) if total_body_rows else 1
    page = positive_int(page, 1, 1, total_pages)
    body_start_row = 2 + (page - 1) * page_size
    body_end_row = min(ws.max_row, body_start_row + page_size - 1) if total_body_rows else 1
    rows = []
    if ws.max_row >= 1:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, max_col=ws.max_column), [])
        rows.append([cell_to_json(cell.value) for cell in header_row])
    if total_body_rows and body_start_row <= body_end_row:
        for row in ws.iter_rows(min_row=body_start_row, max_row=body_end_row, max_col=ws.max_column):
            rows.append([cell_to_json(cell.value) for cell in row])
    payload = {
        "scheme": name,
        "sheet": sheet_name,
        "max_rows": ws.max_row,
        "max_cols": ws.max_column,
        "total_rows": total_body_rows,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "body_start_row": body_start_row if total_body_rows else 0,
        "body_end_row": body_end_row if total_body_rows else 0,
        "preview_rows": len(rows),
        "preview_cols": len(rows[0]) if rows else 0,
        "rows": rows,
        "truncated": total_body_rows > page_size,
    }
    wb.close()
    return payload


def touch_scheme_meta(name: str) -> None:
    meta_path = scheme_dir(name) / META_FILE_NAME
    meta = read_json(meta_path, {}) or {"name": name}
    meta["updated_at"] = now_text()
    write_json(meta_path, meta)


def write_sheet(name: str, sheet_name: str, rows: list[list[Any]], page: int = 1, page_size: int = 200) -> dict[str, Any]:
    if sheet_name == COMPUTE_CONFIG_SHEET_NAME:
        payload = {}
        for row in rows[1:]:
            if not row:
                continue
            key = row[0]
            if key in DEFAULT_COMPUTE_CONFIG:
                payload[str(key)] = parse_compute_sheet_value(row[1] if len(row) > 1 else None)
        write_scheme_compute_config(name, payload)
        return read_sheet(name, sheet_name, page, page_size)
    if not rows:
        raise ValueError("保存内容不能为空")
    path = scheme_params_path(name)
    ensure_cell_current_limit_sheet(path)
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise FileNotFoundError(f"工作表不存在：{sheet_name}")
    ws = wb[sheet_name]
    max_r = max(len(rows), ws.max_row)
    max_c = max(max((len(row) for row in rows), default=0), ws.max_column)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            value = rows[r - 1][c - 1] if r - 1 < len(rows) and c - 1 < len(rows[r - 1]) else None
            ws.cell(row=r, column=c).value = value
    wb.save(path)
    wb.close()
    touch_scheme_meta(name)
    return read_sheet(name, sheet_name, page, page_size)


def write_sheet_cells(name: str, sheet_name: str, updates: list[dict[str, Any]], page: int = 1, page_size: int = 200) -> dict[str, Any]:
    if not updates:
        return read_sheet(name, sheet_name, page, page_size)
    path = scheme_params_path(name)
    ensure_cell_current_limit_sheet(path)
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise FileNotFoundError(f"工作表不存在：{sheet_name}")
    ws = wb[sheet_name]
    for item in updates:
        if item.get("row") is None or item.get("col") is None:
            wb.close()
            raise ValueError("单元格更新需要包含 row 和 col")
        row = positive_int(item.get("row"), 0, 1)
        col = positive_int(item.get("col"), 0, 1)
        ws.cell(row=row, column=col).value = item.get("value")
    wb.save(path)
    wb.close()
    touch_scheme_meta(name)
    return read_sheet(name, sheet_name, page, page_size)


def mutate_sheet_rows(
    name: str,
    sheet_name: str,
    operation: str,
    row: int | None = None,
    after_row: int | None = None,
    values: list[Any] | None = None,
    page: int = 1,
    page_size: int = 200,
) -> dict[str, Any]:
    path = scheme_params_path(name)
    ensure_cell_current_limit_sheet(path)
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise FileNotFoundError(f"工作表不存在：{sheet_name}")
    ws = wb[sheet_name]
    if operation == "insert_row":
        base_row = positive_int(after_row, 1, 1, max(ws.max_row, 1))
        insert_at = min(base_row + 1, ws.max_row + 1)
        ws.insert_rows(insert_at, 1)
        for index, value in enumerate(values or [], start=1):
            ws.cell(row=insert_at, column=index).value = value
        page = max(1, math.ceil(max(insert_at - 1, 1) / positive_int(page_size, 200, 1, 2000)))
    elif operation == "delete_row":
        if row is None:
            wb.close()
            raise ValueError("删除行需要指定正文行号")
        if ws.max_row < 2:
            wb.close()
            raise ValueError("当前工作表没有可删除的正文行")
        delete_at = positive_int(row, 0, 2, ws.max_row)
        ws.delete_rows(delete_at, 1)
    else:
        wb.close()
        raise ValueError("不支持的工作表行操作")
    wb.save(path)
    wb.close()
    touch_scheme_meta(name)
    return read_sheet(name, sheet_name, page, page_size)


DISPATCH_HEADERS: list[tuple[str, str]] = [
    ("hour", "时刻(h)"),
    ("cell_current_a", "电芯电流(A)"),
    ("pack_current_a", "电池包电流(A)"),
    ("u_pi", "内循环泵启停"),
    ("u_po", "外循环泵启停"),
    ("u_lh", "液冷电加热启停"),
    ("p_heat_liquid_kw", "液冷电加热功率(kW)"),
    ("u_ch", "舱体电加热启停"),
    ("p_heat_cont_kw", "舱体电加热功率(kW)"),
    ("soc_ref", "SOC参考值"),
    ("t_bat_ref_c", "电芯温度参考值(℃)"),
    ("t_tank_ref_c", "液冷罐温度参考值(℃)"),
    ("t_cont_ref_c", "舱体温度参考值(℃)"),
    ("pbess_ref_kw", "BESS功率参考值(kW)"),
    ("pv_use_ref_kw", "光伏利用参考值(kW)"),
    ("wt_use_ref_kw", "风电利用参考值(kW)"),
    ("dg_ref_kw", "柴油机功率参考值(kW)"),
]

DISPATCH_HEADER_ALIASES: dict[str, str] = {
    "时刻(h)": "hour",
    "hour": "hour",
    "电芯电流(A)": "cell_current_a",
    "cell_current_a": "cell_current_a",
    "电池包电流(A)": "pack_current_a",
    "pack_current_a": "pack_current_a",
    "内循环泵启停": "u_pi",
    "u_pi": "u_pi",
    "外循环泵启停": "u_po",
    "u_po": "u_po",
    "液冷电加热启停": "u_lh",
    "u_lh": "u_lh",
    "液冷电加热功率(kW)": "p_heat_liquid_kw",
    "p_heat_liquid_kw": "p_heat_liquid_kw",
    "舱体电加热启停": "u_ch",
    "u_ch": "u_ch",
    "舱体电加热功率(kW)": "p_heat_cont_kw",
    "p_heat_cont_kw": "p_heat_cont_kw",
    "SOC参考值": "soc_ref",
    "soc_ref": "soc_ref",
    "电芯温度参考值(℃)": "t_bat_ref_c",
    "t_bat_ref_c": "t_bat_ref_c",
    "液冷罐温度参考值(℃)": "t_tank_ref_c",
    "t_tank_ref_c": "t_tank_ref_c",
    "舱体温度参考值(℃)": "t_cont_ref_c",
    "t_cont_ref_c": "t_cont_ref_c",
    "BESS功率参考值(kW)": "pbess_ref_kw",
    "pbess_ref_kw": "pbess_ref_kw",
    "光伏利用参考值(kW)": "pv_use_ref_kw",
    "pv_use_ref_kw": "pv_use_ref_kw",
    "风电利用参考值(kW)": "wt_use_ref_kw",
    "wt_use_ref_kw": "wt_use_ref_kw",
    "柴油机功率参考值(kW)": "dg_ref_kw",
    "dg_ref_kw": "dg_ref_kw",
}


def dispatch_default_heat_kw(enabled: Any, requested_kw: Any, nominal_w: float) -> float:
    value = float(requested_kw or 0.0)
    if value > 0:
        return value
    return float(enabled or 0.0) * float(nominal_w or 0.0) / 1000.0


def write_dispatch_schedule_workbook(path: Path, rows: list[dict[str, Any]], p: Any | None = None) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = DISPATCH_SCHEDULE_SHEET_NAME
    headers = [label for _, label in DISPATCH_HEADERS]
    ws.append(headers)
    n_p = max(1, int(getattr(p, "N_p", 1))) if p is not None else 1
    for row in rows:
        normalized = dict(row)
        if normalized.get("pack_current_a") in (None, "") and normalized.get("cell_current_a") not in (None, ""):
            normalized["pack_current_a"] = float(normalized["cell_current_a"]) * n_p
        if normalized.get("cell_current_a") in (None, "") and normalized.get("pack_current_a") not in (None, ""):
            normalized["cell_current_a"] = float(normalized["pack_current_a"]) / n_p
        ws.append([normalized.get(key) for key, _ in DISPATCH_HEADERS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
    return path


def read_dispatch_schedule_workbook(path: Path, p: Any | None = None) -> dict[str, Any]:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"调度控制曲线文件不存在：{path.name}")
    wb = load_workbook(path, read_only=True, data_only=True)
    if DISPATCH_SCHEDULE_SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise FileNotFoundError(f"调度控制曲线文件缺少工作表：{DISPATCH_SCHEDULE_SHEET_NAME}")
    ws = wb[DISPATCH_SCHEDULE_SHEET_NAME]
    raw_headers = [cell_to_json(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
    keys = [DISPATCH_HEADER_ALIASES.get(str(header or "").strip(), "") for header in raw_headers]
    rows: list[dict[str, Any]] = []
    n_p = max(1, int(getattr(p, "N_p", 1))) if p is not None else 1
    for excel_row in ws.iter_rows(min_row=2, max_col=len(raw_headers)):
        row: dict[str, Any] = {}
        for key, cell in zip(keys, excel_row):
            if key:
                row[key] = cell_to_json(cell.value)
        if not any(value not in (None, "") for value in row.values()):
            continue
        if row.get("hour") in (None, ""):
            row["hour"] = len(rows)
        if row.get("pack_current_a") in (None, "") and row.get("cell_current_a") not in (None, ""):
            row["pack_current_a"] = float(row["cell_current_a"]) * n_p
        if row.get("cell_current_a") in (None, "") and row.get("pack_current_a") not in (None, ""):
            row["cell_current_a"] = float(row["pack_current_a"]) / n_p
        row["pack_current_a"] = float(row.get("pack_current_a") or 0.0)
        row["cell_current_a"] = float(row.get("cell_current_a") or 0.0)
        for key in ("u_pi", "u_po", "u_lh", "u_ch"):
            row[key] = float(row.get(key) or 0.0)
        row["p_heat_liquid_kw"] = dispatch_default_heat_kw(row.get("u_lh"), row.get("p_heat_liquid_kw"), getattr(p, "P_heat_liquid", 0.0))
        row["p_heat_cont_kw"] = dispatch_default_heat_kw(row.get("u_ch"), row.get("p_heat_cont_kw"), getattr(p, "P_heat_cont", 0.0))
        rows.append(row)
    wb.close()
    return {"path": str(path), "sheet": DISPATCH_SCHEDULE_SHEET_NAME, "rows": rows, "headers": raw_headers}


def dispatch_rows_from_result_data(result_data: dict[str, Any], p: Any | None = None) -> list[dict[str, Any]]:
    source_rows = list(result_data.get("rows") or [])
    unit_by_key = {str(item.get("key")): str(item.get("unit") or "") for item in result_data.get("series") or []}
    n_p = max(1, int(getattr(p, "N_p", 1))) if p is not None else 1
    rows: list[dict[str, Any]] = []
    for raw in source_rows:
        i_pack = raw.get("I_bat", raw.get("pack_current_a", 0.0)) or 0.0
        p_heat_liquid_kw = result_power_kw(raw, "P_heat_liquid_kw", unit_by_key, legacy_w_key="P_heat_liquid_w") or 0.0
        p_heat_cont_kw = result_power_kw(raw, "P_heat_cont_kw", unit_by_key, legacy_w_key="P_heat_cont_w") or 0.0
        rows.append(
            {
                "hour": raw.get("hour", raw.get("时刻(h)", len(rows))),
                "cell_current_a": float(i_pack) / n_p,
                "pack_current_a": float(i_pack),
                "u_pi": float(raw.get("u_pi", 0.0) or 0.0),
                "u_po": float(raw.get("u_po", 0.0) or 0.0),
                "u_lh": float(raw.get("u_lh", 0.0) or 0.0),
                "p_heat_liquid_kw": p_heat_liquid_kw,
                "u_ch": float(raw.get("u_ch", 0.0) or 0.0),
                "p_heat_cont_kw": p_heat_cont_kw,
                "soc_ref": raw.get("SOC"),
                "t_bat_ref_c": raw.get("T_bat"),
                "t_tank_ref_c": raw.get("T_tank"),
                "t_cont_ref_c": raw.get("T_cont"),
                "pbess_ref_kw": result_power_kw(raw, "P_BESS", unit_by_key) or 0.0,
                "pv_use_ref_kw": raw.get("pv_use_kw"),
                "wt_use_ref_kw": raw.get("wt_use_kw"),
                "dg_ref_kw": result_power_kw(raw, "P_dg", unit_by_key) or 0.0,
            }
        )
    return rows


def result_power_kw(raw: dict[str, Any], key: str, unit_by_key: dict[str, str] | None = None, legacy_w_key: str | None = None) -> float | None:
    legacy_w_keys = {"P_BESS", "P_dg"}
    if key in raw and raw.get(key) not in (None, ""):
        value = float(raw.get(key) or 0.0)
        unit = str((unit_by_key or {}).get(key) or "").lower()
        if unit == "w":
            return value / 1000.0
        if not unit and key in legacy_w_keys and abs(value) >= 1000.0:
            return value / 1000.0
        return value
    if legacy_w_key and legacy_w_key in raw and raw.get(legacy_w_key) not in (None, ""):
        return float(raw.get(legacy_w_key) or 0.0) / 1000.0
    return None


def dispatch_schedule_payload(name: str) -> dict[str, Any]:
    p = load_solver_module().load_params(scheme_params_path(name))
    path = scheme_dispatch_schedule_path(name)
    exists = path.exists()
    if exists:
        schedule = read_dispatch_schedule_workbook(path, p)
        rows = schedule["rows"]
    else:
        rows = []
    return {
        "scheme": name,
        "exists": exists,
        "path": str(path),
        "sheet": DISPATCH_SCHEDULE_SHEET_NAME,
        "headers": [{"key": key, "label": label} for key, label in DISPATCH_HEADERS],
        "rows": make_dispatch_rows_json(rows),
        "row_count": len(rows),
    }


def make_dispatch_rows_json(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{key: row.get(key) for key, _ in DISPATCH_HEADERS} for row in rows]


def write_scheme_dispatch_schedule(name: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    p = load_solver_module().load_params(scheme_params_path(name))
    write_dispatch_schedule_workbook(scheme_dispatch_schedule_path(name), rows, p)
    touch_scheme_meta(name)
    return dispatch_schedule_payload(name)


def default_dispatch_schedule_rows(name: str, config: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    solver = load_solver_module()
    p = solver.load_params(scheme_params_path(name))
    cfg = normalize_compute_config(config or read_scheme_compute_config(name))
    dt_minutes = float(cfg.get("dt_minutes") or 60.0)
    mode = str(cfg.get("mode") or "dayahead_24h")
    hours = cfg.get("hours")
    if hours not in (None, ""):
        horizon = float(hours)
    elif mode == "test_1h":
        horizon = 1.0
    elif mode == "test_4h":
        horizon = 4.0
    else:
        horizon = 24.0
    n = int(round(horizon / (dt_minutes / 60.0)))
    return [
        {
            "hour": index * dt_minutes / 60.0,
            "cell_current_a": 0.0,
            "pack_current_a": 0.0,
            "u_pi": 0.0,
            "u_po": 0.0,
            "u_lh": 0.0,
            "p_heat_liquid_kw": 0.0,
            "u_ch": 0.0,
            "p_heat_cont_kw": 0.0,
            "soc_ref": getattr(p, "SOC_init", None),
            "t_bat_ref_c": getattr(p, "T_bat_init", None),
            "t_tank_ref_c": getattr(p, "T_tank_init", None),
            "t_cont_ref_c": getattr(p, "T_cont_init", None),
            "pbess_ref_kw": 0.0,
            "pv_use_ref_kw": None,
            "wt_use_ref_kw": None,
            "dg_ref_kw": None,
        }
        for index in range(n)
    ]


def initialize_scheme_dispatch_schedule(name: str, config: dict[str, Any] | None = None) -> dict[str, Any]:
    rows = default_dispatch_schedule_rows(name, config)
    return write_scheme_dispatch_schedule(name, rows)


def create_dispatch_scheme_from_optimization(task_id: str, target_name: str, description: str = "") -> dict[str, Any]:
    run_dir = RUN_ROOT / safe_name(task_id)
    if not run_dir.exists():
        run_dir = RUN_ROOT / str(task_id)
    manifest = read_json(run_dir / "run_manifest.json", {}) or {}
    source_scheme = str(manifest.get("scheme") or "")
    if not source_scheme:
        raise ValueError("优化任务缺少来源方案信息")
    scheme = copy_scheme(source_scheme, target_name, description or f"由优化任务 {task_id} 生成的调度方案")
    p = load_solver_module().load_params(scheme_params_path(scheme["name"]))
    result_payload = optimization_result_payload(run_dir)
    rows = dispatch_rows_from_result_data(result_payload.get("result_data") or {}, p)
    if not rows:
        raise FileNotFoundError("优化任务没有可读取的结果工作簿，无法生成调度控制文件")
    write_dispatch_schedule_workbook(scheme_dispatch_schedule_path(scheme["name"]), rows, p)
    touch_scheme_meta(scheme["name"])
    return scheme_summary(scheme["name"])


def key_metrics(diag: dict[str, Any], stats: dict[str, Any]) -> dict[str, Any]:
    checks = diag.get("checks") or {}
    return {
        "status": diag.get("status"),
        "success": diag.get("success"),
        "objective": diag.get("objective"),
        "fuel_kg": diag.get("fuel_kg"),
        "curt_kwh": diag.get("curt_kwh"),
        "heat_kwh": diag.get("heat_kwh"),
        "gap": diag.get("gap"),
        "time_s": diag.get("time_s"),
        "node_count": diag.get("node_count"),
        "solver": diag.get("solver_used") or (diag.get("objective_breakdown") or {}).get("solver_name") or stats.get("solver_name"),
        "solver_backend": diag.get("solver_backend") or stats.get("solver_backend"),
        "variables_total": stats.get("variables_total"),
        "binary_variables": stats.get("binary_variables"),
        "constraints_total": stats.get("constraints_total"),
        "dt_minutes": stats.get("dt_minutes") or diag.get("dt_minutes"),
        "steps": stats.get("steps"),
        "soc_min": checks.get("soc_min"),
        "soc_max": checks.get("soc_max"),
        "tbat_min_c": checks.get("tbat_min_c"),
        "tbat_max_c": checks.get("tbat_max_c"),
        "charge_current_limit_violation_max_a": checks.get("charge_current_limit_violation_max_a"),
        "discharge_current_limit_violation_max_a": checks.get("discharge_current_limit_violation_max_a"),
        "ttank_min_c": checks.get("ttank_min_c"),
        "ttank_max_c": checks.get("ttank_max_c"),
        "tcont_min_c": checks.get("tcont_min_c"),
        "tcont_max_c": checks.get("tcont_max_c"),
        "model_balance_max_kw": checks.get("model_balance_max_kw"),
        "pbess_physical_max_kw": checks.get("pbess_physical_max_kw"),
    }


def web_href_for_path(path: Path) -> str:
    try:
        relative = path.resolve().relative_to(WEB_ROOT.resolve())
    except Exception:
        return ""
    return "/" + relative.as_posix()


def file_entry(path: Path, label: str, kind: str) -> dict[str, Any] | None:
    if not path.exists() or not path.is_file():
        return None
    return {
        "label": label,
        "kind": kind,
        "name": path.name,
        "path": str(path),
        "href": web_href_for_path(path),
        "size_bytes": path.stat().st_size,
        "modified_at": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
    }


def first_existing_path(*paths: Any) -> Path | None:
    for value in paths:
        if not value:
            continue
        path = Path(value)
        if path.exists():
            return path
    return None


def newest_matching_file(run_dir: Path, pattern: str) -> Path | None:
    matches = sorted(run_dir.glob(pattern), key=lambda item: item.stat().st_mtime, reverse=True)
    return matches[0] if matches else None


def fallback_statistics(diag: dict[str, Any], stats: dict[str, Any]) -> dict[str, Any]:
    metrics = key_metrics(diag, stats)
    return {
        **metrics,
        "checks": diag.get("checks") or {},
        "model_stats": stats or diag.get("model_stats") or {},
        "objective_breakdown": diag.get("objective_breakdown") or {},
    }


def decode_workbook_json_value(value: Any) -> Any:
    if isinstance(value, str) and value[:1] in {"{", "["}:
        try:
            return json.loads(value)
        except Exception:
            return value
    return cell_to_json(value)


def table_rows_from_worksheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell_to_json(value) or "") for value in rows[0]]
    output: list[dict[str, Any]] = []
    for raw in rows[1:]:
        item = {
            headers[index]: decode_workbook_json_value(value)
            for index, value in enumerate(raw)
            if index < len(headers) and headers[index]
        }
        if any(value not in (None, "") for value in item.values()):
            output.append(item)
    return output


def read_optimization_result_workbook(path: Path | str | None) -> dict[str, Any]:
    if not path:
        return {}
    workbook_path = Path(path)
    if not workbook_path.exists():
        return {}
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "__result_payload" in wb.sheetnames:
            chunks: list[str] = []
            for row in wb["__result_payload"].iter_rows(min_row=2, values_only=True):
                if len(row) >= 2 and row[1] not in (None, ""):
                    chunks.append(str(row[1]))
            if chunks:
                payload = json.loads("".join(chunks))
                if isinstance(payload, dict):
                    return payload

        statistics: dict[str, Any] = {}
        if "统计信息" in wb.sheetnames:
            for row in table_rows_from_worksheet(wb["统计信息"]):
                key = str(row.get("项目") or "")
                if key:
                    statistics[key] = row.get("值")
        series = table_rows_from_worksheet(wb["曲线元数据"]) if "曲线元数据" in wb.sheetnames else []
        rows = table_rows_from_worksheet(wb["调度曲线"]) if "调度曲线" in wb.sheetnames else []
        if not statistics and not series and not rows:
            return {}
        return {
            "version": 1,
            "row_count": len(rows),
            "time_axis": {"key": "hour", "label": "时刻", "unit": "h"},
            "statistics": statistics,
            "series": series,
            "rows": rows,
            "files": {"state_workbook": str(workbook_path)},
        }
    finally:
        wb.close()


def optimization_result_payload(run_dir: Path | str, diag: dict[str, Any] | None = None, stats: dict[str, Any] | None = None) -> dict[str, Any]:
    run_path = Path(run_dir) if run_dir else Path()
    diag = diag or {}
    stats = stats or {}
    if not run_path.exists():
        return {"result_data": {}, "statistics": fallback_statistics(diag, stats), "result_files": []}

    workbook_path = first_existing_path(diag.get("state_workbook"), newest_matching_file(run_path, "optimization_results_*.xlsx"))
    workbook_payload = read_optimization_result_workbook(workbook_path)
    if workbook_payload:
        result_data = workbook_payload
        statistics = workbook_payload.get("statistics") or {}
        files = [file_entry(workbook_path, "结果工作簿", "xlsx")] if workbook_path else []
    else:
        result_data_path = first_existing_path(diag.get("result_data"), run_path / RESULT_DATA_FILE_NAME)
        statistics_path = first_existing_path(diag.get("result_statistics"), run_path / RESULT_STATISTICS_FILE_NAME)
        csv_path = first_existing_path(diag.get("result_timeseries_csv"), run_path / RESULT_TIMESERIES_CSV_FILE_NAME)
        result_data = read_json(result_data_path, {}) if result_data_path else {}
        statistics = read_json(statistics_path, {}) if statistics_path else {}
        if not statistics and isinstance(result_data, dict):
            statistics = result_data.get("statistics") or {}
        if not statistics:
            statistics = fallback_statistics(diag, stats)
        files = []
        for entry in (
            file_entry(workbook_path, "结果工作簿", "xlsx") if workbook_path else None,
            file_entry(result_data_path, "历史结果JSON", "json") if result_data_path else None,
            file_entry(statistics_path, "历史统计JSON", "json") if statistics_path else None,
            file_entry(csv_path, "历史时序CSV", "csv") if csv_path else None,
        ):
            if entry:
                files.append(entry)

    return {
        "result_data": result_data if isinstance(result_data, dict) else {},
        "statistics": statistics if isinstance(statistics, dict) else {},
        "result_files": files,
    }


@dataclass
class OptTask:
    id: str
    scheme: str
    source: str = "single"
    status: str = "排队中"
    created_at: str = field(default_factory=now_text)
    start_time: str = ""
    end_time: str = ""
    elapsed_seconds: float = 0.0
    process_id: int | None = None
    return_code: int | None = None
    latest_log: str = ""
    message: str = ""
    config: dict[str, Any] = field(default_factory=dict)
    run_dir: str = ""
    command: list[str] = field(default_factory=list)
    metrics: dict[str, Any] = field(default_factory=dict)
    diagnostics: dict[str, Any] = field(default_factory=dict)
    model_stats: dict[str, Any] = field(default_factory=dict)
    progress: dict[str, Any] = field(default_factory=dict)
    result_data: dict[str, Any] = field(default_factory=dict)
    result_statistics: dict[str, Any] = field(default_factory=dict)
    result_files: list[dict[str, Any]] = field(default_factory=list)
    summary_text: str = ""
    cancel_requested: bool = False


class TaskManager:
    def __init__(self) -> None:
        self.lock = threading.RLock()
        self.tasks: dict[str, OptTask] = {}
        self.queue: queue.Queue[str] = queue.Queue()
        self.running: dict[str, subprocess.Popen] = {}
        self.max_parallel = 1
        self.worker = threading.Thread(target=self._scheduler_loop, daemon=True)
        self.worker.start()

    def enqueue(self, scheme: str, config: dict[str, Any], source: str = "single") -> OptTask:
        task_id = datetime.now().strftime("%Y%m%d%H%M%S") + "-" + uuid.uuid4().hex[:8]
        task = OptTask(id=task_id, scheme=scheme, config=dict(config), source=source)
        with self.lock:
            self.tasks[task_id] = task
            self.queue.put(task_id)
        return task

    def set_max_parallel(self, value: int) -> None:
        with self.lock:
            self.max_parallel = max(1, min(int(value or 1), max(1, os.cpu_count() or 1)))

    def cancel(self, task_id: str) -> OptTask:
        with self.lock:
            task = self.tasks.get(task_id)
            if not task:
                raise FileNotFoundError("任务不存在")
            task.cancel_requested = True
            proc = self.running.get(task_id)
            if proc and proc.poll() is None:
                task.status = "计算中止"
                task.latest_log = timestamp_log_line("正在终止求解进程")
                proc.terminate()
            elif task.status == "排队中":
                task.status = "退出队列"
                task.end_time = now_text()
                task.latest_log = timestamp_log_line("任务已从队列中移出")
            return task

    def snapshot(self) -> dict[str, Any]:
        with self.lock:
            tasks = [self._task_payload(task) for task in self.tasks.values()]
            tasks.sort(key=lambda item: item["created_at"], reverse=True)
            return {
                "tasks": tasks,
                "max_parallel": self.max_parallel,
                "running": len(self.running),
                "queued": sum(1 for task in self.tasks.values() if task.status == "排队中" and not task.cancel_requested),
            }

    def task_detail(self, task_id: str) -> dict[str, Any]:
        with self.lock:
            task = self.tasks.get(task_id)
            if not task:
                raise FileNotFoundError("任务不存在")
            return self._task_payload(task, detail=True)

    def _scheduler_loop(self) -> None:
        while True:
            try:
                task_id = self.queue.get(timeout=0.5)
            except queue.Empty:
                continue
            while True:
                with self.lock:
                    task = self.tasks.get(task_id)
                    if not task or task.cancel_requested or task.status != "排队中":
                        break
                    if len(self.running) < self.max_parallel:
                        thread = threading.Thread(target=self._run_task, args=(task_id,), daemon=True)
                        task.status = "准备启动"
                        thread.start()
                        break
                time.sleep(0.5)

    def _run_task(self, task_id: str) -> None:
        with self.lock:
            task = self.tasks[task_id]
            task.status = "计算中"
            task.start_time = now_text()
            task.latest_log = timestamp_log_line("准备优化求解输入文件")
        start = time.time()
        try:
            run_dir = RUN_ROOT / task_id
            run_dir.mkdir(parents=True, exist_ok=True)
            params = scheme_params_path(task.scheme)
            params_snapshot = run_dir / PARAM_FILE_NAME
            shutil.copy2(params, params_snapshot)
            config = normalize_compute_config(task.config or read_scheme_compute_config(task.scheme))
            task.config = config
            compute_config_snapshot = run_dir / COMPUTE_CONFIG_FILE_NAME
            write_json(compute_config_snapshot, config)
            diagnostics_json = run_dir / "diagnostics.json"
            model_stats_json = run_dir / "model_stats.json"
            progress_json = run_dir / "progress.json"
            stdout_log = run_dir / "solve_stdout.log"
            command = self._build_command(task, params_snapshot, compute_config_snapshot, diagnostics_json, model_stats_json, progress_json, run_dir)
            with self.lock:
                task.run_dir = str(run_dir)
                task.command = command
                task.latest_log = timestamp_log_line("启动求解器进程")
            write_json(
                run_dir / "run_manifest.json",
                {
                    "id": task.id,
                    "scheme": task.scheme,
                    "source": task.source,
                    "created_at": task.created_at,
                    "start_time": task.start_time,
                    "config": task.config,
                    "command": command,
                },
            )
            proc = subprocess.Popen(
                command,
                cwd=str(PROJECT_ROOT),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
            )
            with self.lock:
                task.process_id = proc.pid
                self.running[task_id] = proc
            lines: list[str] = []
            with stdout_log.open("w", encoding="utf-8", errors="ignore") as log:
                assert proc.stdout is not None
                for line in proc.stdout:
                    clean = line.rstrip()
                    if clean:
                        timestamped = timestamp_log_line(clean)
                        log.write(timestamped + "\n")
                        log.flush()
                        lines.append(timestamped)
                        with self.lock:
                            task.latest_log = timestamped[-300:]
                            progress = read_json(progress_json, {}) if progress_json.exists() else {}
                            if progress:
                                task.progress = progress
            return_code = proc.wait()
            diag = read_json(diagnostics_json, {}) or {}
            stats = read_json(model_stats_json, {}) or {}
            progress = read_json(progress_json, {}) or {}
            summary = self._read_summary(run_dir)
            result_payload = optimization_result_payload(run_dir, diag, stats)
            with self.lock:
                task.return_code = return_code
                task.diagnostics = diag
                task.model_stats = stats
                task.progress = progress
                task.metrics = key_metrics(diag, stats)
                task.result_data = result_payload.get("result_data") or {}
                task.result_statistics = result_payload.get("statistics") or {}
                task.result_files = result_payload.get("result_files") or []
                task.summary_text = summary
                if task.cancel_requested:
                    task.status = "计算中止"
                elif return_code == 0 and diag.get("success") is not False:
                    task.status = "完成计算"
                elif return_code == 0:
                    task.status = "完成计算"
                else:
                    task.status = "计算失败"
                task.message = diag.get("status") or task.latest_log or ("求解完成" if return_code == 0 else "求解失败")
        except Exception as exc:
            with self.lock:
                task.status = "计算失败"
                task.message = str(exc)
                task.latest_log = timestamp_log_line(str(exc))
                task.summary_text = traceback.format_exc()
        finally:
            with self.lock:
                task.end_time = now_text()
                task.elapsed_seconds = round(time.time() - start, 3)
                self.running.pop(task_id, None)

    def _build_command(
        self,
        task: OptTask,
        params: Path,
        compute_config: Path,
        diagnostics_json: Path,
        model_stats_json: Path,
        progress_json: Path,
        run_dir: Path,
    ) -> list[str]:
        cfg = task.config
        cmd = [
            sys.executable,
            str(SOLVER_SCRIPT),
            "--params",
            str(params),
            "--config-file",
            str(compute_config),
            "--diagnostics-json",
            str(diagnostics_json),
            "--model-stats-json",
            str(model_stats_json),
            "--progress-json",
            str(progress_json),
            "--output-dir",
            str(run_dir),
        ]
        return cmd

    def _read_summary(self, run_dir: Path) -> str:
        return ""

    def _read_stdout_logs(self, task: OptTask, limit: int = 600) -> list[dict[str, str]]:
        if not task.run_dir:
            return []
        path = Path(task.run_dir) / "solve_stdout.log"
        if not path.exists():
            return []
        lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
        entries: list[dict[str, str]] = []
        for raw in lines[-limit:]:
            line = raw.strip()
            if not line:
                continue
            level = "info"
            if "error" in line.lower() or "failed" in line.lower() or "traceback" in line.lower() or "失败" in line:
                level = "warn"
            elif "optimal" in line.lower() or "完成" in line or "success" in line.lower():
                level = "ok"
            if len(line) >= 22 and line.startswith("[") and line[20:22] == "] ":
                entries.append({"time": line[1:20], "level": level, "message": line[22:]})
            else:
                entries.append({"time": "", "level": level, "message": line})
        return entries

    def _task_payload(self, task: OptTask, detail: bool = False) -> dict[str, Any]:
        result_payload = None
        if detail and task.run_dir:
            result_payload = optimization_result_payload(Path(task.run_dir), task.diagnostics, task.model_stats)
        payload = {
            "id": task.id,
            "scheme": task.scheme,
            "source": task.source,
            "status": task.status,
            "created_at": task.created_at,
            "start_time": task.start_time,
            "end_time": task.end_time,
            "elapsed_seconds": task.elapsed_seconds,
            "process_id": task.process_id or "",
            "return_code": task.return_code,
            "latest_log": task.latest_log,
            "message": task.message,
            "run_dir": task.run_dir,
            "metrics": task.metrics,
            "config": task.config,
            "progress": task.progress,
            "can_cancel": task.status in {"排队中", "准备启动", "计算中"},
        }
        if detail:
            payload.update(
                {
                    "command": task.command,
                    "diagnostics": task.diagnostics,
                    "model_stats": task.model_stats,
                    "progress": task.progress,
                    "result_data": (result_payload or {}).get("result_data") or task.result_data,
                    "statistics": (result_payload or {}).get("statistics") or task.result_statistics,
                    "result_files": (result_payload or {}).get("result_files") or task.result_files,
                    "logs": self._read_stdout_logs(task),
                    "summary_text": task.summary_text,
                }
            )
        return payload


TASKS = None if IS_WORKER_IMPORT else TaskManager()


@dataclass
class VerificationTask:
    id: str
    scheme: str
    source_task_id: str = ""
    source: str = "single"
    status: str = "排队中"
    created_at: str = field(default_factory=now_text)
    start_time: str = ""
    end_time: str = ""
    elapsed_seconds: float = 0.0
    process_id: int | None = None
    return_code: int | None = None
    latest_log: str = ""
    message: str = ""
    config: dict[str, Any] = field(default_factory=dict)
    output_dir: str = ""
    metrics: dict[str, Any] = field(default_factory=dict)
    rows: list[dict[str, Any]] = field(default_factory=list)
    rows_preview: list[dict[str, Any]] = field(default_factory=list)
    summary_text: str = ""
    cancel_requested: bool = False


class VerificationManager:
    def __init__(self) -> None:
        self.lock = threading.RLock()
        self.tasks: dict[str, VerificationTask] = {}
        self.queue: queue.Queue[str] = queue.Queue()
        self.running: dict[str, subprocess.Popen] = {}
        self.max_parallel = 1
        self.worker = threading.Thread(target=self._scheduler_loop, daemon=True)
        self.worker.start()

    def start(self, scheme: str, config: dict[str, Any], source_task_id: str = "", source: str = "single") -> VerificationTask:
        task_id = datetime.now().strftime("%Y%m%d%H%M%S") + "-" + uuid.uuid4().hex[:8]
        task = VerificationTask(id=task_id, scheme=scheme, config=dict(config), source_task_id=source_task_id, source=source)
        with self.lock:
            self.tasks[task_id] = task
            self.queue.put(task_id)
        return task

    def set_max_parallel(self, value: int) -> None:
        with self.lock:
            self.max_parallel = max(1, min(int(value or 1), max(1, os.cpu_count() or 1)))

    def cancel(self, task_id: str) -> VerificationTask:
        with self.lock:
            task = self.tasks.get(task_id)
            if not task:
                raise FileNotFoundError("校核任务不存在")
            task.cancel_requested = True
            proc = self.running.get(task_id)
            if proc and proc.poll() is None:
                task.status = "校核中止"
                task.latest_log = timestamp_log_line("正在终止校核进程")
                proc.terminate()
            elif task.status == "排队中":
                task.status = "退出队列"
                task.end_time = now_text()
                task.latest_log = timestamp_log_line("任务已从队列中移出")
            elif task.status in {"准备启动", "校核中"}:
                task.status = "校核中止"
                task.latest_log = timestamp_log_line("已请求停止校核任务")
            return task

    def snapshot(self) -> dict[str, Any]:
        with self.lock:
            rows = [self._payload(task) for task in self.tasks.values()]
            rows.sort(key=lambda item: item["created_at"], reverse=True)
            return {
                "verifications": rows,
                "max_parallel": self.max_parallel,
                "running": len(self.running),
                "queued": sum(1 for task in self.tasks.values() if task.status == "排队中" and not task.cancel_requested),
            }

    def detail(self, task_id: str) -> dict[str, Any]:
        with self.lock:
            task = self.tasks.get(task_id)
            if not task:
                raise FileNotFoundError("校核任务不存在")
            return self._payload(task, detail=True)

    def _scheduler_loop(self) -> None:
        while True:
            with self.lock:
                self._reap_finished_locked()
                self._start_queued_locked()
            time.sleep(0.5)

    def _start_queued_locked(self) -> None:
        while len(self.running) < self.max_parallel:
            try:
                task_id = self.queue.get_nowait()
            except queue.Empty:
                return
            task = self.tasks.get(task_id)
            if not task or task.cancel_requested or task.status != "排队中":
                continue
            self._launch_process_locked(task)

    def _launch_process_locked(self, task: VerificationTask) -> None:
        output_dir = VERIFY_ROOT / task.id
        output_dir.mkdir(parents=True, exist_ok=True)
        result_path = output_dir / "verification_process_result.json"
        task.output_dir = str(output_dir)
        task.status = "校核中"
        task.start_time = now_text()
        task.latest_log = timestamp_log_line("校核进程已启动")
        write_json(
            output_dir / "verification_manifest.json",
            {
                "id": task.id,
                "scheme": task.scheme,
                "source_task_id": task.source_task_id,
                "source": task.source,
                "created_at": task.created_at,
                "start_time": task.start_time,
                "config": task.config,
            },
        )
        config_path = output_dir / "verification_config.json"
        write_json(config_path, task.config)
        command = [
            sys.executable,
            str(VERIFY_WORKER_SCRIPT),
            "--task-id",
            task.id,
            "--scheme",
            task.scheme,
            "--config-file",
            str(config_path),
            "--output-dir",
            str(output_dir),
            "--result-json",
            str(result_path),
        ]
        proc = subprocess.Popen(
            command,
            cwd=str(PROJECT_ROOT),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            text=True,
        )
        task.process_id = proc.pid
        self.running[task.id] = proc

    def _reap_finished_locked(self) -> None:
        for task_id, proc in list(self.running.items()):
            return_code = proc.poll()
            if return_code is None:
                continue
            task = self.tasks.get(task_id)
            self.running.pop(task_id, None)
            if not task:
                continue
            task.return_code = return_code
            task.end_time = now_text()
            if task.start_time:
                task.elapsed_seconds = round((datetime.strptime(task.end_time, "%Y-%m-%d %H:%M:%S") - datetime.strptime(task.start_time, "%Y-%m-%d %H:%M:%S")).total_seconds(), 3)
            if task.cancel_requested:
                task.status = "校核中止"
                task.latest_log = timestamp_log_line("方案校核已中止")
                task.message = "方案校核已中止"
                continue
            result_path = Path(task.output_dir) / "verification_process_result.json"
            result = read_json(result_path, {}) if result_path.exists() else {}
            if result.get("ok"):
                payload = result.get("payload") or {}
                task.status = "完成校核"
                task.metrics = payload.get("metrics") or {}
                task.rows = payload.get("rows") or []
                task.rows_preview = task.rows[:80]
                task.summary_text = payload.get("summary_text") or ""
                task.latest_log = timestamp_log_line("方案校核完成")
                task.message = task.summary_text.splitlines()[0] if task.summary_text else "方案校核完成"
            else:
                task.status = "校核失败"
                task.message = result.get("message") or f"校核进程异常退出：{return_code}"
                task.latest_log = timestamp_log_line(task.message)
                task.summary_text = result.get("traceback") or task.message

    def _payload(self, task: VerificationTask, detail: bool = False) -> dict[str, Any]:
        payload = {
            "id": task.id,
            "scheme": task.scheme,
            "source_task_id": task.source_task_id,
            "source": task.source,
            "status": task.status,
            "created_at": task.created_at,
            "start_time": task.start_time,
            "end_time": task.end_time,
            "elapsed_seconds": task.elapsed_seconds,
            "process_id": task.process_id or "",
            "return_code": task.return_code,
            "latest_log": task.latest_log,
            "message": task.message,
            "config": task.config,
            "output_dir": task.output_dir,
            "metrics": task.metrics,
            "can_cancel": task.status in {"排队中", "准备启动", "校核中"},
        }
        if detail:
            payload.update({"rows": task.rows, "rows_preview": task.rows_preview, "summary_text": task.summary_text})
        return payload


def load_solver_module():
    spec = importlib.util.spec_from_file_location("bess_solver_for_estore_web", SOLVER_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"无法加载求解脚本：{SOLVER_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def run_scheme_verification(params_path: Path, config: dict[str, Any], output_dir: Path, dispatch_path: Path | None = None) -> dict[str, Any]:
    solver = load_solver_module()
    p = solver.load_params(params_path)
    experiment = str(config.get("experiment") or "perspective_i2r_block20")
    mode = str(config.get("mode") or "dayahead_24h")
    solver.apply_diagnostic_experiment(p, experiment)
    if config.get("tight_temp_bounds", True):
        solver.apply_tight_temp_bounds(p, use_defaults=True)
    else:
        solver.validate_temperature_bounds(p)
        p.diag_tight_temp_bounds = False
        p.diag_temperature_bounds = solver.temperature_bounds_snapshot(p)
    p.diag_mip_focus = int(config.get("mip_focus") or 1)
    p.diag_cuts = config.get("cuts")
    p.diag_heuristics = float(config.get("heuristics") or 0.3)
    p.diag_threads = int(config.get("threads") or max(1, os.cpu_count() or 1))
    p.diag_strict_current_sos2 = bool(config.get("strict_current_sos2", False))
    p.diag_current_mode = str(config.get("current_mode") or "continuous")
    p.diag_log_file = None
    p.diag_model_stats_json = None

    dt_minutes = float(config.get("dt_minutes") or 60)
    hours = config.get("hours")
    if hours not in (None, ""):
        horizon = float(hours)
    elif mode == "test_1h":
        horizon = 1.0
    elif mode == "test_4h":
        horizon = 4.0
    else:
        horizon = 24.0
    dt_hours = dt_minutes / 60.0
    data = solver.interpolate_profiles(p, dt_hours, horizon)
    if dispatch_path is None:
        try:
            scheme_name = Path(params_path).resolve().parent.name
            dispatch_path = scheme_dispatch_schedule_path(scheme_name)
        except Exception:
            dispatch_path = Path(params_path).with_name(DISPATCH_SCHEDULE_FILE_NAME)
    schedule = read_dispatch_schedule_workbook(dispatch_path, p)
    verification = run_dispatch_time_domain_verification(solver, p, data, schedule["rows"])
    verification["config"] = {
        "mode": mode,
        "experiment": experiment,
        "dt_minutes": float(data["dt"]) / 60.0,
        "horizon_hours": horizon,
        "dispatch_schedule": str(dispatch_path),
        "calculation": "time_domain_simulation",
    }
    write_json(output_dir / "verification.json", json_safe(verification))
    write_verification_csv(output_dir / "verification_timeseries.csv", verification["rows"])
    summary_text = verification_summary_text(verification)
    (output_dir / "verification_summary.md").write_text(summary_text, encoding="utf-8")
    return {"metrics": verification["metrics"], "rows": verification["rows"], "summary_text": summary_text}


def _interp_profile(data: dict[str, Any], key: str, index: int, default: float = 0.0) -> float:
    values = np.asarray(data.get(key, []), dtype=float).reshape(-1)
    if values.size == 0:
        return float(default)
    return float(values[min(index, values.size - 1)])


def _reference_error(row: dict[str, Any], actual_key: str, ref_key: str, out_key: str) -> None:
    ref = row.get(ref_key)
    if ref in (None, ""):
        row[out_key] = None
        return
    row[out_key] = float(row[actual_key]) - float(ref)


def dispatch_diesel_for_load(p: Any, demand_kw: float, dt_s: float) -> dict[str, Any]:
    remaining_w = max(0.0, float(demand_kw) * 1000.0)
    units: list[dict[str, Any]] = []
    fuel_kg = 0.0
    total_w = 0.0
    for unit in getattr(p, "diesel_units", []) or []:
        if remaining_w <= 1e-9:
            power_w = 0.0
        else:
            power_w = min(remaining_w, float(unit.get("p_max_w", 0.0)))
            if power_w > 1e-9:
                power_w = max(power_w, float(unit.get("p_min_w", 0.0)))
        remaining_w = max(0.0, remaining_w - power_w)
        if power_w > 1e-9:
            fuel_rate = float(np.interp(power_w, np.asarray(unit["powers_w"], dtype=float), np.asarray(unit["fuel_kg_h"], dtype=float)))
        else:
            fuel_rate = 0.0
        fuel_kg += fuel_rate * dt_s / 3600.0
        total_w += power_w
        units.append({"id": unit.get("id"), "power_kw": power_w / 1000.0, "fuel_kg_h": fuel_rate})
    return {
        "diesel_kw": total_w / 1000.0,
        "fuel_kg": fuel_kg,
        "unserved_kw": max(0.0, remaining_w / 1000.0),
        "units": units,
    }


def run_dispatch_time_domain_verification(solver, p: Any, data: dict[str, Any], schedule_rows: list[dict[str, Any]]) -> dict[str, Any]:
    dt_s = float(data["dt"])
    n_data = int(data.get("N") or len(data.get("hours", [])))
    n = min(len(schedule_rows), n_data) if n_data else len(schedule_rows)
    if n <= 0:
        raise ValueError("调度控制曲线没有可仿真的时刻数据")

    soc = float(p.SOC_init)
    tbat = float(p.T_bat_init)
    ttank = float(p.T_tank_init)
    tcont = float(p.T_cont_init)
    rows: list[dict[str, Any]] = []
    totals = {"fuel_kg": 0.0, "pv_use_kwh": 0.0, "wt_use_kwh": 0.0, "pv_curt_kwh": 0.0, "wt_curt_kwh": 0.0, "unserved_kwh": 0.0}

    for t in range(n):
        control = schedule_rows[t]
        hour = float(control.get("hour") if control.get("hour") not in (None, "") else _interp_profile(data, "hours", t, t * dt_s / 3600.0))
        i_bat = float(control.get("pack_current_a") or 0.0)
        u_pi = float(control.get("u_pi") or 0.0)
        u_po = float(control.get("u_po") or 0.0)
        u_lh = float(control.get("u_lh") or 0.0)
        u_ch = float(control.get("u_ch") or 0.0)
        p_heat_liquid_w = max(0.0, float(control.get("p_heat_liquid_kw") or 0.0) * 1000.0)
        p_heat_cont_w = max(0.0, float(control.get("p_heat_cont_kw") or 0.0) * 1000.0)
        if u_lh <= 0:
            p_heat_liquid_w = 0.0
        if u_ch <= 0:
            p_heat_cont_w = 0.0

        r0 = float(solver.bilinear_interp(soc, tbat, p.soc_pts, p.temp_pts, p.r0_table))
        ocv = float(np.interp(soc, p.ocv_soc, p.ocv_1d))
        terminal_v = ocv - i_bat * r0
        qgen_w = i_bat * i_bat * r0
        pdc_w = ocv * i_bat - qgen_w
        aux_w = u_pi * p.P_pump_in + u_po * p.P_pump_out + p_heat_liquid_w + p_heat_cont_w
        pbess_w = pdc_w - p.mu_pcs * abs(pdc_w) - aux_w
        pbess_kw = pbess_w / 1000.0

        load_kw = _interp_profile(data, "P_load", t) / 1000.0
        pv_avail_kw = _interp_profile(data, "P_pv", t) / 1000.0
        wt_avail_kw = _interp_profile(data, "P_wt", t) / 1000.0
        net_after_bess_kw = load_kw - pbess_kw
        renewable_need_kw = max(0.0, net_after_bess_kw)
        pv_use_kw = min(pv_avail_kw, renewable_need_kw)
        wt_use_kw = min(wt_avail_kw, max(0.0, renewable_need_kw - pv_use_kw))
        diesel_need_kw = max(0.0, net_after_bess_kw - pv_use_kw - wt_use_kw)
        diesel = dispatch_diesel_for_load(p, diesel_need_kw, dt_s)
        unserved_kw = float(diesel["unserved_kw"])
        oversupply_kw = max(0.0, diesel["diesel_kw"] - diesel_need_kw)
        pv_curt_kw = max(0.0, pv_avail_kw - pv_use_kw)
        wt_curt_kw = max(0.0, wt_avail_kw - wt_use_kw)
        balance_kw = pv_use_kw + wt_use_kw + diesel["diesel_kw"] + pbess_kw - load_kw - oversupply_kw - unserved_kw

        charge_limit = float(np.interp(tbat, p.current_limit_temps, p.charge_current_limit_pack, left=p.charge_current_limit_pack[0], right=p.charge_current_limit_pack[-1]))
        discharge_limit = float(np.interp(tbat, p.current_limit_temps, p.discharge_current_limit_pack, left=p.discharge_current_limit_pack[0], right=p.discharge_current_limit_pack[-1]))
        soc_violation = max(float(p.SOC_min) - soc, 0.0) + max(soc - float(p.SOC_max), 0.0)
        charge_violation = max(-i_bat - charge_limit, 0.0)
        discharge_violation = max(i_bat - discharge_limit, 0.0)
        tbat_violation = max(float(p.T_bat_min) - tbat, 0.0) + max(tbat - float(p.T_bat_max), 0.0)
        ttank_violation = max(float(p.T_tank_min) - ttank, 0.0) + max(ttank - float(p.T_tank_max), 0.0)
        tcont_violation = max(float(p.T_cont_min) - tcont, 0.0) + max(tcont - float(p.T_cont_max), 0.0)

        row = {
            "step": t,
            "hour": hour,
            "i_pack_a": i_bat,
            "i_cell_a": i_bat / max(1, int(getattr(p, "N_p", 1))),
            "soc_sim": soc,
            "t_bat_sim_c": tbat,
            "t_tank_sim_c": ttank,
            "t_cont_sim_c": tcont,
            "u_terminal_sim_v": terminal_v,
            "r0_sim_ohm": r0,
            "qgen_sim_kw": qgen_w / 1000.0,
            "pbess_sim_kw": pbess_kw,
            "load_kw": load_kw,
            "pv_available_kw": pv_avail_kw,
            "wt_available_kw": wt_avail_kw,
            "pv_use_actual_kw": pv_use_kw,
            "wt_use_actual_kw": wt_use_kw,
            "pv_curt_actual_kw": pv_curt_kw,
            "wt_curt_actual_kw": wt_curt_kw,
            "diesel_actual_kw": diesel["diesel_kw"],
            "diesel_fuel_kg": diesel["fuel_kg"],
            "unserved_kw": unserved_kw,
            "oversupply_kw": oversupply_kw,
            "power_balance_kw": balance_kw,
            "soc_ref": control.get("soc_ref"),
            "t_bat_ref_c": control.get("t_bat_ref_c"),
            "t_tank_ref_c": control.get("t_tank_ref_c"),
            "t_cont_ref_c": control.get("t_cont_ref_c"),
            "pbess_ref_kw": control.get("pbess_ref_kw"),
            "soc_violation": soc_violation,
            "charge_current_violation_a": charge_violation,
            "discharge_current_violation_a": discharge_violation,
            "t_bat_violation_c": tbat_violation,
            "t_tank_violation_c": ttank_violation,
            "t_cont_violation_c": tcont_violation,
            "u_pi": u_pi,
            "u_po": u_po,
            "u_lh": u_lh,
            "u_ch": u_ch,
            "p_heat_liquid_kw": p_heat_liquid_w / 1000.0,
            "p_heat_cont_kw": p_heat_cont_w / 1000.0,
        }
        _reference_error(row, "soc_sim", "soc_ref", "soc_error")
        _reference_error(row, "t_bat_sim_c", "t_bat_ref_c", "t_bat_error_c")
        _reference_error(row, "t_tank_sim_c", "t_tank_ref_c", "t_tank_error_c")
        _reference_error(row, "t_cont_sim_c", "t_cont_ref_c", "t_cont_error_c")
        _reference_error(row, "pbess_sim_kw", "pbess_ref_kw", "pbess_error_kw")
        rows.append(row)

        totals["fuel_kg"] += diesel["fuel_kg"]
        totals["pv_use_kwh"] += pv_use_kw * dt_s / 3600.0
        totals["wt_use_kwh"] += wt_use_kw * dt_s / 3600.0
        totals["pv_curt_kwh"] += pv_curt_kw * dt_s / 3600.0
        totals["wt_curt_kwh"] += wt_curt_kw * dt_s / 3600.0
        totals["unserved_kwh"] += unserved_kw * dt_s / 3600.0

        qbt = u_pi * p.K_bt * (tbat - ttank)
        qtamb = u_po * p.K_t_amb * (ttank - _interp_profile(data, "T_amb", t))
        soc = soc - i_bat * dt_s / (p.Q_nom * 3600.0)
        tbat, ttank, tcont = solve_implicit_temperature_step(
            p,
            dt_s,
            tbat,
            ttank,
            tcont,
            _interp_profile(data, "T_amb", t),
            qgen_w,
            qbt,
            qtamb,
            u_lh if p_heat_liquid_w > 0 else 0.0,
            u_ch if p_heat_cont_w > 0 else 0.0,
            p_heat_liquid_w=p_heat_liquid_w,
            p_heat_cont_w=p_heat_cont_w,
        )

    metrics = {
        "status": "SIMULATED",
        "steps": n,
        "dt_minutes": dt_s / 60.0,
        "soc": error_stats([row.get("soc_error") for row in rows]),
        "t_bat_c": error_stats([row.get("t_bat_error_c") for row in rows]),
        "t_tank_c": error_stats([row.get("t_tank_error_c") for row in rows]),
        "t_cont_c": error_stats([row.get("t_cont_error_c") for row in rows]),
        "pbess_kw": error_stats([row.get("pbess_error_kw") for row in rows]),
        "diesel": {"fuel_kg": totals["fuel_kg"], "max_kw": max((row["diesel_actual_kw"] for row in rows), default=0.0)},
        "renewable": {
            "pv_use_kwh": totals["pv_use_kwh"],
            "wt_use_kwh": totals["wt_use_kwh"],
            "pv_curt_kwh": totals["pv_curt_kwh"],
            "wt_curt_kwh": totals["wt_curt_kwh"],
            "unserved_kwh": totals["unserved_kwh"],
        },
        "violations": {
            "soc_max": max((row["soc_violation"] for row in rows), default=0.0),
            "charge_current_max_a": max((row["charge_current_violation_a"] for row in rows), default=0.0),
            "discharge_current_max_a": max((row["discharge_current_violation_a"] for row in rows), default=0.0),
            "t_bat_max_c": max((row["t_bat_violation_c"] for row in rows), default=0.0),
            "t_tank_max_c": max((row["t_tank_violation_c"] for row in rows), default=0.0),
            "t_cont_max_c": max((row["t_cont_violation_c"] for row in rows), default=0.0),
            "power_balance_max_kw": max((abs(row["power_balance_kw"]) for row in rows), default=0.0),
        },
    }
    return {"metrics": metrics, "rows": rows}


def detailed_replay_verification(solver, p, result: dict[str, Any]) -> dict[str, Any]:
    n = len(result["I_bat"])
    dt_s = float(result["dt_s"])
    sim_soc = np.zeros(n + 1)
    sim_tbat = np.zeros(n + 1)
    sim_ttank = np.zeros(n + 1)
    sim_tcont = np.zeros(n + 1)
    sim_soc[0] = p.SOC_init
    sim_tbat[0] = p.T_bat_init
    sim_ttank[0] = p.T_tank_init
    sim_tcont[0] = p.T_cont_init
    rows: list[dict[str, Any]] = []
    sim_voltage = np.zeros(n)
    opt_voltage = np.zeros(n)
    sim_pbess = np.zeros(n)
    opt_pbess_detail = np.zeros(n)

    for t in range(n):
        i_bat = float(result["I_bat"][t])
        opt_r0 = float(solver.bilinear_interp(float(result["SOC"][t]), float(result["T_bat"][t]), p.soc_pts, p.temp_pts, p.r0_table))
        opt_ocv = float(np.interp(float(result["SOC"][t]), p.ocv_soc, p.ocv_1d))
        opt_voltage[t] = opt_ocv - i_bat * opt_r0
        opt_pdc = opt_ocv * i_bat - i_bat * i_bat * opt_r0
        aux = (
            float(result["u_pi"][t]) * p.P_pump_in
            + float(result["u_po"][t]) * p.P_pump_out
            + float(result["u_lh"][t]) * p.P_heat_liquid
            + float(result["u_ch"][t]) * p.P_heat_cont
        )
        opt_pbess_detail[t] = opt_pdc - p.mu_pcs * abs(opt_pdc) - aux

        sim_r0 = float(solver.bilinear_interp(sim_soc[t], sim_tbat[t], p.soc_pts, p.temp_pts, p.r0_table))
        sim_ocv = float(np.interp(sim_soc[t], p.ocv_soc, p.ocv_1d))
        sim_voltage[t] = sim_ocv - i_bat * sim_r0
        sim_qgen = i_bat * i_bat * sim_r0
        sim_pdc = sim_ocv * i_bat - sim_qgen
        sim_pbess[t] = sim_pdc - p.mu_pcs * abs(sim_pdc) - aux

        rows.append(
            {
                "step": t,
                "hour": t * dt_s / 3600.0,
                "i_pack_a": i_bat,
                "soc_opt": float(result["SOC"][t]),
                "soc_sim": float(sim_soc[t]),
                "soc_error": float(sim_soc[t] - result["SOC"][t]),
                "t_bat_opt_c": float(result["T_bat"][t]),
                "t_bat_sim_c": float(sim_tbat[t]),
                "t_bat_error_c": float(sim_tbat[t] - result["T_bat"][t]),
                "t_tank_opt_c": float(result["T_tank"][t]),
                "t_tank_sim_c": float(sim_ttank[t]),
                "t_tank_error_c": float(sim_ttank[t] - result["T_tank"][t]),
                "t_cont_opt_c": float(result["T_cont"][t]),
                "t_cont_sim_c": float(sim_tcont[t]),
                "t_cont_error_c": float(sim_tcont[t] - result["T_cont"][t]),
                "u_terminal_opt_v": float(opt_voltage[t]),
                "u_terminal_sim_v": float(sim_voltage[t]),
                "u_terminal_error_v": float(sim_voltage[t] - opt_voltage[t]),
                "r0_opt_ohm": opt_r0,
                "r0_sim_ohm": sim_r0,
                "qgen_opt_kw": float(result["Q_gen_pack"][t]) / 1000.0,
                "qgen_sim_kw": sim_qgen / 1000.0,
                "pbess_opt_kw": float(result["P_BESS"][t]) / 1000.0,
                "pbess_sim_kw": sim_pbess[t] / 1000.0,
                "pbess_error_kw": float(sim_pbess[t] - result["P_BESS"][t]) / 1000.0,
            }
        )

        sim_soc[t + 1] = float(np.clip(sim_soc[t] - i_bat * dt_s / (p.Q_nom * 3600.0), p.SOC_min, p.SOC_max))
        qbt = float(result["u_pi"][t]) * p.K_bt * (sim_tbat[t] - sim_ttank[t])
        qtamb = float(result["u_po"][t]) * p.K_t_amb * (sim_ttank[t] - float(result["T_amb"][t]))
        sim_tbat[t + 1], sim_ttank[t + 1], sim_tcont[t + 1] = solve_implicit_temperature_step(
            p,
            dt_s,
            sim_tbat[t],
            sim_ttank[t],
            sim_tcont[t],
            float(result["T_amb"][t]),
            sim_qgen,
            qbt,
            qtamb,
            float(result["u_lh"][t]),
            float(result["u_ch"][t]),
        )

    rows.append(
        {
            "step": n,
            "hour": n * dt_s / 3600.0,
            "i_pack_a": None,
            "soc_opt": float(result.get("SOC_end", result["SOC"][-1])),
            "soc_sim": float(sim_soc[n]),
            "soc_error": float(sim_soc[n] - result.get("SOC_end", result["SOC"][-1])),
            "t_bat_opt_c": float(result.get("T_bat_end", result["T_bat"][-1])),
            "t_bat_sim_c": float(sim_tbat[n]),
            "t_bat_error_c": float(sim_tbat[n] - result.get("T_bat_end", result["T_bat"][-1])),
            "t_tank_opt_c": float(result.get("T_tank_end", result["T_tank"][-1])),
            "t_tank_sim_c": float(sim_ttank[n]),
            "t_tank_error_c": float(sim_ttank[n] - result.get("T_tank_end", result["T_tank"][-1])),
            "t_cont_opt_c": float(result.get("T_cont_end", result["T_cont"][-1])),
            "t_cont_sim_c": float(sim_tcont[n]),
            "t_cont_error_c": float(sim_tcont[n] - result.get("T_cont_end", result["T_cont"][-1])),
            "u_terminal_opt_v": None,
            "u_terminal_sim_v": None,
            "u_terminal_error_v": None,
            "pbess_opt_kw": None,
            "pbess_sim_kw": None,
            "pbess_error_kw": None,
        }
    )

    metrics = {
        "soc": error_stats([row["soc_error"] for row in rows]),
        "t_bat_c": error_stats([row["t_bat_error_c"] for row in rows]),
        "t_tank_c": error_stats([row["t_tank_error_c"] for row in rows]),
        "t_cont_c": error_stats([row["t_cont_error_c"] for row in rows]),
        "u_terminal_v": error_stats([row["u_terminal_error_v"] for row in rows if row["u_terminal_error_v"] is not None]),
        "pbess_kw": error_stats([row["pbess_error_kw"] for row in rows if row["pbess_error_kw"] is not None]),
        "objective": result.get("objective"),
        "fuel_kg": result.get("fuel_kg"),
        "gap": result.get("gap"),
        "status": result.get("status"),
        "steps": n,
        "dt_minutes": dt_s / 60.0,
    }
    return {"metrics": metrics, "rows": rows}


def solve_implicit_temperature_step(
    p,
    dt_s,
    tb,
    tt,
    tc,
    tamb,
    qgen,
    qbt,
    qtamb,
    u_lh,
    u_ch,
    *,
    p_heat_liquid_w: float | None = None,
    p_heat_cont_w: float | None = None,
) -> tuple[float, float, float]:
    ab = dt_s / p.C_bat
    at = dt_s / p.C_tank
    ac = dt_s / p.C_cont
    mat = np.array(
        [
            [1.0 + ab * p.K_b_cont, 0.0, -ab * p.K_b_cont],
            [0.0, 1.0 + at * p.K_t_cont, -at * p.K_t_cont],
            [-ac * p.K_b_cont, -ac * p.K_t_cont, 1.0 + ac * (p.K_b_cont + p.K_t_cont + p.K_cont_amb)],
        ],
        dtype=float,
    )
    rhs = np.array(
        [
            tb + ab * (qgen - qbt),
            tt + at * (qbt + (u_lh * p.P_heat_liquid if p_heat_liquid_w is None else p_heat_liquid_w) - qtamb),
            tc + ac * ((u_ch * p.P_heat_cont if p_heat_cont_w is None else p_heat_cont_w) + p.K_cont_amb * tamb),
        ],
        dtype=float,
    )
    vals = np.linalg.solve(mat, rhs)
    return (
        float(np.clip(vals[0], p.T_bat_min, p.T_bat_max)),
        float(np.clip(vals[1], p.T_tank_min, p.T_tank_max)),
        float(np.clip(vals[2], p.T_cont_min, p.T_cont_max)),
    )


def error_stats(values: list[Any]) -> dict[str, Any]:
    arr = np.asarray([float(v) for v in values if v is not None], dtype=float)
    if arr.size == 0:
        return {"max_abs": None, "mae": None, "rmse": None, "final": None}
    return {
        "max_abs": float(np.max(np.abs(arr))),
        "mae": float(np.mean(np.abs(arr))),
        "rmse": float(np.sqrt(np.mean(arr * arr))),
        "final": float(arr[-1]),
    }


def write_verification_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def verification_summary_text(verification: dict[str, Any]) -> str:
    metrics = verification["metrics"]
    cfg = verification.get("config", {})
    diesel = metrics.get("diesel") or {}
    renewable = metrics.get("renewable") or {}
    violations = metrics.get("violations") or {}
    pv_use = renewable.get("pv_use_kwh")
    wt_use = renewable.get("wt_use_kwh")
    pv_curt = renewable.get("pv_curt_kwh")
    wt_curt = renewable.get("wt_curt_kwh")
    renewable_use = None if pv_use is None and wt_use is None else float(pv_use or 0.0) + float(wt_use or 0.0)
    renewable_curt = None if pv_curt is None and wt_curt is None else float(pv_curt or 0.0) + float(wt_curt or 0.0)
    lines = [
        "# 方案校核结果",
        "",
        "- 计算方式: `时域仿真`",
        f"- 模式: `{cfg.get('mode')}`",
        f"- 步长: `{cfg.get('dt_minutes')}` 分钟",
        f"- 周期: `{cfg.get('horizon_hours')}` 小时",
        f"- 调度文件: `{Path(str(cfg.get('dispatch_schedule') or DISPATCH_SCHEDULE_FILE_NAME)).name}`",
        f"- 仿真状态: `{metrics.get('status')}`",
        f"- 时刻数: `{metrics.get('steps')}`",
        "",
        "## 参考曲线偏差",
        "",
        "| 指标 | 最大绝对偏差 | 平均绝对偏差 | RMSE | 末端偏差 |",
        "|---|---:|---:|---:|---:|",
    ]
    labels = {
        "soc": "SOC",
        "t_bat_c": "电芯温度(℃)",
        "t_tank_c": "液冷罐温度(℃)",
        "t_cont_c": "舱体温度(℃)",
        "u_terminal_v": "端口电压(V)",
        "pbess_kw": "BESS端口功率(kW)",
    }
    for key, label in labels.items():
        stat = metrics.get(key) or {}
        lines.append(
            f"| {label} | {format_metric(stat.get('max_abs'))} | {format_metric(stat.get('mae'))} | "
            f"{format_metric(stat.get('rmse'))} | {format_metric(stat.get('final'))} |"
        )
    lines += [
        "",
        "## 运行效果",
        "",
        "| 类别 | 指标 | 数值 |",
        "|---|---|---:|",
        f"| 柴油 | 柴油实际耗油(kg) | {format_metric(diesel.get('fuel_kg'))} |",
        f"| 柴油 | 柴油最大出力(kW) | {format_metric(diesel.get('max_kw'))} |",
        f"| 新能源实际消纳 | 光伏(kWh) | {format_metric(pv_use)} |",
        f"| 新能源实际消纳 | 风电(kWh) | {format_metric(wt_use)} |",
        f"| 新能源实际消纳 | 合计(kWh) | {format_metric(renewable_use)} |",
        f"| 新能源弃电 | 弃光(kWh) | {format_metric(pv_curt)} |",
        f"| 新能源弃电 | 弃风(kWh) | {format_metric(wt_curt)} |",
        f"| 新能源弃电 | 合计(kWh) | {format_metric(renewable_curt)} |",
        f"| 供电 | 未供电量(kWh) | {format_metric(renewable.get('unserved_kwh'))} |",
        "",
        "## 越限校核",
        "",
        "| 指标 | 最大越限 |",
        "|---|---:|",
        f"| SOC | {format_metric(violations.get('soc_max'))} |",
        f"| 充电电流(A) | {format_metric(violations.get('charge_current_max_a'))} |",
        f"| 放电电流(A) | {format_metric(violations.get('discharge_current_max_a'))} |",
        f"| 电芯温度(℃) | {format_metric(violations.get('t_bat_max_c'))} |",
        f"| 液冷罐温度(℃) | {format_metric(violations.get('t_tank_max_c'))} |",
        f"| 舱体温度(℃) | {format_metric(violations.get('t_cont_max_c'))} |",
        f"| 功率平衡残差(kW) | {format_metric(violations.get('power_balance_max_kw'))} |",
        "",
        "说明：校核以独立调度控制曲线文件中的电芯电流、循环泵启停、电加热启停与功率为输入，按详细 R0/OCV 查表、端口电压、I²R 发热、SOC 积分和三节点热平衡公式逐步推进，评估该方案执行后的电热边界与能源消纳效果。",
    ]
    return "\n".join(lines) + "\n"


def format_metric(value: Any) -> str:
    if value is None:
        return "-"
    return f"{float(value):.6g}"


VERIFY_TASKS = None if IS_WORKER_IMPORT else VerificationManager()


def latest_by_scheme(tasks: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    for task in sorted(tasks, key=lambda item: item.get("created_at", ""), reverse=True):
        scheme = str(task.get("scheme") or "")
        if scheme and scheme not in latest:
            latest[scheme] = task
    return latest


def task_board_snapshot() -> dict[str, Any]:
    schemes = list_schemes()
    opt_snapshot = TASKS.snapshot()
    verify_snapshot = VERIFY_TASKS.snapshot()
    opt_latest = latest_by_scheme(opt_snapshot.get("tasks", []))
    verify_latest = latest_by_scheme(verify_snapshot.get("verifications", []))
    return {
        "tasks": [
            *(task_board_rows("optimization", schemes, opt_latest)),
            *(task_board_rows("verification", schemes, verify_latest)),
        ],
        "optimization": opt_snapshot,
        "verification": verify_snapshot,
    }


def task_board_result_info(task_type: str, task: dict[str, Any]) -> dict[str, str]:
    defaults = {
        "optimization": ("opt_results.xlsx", "xlsx"),
        "verification": ("verification_timeseries.csv", "csv"),
    }
    default_name, default_kind = defaults.get(task_type, ("results.xlsx", "xlsx"))
    preferred_exts = [".xlsx"] if task_type == "optimization" else [".csv", ".json", ".xlsx"]
    for ext in preferred_exts:
        for item in task.get("result_files") or []:
            raw_path = str(item.get("path") or item.get("href") or item.get("name") or "")
            if raw_path.lower().endswith(ext):
                return {"result_name": Path(raw_path).name, "result_kind": str(item.get("kind") or ext.lstrip("."))}
    folder_key = "run_dir" if task_type == "optimization" else "output_dir"
    folder_value = str(task.get(folder_key) or "")
    folder = Path(folder_value) if folder_value else None
    if folder and folder.exists():
        patterns = ["optimization_results_*.xlsx", "*.xlsx"] if task_type == "optimization" else ["verification_timeseries.csv", "verification.json", "*.xlsx"]
        for pattern in patterns:
            matches = sorted(folder.glob(pattern), key=lambda path: path.stat().st_mtime, reverse=True)
            if matches:
                return {"result_name": matches[0].name, "result_kind": matches[0].suffix.lstrip(".") or default_kind}
    return {"result_name": default_name, "result_kind": default_kind}


def task_board_rows(task_type: str, schemes: list[dict[str, Any]], latest: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for scheme in schemes:
        name = str(scheme.get("name") or "")
        if not name:
            continue
        task = latest.get(name) or {}
        result_info = task_board_result_info(task_type, task)
        status = str(task.get("status") or "未计算")
        active = status in {"排队中", "准备启动", "计算中", "校核中"}
        queued = status == "排队中"
        task_type_label = "优化求解" if task_type == "optimization" else "方案校核"
        rows.append(
            {
                "id": f"{task_type}::{name}",
                "task_id": task.get("id") or "",
                "task_type_key": task_type,
                "task_type": task_type_label,
                "scheme": name,
                "scheme_description": scheme.get("description") or "",
                **result_info,
                "status": status,
                "queued": queued,
                "process_id": task.get("process_id") or "",
                "start_time": task.get("start_time") or "",
                "end_time": task.get("end_time") or "",
                "elapsed_seconds": task.get("elapsed_seconds") or 0,
                "latest_log": task.get("latest_log") or "",
                "metrics": task.get("metrics") or {},
                "can_start": not active,
                "can_queue": not active,
                "can_stop": status in {"排队中", "准备启动", "计算中", "校核中"},
            }
        )
    return rows


def latest_active_task_id(task_type: str, scheme: str) -> str:
    manager_snapshot = TASKS.snapshot() if task_type == "optimization" else VERIFY_TASKS.snapshot()
    key = "tasks" if task_type == "optimization" else "verifications"
    statuses = {"排队中", "准备启动", "计算中", "校核中"}
    for task in manager_snapshot.get(key, []):
        if task.get("scheme") == scheme and task.get("status") in statuses:
            return str(task.get("id") or "")
    return ""


def control_task(action: str, task_type: str, scheme: str, task_id: str = "") -> dict[str, Any]:
    normalized_type = "verification" if task_type in {"verification", "verify", "方案校核"} else "optimization"
    normalized_action = str(action or "").strip().lower()
    if normalized_action in {"start", "queue"}:
        cfg = normalize_compute_config(read_scheme_compute_config(scheme))
        if normalized_type == "optimization":
            source = "batch-start" if normalized_action == "start" else "batch-queue"
            task = TASKS.enqueue(scheme, cfg, source)
            detail = TASKS.task_detail(task.id)
        else:
            source = "batch-start" if normalized_action == "start" else "batch-queue"
            task = VERIFY_TASKS.start(scheme, cfg, source=source)
            detail = VERIFY_TASKS.detail(task.id)
        return {"ok": True, "task": detail, **task_board_snapshot()}
    if normalized_action in {"stop", "cancel", "cancel_queue"}:
        selected_id = task_id or latest_active_task_id(normalized_type, scheme)
        if not selected_id:
            raise FileNotFoundError("没有可停止的任务")
        if normalized_type == "optimization":
            task = TASKS.cancel(selected_id)
            detail = TASKS.task_detail(task.id)
        else:
            task = VERIFY_TASKS.cancel(selected_id)
            detail = VERIFY_TASKS.detail(task.id)
        return {"ok": True, "task": detail, **task_board_snapshot()}
    raise ValueError("任务操作必须是 start、queue 或 stop")


def comparison_items() -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    items.extend(optimization_comparison_items())
    items.extend(verification_comparison_items())
    items.sort(key=lambda item: item.get("created_at") or item.get("mtime") or "", reverse=True)
    return items


def fallback_scheme_name(result_dir: Path | None = None) -> str:
    try:
        schemes = list_schemes()
        if result_dir is not None:
            params = result_dir / PARAM_FILE_NAME
            if params.exists():
                for scheme in schemes:
                    try:
                        if params.read_bytes() == Path(scheme["file"]).read_bytes():
                            return scheme["name"]
                    except Exception:
                        continue
        if len(schemes) == 1:
            return schemes[0]["name"]
    except Exception:
        pass
    return "未知方案"


def optimization_comparison_items() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not RUN_ROOT.exists():
        return rows
    for folder in RUN_ROOT.iterdir():
        if not folder.is_dir():
            continue
        diag_path = folder / "diagnostics.json"
        if not diag_path.exists():
            continue
        diag = read_json(diag_path, {}) or {}
        stats = read_json(folder / "model_stats.json", {}) or {}
        manifest = read_json(folder / "run_manifest.json", {}) or {}
        metrics = key_metrics(diag, stats)
        rows.append(
            {
                "id": f"opt:{folder.name}",
                "raw_id": folder.name,
                "type": "optimization",
                "type_label": "优化求解",
                "scheme": manifest.get("scheme") or fallback_scheme_name(folder),
                "status": diag.get("status") or "",
                "success": diag.get("success"),
                "created_at": manifest.get("created_at") or datetime.fromtimestamp(folder.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "mtime": datetime.fromtimestamp(folder.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "output_dir": str(folder),
                "config": manifest.get("config") or {},
                "metrics": {
                    "objective": metrics.get("objective"),
                    "fuel_kg": metrics.get("fuel_kg"),
                    "curt_kwh": metrics.get("curt_kwh"),
                    "heat_kwh": metrics.get("heat_kwh"),
                    "gap": metrics.get("gap"),
                    "time_s": metrics.get("time_s"),
                    "node_count": metrics.get("node_count"),
                    "variables_total": metrics.get("variables_total"),
                    "binary_variables": metrics.get("binary_variables"),
                    "constraints_total": metrics.get("constraints_total"),
                    "soc_min": metrics.get("soc_min"),
                    "soc_max": metrics.get("soc_max"),
                    "tbat_min_c": metrics.get("tbat_min_c"),
                    "tbat_max_c": metrics.get("tbat_max_c"),
                    "charge_current_limit_violation_max_a": metrics.get("charge_current_limit_violation_max_a"),
                    "discharge_current_limit_violation_max_a": metrics.get("discharge_current_limit_violation_max_a"),
                    "ttank_min_c": metrics.get("ttank_min_c"),
                    "ttank_max_c": metrics.get("ttank_max_c"),
                    "tcont_min_c": metrics.get("tcont_min_c"),
                    "tcont_max_c": metrics.get("tcont_max_c"),
                    "model_balance_max_kw": metrics.get("model_balance_max_kw"),
                    "pbess_physical_max_kw": metrics.get("pbess_physical_max_kw"),
                },
            }
        )
    with TASKS.lock:
        for task in TASKS.tasks.values():
            if task.run_dir and not any(row["raw_id"] == task.id for row in rows):
                rows.append(
                    {
                        "id": f"opt:{task.id}",
                        "raw_id": task.id,
                        "type": "optimization",
                        "type_label": "优化求解",
                        "scheme": task.scheme,
                        "status": task.status,
                        "success": task.status == "完成计算",
                        "created_at": task.created_at,
                        "mtime": task.end_time or task.start_time or task.created_at,
                        "output_dir": task.run_dir,
                        "config": task.config,
                        "metrics": task.metrics,
                    }
                )
    return rows


def verification_comparison_items() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not VERIFY_ROOT.exists():
        return rows
    for folder in VERIFY_ROOT.iterdir():
        if not folder.is_dir():
            continue
        verification_path = folder / "verification.json"
        if not verification_path.exists():
            continue
        payload = read_json(verification_path, {}) or {}
        metrics = payload.get("metrics") or {}
        cfg = payload.get("config") or {}
        manifest = read_json(folder / "verification_manifest.json", {}) or {}
        rows.append(
            {
                "id": f"verify:{folder.name}",
                "raw_id": folder.name,
                "type": "verification",
                "type_label": "方案校核",
                "scheme": manifest.get("scheme") or fallback_scheme_name(),
                "status": metrics.get("status") or "完成校核",
                "success": True,
                "created_at": manifest.get("created_at") or datetime.fromtimestamp(folder.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "mtime": datetime.fromtimestamp(folder.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "output_dir": str(folder),
                "config": cfg,
                "metrics": flatten_verification_metrics(metrics),
            }
        )
    with VERIFY_TASKS.lock:
        for task in VERIFY_TASKS.tasks.values():
            if task.output_dir and not any(row["raw_id"] == task.id for row in rows):
                rows.append(
                    {
                        "id": f"verify:{task.id}",
                        "raw_id": task.id,
                        "type": "verification",
                        "type_label": "方案校核",
                        "scheme": task.scheme,
                        "status": task.status,
                        "success": task.status == "完成校核",
                        "created_at": task.created_at,
                        "mtime": task.end_time or task.start_time or task.created_at,
                        "output_dir": task.output_dir,
                        "config": task.config,
                        "metrics": flatten_verification_metrics(task.metrics),
                    }
                )
    return rows


def flatten_verification_metrics(metrics: dict[str, Any]) -> dict[str, Any]:
    def stat(name: str, key: str) -> Any:
        group = metrics.get(name) if isinstance(metrics, dict) else {}
        return group.get(key) if isinstance(group, dict) else None

    def group(name: str) -> dict[str, Any]:
        value = metrics.get(name) if isinstance(metrics, dict) else {}
        return value if isinstance(value, dict) else {}

    def sum_available(*values: Any) -> Any:
        present = [float(value) for value in values if value is not None]
        return sum(present) if present else None

    diesel = group("diesel")
    renewable = group("renewable")
    violations = group("violations")
    return {
        "status": metrics.get("status"),
        "objective": metrics.get("objective"),
        "fuel_kg": diesel.get("fuel_kg", metrics.get("fuel_kg")),
        "diesel_max_kw": diesel.get("max_kw"),
        "gap": metrics.get("gap"),
        "steps": metrics.get("steps"),
        "dt_minutes": metrics.get("dt_minutes"),
        "pv_use_kwh": renewable.get("pv_use_kwh"),
        "wt_use_kwh": renewable.get("wt_use_kwh"),
        "renewable_use_kwh": sum_available(renewable.get("pv_use_kwh"), renewable.get("wt_use_kwh")),
        "pv_curt_kwh": renewable.get("pv_curt_kwh"),
        "wt_curt_kwh": renewable.get("wt_curt_kwh"),
        "renewable_curt_kwh": sum_available(renewable.get("pv_curt_kwh"), renewable.get("wt_curt_kwh")),
        "unserved_kwh": renewable.get("unserved_kwh"),
        "soc_violation_max": violations.get("soc_max"),
        "charge_current_violation_max_a": violations.get("charge_current_max_a"),
        "discharge_current_violation_max_a": violations.get("discharge_current_max_a"),
        "t_bat_violation_max_c": violations.get("t_bat_max_c"),
        "t_tank_violation_max_c": violations.get("t_tank_max_c"),
        "t_cont_violation_max_c": violations.get("t_cont_max_c"),
        "power_balance_max_kw": violations.get("power_balance_max_kw"),
        "soc_max_abs": stat("soc", "max_abs"),
        "soc_mae": stat("soc", "mae"),
        "soc_rmse": stat("soc", "rmse"),
        "t_bat_max_abs_c": stat("t_bat_c", "max_abs"),
        "t_bat_mae_c": stat("t_bat_c", "mae"),
        "t_tank_max_abs_c": stat("t_tank_c", "max_abs"),
        "t_cont_max_abs_c": stat("t_cont_c", "max_abs"),
        "u_terminal_max_abs_v": stat("u_terminal_v", "max_abs"),
        "pbess_max_abs_kw": stat("pbess_kw", "max_abs"),
        "pbess_mae_kw": stat("pbess_kw", "mae"),
    }


def comparison_table(selected_ids: list[str]) -> dict[str, Any]:
    all_items = comparison_items()
    if selected_ids:
        by_id = {item["id"]: item for item in all_items}
        selected = [by_id[item_id] for item_id in selected_ids if item_id in by_id]
    else:
        selected = all_items
    fields = comparison_fields(selected)
    return {"items": all_items, "selected": selected, "fields": fields}


def comparison_fields(items: list[dict[str, Any]]) -> list[dict[str, str]]:
    has_opt = any(item.get("type") == "optimization" for item in items)
    has_verify = any(item.get("type") == "verification" for item in items)
    fields = [
        {"key": "objective", "label": "目标函数"},
        {"key": "fuel_kg", "label": "燃油(kg)"},
        {"key": "gap", "label": "Gap"},
    ]
    if has_opt:
        fields.extend(
            [
                {"key": "curt_kwh", "label": "弃风光(kWh)"},
                {"key": "heat_kwh", "label": "电加热(kWh)"},
                {"key": "time_s", "label": "求解时间(s)"},
                {"key": "variables_total", "label": "变量数"},
                {"key": "binary_variables", "label": "二进制变量"},
                {"key": "constraints_total", "label": "约束数"},
                {"key": "pbess_physical_max_kw", "label": "P_BESS后验偏差(kW)"},
                {"key": "model_balance_max_kw", "label": "功率平衡残差(kW)"},
                {"key": "soc_min", "label": "SOC最小"},
                {"key": "soc_max", "label": "SOC最大"},
                {"key": "tbat_min_c", "label": "电芯最低温(℃)"},
                {"key": "tbat_max_c", "label": "电芯最高温(℃)"},
                {"key": "charge_current_limit_violation_max_a", "label": "充电限值越限(A)"},
                {"key": "discharge_current_limit_violation_max_a", "label": "放电限值越限(A)"},
            ]
        )
    if has_verify:
        fields.extend(
            [
                {"key": "diesel_max_kw", "label": "柴油最大出力(kW)"},
                {"key": "renewable_use_kwh", "label": "新能源消纳(kWh)"},
                {"key": "renewable_curt_kwh", "label": "新能源弃电(kWh)"},
                {"key": "unserved_kwh", "label": "未供电量(kWh)"},
                {"key": "soc_violation_max", "label": "SOC越限"},
                {"key": "charge_current_violation_max_a", "label": "充电电流越限(A)"},
                {"key": "discharge_current_violation_max_a", "label": "放电电流越限(A)"},
                {"key": "t_bat_violation_max_c", "label": "电芯温度越限(℃)"},
                {"key": "t_tank_violation_max_c", "label": "液冷罐温度越限(℃)"},
                {"key": "t_cont_violation_max_c", "label": "舱体温度越限(℃)"},
                {"key": "power_balance_max_kw", "label": "功率平衡残差(kW)"},
                {"key": "soc_max_abs", "label": "SOC最大偏差"},
                {"key": "soc_mae", "label": "SOC MAE"},
                {"key": "t_bat_max_abs_c", "label": "电芯温度最大偏差(℃)"},
                {"key": "t_tank_max_abs_c", "label": "液冷罐温度最大偏差(℃)"},
                {"key": "t_cont_max_abs_c", "label": "舱体温度最大偏差(℃)"},
                {"key": "u_terminal_max_abs_v", "label": "端口电压最大偏差(V)"},
                {"key": "pbess_max_abs_kw", "label": "BESS功率最大偏差(kW)"},
            ]
        )
    return fields


class EstoreHandler(SimpleHTTPRequestHandler):
    server_version = "EStoreOptWeb/0.1"
    public_pages = {"/login.html", "/register.html"}
    public_api = {"/api/auth/me", "/api/auth/login", "/api/auth/register", "/api/auth/logout", "/api/health"}
    public_prefixes = ("/assets/",)

    def end_headers(self) -> None:
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def translate_path(self, path: str) -> str:
        parsed = urlparse(path)
        clean = unquote(parsed.path.lstrip("/"))
        if not clean:
            clean = "index.html"
        return str((WEB_ROOT / clean).resolve())

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        try:
            if not self.ensure_authenticated(parsed.path):
                return
            if parsed.path.startswith("/api/"):
                self.handle_api_get(parsed.path, parsed.query)
            else:
                super().do_GET()
        except Exception as exc:
            json_response(self, {"ok": False, "message": str(exc)}, 500)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            if not self.ensure_authenticated(parsed.path):
                return
            if parsed.path.startswith("/api/"):
                self.handle_api_post(parsed.path, self.read_body())
            else:
                self.send_error(HTTPStatus.NOT_FOUND)
        except Exception as exc:
            json_response(self, {"ok": False, "message": str(exc), "traceback": traceback.format_exc()}, 500)

    def do_PUT(self) -> None:
        parsed = urlparse(self.path)
        try:
            if not self.ensure_authenticated(parsed.path):
                return
            if parsed.path.startswith("/api/"):
                self.handle_api_put(parsed.path, self.read_body())
            else:
                self.send_error(HTTPStatus.NOT_FOUND)
        except Exception as exc:
            json_response(self, {"ok": False, "message": str(exc)}, 500)

    def do_DELETE(self) -> None:
        parsed = urlparse(self.path)
        try:
            if not self.ensure_authenticated(parsed.path):
                return
            if parsed.path.startswith("/api/"):
                self.handle_api_delete(parsed.path, parsed.query)
            else:
                self.send_error(HTTPStatus.NOT_FOUND)
        except Exception as exc:
            json_response(self, {"ok": False, "message": str(exc)}, 500)

    def read_body(self) -> bytes:
        length = int(self.headers.get("Content-Length") or 0)
        return self.rfile.read(length) if length else b""

    def read_json_body(self, body: bytes) -> dict[str, Any]:
        if not body:
            return {}
        return json.loads(body.decode("utf-8"))

    def is_public_path(self, path: str) -> bool:
        if path in self.public_api or path in self.public_pages:
            return True
        if path == "/" or path == "":
            return False
        return any(path.startswith(prefix) for prefix in self.public_prefixes)

    def session_token(self) -> str:
        raw = self.headers.get("Cookie", "")
        if not raw:
            return ""
        cookie = SimpleCookie()
        try:
            cookie.load(raw)
        except Exception:
            return ""
        morsel = cookie.get(SESSION_COOKIE_NAME)
        return morsel.value if morsel else ""

    def current_user(self) -> dict[str, Any] | None:
        if LOCAL_AUTH_BYPASS_ENABLED:
            return {"id": 0, "username": "local-admin", "role": "admin", "created_at": now_text()}
        return USER_STORE.user_for_session(self.session_token())

    def ensure_authenticated(self, path: str) -> bool:
        if self.is_public_path(path):
            return True
        if self.current_user():
            return True
        if path.startswith("/api/"):
            json_response(self, {"ok": False, "message": "请先登录"}, 401)
            return False
        next_path = path if path and path != "/" else "/index.html"
        target = f"/login.html?next={next_path}"
        self.send_response(HTTPStatus.FOUND)
        self.send_header("Location", target)
        self.send_header("Content-Length", "0")
        self.end_headers()
        return False

    def set_session_cookie(self, token: str) -> None:
        self.send_header(
            "Set-Cookie",
            f"{SESSION_COOKIE_NAME}={token}; Path=/; Max-Age={SESSION_MAX_AGE_SECONDS}; SameSite=Lax; HttpOnly",
        )

    def clear_session_cookie(self) -> None:
        self.send_header("Set-Cookie", f"{SESSION_COOKIE_NAME}=; Path=/; Max-Age=0; SameSite=Lax; HttpOnly")

    def send_auth_json(self, payload: dict[str, Any], status: int = 200, token: str | None = None, clear: bool = False) -> None:
        data = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        if token:
            self.set_session_cookie(token)
        if clear:
            self.clear_session_cookie()
        self.end_headers()
        self.wfile.write(data)

    def handle_api_get(self, path: str, query: str) -> None:
        params = parse_qs(query)
        if path == "/api/auth/me":
            user = self.current_user()
            json_response(self, {"ok": True, "authenticated": bool(user), "user": user})
        elif path == "/api/health":
            json_response(self, {"ok": True, "time": now_text(), "solver_script": str(SOLVER_SCRIPT), "default_params": str(DEFAULT_PARAMS)})
        elif path == "/api/schemes":
            json_response(self, {"schemes": list_schemes()})
        elif path == "/api/scheme":
            name = params.get("name", [""])[0]
            json_response(self, workbook_overview(name))
        elif path == "/api/sheet":
            name = params.get("scheme", [""])[0]
            sheet = params.get("sheet", [""])[0]
            page = positive_int(params.get("page", ["1"])[0], 1, 1)
            page_size = positive_int(params.get("page_size", ["200"])[0], 200, 1, 2000)
            json_response(self, read_sheet(name, sheet, page, page_size))
        elif path == "/api/compute-config":
            name = params.get("scheme", [""])[0]
            json_response(self, {"scheme": name, "config": read_scheme_compute_config(name)})
        elif path == "/api/dispatch-schedule":
            name = params.get("scheme", [""])[0]
            json_response(self, dispatch_schedule_payload(name))
        elif path == "/api/tasks":
            json_response(self, TASKS.snapshot())
        elif path == "/api/task-board":
            json_response(self, task_board_snapshot())
        elif path == "/api/task":
            json_response(self, TASKS.task_detail(params.get("id", [""])[0]))
        elif path == "/api/verification":
            json_response(self, VERIFY_TASKS.snapshot())
        elif path == "/api/verification/item":
            json_response(self, VERIFY_TASKS.detail(params.get("id", [""])[0]))
        elif path == "/api/comparison/items":
            json_response(self, {"items": comparison_items()})
        elif path == "/api/comparison/data":
            ids = params.get("id", [])
            json_response(self, comparison_table(ids))
        else:
            json_response(self, {"ok": False, "message": "接口不存在"}, 404)

    def handle_api_post(self, path: str, body: bytes) -> None:
        if path == "/api/auth/register":
            data = self.read_json_body(body)
            user = USER_STORE.register(data.get("username", ""), data.get("password", ""))
            token = USER_STORE.create_session(int(user["id"]))
            self.send_auth_json({"ok": True, "user": user}, token=token)
        elif path == "/api/auth/login":
            data = self.read_json_body(body)
            user = USER_STORE.authenticate(data.get("username", ""), data.get("password", ""))
            if not user:
                json_response(self, {"ok": False, "message": "用户名或密码不正确"}, 401)
                return
            token = USER_STORE.create_session(int(user["id"]))
            self.send_auth_json({"ok": True, "user": user}, token=token)
        elif path == "/api/auth/logout":
            USER_STORE.delete_session(self.session_token())
            self.send_auth_json({"ok": True}, clear=True)
        elif path == "/api/schemes":
            data = self.read_json_body(body)
            scheme = create_scheme(data.get("name", ""), data.get("description", ""))
            json_response(self, {"ok": True, "scheme": scheme, "schemes": list_schemes()})
        elif path == "/api/schemes/copy":
            data = self.read_json_body(body)
            scheme = copy_scheme(data.get("source", ""), data.get("name", ""), data.get("description", ""))
            json_response(self, {"ok": True, "scheme": scheme, "schemes": list_schemes()})
        elif path == "/api/schemes/update":
            data = self.read_json_body(body)
            scheme = update_scheme_meta(data.get("source", ""), data.get("name", ""), data.get("description", None))
            json_response(self, {"ok": True, "scheme": scheme, "schemes": list_schemes()})
        elif path == "/api/schemes/upload":
            self.handle_upload(body)
        elif path == "/api/tasks":
            data = self.read_json_body(body)
            scheme = data.get("scheme", "")
            cfg = read_scheme_compute_config(scheme)
            cfg.update(data.get("config") or {})
            task = TASKS.enqueue(scheme, normalize_compute_config(cfg), data.get("source") or "single")
            json_response(self, {"ok": True, "task": TASKS.task_detail(task.id), **TASKS.snapshot()})
        elif path == "/api/tasks/batch":
            data = self.read_json_body(body)
            TASKS.set_max_parallel(int(data.get("max_parallel") or 1))
            tasks = []
            for scheme in data.get("schemes") or []:
                cfg = read_scheme_compute_config(str(scheme))
                cfg.update(data.get("config") or {})
                tasks.append(TASKS.enqueue(str(scheme), normalize_compute_config(cfg), "batch").id)
            json_response(self, {"ok": True, "task_ids": tasks, **TASKS.snapshot()})
        elif path == "/api/tasks/cancel":
            data = self.read_json_body(body)
            task = TASKS.cancel(data.get("id", ""))
            json_response(self, {"ok": True, "task": TASKS.task_detail(task.id), **TASKS.snapshot()})
        elif path == "/api/tasks/control":
            data = self.read_json_body(body)
            payload = control_task(data.get("action", ""), data.get("task_type", ""), data.get("scheme", ""), data.get("task_id", ""))
            json_response(self, payload)
        elif path == "/api/verification/start":
            data = self.read_json_body(body)
            scheme = data.get("scheme", "")
            cfg = read_scheme_compute_config(scheme)
            cfg.update(data.get("config") or {})
            task = VERIFY_TASKS.start(scheme, normalize_compute_config(cfg), data.get("source_task_id", ""), data.get("source") or "single")
            json_response(self, {"ok": True, "verification": VERIFY_TASKS.detail(task.id), **VERIFY_TASKS.snapshot()})
        elif path == "/api/verification/cancel":
            data = self.read_json_body(body)
            task = VERIFY_TASKS.cancel(data.get("id", ""))
            json_response(self, {"ok": True, "verification": VERIFY_TASKS.detail(task.id), **VERIFY_TASKS.snapshot()})
        elif path == "/api/dispatch-schedule/init":
            data = self.read_json_body(body)
            payload = initialize_scheme_dispatch_schedule(data.get("scheme", ""), data.get("config") or None)
            json_response(self, {"ok": True, **payload, "schemes": list_schemes()})
        elif path == "/api/dispatch-schedule/from-optimization":
            data = self.read_json_body(body)
            scheme = create_dispatch_scheme_from_optimization(data.get("task_id", ""), data.get("name", ""), data.get("description", ""))
            json_response(self, {"ok": True, "scheme": scheme, "dispatch": dispatch_schedule_payload(scheme["name"]), "schemes": list_schemes()})
        elif path == "/api/compute-config":
            data = self.read_json_body(body)
            name = data.get("scheme", "")
            cfg = write_scheme_compute_config(name, data.get("config") or {})
            json_response(self, {"ok": True, "scheme": name, "config": cfg, "schemes": list_schemes()})
        else:
            json_response(self, {"ok": False, "message": "接口不存在"}, 404)

    def handle_api_put(self, path: str, body: bytes) -> None:
        if path == "/api/sheet":
            data = self.read_json_body(body)
            page = positive_int(data.get("page"), 1, 1)
            page_size = positive_int(data.get("page_size"), 200, 1, 2000)
            if data.get("operation"):
                payload = mutate_sheet_rows(
                    data.get("scheme", ""),
                    data.get("sheet", ""),
                    data.get("operation", ""),
                    row=data.get("row"),
                    after_row=data.get("after_row"),
                    values=data.get("values") or [],
                    page=page,
                    page_size=page_size,
                )
            elif "updates" in data:
                payload = write_sheet_cells(data.get("scheme", ""), data.get("sheet", ""), data.get("updates") or [], page, page_size)
            else:
                payload = write_sheet(data.get("scheme", ""), data.get("sheet", ""), data.get("rows") or [], page, page_size)
            json_response(self, {"ok": True, **payload})
        elif path == "/api/dispatch-schedule":
            data = self.read_json_body(body)
            payload = write_scheme_dispatch_schedule(data.get("scheme", ""), data.get("rows") or [])
            json_response(self, {"ok": True, **payload, "schemes": list_schemes()})
        else:
            json_response(self, {"ok": False, "message": "接口不存在"}, 404)

    def handle_api_delete(self, path: str, query: str) -> None:
        if path == "/api/schemes":
            name = parse_qs(query).get("name", [""])[0]
            target = scheme_dir(name)
            if target.exists():
                shutil.rmtree(target)
            json_response(self, {"ok": True, "schemes": list_schemes()})
        else:
            json_response(self, {"ok": False, "message": "接口不存在"}, 404)

    def handle_upload(self, body: bytes) -> None:
        form = cgi.FieldStorage(
            fp=io.BytesIO(body),
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type", "")},
        )
        name = safe_name(form.getfirst("name", ""))
        description = form.getfirst("description", "")
        file_item = form["file"] if "file" in form else None
        if not name or file_item is None or not getattr(file_item, "file", None):
            raise ValueError("需要提供方案名称和 Excel 文件")
        target = scheme_dir(name)
        if target.exists():
            raise ValueError(f"方案已存在：{name}")
        target.mkdir(parents=True, exist_ok=True)
        with (target / PARAM_FILE_NAME).open("wb") as f:
            shutil.copyfileobj(file_item.file, f)
        write_json(target / COMPUTE_CONFIG_FILE_NAME, normalize_compute_config())
        meta = {"name": name, "description": description, "created_at": now_text(), "updated_at": now_text(), "source": "upload"}
        write_json(target / META_FILE_NAME, meta)
        json_response(self, {"ok": True, "scheme": scheme_summary(name), "schemes": list_schemes()})


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="EStore optimization local web system")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8877)
    return parser.parse_args()


def main() -> int:
    ensure_default_scheme()
    args = parse_args()
    server = ThreadingHTTPServer((args.host, args.port), EstoreHandler)
    print(timestamp_log_line(f"EStore Opt Web running at http://{args.host}:{args.port}/"), flush=True)
    server.serve_forever()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
