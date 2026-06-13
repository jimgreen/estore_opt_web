import os
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("ESTORE_OPT_WORKER", "1")

import server  # noqa: E402
import solve  # noqa: E402


class InitialStateConfigTests(unittest.TestCase):
    def test_compute_config_normalizes_initial_state_fields(self):
        cfg = server.normalize_compute_config(
            {
                "initial_soc": "0.62",
                "initial_t_bat_c": "8.5",
                "initial_t_tank_c": "7.5",
                "initial_t_cont_c": "6.5",
            }
        )

        self.assertIn("initial_soc", cfg)
        self.assertEqual(cfg["initial_soc"], 0.62)
        self.assertEqual(cfg["initial_t_bat_c"], 8.5)
        self.assertEqual(cfg["initial_t_tank_c"], 7.5)
        self.assertEqual(cfg["initial_t_cont_c"], 6.5)
        self.assertEqual(server.COMPUTE_CONFIG_LABELS["initial_soc"], "电池初始SOC")

    def test_existing_compute_config_sheet_is_migrated_with_initial_state_rows(self):
        original_root = server.SCHEME_ROOT
        try:
            with tempfile.TemporaryDirectory() as tmp:
                server.SCHEME_ROOT = Path(tmp)
                scheme_dir = server.scheme_dir("demo")
                scheme_dir.mkdir(parents=True)
                workbook_path = scheme_dir / server.PARAM_FILE_NAME
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = server.COMPUTE_CONFIG_SHEET_NAME
                sheet.append(["参数", "值", "说明"])
                sheet.append(["solver", "auto", "求解器"])
                workbook.save(workbook_path)
                workbook.close()

                cfg = server.ensure_scheme_compute_config("demo")

                self.assertEqual(cfg["initial_soc"], 0.5)
                migrated = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
                try:
                    keys = [row[0] for row in migrated[server.COMPUTE_CONFIG_SHEET_NAME].iter_rows(min_row=2, values_only=True)]
                finally:
                    migrated.close()
                self.assertIn("initial_soc", keys)
                self.assertIn("initial_t_bat_c", keys)
                self.assertIn("initial_t_tank_c", keys)
                self.assertIn("initial_t_cont_c", keys)
        finally:
            server.SCHEME_ROOT = original_root

    def test_solver_applies_initial_state_config_to_loaded_params(self):
        self.assertTrue(hasattr(solve, "apply_initial_state_config"))
        p = SimpleNamespace(
            raw={"initial_state": {}},
            SOC_min=0.1,
            SOC_max=0.9,
            T_bat_min=-20.0,
            T_bat_max=45.0,
            T_tank_min=-20.0,
            T_tank_max=45.0,
            T_cont_min=-25.0,
            T_cont_max=45.0,
        )
        args = SimpleNamespace(initial_soc=0.62, initial_t_bat_c=8.5, initial_t_tank_c=7.5, initial_t_cont_c=6.5)

        solve.apply_initial_state_config(p, args)

        self.assertEqual(p.SOC_init, 0.62)
        self.assertEqual(p.T_bat_init, 8.5)
        self.assertEqual(p.T_tank_init, 7.5)
        self.assertEqual(p.T_cont_init, 6.5)
        self.assertEqual(
            p.raw["initial_state"],
            {"soc": 0.62, "t_bat_c": 8.5, "t_tank_c": 7.5, "t_cont_c": 6.5},
        )

    def test_solver_reads_bom_prefixed_compute_config(self):
        with tempfile.TemporaryDirectory() as tmp:
            config_path = Path(tmp) / "compute_config.json"
            config_path.write_text('{"initial_soc": 0.62, "dt_minutes": 30}', encoding="utf-8-sig")

            args = solve.parse_args(["--config-file", str(config_path)])

        self.assertEqual(args.initial_soc, 0.62)
        self.assertEqual(args.dt_minutes, 30)

    def test_frontend_exposes_initial_state_fields(self):
        script = (ROOT / "assets" / "schemes.js").read_text(encoding="utf-8-sig")

        for key in ("initial_soc", "initial_t_bat_c", "initial_t_tank_c", "initial_t_cont_c"):
            self.assertIn(f'key: "{key}"', script)

    def test_optimization_terminal_targets_return_to_initial_state(self):
        source = (ROOT / "solve.py").read_text(encoding="utf-8-sig")

        self.assertEqual(source.count('"soc_terminal_target"'), 3)
        self.assertIn('name="soc_terminal_target"', source)
        self.assertEqual(source.count('"ttank_terminal_target"'), 3)
        self.assertIn('name="ttank_terminal_target"', source)
        self.assertEqual(source.count('"tcont_terminal_target"'), 3)
        self.assertIn('name="tcont_terminal_target"', source)
        self.assertNotIn("SOC_end - 0.5", source)
        self.assertNotIn("0.5 - SOC_end", source)


if __name__ == "__main__":
    unittest.main()
