import json
import sys
import unittest
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))


PARAMS_PATH = ROOT / "estore_schemes" / "2" / "params.xlsx"
CONFIG_PATH = ROOT / "estore_schemes" / "2" / "compute_config.json"
SYSTEM_SHEET = "\u7cfb\u7edf\u5b9a\u4e49"
CURVE_SHEET = "\u8fd0\u884c\u66f2\u7ebf"


def rows_by_hour():
    wb = openpyxl.load_workbook(PARAMS_PATH, data_only=True, read_only=True)
    try:
        ws = wb[CURVE_SHEET]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in row[:5]):
                continue
            hour = float(row[0])
            rows.append(
                {
                    "hour": hour,
                    "wind": float(row[1]),
                    "ambient": float(row[2]),
                    "solar": float(row[3]),
                    "load": float(row[4]),
                }
            )
        return rows
    finally:
        wb.close()


class OperationCurveTests(unittest.TestCase):
    def test_operation_curve_is_three_days_with_requested_wind_solar_and_load(self):
        wb = openpyxl.load_workbook(PARAMS_PATH, data_only=True, read_only=True)
        try:
            self.assertEqual(float(wb[SYSTEM_SHEET].cell(2, 5).value), 4320.0)
        finally:
            wb.close()

        config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        self.assertEqual(float(config["hours"]), 72.0)

        rows = rows_by_hour()
        self.assertEqual(len(rows), 289)
        self.assertEqual(rows[0]["hour"], 0.0)
        self.assertEqual(rows[-1]["hour"], 72.0)
        self.assertTrue(all(abs(row["hour"] - idx * 0.25) < 1e-9 for idx, row in enumerate(rows)))

        by_hour = {round(row["hour"], 2): row for row in rows}
        samples = [idx * 0.25 for idx in range(96)]
        for sample in samples:
            self.assertAlmostEqual(by_hour[round(sample, 2)]["load"], by_hour[round(sample + 24.0, 2)]["load"], places=9)
            self.assertAlmostEqual(by_hour[round(sample, 2)]["load"], by_hour[round(sample + 48.0, 2)]["load"], places=9)

        day2 = [row for row in rows if 24.0 <= row["hour"] < 48.0]
        self.assertTrue(day2)
        self.assertTrue(all(row["wind"] == 0.0 for row in day2))
        self.assertTrue(all(row["solar"] == 0.0 for row in day2))

        for start in (0.0, 48.0):
            day_rows = [row for row in rows if start <= row["hour"] < start + 24.0]
            self.assertEqual(min(row["wind"] for row in day_rows), 3.0)
            self.assertEqual(max(row["wind"] for row in day_rows), 20.0)
            for row in day_rows:
                phase = (row["hour"] - start) % 4.0
                expected = 20.0 if phase < 2.0 else 3.0
                self.assertEqual(row["wind"], expected)


if __name__ == "__main__":
    unittest.main()
