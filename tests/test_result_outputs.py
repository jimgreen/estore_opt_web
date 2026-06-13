import os
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

import numpy as np
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("ESTORE_OPT_WORKER", "1")

import solve  # noqa: E402
import server  # noqa: E402


CELL_STATE_SHEET = "\u7535\u82af\u72b6\u6001"


def sample_params() -> SimpleNamespace:
    return SimpleNamespace(
        N_s=2,
        N_p=3,
        ocv_soc=np.asarray([0.0, 1.0], dtype=float),
        ocv_1d=np.asarray([600.0, 720.0], dtype=float),
        ocv_cell_1d=np.asarray([3.0, 3.6], dtype=float),
        mu_pcs=0.02,
        P_pump_in=100.0,
        P_pump_out=120.0,
        P_heat_liquid=500.0,
        P_heat_cont=600.0,
        m_in=0.5,
        m_out=0.4,
        c_liq=4200.0,
    )


def sample_result() -> dict:
    n = 3
    return {
        "success": True,
        "status": "OPTIMAL",
        "objective": 12.5,
        "gap": 0.01,
        "best_bound": 12.4,
        "time_s": 3.2,
        "node_count": 7,
        "fuel_kg": 4.5,
        "curt_kwh": 1.2,
        "heat_kwh": 0.8,
        "solver_used": "unit-test",
        "solver_backend": "mock",
        "hours": np.asarray([0.0, 0.25, 0.5], dtype=float),
        "SOC": np.asarray([0.2, 0.3, 0.4], dtype=float),
        "T_bat": np.asarray([10.0, 11.0, 12.0], dtype=float),
        "T_tank": np.asarray([5.0, 5.5, 6.0], dtype=float),
        "T_cont": np.asarray([2.0, 2.5, 3.0], dtype=float),
        "T_amb": np.asarray([-8.0, -7.5, -7.0], dtype=float),
        "I_bat": np.asarray([0.0, 10.0, -5.0], dtype=float),
        "R0": np.asarray([0.08, 0.081, 0.082], dtype=float),
        "P_dc_pack": np.asarray([0.0, 5900.0, -3100.0], dtype=float),
        "P_dc_abs": np.asarray([0.0, 5900.0, 3100.0], dtype=float),
        "Q_gen_pack": np.asarray([0.0, 8.1, 2.05], dtype=float),
        "P_BESS": np.asarray([0.0, 5682.0, -3162.0], dtype=float),
        "Q_bt": np.asarray([0.0, 20.0, 25.0], dtype=float),
        "Q_tamb": np.asarray([0.0, 15.0, 18.0], dtype=float),
        "Q_tamb_dump": np.asarray([0.0, 1.0, 2.0], dtype=float),
        "u_pi": np.asarray([0.0, 1.0, 1.0], dtype=float),
        "u_po": np.asarray([0.0, 0.0, 1.0], dtype=float),
        "u_lh": np.asarray([0.0, 1.0, 0.0], dtype=float),
        "u_ch": np.asarray([0.0, 0.0, 1.0], dtype=float),
        "pv_use_kw": np.asarray([0.0, 3.0, 4.0], dtype=float),
        "wt_use_kw": np.asarray([0.0, 2.0, 0.0], dtype=float),
        "P_dg": np.asarray([1000.0, 2000.0, 1500.0], dtype=float),
        "load_kw": np.asarray([1.0, 10.0, 8.0], dtype=float),
        "P_dg_units": np.asarray([[1000.0, 2000.0, 1500.0], [0.0, 0.0, 0.0]], dtype=float),
        "u_dg": np.asarray([[1.0, 1.0, 1.0], [0.0, 0.0, 0.0]], dtype=float),
        "custom_curve": np.asarray([3.0, 4.0, 5.0], dtype=float),
        "model_stats": {"variables_total": 30, "constraints_total": 12, "steps": n, "dt_minutes": 15},
        "checks": {"soc_min": 0.2, "soc_max": 0.4, "model_balance_max_kw": 0.001},
    }


class ResultOutputTests(unittest.TestCase):
    def test_result_data_payload_exports_all_time_series_tables_and_statistics(self):
        payload = solve.build_result_data_payload(sample_params(), sample_result())

        series_by_key = {item["key"]: item for item in payload["series"]}
        self.assertEqual(payload["row_count"], 3)
        self.assertIn("SOC", series_by_key)
        self.assertIn("P_BESS", series_by_key)
        self.assertIn("custom_curve", series_by_key)
        self.assertIn("P_dg_units_1", series_by_key)
        self.assertEqual(series_by_key["P_BESS"]["unit"], "kW")
        self.assertEqual(series_by_key["P_dc_pack"]["unit"], "kW")
        self.assertEqual(series_by_key["P_dg_units_1"]["unit"], "kW")
        self.assertNotIn("W", {item["unit"] for item in payload["series"]})
        self.assertAlmostEqual(series_by_key["P_BESS"]["values"][1], 5.682)
        self.assertAlmostEqual(series_by_key["P_dg_units_1"]["values"][1], 2.0)
        self.assertAlmostEqual(payload["rows"][1]["P_BESS"], 5.682)
        self.assertAlmostEqual(payload["rows"][1]["P_dg_units_1"], 2.0)
        self.assertEqual(payload["rows"][2]["SOC"], 0.4)
        self.assertEqual(payload["statistics"]["objective"], 12.5)
        self.assertEqual(payload["statistics"]["checks"]["soc_min"], 0.2)
        self.assertEqual(payload["statistics"]["model_stats"]["variables_total"], 30)
        self.assertEqual(len(payload["tables"][CELL_STATE_SHEET]["rows"]), 3)

    def test_result_workbook_embeds_statistics_curves_and_payload_without_sidecars(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_dir = Path(tmp)
            workbook_path = output_dir / "optimization_results_perspective_i2r_15min_20260530.xlsx"
            result = sample_result()
            result["state_workbook"] = str(workbook_path)

            solve.write_detailed_results_workbook(sample_params(), result, workbook_path)

            self.assertTrue(workbook_path.exists())
            unexpected = {
                path.suffix
                for path in output_dir.iterdir()
                if path.suffix.lower() in {".png", ".csv", ".json", ".md"}
            }
            self.assertEqual(unexpected, set())
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            try:
                self.assertIn("统计信息", workbook.sheetnames)
                self.assertIn("调度曲线", workbook.sheetnames)
                self.assertIn("曲线元数据", workbook.sheetnames)
                self.assertIn("__result_payload", workbook.sheetnames)
                self.assertEqual(workbook["统计信息"]["A1"].value, "项目")
                self.assertEqual(workbook["调度曲线"]["A1"].value, "step")
                curve_headers = [cell.value for cell in workbook["调度曲线"][1]]
                self.assertIn("custom_curve", curve_headers)
                pbess_col = curve_headers.index("P_BESS") + 1
                self.assertAlmostEqual(workbook["调度曲线"].cell(row=3, column=pbess_col).value, 5.682)
                metadata_rows = {
                    workbook["曲线元数据"].cell(row=row, column=1).value: workbook["曲线元数据"].cell(row=row, column=4).value
                    for row in range(2, workbook["曲线元数据"].max_row + 1)
                }
                self.assertEqual(metadata_rows["P_BESS"], "kW")
                self.assertEqual(metadata_rows["P_dg_units_1"], "kW")
                self.assertNotIn("W", set(metadata_rows.values()))
            finally:
                workbook.close()

    def test_server_reads_result_data_and_file_links_from_result_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            run_dir = Path(tmp)
            workbook_path = run_dir / "optimization_results_perspective_i2r_15min_20260530.xlsx"
            result = sample_result()
            result["state_workbook"] = str(workbook_path)
            solve.write_detailed_results_workbook(sample_params(), result, workbook_path)

            payload = server.optimization_result_payload(run_dir, {"state_workbook": str(workbook_path)}, {"variables_total": 5})

            series_keys = {item["key"] for item in payload["result_data"]["series"]}
            self.assertIn("SOC", series_keys)
            self.assertIn("custom_curve", series_keys)
            self.assertEqual(payload["statistics"]["objective"], 12.5)
            labels = {item["label"] for item in payload["result_files"]}
            self.assertEqual(labels, {"结果工作簿"})

    def test_output_paths_only_include_result_workbook(self):
        paths = solve.get_output_paths("dayahead_24h", Path("runs/demo"), 60.0)

        self.assertEqual(set(paths), {"state_workbook"})
        self.assertEqual(paths["state_workbook"].suffix, ".xlsx")


if __name__ == "__main__":
    unittest.main()
