import sys
import unittest
from pathlib import Path

import numpy as np
import openpyxl


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import solve  # noqa: E402


CURRENT_LIMIT_SHEET = "\u7535\u82af\u7535\u6d41\u9650\u503c"
PARAMS_PATH = ROOT / "estore_schemes" / "2" / "params.xlsx"


class CellCurrentLimitTests(unittest.TestCase):
    def test_excel_current_limit_table_drives_pack_breakpoints(self):
        wb = openpyxl.load_workbook(PARAMS_PATH, data_only=True)
        self.assertIn(CURRENT_LIMIT_SHEET, wb.sheetnames)

        ws = wb[CURRENT_LIMIT_SHEET]
        headers = [ws.cell(1, col).value for col in range(1, 4)]
        self.assertEqual(headers, ["\u6e29\u5ea6(\u2103)", "\u6700\u5927\u5145\u7535\u7535\u6d41(A)", "\u6700\u5927\u653e\u7535\u7535\u6d41(A)"])

        raw = solve.load_raw_params_from_excel(PARAMS_PATH)
        limits = raw["cell_current_limits"]
        self.assertGreaterEqual(len(limits["temperatures_c"]), 2)
        self.assertEqual(len(limits["temperatures_c"]), len(limits["charge_max_a"]))
        self.assertEqual(len(limits["temperatures_c"]), len(limits["discharge_max_a"]))

        p = solve.load_params(PARAMS_PATH)
        n_p = raw["battery_cabinet"]["n_p"]
        np.testing.assert_allclose(p.current_limit_temps, limits["temperatures_c"])
        np.testing.assert_allclose(p.charge_current_limit_pack, np.asarray(limits["charge_max_a"], dtype=float) * n_p)
        np.testing.assert_allclose(p.discharge_current_limit_pack, np.asarray(limits["discharge_max_a"], dtype=float) * n_p)
        self.assertLessEqual(p.I_charge_max, p.pcs_i_max)
        self.assertLessEqual(p.I_discharge_max, p.pcs_i_max)

        bp = solve.make_breakpoints(p, "test_1h", current_segments=4, dt_minutes=15.0)
        expected_charge = np.interp(
            bp.temp,
            p.current_limit_temps,
            p.charge_current_limit_pack,
            left=p.charge_current_limit_pack[0],
            right=p.charge_current_limit_pack[-1],
        )
        expected_discharge = np.interp(
            bp.temp,
            p.current_limit_temps,
            p.discharge_current_limit_pack,
            left=p.discharge_current_limit_pack[0],
            right=p.discharge_current_limit_pack[-1],
        )
        np.testing.assert_allclose(bp.charge_current_limit, expected_charge)
        np.testing.assert_allclose(bp.discharge_current_limit, expected_discharge)


if __name__ == "__main__":
    unittest.main()
